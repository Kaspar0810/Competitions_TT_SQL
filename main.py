
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.lib.styles import ParagraphStyle as PS
from reportlab.lib import colors
from reportlab.lib.colors import *
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, TableStyle, SimpleDocTemplate, Frame
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from reportlab.pdfgen.canvas import Canvas
from PyPDF2 import PdfMerger
from main_window import Ui_MainWindow
from start_form import Ui_Form
from datetime import *
from PyQt5 import *
from PyQt5.QtCore import QAbstractTableModel
from PyQt5.QtGui import QIcon, QBrush, QColor, QFont, QPalette
from PyQt5.QtWidgets import QPushButton, QRadioButton, QHeaderView, QComboBox, QListWidgetItem, QItemDelegate, QStyledItemDelegate
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QMenu, QInputDialog, QTableWidgetItem, QLineEdit
from PyQt5.QtWidgets import QAbstractItemView, QFileDialog, QProgressBar, QAction, QDesktopWidget, QTableView, QColorDialog
from PyQt5 import QtGui, QtWidgets, QtCore
from models import *
from collections import Counter
from itertools import *
import os
import openpyxl as op
import pandas as pd
import numpy as np
import contextlib
import sys
import sqlite3
import pathlib
from pathlib import Path
import random
import time
# import collections
# from playhouse.migrate import *

if not os.path.isdir("table_pdf"):  # создает папку 
    os.mkdir("table_pdf")
if not os.path.isdir("competition_pdf"):  # создает папку 
    os.mkdir("competition_pdf")


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.

if __name__ == '__main__':
    print_hi('PyCharm and Alex')

# from playhouse.sqlite_ext import SqliteExtDatabase, backup_to_file, backup

registerFontFamily('DejaVuSerif', normal='DejaVuSerif',
                   bold='DejaVuSerif-Bold', italic='DejaVuSerif-Italic')
enc = 'UTF-8'

TTFSearchPath = (
    'c:/winnt/fonts',
    'c:/windows/fonts',
    '%(REPORTLAB_DIR)s/fonts',  # special
    '%(REPORTLAB_DIR)s/../fonts',  # special
    '%(REPORTLAB_DIR)s/../../fonts',  # special
    '%(CWD)s/fonts',  # special
    '~/fonts',
    '~/.fonts',
    '%(XDG_DATA_HOME)s/fonts',
    '~/.local/share/fonts',
    # mac os X - from
    '~/Library/Fonts',
    '/Library/Fonts',
    '/System/Library/Fonts',
)
pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf', enc))
pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf', enc))
pdfmetrics.registerFont(TTFont('DejaVuSerif', 'DejaVuSerif.ttf', enc))
pdfmetrics.registerFont(TTFont('DejaVuSerif-Bold', 'DejaVuSerif-Bold.ttf', enc))
pdfmetrics.registerFont(TTFont('DejaVuSerif-Italic', 'DejaVuSerif-Italic.ttf', enc))

class MyTableModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data
        self.horizontalHeaderLabels = []
 
    def setHorizontalHeaderLabels(self, horizontalHeaderLabels):
        self.horizontalHeaderLabels = horizontalHeaderLabels
 
    def headerData(self, section: int, orientation: QtCore.Qt.Orientation, role: QtCore.Qt.ItemDataRole.DisplayRole):
        if (orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole and len(self.horizontalHeaderLabels) == self.columnCount(None)):
            return self.horizontalHeaderLabels[section]
        return super().headerData(section, orientation, role)
    
    def rowCount(self, parent):
        return len(self._data)
    
    def columnCount(self, parent):
        if len(self._data) > 0:
            return len(self._data[0])
        else:
            return 0

    def data(self, index, role):
        if role == QtCore.Qt.ItemDataRole.DisplayRole:
            return str(self._data[index.row()][index.column()])
        return 
        
    def flags(self, index):
        if (index.column() == 1) or (index.column() == 2):
            return QtCore.Qt.ItemIsEditable | QtCore.Qt.ItemIsEnabled
        else:
            return QtCore.Qt.ItemIsEnabled
        
    def setData(self, index, text):
        self.items[index.row()] = text
# ++++++++++++++++++++++++++++++++++++++++++++++++++++
# class MyComboDelegate(QItemDelegate):

#     def __init__(self, parent=None):
#         super().__init__(parent)

#     def createEditor(self, parent, option, index):
#         # if index.column() == 2:
#         editor = QComboBox(parent)
#         post_list = ["-выберите должность-", "Зам. Главного судьи", "Зам. Главного секретаря", "Ведущий судья"]
#         editor.addItems(post_list)
#         self.connect(editor, QtCore.SIGNAL("currentIndexChanged(int)"), 
#                  self, QtCore.SLOT("currentIndexChanged()"))
#         return editor

#     def setEditorData(self, editor, index):
#         value = index.model().data(index, QtCore.Qt.EditRole)
#         if value:
#             editor.setCurrentIndex(int(value))

#     @QtCore.pyqtSlot()
#     def currentIndexChanged(self):
#         self.commitData.emit(self.sender())
    
#     # def setModelData(self, editor, model, index):
#     #     model.setData(index, editor.currentIndex())  

#     def setModelData(self, editor, model, index):
#         comboIndex = editor.currentIndex()
#         text = self.comboItems[comboIndex]
#         model.setData(index, text)

#     def setData(self, index, value, role=QtCore.Qt.ItemDataRole.DisplayRole):
#         print ("setData", index.row(), index.column(), value)
#         # todo: remember the data
#  ===
# class ComboDelegate(QItemDelegate):
#     """
#     A delegate to add QComboBox in every cell of the given column
#     """

#     def __init__(self, options, parent=None):
#         super(ComboDelegate, self).__init__(parent)
#         self.parent = parent
#         self.options = options

#     def createEditor(self, parent, option, index):
#         editor = QComboBox(parent)
#         # post_list = ["-выберите должность-", "Зам. Главного судьи", "Зам. Главного секретаря", "Ведущий судья"]
#         # editor.addItems(post_list)
#         # self.connect(editor, QtCore.SIGNAL("currentIndexChanged(int)"), 
#                 #  self, QtCore.SLOT("currentIndexChanged()"))
#         editor.currentTextChanged.connect(lambda value: self.currentIndexChanged(index, value))
#         return editor
    
#     @QtCore.pyqtSlot()
#     def currentIndexChanged(self):
#         self.commitData.emit(self.sender())

#     def setModelData(self, editor, model, index):
#         value = editor.currentText()
#         # text = self.comboItems[comboIndex]
#         model.setData(index, value, QtCore.Qt.EditRole)   
#     # def setEditorData(self, editor, index):
#     #     value = index.data()
#     #     if value:
#     #         maxval = len(value)
#     #         editor.setCurrentIndex(maxval - 1)
class MyColorDelegate(QItemDelegate):

    def __init__(self, parent=None):
        super().__init__(parent)

    def paint(self, painter, option, index):
        if index.isValid():
            color = index.data(QtCore.Qt.ItemDataRole.DisplayRole)
        # if color.isValid():
        #     painter.fillRect(option.rect, color)

    def createEditor(self, parent, option, index):
        color_dialog = QColorDialog(parent)
        color_dialog.setOption(QColorDialog.ShowAlphaChannel, True)
        return color_dialog

    def setEditorData(self, editor, index):
        value = index.data(QtCore.Qt.ItemDataRole.DisplayRole)
        if value.isValid():
            editor.setCurrentColor(value)

    def setModelData(self, editor, model, index):
        value = editor.currentColor()
        model.setData(index, value, QtCore.Qt.ItemDataRole.DisplayRole)


# =====================================  
class ComboBoxDelegate(QItemDelegate):
    def __init__(self, options, parent=None):
        super().__init__(parent)
        self._options = options
    
    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        editor.addItems(self._options)
        return editor
    
    def setEditorData(self, editor, index):
        value = index.data(QtCore.Qt.EditRole)
        editor.setCurrentIndex(editor.findText(str(value)))
    
    def setModelData(self, editor, model, index):
        value = editor.currentText()
        model.setData(index, value, QtCore.Qt.EditRole)
    
class LineDelegate(QItemDelegate):
    def __init__(self, options, parent=None):
        super().__init__(parent)

    
    def createEditor(self, parent, index):
        lineEd = QLineEdit(parent)
        return lineEd
    
    def setEditorData(self, lineEd, index):
        value = index.data(QtCore.Qt.EditRole)
        # editor.setCurrentIndex(editor.findText(str(value)))
    
    def setModelData(self, lineEd, model, index):
        value = lineEd.text()
        model.setData(index, value, QtCore.Qt.EditRole)
# ++++++  
class MainWindow(QMainWindow, Ui_MainWindow):

    def __init__(self, parent=None, *args, **kwargs) -> object:
        QMainWindow.__init__(self)
        self.setupUi(self)

        self._createAction()
        self._createMenuBar()
        self._connectActions()

        self.menuBar()

        self.Button_title_made.setEnabled(False)
        self.Button_system_made.setEnabled(False)
        self.tabWidget.setCurrentIndex(0)  # включает вкладку титул
        # ++ отключение страниц
        self.tabWidget.setTabEnabled(1, True)
        self.tabWidget.setTabEnabled(2, False)
        self.tabWidget.setTabEnabled(3, False)
        self.tabWidget.setTabEnabled(4, False)
        self.tabWidget.setTabEnabled(5, False)
        self.tabWidget.setTabEnabled(6, False)
        self.tabWidget.setTabEnabled(7, True)

        self.toolBox.setItemEnabled(0, False)
        self.toolBox.setItemEnabled(1, False)
        self.toolBox.setItemEnabled(2, False)
        self.toolBox.setItemEnabled(3, False)
        self.toolBox.setItemEnabled(4, False)
        self.toolBox.setItemEnabled(5, False)
        self.toolBox.setItemEnabled(6, False)
        self.toolBox.setItemEnabled(7, True)


    # ====== создание строки меню ===========
    def _createMenuBar(self):
        menuBar = self.menuBar()
        menuBar.setNativeMenuBar(False)  # разрешает показ менюбара

        # меню Соревнования
        fileMenu = QMenu("Соревнования", self)  # основное
        menuBar.addMenu(fileMenu)
        # подменю с выбором (addMenu добавляет к пункту возможность выбора)
        new_comp = fileMenu.addMenu("Новые")
        fileMenu.addSeparator()  # вставляет разделительную черту
        go_to = fileMenu.addMenu("Перейти к")
        fileMenu.addSeparator()  # вставляет разделительную черту
        # подменю без выбора (addAction создает сразу действие)
        system = fileMenu.addMenu("Система")
        choice = fileMenu.addMenu("Жеребьевка")
        # saveList = fileMenu.addMenu("Сохранить")
        fileMenu.addSeparator()
        last_comp = fileMenu.addMenu("Последние")
        fileMenu.addSeparator()
        fileMenu.addAction(self.exitAction)
        # меню Редактировать
        editMenu = menuBar.addMenu("Редактировать")  # основное
        # меню Печать
        printMenu = menuBar.addMenu("Печать") # основное

        # ============ создание подменю

        go_to.addAction(self.go_to_Action)  # подменю выбора соревнования
        system.addAction(self.system_made_Action)  # подменю создание системы
        system.addAction(self.system_edit_Action)  # подменю редактирование системы
        choice.addAction(self.choice_one_table_Action) # подменю одна таблица
        choice.addAction(self.choice_gr_Action)  # подменю группы
        choice.addAction(self.choice_pf_Action)  # подменю полуфиналы
        choice.addAction(self.choice_fin_Action)  # подменю финалы
    
        last_comp.addAction(self.first_comp_Action)
        last_comp.addAction(self.second_comp_Action)
        last_comp.addAction(self.third_comp_Action)
        last_comp.addAction(self.fourth_comp_Action)
        last_comp.addAction(self.fifth_comp_Action)
        ed_Menu = editMenu.addMenu("Жеребьевка")
        ed_Menu.addAction(self.ed_one_table_Action)
        ed_Menu.addAction(self.ed_etap_Action)

        editMenu.addAction(self.vid_edit_Action)  #в осн меню -Редактировать- добавлен пункт сразу с акцией -Вид страницы этапов
        # меню Рейтинг
        rank_Menu = menuBar.addMenu("Рейтинг")  # основное
        rank_Menu.addAction(self.rAction)
        rank_Menu.addAction(self.r1Action)
        # меню печать
        print_Menu = printMenu.addMenu("Чистые таблицы") 
        print_Menu.addAction(self.clear_s8_full_Action)  
        print_Menu.addAction(self.clear_s8_2_Action)       
        print_Menu.addAction(self.clear_s16_Action)
        print_Menu.addAction(self.clear_s16_2_Action)
        print_Menu.addAction(self.clear_s32_Action)
        print_Menu.addAction(self.clear_s32_full_Action)
        print_Menu.addAction(self.clear_s32_2_Action)

        # меню просмотр (последовательность вида в меню)
        view_Menu = menuBar.addMenu("Просмотр")
        view_Menu.addAction(self.view_all_comp_Action)
        view_Menu.addSeparator()
        view_Menu.addAction(self.view_title_Action)
        view_Menu.addAction(self.view_regions_list_Action)
        view_Menu.addAction(self.view_referee_list_Action)
        view_Menu.addAction(self.view_winners_list_Action)
        view_Menu.addAction(self.view_list_Action)
        view_Menu.addSeparator()
        view_Menu.addAction(self.view_gr_Action)
        pf_view_Menu = view_Menu.addMenu("Полуфиналы")
        pf_view_Menu.addAction(self.view_pf1_Action)
        pf_view_Menu.addAction(self.view_pf2_Action)
        view_Menu.addAction(self.view_one_table_Action)
        v_Menu = view_Menu.addMenu("Финалы")
        v_Menu.addAction(self.view_fin1_Action)
        v_Menu.addAction(self.view_fin2_Action)
        v_Menu.addAction(self.view_fin3_Action)
        v_Menu.addAction(self.view_fin4_Action)
        v_Menu.addAction(self.view_fin5_Action)
        v_Menu.addAction(self.view_fin6_Action)
        v_Menu.addAction(self.view_fin7_Action)
        v_Menu.addAction(self.view_fin8_Action)

        # меню помощь
        help_Menu = menuBar.addMenu("Помощь")  # основное
        help_Menu.addAction(self.copy_db_Action)
    #  создание действий меню

    def _createAction(self):
        self.helpAction = QAction(self)
        self.system_edit_Action = QAction("Редактировать")
        self.system_made_Action = QAction("Создать")
        self.exitAction = QAction("Выход")
        self.rAction = QAction("Текущий рейтинг")
        self.r1Action = QAction("Рейтинг за январь")
        self.first_comp_Action = QAction("пусто")
        self.second_comp_Action = QAction("пусто")
        self.third_comp_Action = QAction("пусто")
        self.fourth_comp_Action = QAction("пусто")
        self.fifth_comp_Action = QAction("пусто")
        self.ed_one_table_Action = QAction("Редакитровать таблицу")

        self.ed_etap_Action = QAction("Редактирование этапов")  # подменю редактор
        self.vid_edit_Action = QAction("Вид страницы этапов")

        self.choice_one_table_Action = QAction("Одна таблица")
        # подменю жеребьевка -группы-
        self.choice_gr_Action = QAction("Группы")
        # подменю жеребьевка -полуфиналы-
        self.choice_pf_Action = QAction("Полуфиналы")
        self.choice_fin_Action = QAction("Финалы")  # подменю жеребьевка -финалы-
        self.view_all_comp_Action = QAction("Полные соревнования")
        self.view_title_Action = QAction("Титульный лист")
        self.view_referee_list_Action = QAction("Список ГСК")
        self.view_regions_list_Action = QAction("Список субъектов РФ")
        self.view_winners_list_Action = QAction("Список победителей")
        self.view_list_Action = QAction("Список участников")
        self.view_gr_Action = QAction("Группы")
        self.view_pf1_Action = QAction("1-й полуфинал")
        self.view_pf2_Action = QAction("2-й полуфинал")

        self.view_one_table_Action = QAction("Одна таблица")
        self.go_to_Action = QAction("пусто")
        # подменю -печать-
        self.clear_s8_full_Action = QAction("Сетка 8")
        self.clear_s8_2_Action = QAction("Сетка 8 минус 2")
        self.clear_s16_Action = QAction("Сетка 16")
        self.clear_s16_2_Action = QAction("Сетка 16 минус 2")
        self.clear_s32_2_Action = QAction("Сетка 32 минус 2")
        self.clear_s32_full_Action = QAction("Сетка 32 прогрессивная")
        self.clear_s32_Action = QAction("Сетка 32 (1-3 места)")
        # ======== подменю финалы ============= сделать в зависимости от кол-во финалов остальные невидимые
        self.view_fin1_Action = QAction("1-финал")
        self.view_fin2_Action = QAction("2-финал")
        self.view_fin3_Action = QAction("3-финал")
        self.view_fin4_Action = QAction("4-финал")
        self.view_fin5_Action = QAction("5-финал")
        self.view_fin6_Action = QAction("6-финал")
        self.view_fin7_Action = QAction("7-финал")
        self.view_fin8_Action = QAction("8-финал")
        # выключает пункты меню пока не создана система
        self.choice_one_table_Action.setEnabled(False)
        self.choice_gr_Action.setEnabled(False)
        self.choice_pf_Action.setEnabled(False)
        self.choice_fin_Action.setEnabled(False)

        # self.view_all_comp_Action
        self.view_one_table_Action.setEnabled(False)
        self.view_gr_Action.setEnabled(False)
        self.view_pf1_Action.setEnabled(False)
        self.view_pf2_Action.setEnabled(False)
        # self.v_menu
        self.view_fin1_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin2_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin3_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin4_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin5_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin6_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin7_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin8_Action.setEnabled(False)  # делает пункт меню не видимым
        # пункты меню редактирование жеребьевки
        self.ed_one_table_Action.setEnabled(False)  # делает пункт меню не видимым
        self.ed_etap_Action.setEnabled(False)  # делает пункт меню не видимым

        self.copy_db_Action = QAction("Импорт из базы данных")

    def _connectActions(self):
        # Connect File actions
        self.system_made_Action.triggered.connect(self.system_made)
        self.system_edit_Action.triggered.connect(self.system_made)
        self.vid_edit_Action.triggered.connect(self.vid_edit)
        self.exitAction.triggered.connect(self.exit)
        self.choice_one_table_Action.triggered.connect(self.choice)
        self.choice_gr_Action.triggered.connect(self.choice)
        self.choice_pf_Action.triggered.connect(self.choice)
        self.choice_fin_Action.triggered.connect(self.choice)
        self.view_all_comp_Action.triggered.connect(self.view)
        self.view_title_Action.triggered.connect(self.view)
        self.view_referee_list_Action.triggered.connect(self.view)
        self.view_regions_list_Action.triggered.connect(self.view)
        self.view_winners_list_Action.triggered.connect(self.view)
        self.view_list_Action.triggered.connect(self.view)
        self.view_one_table_Action.triggered.connect(self.view)
        self.view_gr_Action.triggered.connect(self.view)
        self.view_pf1_Action.triggered.connect(self.view)
        self.view_pf2_Action.triggered.connect(self.view)
        self.view_fin1_Action.triggered.connect(self.view)
        self.view_fin2_Action.triggered.connect(self.view)
        self.view_fin3_Action.triggered.connect(self.view)
        self.view_fin4_Action.triggered.connect(self.view)
        self.view_fin5_Action.triggered.connect(self.view)
        self.view_fin6_Action.triggered.connect(self.view)
        self.view_fin7_Action.triggered.connect(self.view)
        self.view_fin8_Action.triggered.connect(self.view)
        self.clear_s8_full_Action.triggered.connect(self.print_clear)
        self.clear_s8_2_Action.triggered.connect(self.print_clear)
        self.clear_s16_Action.triggered.connect(self.print_clear)
        self.clear_s16_2_Action.triggered.connect(self.print_clear)
        self.clear_s32_full_Action.triggered.connect(self.print_clear)
        self.clear_s32_Action.triggered.connect(self.print_clear)
        self.clear_s32_2_Action.triggered.connect(self.print_clear)

        self.first_comp_Action.triggered.connect(self.last)
        self.second_comp_Action.triggered.connect(self.last)
        self.third_comp_Action.triggered.connect(self.last)
        self.fourth_comp_Action.triggered.connect(self.last)
        self.fifth_comp_Action.triggered.connect(self.last)

        self.ed_etap_Action.triggered.connect(self.edit_etap)

        self.go_to_Action.triggered.connect(self.open)
        # Connect Рейтинг actions
        self.rAction.triggered.connect(self.r_File)
        self.r1Action.triggered.connect(self.r1_File)

        self.copy_db_Action.triggered.connect(self.import_db)

    def newFile(self):
        # Logic for creating a new file goes here...
        my_win.textEdit.setText("Нажата кнопка меню соревнования")
        gamer = db_select_title()

    def r_File(self):
        # Logic for creating a new file goes here...
        self.statusbar.showMessage("Загружен рейтинг-лист на текущий месяц")
        my_win.tabWidget.setTabEnabled(6, True)
        my_win.tabWidget.setCurrentIndex(6)
        fill_table_R_list()
        my_win.comboBox_choice_R.setCurrentIndex(0)
        my_win.lineEdit_find_player_in_R.setFocus()

    def r1_File(self):
        self.statusbar.showMessage("Загружен рейтинг-лист на январь месяц")
        my_win.tabWidget.setTabEnabled(6, True)
        my_win.tabWidget.setCurrentIndex(6)
        fill_table_R1_list()
        my_win.comboBox_choice_R.setCurrentIndex(1)
        my_win.lineEdit_find_player_in_R.setFocus()

    def import_db(self):
        """Импорт из бэкап в базу данных"""
        fname = QFileDialog.getOpenFileName(my_win, "Выбрать файл базы данных", "", "comp_db_backup.db")
        filepath = str(fname[1])
        try:
            db = sqlite3.connect('comp_db.db')
            db_backup = sqlite3.connect(filepath)
            with db_backup:
                db_backup.backup(db, pages=3, progress=None)
            # показывает статус бар на 5 секунд
            my_win.statusbar.showMessage(
                "Импорт базы данных завершен успешно", 5000)
        except sqlite3.Error as error:
            # показывает статус бар на 5 секунд
            my_win.statusbar.showMessage(
                "Ошибка при копировании базы данных", 5000)
        finally:
            if (db_backup):
                db_backup.close()
                my_win.activateWindow()

    def exit(self):
        exit_comp()

    def choice(self):
        msg = QMessageBox
        sender = self.sender()
        system = System.select().where(System.title_id == title_id())
        if sender == self.choice_one_table_Action: # одна таблица
            sys = system.select().where(System.stage == "Одна таблица").get()
            type = sys.type_table
            group = sys.total_group
            fin = "Одна таблица"
            check_flag = check_choice(fin)
            if check_flag  is True:
                reply = msg.information(my_win, 'Уведомление', f"Жеребъевка {fin} была произведена,"
                                                                            f"\nесли хотите сделать "
                                                                            "повторно\nнажмите-ОК-, "
                                                                            "если нет то - Cancel-",
                                                msg.Ok,
                                                msg.Cancel)
                if reply == msg.Ok:
                    if type == "круг":
                        one_table(fin, group)
                    else:
                        clear_db_before_choice_final(fin)
                        posev_data = player_choice_in_setka(fin)
                        player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                        load_combobox_filter_final()
                else:
                    return
            else:
                if type == "круг":
                    player_fin_on_circle(fin)
                else:
                    posev_data = player_choice_in_setka(fin)
                    player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                    load_combobox_filter_final()
            add_open_tab(tab_page="Финалы")
        elif sender == self.choice_gr_Action:  # нажат подменю жеребъевка групп
            flag_real = check_real_player()
            if flag_real is True:
                reply = msg.information(my_win, 'Уведомление',
                                                "В списке присутствуют спортсмены,\nиз предварительной заявке"
                                                "\nне подтвержденые о своем участии!",
                                        msg.Ok)
                if reply == msg.Ok:
                    my_win.tabWidget.setCurrentIndex(2)                                                        
            for stage_sys in system:
                stage = stage_sys.stage
                if stage == "Предварительный":
                    if stage_sys.choice_flag == True:
                        reply = msg.information(my_win, 'Уведомление',
                                                        "Жеребъевка была произведена,\nесли хотите сделать "
                                                        "повторно\nнажмите -ОК-, если нет то - Cancel-",
                                                        msg.Ok, msg.Cancel)

                        if reply == msg.Ok:
                            my_win.tabWidget.setCurrentIndex(2)
                            clear_db_before_choice(stage)
                            # === вставить ручной вид жеребьевки
                            choice_gr_automat()
                            add_open_tab(tab_page="Группы")
                            my_win.tabWidget.setCurrentIndex(3)
                            my_win.ed_etap_Action.setEnabled(True) # включает меню - редактирование жеребьеввки групп
                            return
                        else:
                            return
                    else:
                        my_win.tabWidget.setCurrentIndex(2)
                        choice_gr_automat()
        elif sender == self.choice_pf_Action: # подменю полуфиналы            
            stage = select_choice_semifinal()
            system_stage = system.select().where(System.stage == stage).get()
            choice_flag = system_stage.choice_flag
            if stage is None: # если отмена при выборе жеребьевки
                return
            if choice_flag is True:
                reply = msg.information(my_win, 'Уведомление',
                                                "Жеребъевка была произведена,\nесли хотите сделать "
                                                "повторно\nнажмите -ОК-, если нет то - Cancel-",
                                                msg.Ok, msg.Cancel)

                if reply == msg.Ok:
                    clear_db_before_choice_semifinal(stage)
                    # === вставить ручной вид жеребьевки
                    choice_semifinal_automat(stage)
# ======= заполнение сыграныыми играми в группах
                    reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить {stage} результатами "
                                                                            f"встреч, сыгранных в группах.",                                                                            
                                            msg.Ok,
                                            msg.Cancel)
                    if reply == msg.Ok:
                        load_playing_game_in_table_for_semifinal(stage)
                    else:
                        return
                    add_open_tab(tab_page="Полуфиналы")
                    my_win.tabWidget.setCurrentIndex(4)
                    my_win.ed_etap_Action.setEnabled(True) # включает меню - редактирование жеребьеввки групп
                    return
                else:
                    return
            else:
                # проверяет все или игры в группе сыграны
                result_all = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == "Предварительный"))
                all_game = len(result_all)
                result_gameing = Result.select().where((Result.title_id == title_id()) & (Result.winner != ""))
                playing_games = len(result_gameing)
                remains = all_game - playing_games
                if remains == 0:
                    choice_semifinal_automat(stage)
                    reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить {stage} результатами "
                                                                            f"встреч, сыгранных в группах.",                                                                            
                                            msg.Ok,
                                            msg.Cancel)
                    if reply == msg.Ok:
                        load_playing_game_in_table_for_semifinal(stage)
                    else:
                        return
                    add_open_tab(tab_page="Полуфиналы")
                    my_win.tabWidget.setCurrentIndex(4)
                    my_win.ed_etap_Action.setEnabled(True) # включает меню - редактирование жеребьеввки групп
        elif sender == self.choice_fin_Action:  # нажат подменю жеребьевка финалов

            fin = select_choice_final()
            if fin is None: # если отмена при выборе жеребьевки
                return
            # ======
            fin_list = []
            stage_list = ["Одна таблица", "Предварительный", "1-й полуфинал", "2-й полуфинал"]
            for k in system:
                stage = k.stage
                if stage not in stage_list:
                    fin_list.append(stage)
            count_fin = len(fin_list) 
            if count_fin == 1:
                title = Title.get(Title.id == title_id())
                tab_str = title.tab_enabled       
            # ======
            sys = system.select().where(System.stage == fin).get()
            type = sys.type_table
            kol_player_exit = sys.mesta_exit
            etap_exit = sys.stage_exit
            if etap_exit == "Предварительный":
                etap_replacing = etap_exit.replace("ый", "ом")
            elif etap_exit == "1-й полуфинал" or etap_exit == "2-й полуфинал":
                etap_replacing = etap_exit + "е"
            fin_replacing = fin.replace("й", "ого") + "а"
            if fin is not None:
                checking_flag = checking_possibility_choice(fin) # флаг жеребьевки этапа, если True значит все игры предварительного или полуфиналов сыграны
                if checking_flag is False:
                    return
                check_flag = check_choice(fin) # была ли сделана жеребьевка
                if check_flag is True:
                    reply = msg.information(my_win, 'Уведомление', f"Жеребъевка {fin} была произведена,"
                                                                        f"\nесли хотите сделать "
                                                                        "повторно\nнажмите-ОК-, "
                                                                        "если нет то - Cancel-",
                                                    msg.Ok,
                                                    msg.Cancel)
                    if reply == msg.Ok:
                        if type == "круг":
                            clear_db_before_choice_final(fin)
                            player_fin_on_circle(fin)
                            if kol_player_exit > 1:
                                reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить игры {fin_replacing} результатами "
                                                                        f"встреч, сыгранных в {etap_replacing} этапе.",                                                                            
                                                    msg.Ok,
                                                    msg.Cancel)
                                if reply == msg.Ok:
                                    load_playing_game_in_table_for_final(fin)
                                else:
                                    return
                            add_open_tab(tab_page="Финалы")
                        else:
                            clear_db_before_choice_final(fin)
                            posev_data = player_choice_in_setka(fin)
                            player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                            load_combobox_filter_final()
                            add_open_tab(tab_page="Финалы")
                            load_name_net_after_choice_for_wiev(fin)
                            my_win.tabWidget.setCurrentIndex(5) 
                    else:
                        return
                else:
                    if type == "круг":
                        player_fin_on_circle(fin)
                        if kol_player_exit > 1:
                            reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить игры {fin_replacing} результатами "
                                                                        f"встреч, сыгранных в {etap_replacing} этапе.",
                                                                            
                                                    msg.Ok,
                                                    msg.Cancel)
                            if reply == msg.Ok:
                                load_playing_game_in_table_for_final(fin)
                            else:
                                return                   
                    else:
                        posev_data = player_choice_in_setka(fin)
                        player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                        load_combobox_filter_final()
                        add_open_tab(tab_page="Финалы")
                        load_name_net_after_choice_for_wiev(fin)
                        my_win.tabWidget.setCurrentIndex(5) 
            else:
                return
        enabled_menu_after_choice()

    def system_made(self):
        system_competition()

    def help(self):
        pass

    def edit_etap(self):
        """редактирование жеребьевки этапов соревнования"""
        my_win.tabWidget.setCurrentIndex(7)
        my_win.comboBox_first_group.clear()
        my_win.comboBox_second_group.clear()
        my_win.tableView.hide()

    def open(self):
        go_to()

    def view(self):
        view()
    
    def vid_edit(self):
        change_page_vid()

    def print_clear(self):
        """Печать чистых таблиц"""
        sender = self.sender()

        if sender == self.clear_s32_Action:
            setka_32_made(fin="1-й финал")
        elif sender == self.clear_s32_full_Action:
            setka_32_full_made(fin="1-й финал")
        elif sender == self.clear_s32_2_Action:
            setka_32_2_made(fin="1-й финал")
        elif sender == self.clear_s16_Action:
            setka_16_full_made(fin="1-й финал")
        elif sender == self.clear_s16_2_Action:
            setka_16_2_made(fin="1-й финал")
        elif sender == self.clear_s8_full_Action:
            setka_8_full_made(fin="1-й финал")
        elif sender == self.clear_s8_2_Action:
            setka_8_2_made(fin="1-й финал")
        view()

    def last(self):
        """открыте соревнований из пункта меню - последние-"""
        sender = self.sender()
        if sender == self.first_comp_Action:
            go_to()
        elif sender == self.second_comp_Action:
            go_to()
        elif sender == self.third_comp_Action:
            go_to()
        elif sender == self.fourth_comp_Action:
            go_to()
        elif sender == self.fifth_comp_Action:
            go_to()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


app = QApplication(sys.argv)
my_win = MainWindow()
my_win.setWindowTitle("Соревнования по настольному теннису")
my_win.setWindowIcon(QIcon("CTT.png"))
my_win.resize(1390, 804)
my_win.center()

class StartWindow(QMainWindow, Ui_Form):
    """Стартовое окно приветствия"""
    def __init__(self):
        super(StartWindow, self).__init__()
        self.setupUi(self)  # загружает настройки формы(окна) из QT
        self.setWindowTitle('Добро пожаловать в COMPETITIONS_TT')
        self.setWindowIcon(QIcon("CTT.png"))
        self.Button_open.clicked.connect(self.open)
        self.Button_new.clicked.connect(self.new)
        self.Button_old.clicked.connect(self.load_old)
        self.Button_R.clicked.connect(self.r_load)
        self.LinkButton.clicked.connect(self.last_comp)
        self.Button_open.setEnabled(False)

        self.pb = QProgressBar()
        self.pb.setMinimum(0)
        self.pb.setMaximum(100)

        dbase()
        count = len(Title.select())
        if count != 0:
            # получение последней записи в таблице
            t_id = Title.select().order_by(Title.id.desc()).get()
            id = t_id.id
            old_title = Title.get(Title.id == id)
            last_comp = old_title.full_name_comp
            self.LinkButton.setText(f"{last_comp}")
        else:
            self.LinkButton.setText("Список прошедших соревнований пуст")
            self.LinkButton.setEnabled(False)
            self.Button_open.setEnabled(False)
            self.Button_old.setEnabled(False)   

    def last_comp(self):
        """открытие последних соревнований"""
        gamer = db_select_title()
        tab_enabled(gamer)
        self.close()
        my_win.show()

    def open(self):
        full_name = db_select_title()
        self.close()
        my_win.setWindowTitle(f"Соревнования по настольному теннису. {full_name}")
        my_win.show()

    def new(self):
        """запускает новые соревнования"""
        msgBox = QMessageBox
        result = msgBox.question(my_win, "", "Вы действительно хотите создать новые соревнования?",
                                 msgBox.Ok, msgBox.Cancel)
        if result == msgBox.Ok:
            gamer = ("Мальчики", "Девочки", "Юноши",
                     "Девушки", "Мужчины", "Женщины")
            gamer, ok = QInputDialog.getItem(
                my_win, "Участники", "Выберите категорию спортсменов", gamer, 0, False)

            title = Title(name="", sredi="", vozrast="", data_start="", data_end="", mesto="", referee="",
                          kat_ref="", secretary="", kat_sec="", gamer=gamer, full_name_comp="", pdf_comp="",
                          short_name_comp="", tab_enabled="Титул", multiregion="").save()
            # получение последней записи в таблице
            t_id = Title.select().order_by(Title.id.desc()).get()
            title_id = t_id.id
            my_win.lineEdit_title_gamer.setText(gamer)
            db_r(gamer)
            system = System(title_id=title_id, total_athletes=0, total_group=0, max_player=0, stage="", type_table="",
                            page_vid="", label_string="", kol_game_string="", choice_flag=False, score_flag=5,
                            visible_game=False, stage_exit="", mesta_exit=0).save()
            self.close()
            tab_enabled(gamer)
            my_win.show()
        else:
            return

    def last_competition():
        """заполняе меню -последние- прошедшими соревнованиями 5 штук"""
        go_to()
    

    def r_load(self):
        pass


    def load_old(self):
        """загружает в комбобокс архивные соревнования"""
        self.label_4.show()
        comp_list = []
        # получение последней записи в таблице
        t_id = Title.select().order_by(Title.id.desc())
        count = len(t_id)
        n = 7
        if count != 0: 
            for i in t_id:
                n -= 1
                if n == 0:
                    break
                elif  n < 6:                   
                    old_comp = i.name
                    gamer = i.gamer
                    if old_comp != "":
                        name_comp = f"{old_comp}.{gamer}"
                        self.comboBox.addItem(name_comp)
                        full_name = i.full_name_comp
                    else:
                        full_name = "пусто"
                    comp_list.append(full_name)
 
        else:       
            print("нет соревнований")
            
        my_win.first_comp_Action.setText(comp_list[0])   
        my_win.second_comp_Action.setText(comp_list[1])
        my_win.third_comp_Action.setText(comp_list[2])
        my_win.fourth_comp_Action.setText(comp_list[3])
        my_win.fifth_comp_Action.setText(comp_list[4])
               
        if fir_window.comboBox.currentText() != "":
            fir_window.Button_open.setEnabled(True)

class ToolTip(): # создание всплывающих подсказок
    my_win.Button_made_R_file.setToolTip("Создание файла Excel для обсчета рейтинга")
    my_win.Button_made_one_file_pdf.setToolTip("Перед созданием одного файла, передвиньте строки с названием этапаов в необходимом порядке")

# class ProgressBarThread(QThread):
#     def __init__(self, fir_window, parent=None):
#         super().__init__()
#         countChanged = pyqtSignal(int)
#         # self.fir_window = fir_window
#         # self.ProgressBarThread_instance = ProgressBarThread(fir_window=self)

#     def run(value):
#         # value = self.fir_window.progressBar.value()
#         if value < 100:
#             # value = value + 1
#             fir_window.progressBar.setValue(value)
#             # time.sleep(0.2)
# class Actions(QDialog):
#     def __init__(self):
#         super().__init__()

#     def run(value): 
#         fir_window = fir_window
#         calc = ProgressBarThread()
#         calc.countChanged.connect(onCountChanged)
#         calc.start()

#     def ocCountChanged(self, value):
#         self.fir_window.progressBar.setValue(value)


    # def progress_bar_start_form(step):
    #     """Прогресс бар стартового окна"""
    #     # msgBox = QMessageBox
    #     # fir_window.activate()
    #     fir_window.pb.setValue(step)
    #     # if step >= 99:
    #     #     result = fir_window.msgBox.information(my_win, "Уведомление", "Загрузка рейтинг листа завершена.", msgBox.Ok)
    #     #     if result == fir_window.msgBox.Ok:
    #     #             fir_window.pb.setValue(0)
    #     return step

def dbase():
    """Создание DB и таблиц"""
    with db:
        db.create_tables([Title, R_list_m, R_list_d, Region, City, Player, R1_list_m, R1_list_d, Coach, System,
                          Result, Game_list, Choice, Delete_player, Referee])


def db_r(gamer):  # table_db присваивает по умолчанию значение R_list
    """переходит на функцию выбора файла рейтинга в зависимости от текущего или январского,
     а потом загружает список регионов базу данных"""
    msgbox = QMessageBox
    gamer_list = ["Мальчики", "Юноши", "Мужчины"]
    if gamer in gamer_list:
        table_db = R_list_m
    else:
        table_db = R_list_d
    reply = msgbox.information(my_win, 'Уведомление', "Выберите файл с текущим рейтингом, \nзатем файл рейтинга за январь месяц.",
                                                  msgbox.Ok)
 
    fname = QFileDialog.getOpenFileName(
        my_win, "Выбрать файл R-листа", "", "Excel files(*.xls *.xlsx)")
    if fname == ("", ""):
        # получение последней записи в таблице
        title = Title.select().order_by(Title.id.desc()).get()
        system = System.get(Title.id == title)
        system.delete_instance()
        title.delete_instance()
        return
    control_R_list(fname, gamer)
    load_listR_in_db(fname, table_db)
    my_win.statusbar.showMessage("Текущий рейтинг загружен")
    if gamer in gamer_list:
        table_db = R1_list_m
        ext = "(*01_m.xlsx *01_m.xls)"
    else:
        table_db = R1_list_d
        ext = "(*01_w.xlsx *01_w.xls)"
    fname = QFileDialog.getOpenFileName(
        my_win, "Выбрать файл R-листа", "", f"Excels files {ext}")
    load_listR_in_db(fname, table_db)
    my_win.statusbar.showMessage("Январский рейтинг загружен")
    # добавляет в таблицу регионы
    # получение последней записи в таблице
    t = Title.select().order_by(Title.id.desc()).get()
    title = t.id
    if title == 1:
        wb = op.load_workbook("регионы.xlsx")
        s = wb.sheetnames[0]
        sheet = wb[s]
        reg = []
        for i in range(1, 86):
            a = sheet['B%s' % i].value
            reg.append([a])
        with db:
            Region.insert_many(reg).execute()
    region()
    # показывает статус бар на 5 секунд
    my_win.statusbar.showMessage("Список регионов загружен", 5000)
    my_win.lineEdit_title_nazvanie.hasFocus()


def control_R_list(fname, gamer):
    """проверка рейтинга текущему месяцу"""
    filepatch = str(fname[0])
    znak = filepatch.rfind("/")
    month_vybor = filepatch[znak + 6:znak + 8]
    d = date.today()
    current_month = d.strftime("%m")
    if current_month != month_vybor:
        message = "Вы выбрали файл с не актуальным рейтингом!\nесли все равно хотите его использовать, нажмите <Ок>\nесли хотите вернуться, нажмите <Cancel>"
        reply = QtWidgets.QMessageBox.information(my_win, 'Уведомление', message,
                                                  QtWidgets.QMessageBox.Ok,
                                                  QtWidgets.QMessageBox.Cancel)
        if reply == QMessageBox.Ok:
            return
        else:
            db_r(gamer)
    else:
        return


def load_listR_in_db(fname, table_db):
    """при отсутствии выбора файла рейтинга, позволяет выбрать вторично или выйти из диалога
    если выбор был сделан загружает в базу данных"""
    msgBox = QMessageBox
    step = 0
    filepatch = str(fname[0])
    if table_db == R_list_m or table_db == R_list_d:
        r = "текущим"
    elif table_db == R1_list_m or table_db == R1_list_d:
        r = "январским"
    if filepatch == "":
        message = f"Вы не выбрали файл с {r} рейтингом!\nесли хотите выйти, нажмите <Ок>\nесли хотите вернуться, нажмите <Cancel>"
        reply = msgBox.information(my_win, 'Уведомление', message,
                                                  msgBox.Ok,
                                                  msgBox.Cancel)
        if reply == msgBox.Ok:
            return
        else:
            db_r(table_db)
    else:
        data = []
        data_tmp = []

        rlist = table_db.delete().execute()

        excel_data = pd.read_excel(filepatch)  # читает  excel файл Pandas
        data_pandas = pd.DataFrame(excel_data)  # получает Dataframe
        # создает список заголовков столбцов
        column = data_pandas.columns.ravel().tolist()

        count = len(data_pandas)  # кол-во строк в excel файле

        count_column = len(column)
        if count_column == 5:
            data_list_new = []
            data_list = [""]
            data_list_new = data_list * count
            data_pandas["Субъект РФ"] = data_list_new
            data_pandas["Федеральный округ"] = data_list_new
            column = data_pandas.columns.ravel().tolist()
        count = len(data_pandas)  # кол-во строк в excel файле

        for i in range(0, count):  # цикл по строкам
            # ProgressBarThread.run(value=i)
            for col in column:  # цикл по столбцам
                player_data = data_pandas.iloc[i][col]
                # заменяет пустые строки рейтинга на ноль и преобразовывает в тип int
                data_pandas['Рейтинг'] = data_pandas['Рейтинг'].fillna (0)
                data_pandas['Рейтинг'] = data_pandas['Рейтинг'].astype(int)
                data_tmp.append(player_data)  # получает временный список строки
            data.append(data_tmp.copy())  # добавляет в список Data
            data_tmp.clear()  # очищает временный список
        with db.atomic():
            for idx in range(0, len(data), 100):
                table_db.insert_many(data[idx:idx+100]).execute()


def region():
    """добавляет из таблицы в комбобокс регионы"""
    count = len(Region.select())
    if my_win.comboBox_region.currentIndex() > 0:  # проверка на заполненность комбобокса данными
        return
    else:
        with db:
            for r in range(1, count + 1):
                reg = Region.get(Region.id == r)
                my_win.comboBox_region.addItem(reg.region)


fir_window = StartWindow()  # Создаём объект класса ExampleApp
fir_window.show()  # Показываем окно


def change_sroki():
    """изменение текста label формы стартового окна в зависимости от выбора соревнования"""
    comp_data = {}
    data_comp = []
    data_comp_tmp = []
    t_id = Title.select().order_by(Title.id.desc())
    count = len(t_id)
    i = 0
    for k in t_id:
        data_st = k.data_start
        data_end = k.data_end
        data_comp.append(data_st)
        data_comp.append(data_end)
        data_comp_tmp = data_comp.copy()
        data_comp.clear()
        if i != 0:
            comp_data[i - 1] = data_comp_tmp
        i += 1
        if i == 5 or i == count:
            break
    index = fir_window.comboBox.currentIndex()
    data_list = comp_data[index]
    fir_window.label_4.setText(f"сроки: с {data_list[0]} по {data_list[1]}")


#  ==== наполнение комбобоксов ==========
page_orient = ("альбомная", "книжная")
kategoria_list = ("-выбор категории-", "2-я кат.", "1-я кат.", " ССВК")
mylist = ('мальчиков и девочек', 'юношей и девушек', 'мужчин и женщин')
raz = ("б/р", "3-юн", "2-юн", "1-юн", "3-р",
       "2-р", "1-р", "КМС", "МС", "МСМК", "ЗМС")
res = ("все игры", "завершенные", "не сыгранные")
vid_setki_one_table = ("-выбор типа таблицы-", "Сетка (-2)", "Сетка (с розыгрышем всех мест)",
             "Сетка (за 1-3 место)", "Круговая система")

my_win.comboBox_page_vid.addItems(page_orient)
my_win.comboBox_kategor_ref.addItems(kategoria_list)
my_win.comboBox_kategor_sec.addItems(kategoria_list)
my_win.comboBox_sredi.addItems(mylist)
my_win.comboBox_razryad.addItems(raz)
my_win.comboBox_filter_played.addItems(res)
my_win.comboBox_filter_played_sf.addItems(res)
my_win.comboBox_filter_played_fin.addItems(res)

my_win.comboBox_table_1.addItems(vid_setki_one_table)
my_win.comboBox_table_2.addItems(vid_setki_one_table)
my_win.comboBox_table_3.addItems(vid_setki_one_table)
my_win.comboBox_table_4.addItems(vid_setki_one_table)
my_win.comboBox_table_5.addItems(vid_setki_one_table)
my_win.comboBox_table_6.addItems(vid_setki_one_table)
my_win.comboBox_table_7.addItems(vid_setki_one_table)
my_win.comboBox_table_8.addItems(vid_setki_one_table)
my_win.comboBox_table_9.addItems(vid_setki_one_table)
my_win.comboBox_table_10.addItems(vid_setki_one_table)
my_win.comboBox_table_11.addItems(vid_setki_one_table)
# ставит сегодняшнюю дату в виджете календарь
my_win.dateEdit_start.setDate(date.today())
my_win.dateEdit_end.setDate(date.today())


def tab_enabled(gamer):
    """Включает вкладки в зависимости от создании системы и жеребьевки"""
    # включает вкладки меню системы
    title_list = []
    my_win.system_edit_Action.setEnabled(True) # делает меню  -редактировать- видиммым
    my_win.system_made_Action.setEnabled(True) # делает меню  -редактировать- видиммым

    sender = my_win.sender()
    tab_index = ["Титул", "Участники", "Система", "Группы", "Полуфиналы", "Финалы"]
    titles = Title.select().order_by(Title.id.desc())  # получает все title.id по убыванию
    title = Title.get(Title.id == title_id()) # текущий title
    n = 2
    for k in titles:
        if n != 0: 
            title_list.append(k.id)
            n -= 1
        else:
            break
    count_title = len(Title.select())
    title_id_current = title_list[0]
    title_id_last = title_list[1]# последний ид соревнования
 
    if count_title > 0: # если соревнования не первые
        my_win.setWindowTitle(f"Соревнования по настольному теннису. {gamer}")
        if sender == fir_window.LinkButton or sender == my_win.toolBox:  # если переход со стартового окна последение соревнование
            title_current = title.id
            if title_current == title_id_current:
                tit_id = Title.get(Title.id == title_id_last)
            else:
                tit_id = Title.get(Title.id == title_id_current)
            old_comp = tit_id.name
            old_data = tit_id.data_start
            old_gamer = tit_id.gamer
            comp = f"{old_comp}.{old_data}.{old_gamer}"
            my_win.go_to_Action.setText(comp)
            last_competition()
    my_win.tabWidget.setTabEnabled(1, False)        
    my_win.tabWidget.setTabEnabled(2, False)
    my_win.tabWidget.setTabEnabled(3, False)
    my_win.tabWidget.setTabEnabled(4, False)
    my_win.tabWidget.setTabEnabled(5, False)
    my_win.tabWidget.setTabEnabled(6, False)
    my_win.tabWidget.setTabEnabled(7, True)
# включает вкладки записаные в Титул
    tab_str = title.tab_enabled
    tab_list = tab_str.split(" ")
    for k in tab_list:
        ind = tab_index.index(k)
        my_win.tabWidget.setTabEnabled(ind, True)
        my_win.toolBox.setItemEnabled(ind, True)
    if gamer == "":
        gamer = my_win.lineEdit_title_gamer.text()
    my_win.toolBox.setCurrentIndex(0) # включает toolbox вкладку титул
    # Скрывает подменю системы в зависимости от созданной системы или нет
    if "Система" not in tab_list:
        my_win.system_edit_Action.setEnabled(False) # делает меню  -редактировать- не видиммым
    else:
        my_win.system_made_Action.setEnabled(False) # делает меню - создать- не видиммым
    enabled_menu_after_choice()


def add_open_tab(tab_page):
    """добавляет в таблицу -Title- список открытых вкладок"""
    tab_index = ["Титул", "Участники", "Система", "Группы", "Полуфиналы", "Финалы"]
    titles = Title.select().where(Title.id == title_id()).get()

    if tab_page != "":
        tab_str = titles.tab_enabled
        tab_list = tab_str.split(" ")

        if tab_page not in tab_list:
            tab_list.append(tab_page)        

        for k in tab_list:
            ind = tab_index.index(k)
            my_win.tabWidget.setTabEnabled(ind, True)
        tab_str = (' '.join(tab_list))
        titles.tab_enabled = tab_str
        titles.save()


def enabled_menu_after_choice():
    """Скрывает меню если еще не сделана жеребьевка"""
    systems = System.select().where(System.title_id == title_id())
    for k in systems:
        choice = k.choice_flag
        if choice is True:
            stage = k.stage
            if stage == "Одна таблица":
                my_win.view_one_table_Action.setEnabled(True)
            elif stage == "Предварительный":
                my_win.view_gr_Action.setEnabled(True)
                # my_win.ed_gr_Action.setEnabled(True) # включает меню - редакирование жеребьевки групп
            elif stage == "1-й полуфинал":
                my_win.view_pf1_Action.setEnabled(True)
                # my_win.ed_pf_Action.setEnabled(True)
            elif stage == "2-й полуфинал":
                my_win.view_pf2_Action.setEnabled(True)
                # my_win.ed_pf2_Action.setEnabled(True)
            elif stage == "1-й финал":
                my_win.view_fin1_Action.setEnabled(True)
            elif stage == "2-й финал":
                my_win.view_fin2_Action.setEnabled(True)
            elif stage == "3-й финал":
                my_win.view_fin3_Action.setEnabled(True)
            elif stage == "4-й финал":
                my_win.view_fin4_Action.setEnabled(True)
            elif stage == "5-й финал":
                my_win.view_fin5_Action.setEnabled(True)
            elif stage == "6-й финал":
                my_win.view_fin6_Action.setEnabled(True)
            elif stage == "7-й финал":
                my_win.view_fin7_Action.setEnabled(True)
            elif stage == "8-й финал":
                my_win.view_fin8_Action.setEnabled(True)
            my_win.ed_etap_Action.setEnabled(True)
        stage = k.stage

        if stage == "Одна таблица":
            my_win.choice_one_table_Action.setEnabled(True)
        elif stage == "Предварительный":
            my_win.choice_gr_Action.setEnabled(True)
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            my_win.choice_pf_Action.setEnabled(True)
        else:
            my_win.choice_fin_Action.setEnabled(True)


def db_insert_title(title_str):
    """Вставляем запись в таблицу титул"""
    msgBox = QMessageBox()
    nm = title_str[0]
    sr = title_str[1]
    vz = title_str[2]
    ds = title_str[3]
    de = title_str[4]
    ms = title_str[5]
    ref = title_str[6]
    kr = title_str[7]
    sek = title_str[8]
    ks = title_str[9]
    gm = title_str[10]
    fn = title_str[11]

    reply = msgBox.question(my_win, "Уведомление", "Если соревнования межрегиональные\n нажмите -YES-, \n"
                                            "Если принимают участие спортсмены одного реиона\n нажмите -NO-", msgBox.Yes, msgBox.No)
    if msgBox.Yes:
        mr = 1
    else:
        mr = 0                                        

    short_name, ok = QInputDialog.getText(my_win, "Краткое имя соревнования", "Создайте краткое имя соревнования,\nдля"
                                          " отбражения в названии файла при "
                                          "сохранении,\nиспользуете латинские буквы"
                                          " без пробелов.\n"
                                          "В формате название, возраст участников_дата,"
                                          " месяц, год и кто "
                                          "играет.")
    if ok:
        # получение последней записи в таблице
        t = Title.select().order_by(Title.id.desc()).get()
        title = Title(id=t, name=nm, sredi=sr, vozrast=vz, data_start=ds, data_end=de, mesto=ms, referee=ref,
                     kat_ref=kr, secretary=sek, kat_sec=ks, gamer=gm, full_name_comp=fn, pdf_comp="",
                     short_name_comp=short_name, multiregion=mr).save()
    else:
        return


def go_to():
    """переход на предыдущие соревнования и обратно при нажатии меню -перейти к- или из меню -последние-"""
    msgBox = QMessageBox
    sender = my_win.sender()
    tit = Title.get(Title.id == title_id())
    name = tit.name
    data = tit.data_start
    gamer_current = tit.gamer
    # полное название текущих соревнований
    full_name_current = f"{name}.{data}.{gamer_current}"

    if sender == my_win.first_comp_Action:
        full_name = my_win.first_comp_Action.text()
    elif sender == my_win.second_comp_Action:
        full_name = my_win.second_comp_Action.text()
    elif sender == my_win.third_comp_Action:
        full_name = my_win.third_comp_Action.text()
    elif sender == my_win.fourth_comp_Action:
        full_name = my_win.fourth_comp_Action.text()
    elif sender == my_win.fifth_comp_Action:
        full_name = my_win.fifth_comp_Action.text()
    elif sender == my_win.go_to_Action:
        full_name = my_win.go_to_Action.text()  # полное название к которым переходим 
        # присваиваем новый текст соревнований в меню -перейти к-
        my_win.go_to_Action.setText(full_name_current)

    if full_name == full_name_current:
        reply = msgBox.information(my_win, 'Уведомление', 'Данные соревнования уже открыты.',
                                    msgBox.Ok)
  
    titles = Title.get(Title.full_name_comp == full_name)
    gamer = titles.gamer
    my_win.lineEdit_title_nazvanie.setText(titles.name)
    my_win.lineEdit_title_vozrast.setText(titles.vozrast)
    my_win.dateEdit_start.setDate(titles.data_start)
    my_win.dateEdit_end.setDate(titles.data_end)
    my_win.lineEdit_city_title.setText(titles.mesto)
    my_win.comboBox_referee.setCurrentText(titles.referee)
    my_win.comboBox_kategor_ref.setCurrentText(titles.kat_ref)
    my_win.comboBox_secretary.setCurrentText(titles.secretary)
    my_win.comboBox_kategor_sec.setCurrentText(titles.kat_sec)
    my_win.lineEdit_title_gamer.setText(titles.gamer)
    my_win.tabWidget.setCurrentIndex(0)  # открывает вкладку титул
    tab_enabled(gamer)
    player_list = Player.select().where(Player.title_id == title_id())
    count_player = len(player_list)
    my_win.label_46.setText(f"Всего: {count_player} участников")
    list_player_pdf(player_list)


def db_select_title():
    """извлекаем из таблицы данные и заполняем поля титула для редактирования или просмотра"""
    sender = fir_window.sender()  # от какой кнопки сигнал
    if sender == my_win.go_to_Action:  # переход к соревнованиям из меню основного окна
        title = Title.get(Title.id == title_id())
        name = title.name
        data = title.data_start
        gamer_current = title.gamer
        # полное название текущих соревнований
        full_name_current = f"{name}.{data}.{gamer_current}"
        # присваиваем новый текст соревнований в меню -перейти к-
        my_win.go_to_Action.setText(full_name_current)
        gamer = title.gamer
    elif sender == my_win.toolBox or sender.text() != "Открыть":
        title = Title.get(Title.id == title_id())
        name = title.name
        gamer = title.gamer
    # сигнал от кнопки с текстом -открыть- соревнования из архива (стартовое окно)
    else:
        change_sroki()
        txt = fir_window.comboBox.currentText()
        key = txt.rindex(".")
        gamer = txt[key +  1:]
        name = txt[:key]
        sroki = fir_window.label_4.text()
        data = sroki[9:19]
        titles = Title.select()
        for title in titles:
            name_title = title.name
            gamer_title = title.gamer
            data_title = str(title.data_start)
            if name == name_title and gamer == gamer_title:
                if data == data_title:
                    break
    if name != "":
        my_win.lineEdit_title_nazvanie.setText(title.name)
        my_win.lineEdit_title_vozrast.setText(title.vozrast)
        my_win.dateEdit_start.setDate(title.data_start)
        my_win.dateEdit_end.setDate(title.data_end)
        my_win.lineEdit_city_title.setText(title.mesto)
        my_win.comboBox_referee.setCurrentText(title.referee)
        my_win.comboBox_kategor_ref.setCurrentText(title.kat_ref)
        my_win.comboBox_secretary.setCurrentText(title.secretary)
        my_win.comboBox_kategor_sec.setCurrentText(title.kat_sec)
        my_win.lineEdit_title_gamer.setText(title.gamer)
    else:
        load_comboBox_referee()
    tab_enabled(gamer)

    return gamer


def system_edit():
    """редактирование системы"""
    system_made()


def system_made():
    """Заполняет таблицу система кол-во игроков, кол-во групп и прочее"""
    systems  = System.select().where(System.title_id == title_id()).get()
    count_system = len(systems)  # получение количества записей (этапов) в системе
    sg = my_win.comboBox_table.currentText()
    page_v = my_win.comboBox_page_1.currentText()
    total_group = systems.total_group
    total_athletes = systems.total_athletes
    max_player = systems.max_player
    if sg == "одна таблица":
        system = System(id=systems, title_id=title_id(), total_athletes=total_athletes, total_group=0,
                        max_player=0, stage=sg, page_vid=page_v, label_string="", kol_game_string="",
                        choice_flag=False, score_flag=5, visible_game=True).save()
    else:  # предварительный этап
        for i in range(1, count_system + 1):
            system = System(id=systems, title_id=title_id(), total_athletes=total_athletes, total_group=total_group,
                            max_player=max_player, stage=sg, page_vid=page_v, label_string="", kol_game_string="",
                            choice_flag=False, score_flag=5, visible_game=True).save()
    player_in_table_group_and_write_Game_list_Result()
    my_win.label_33.setText("Всего: 0 игр.")
    my_win.checkBox_2.setChecked(False)
    my_win.checkBox_3.setChecked(False)
    my_win.Button_system_made.setEnabled(False)
    my_win.Button_1etap_made.setEnabled(False)
    my_win.Button_2etap_made.setEnabled(False)
    my_win.Button_3etap_made.setEnabled(False)
    my_win.Button_4etap_made.setEnabled(False)
    my_win.Button_5etap_made.setEnabled(False)
    my_win.Button_6etap_made.setEnabled(False)
    my_win.Button_7etap_made.setEnabled(False)
    my_win.Button_8etap_made.setEnabled(False)


def r_list_load_tableView():
    my_win.lineEdit_find_player_in_R.clear()
    r_combo_index = my_win.comboBox_choice_R.currentIndex()  
    if r_combo_index == 0:
        fill_table_R_list() 
    else:
        fill_table_R1_list() 


def title_string():
    """ переменные строк титульного листа """
    title_str = []
    # получение последней записи в таблице
    title = Title.select().order_by(Title.id.desc()).get()

    nm = my_win.lineEdit_title_nazvanie.text()
    sr = my_win.comboBox_sredi.currentText()
    vz = my_win.lineEdit_title_vozrast.text()
    ds = my_win.dateEdit_start.text()
    de = my_win.dateEdit_end.text()
    ms = my_win.lineEdit_city_title.text()
    ref = my_win.comboBox_referee.currentText()
    sek = my_win.comboBox_secretary.currentText()
    kr = my_win.comboBox_kategor_ref.currentText()
    ks = my_win.comboBox_kategor_sec.currentText()
    gm = title.gamer
    fn = f"{nm}.{ds}.{gm}"

    title_str = [nm, sr, vz, ds, de, ms, ref, sek, kr, ks, gm, fn]
    return title_str


def title_pdf():
    """сохранение в PDF формате титульной страницы"""
    msgBox = QMessageBox
    string_data = data_title_string()
    nz = my_win.lineEdit_title_nazvanie.text()
    sr = my_win.comboBox_sredi.currentText()
    vz = my_win.lineEdit_title_vozrast.text()
    ct = my_win.lineEdit_city_title.text()

    message = "Хотите добавить изображение в титульный лист?"
    reply = msgBox.question(my_win, 'Уведомление', message,
                                           msgBox.Yes,
                                           msgBox.No)
    if reply == msgBox.Yes:
        fname = QFileDialog.getOpenFileName(
            my_win, "Выбрать изображение", "/desktop", "Image files (*.jpg, *.png)")
        if fname[0] == "":
            return
        filepatch = str(fname[0])
    else:
        filepatch = None

    tit_id = Title.select().where(Title.id == title_id()).get()
    short_name = tit_id.short_name_comp
    canvas = Canvas(f"{short_name}_title.pdf", pagesize=A4)

    if filepatch == None:
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5 * cm, 28 * cm, "Федерация настольного тенниса России")
        canvas.drawString(3 * cm, 27 * cm, "Федерация настольного тенниса Нижегородской области")
        canvas.setFont("DejaVuSerif-Italic", 20)
        canvas.drawString(2 * cm, 23 * cm, nz)
        canvas.setFont("DejaVuSerif-Italic", 16)
        canvas.drawString(2.5 * cm, 22 * cm, f"среди {sr} {vz}")
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5.5 * cm, 5 * cm, f"г. {ct} Нижегородская область")
        canvas.drawString(7.5 * cm, 4 * cm, string_data)
    else:
        canvas.drawImage(filepatch, 7 * cm, 12 * cm, 6.9 * cm, 4.9 * cm,
                         mask=[0, 2, 0, 2, 0, 2])  # делает фон прозрачным
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5 * cm, 28 * cm, "Федерация настольного тенниса России")
        canvas.drawString(3 * cm, 27 * cm, "Федерация настольного тенниса Нижегородской области")
        canvas.setFont("DejaVuSerif-Italic", 20)
        canvas.drawString(2 * cm, 23 * cm, nz) # попробовать выравнить титул
        canvas.setFont("DejaVuSerif-Italic", 16)
        canvas.drawString(2.5 * cm, 22 * cm, f"среди {sr} {vz}")
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5.5 * cm, 5 * cm, f"г. {ct} Нижегородская область")
        canvas.drawString(7.5 * cm, 4 * cm, string_data)
    catalog = 1 # файл сохраяняется в каталоге /table_pdf
    change_dir(catalog)
    canvas.save()
    os.chdir("..")


def title_made():
    """создание титульного листа соревнования"""
    title_str = title_string()
    if my_win.Button_title_made.text() == "Редактировать":
        title_update()
        my_win.checkBox.setChecked(False)
        add_open_tab(tab_page="Участники")
        return
    else:
        db_insert_title(title_str)
    title_pdf()
    # после заполнения титула выключает чекбокс
    my_win.checkBox.setChecked(False)
    my_win.Button_title_made.setText("Создать")
    region()
    # получение последней записи в таблице
    title = Title.select().order_by(Title.id.desc()).get()

    # получение последнего id системы соревнования
    s = System.select().order_by(System.id.desc()).get()
    add_open_tab(tab_page="Участники")
    with db:
        System.create_table()
        sys = System(id=s, title_id=title, total_athletes=0, total_group=0, max_player=0, stage="", page_vid="",
                     label_string="", kol_game_string="", choice_flag=False, score_flag=5, visible_game=True).save()


def data_title_string():
    """получение строки начало и конец соревнований для вставки в титульный лист"""
    months_list = ("января", "февраля", "марта", "апреля", "мая", "июня", "июля",
                   "августа", "сентября", "октября", "ноября", "декабря")
    # получение последней записи в таблице
    title = Title.select().order_by(Title.id.desc()).get()
    datastart = str(title.data_start)
    dataend = str(title.data_end)
    ds = datastart[8:10]  # получаем число день из календаря
    ms = datastart[5:7]  # получаем число месяц из календаря
    ys = datastart[0:4]  # получаем число год из календаря
    # ye = int(dataend[0:4])
    me = dataend[5:7]
    de = dataend[8:10]
    month_st = months_list[int(ms) - 1]
    if de > ds:  # получаем строку начало и конец соревнования в
        # одном месяце или два месяца если начало и конец в разных месяцах
        return f"{ds}-{de} {month_st} {ys} г."
    elif de == ds:
        return f"{ds} {month_st} {ys} г."
    else:
        month_end = months_list[int(me) - 1]
        return f"{ds} {month_st}-{de} {month_end} {ys} г."


def title_update():
    """обновляет запись титула, если был он изменен"""
    title_str = title_string()
    nm = title_str[0]
    sr = title_str[1]
    vz = title_str[2]
    ds = title_str[3]
    de = title_str[4]
    ms = title_str[5]
    ref = title_str[6]
    kr = title_str[7]
    sek = title_str[8]
    ks = title_str[9]

    nazv = Title.select().order_by(Title.id.desc()).get()
    nazv.name = nm
    nazv.vozrast = vz
    nazv.data_start = ds
    nazv.data_end = de
    nazv.mesto = ms
    nazv.referee = ref
    nazv.kat_ref = kr
    nazv.secretary = sek
    nazv.kat_sec = ks
    nazv.save()
    title_pdf()


def clear_filter_rejting_list():
    """сбрасывает данные фильтра на вкладкк -рейтинг-"""
    my_win.lineEdit_find_player_in_R.clear()
    my_win.comboBox_filter_region_in_R.setCurrentIndex(0)
    my_win.comboBox_filter_city_in_R.setCurrentIndex(0)
    my_win.comboBox_filter_date_in_R.setCurrentIndex(0)
    filter_rejting_list()


def find_in_rlist():
    """при создании списка участников ищет спортсмена в текущем R-листе"""
    tb = my_win.tabWidget.currentIndex()
    if my_win.checkBox_find_player.isChecked():
        find_in_player_list()
    else:
        r_data_m = [R_list_m, R1_list_m]
        r_data_w = [R_list_d, R1_list_d]
        t_id = Title.get(Title.id == title_id())
        gamer = t_id.gamer
        my_win.listWidget.clear()
        if tb == 6:
            cur_index = my_win.comboBox_choice_R.currentIndex()
            txt = my_win.lineEdit_find_player_in_R.text()
        else:
            my_win.textEdit.clear()
            txt = my_win.lineEdit_Family_name.text()

        zn = txt.find(" ")
        if zn != -1:
            family = txt[:zn]
            name = txt[zn + 1:]
            if name != "":
                family = family.capitalize()
                name = name.capitalize()  # Переводит первую букву в заглавную
                txt = f"{family} {name}"
        else:
            txt = txt.capitalize()  # Переводит первую букву в заглавную
        if gamer == "Девочки" or gamer == "Девушки" or gamer == "Женщины":
            if tb == 6 and cur_index == 0:
                r_data = r_data_w[0]
            elif tb == 6 and cur_index == 1:
                r_data = r_data_w[1]
            else:
                r_data = r_data_w
        else:
            if tb == 6 and cur_index == 0:
                r_data = r_data_m[0]
            elif tb == 6 and cur_index == 1:
                r_data = r_data_m[1]
            else:
                r_data = r_data_m
        
        r = 0
        if tb == 6:
            if cur_index == 0:
                player_list = r_data.select().where(r_data.r_fname ** f'{txt}%')  # like поиск в текущем рейтинге
            else:
                player_list = r_data.select().where(r_data.r1_fname ** f'{txt}%')  # like поиск в текущем рейтинге
            # fill_table(player_list)
        else:
            for r_list in r_data:
                p = r_list.select()
                if r == 0:
                    my_win.label_63.setText("Поиск в текущем рейтинг листе.")
                    p = p.where(r_list.r_fname ** f'{txt}%')  # like поиск в текущем рейтинге
                    if r == 0  and len(p) != 0:
                        for pl in p:
                            full_stroka = f"{pl.r_fname}, {str(pl.r_list)}, {pl.r_bithday}, {pl.r_city}"
                            my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
                        return
                    elif r == 0:
                        r = 1
                        continue
                else:
                    my_win.label_63.setText("Поиск в январском рейтинге.")
                    p = p.where(r_list.r1_fname ** f'{txt}%')  # like поиск в январском рейтинге
                    if len(p) > 0:
                        for pl in p:
                            full_stroka = f"{pl.r1_fname}, {str(pl.r1_list)}, {pl.r1_bithday}, {pl.r1_city}"
                            my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
                    else:
                        full_stroka = ""
                        my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
                    return
      

def input_player():
    """Ввод нового игрока если его нет в рейтинг листе текущем и январском"""
    text = my_win.lineEdit_Family_name.text()
    zn = text.find(" ")
    family = text[:zn]
    name = text[zn + 1:]
    family = family.capitalize()
    name = name.capitalize()  # Переводит первую букву в заглавную  
    family = family.upper()
    my_win.lineEdit_Family_name.setText(f"{family} {name}")
    my_win.lineEdit_bday.setFocus()
    my_win.lineEdit_bday.setInputMask('00.00.0000')


def next_field():
    """переход к следующему полю ввода спортсмена"""
    my_win.lineEdit_R.setText('0')
    pl = my_win.lineEdit_Family_name.text()
    check_rejting_pay(pl)
    my_win.label_63.setText("Список городов.")
    my_win.lineEdit_city_list.setFocus()


def find_city():
    """Поиск городов и область"""
    sender = my_win.sender()
    my_win.listWidget.clear()
    txt = my_win.label_63.text()
    city_field = my_win.lineEdit_city_list.text()
    if txt == "Список городов.":
        city_field = city_field.capitalize()  # Переводит первую букву в заглавную
        index = city_field.find(".")
        if index != -1:
            second_word = city_field[index + 1:]
            second_word = second_word.capitalize()
            city_field = city_field[:index + 1] + second_word
    
    c = City.select()
    c = c.where(City.city ** f'{city_field}%')  # like
    if sender != my_win.comboBox_region:
        if (len(c)) == 0:
            my_win.textEdit.setText("Нет такого города в базе")
        else:
            for pl in c:
                full_stroka = f"{pl.city}"
                my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
            return
    else:  # вставляет регион соответсвующий городу
        if city_field != "":
            ir = my_win.comboBox_region.currentIndex()
            ir = ir + 1
            ct = my_win.lineEdit_city_list.text()
            with db:
                city = City(city=ct, region_id=ir).save()


def fill_table(player_list):
    """заполняет таблицу со списком участников QtableView спортсменами из db"""
    data = []
    data_table_tmp = []
    data_table_list = []
    model = MyTableModel(data)
    tb = my_win.tabWidget.currentIndex()
    player_selected = player_list.dicts().execute()
    row_count = len(player_selected)  # кол-во строк в таблице
    num_columns = [0, 1, 2, 3, 4, 5, 6]
    # кол-во наваний должно совпадать со списком столбцов
    if tb == 1:
        if my_win.checkBox_6.isChecked():
            num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер', 'Место', 'id_del'])
        else:
            num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер', 'Место']) 
    elif tb == 2:
        stage = my_win.comboBox_filter_choice_stage.currentText()
        if my_win.comboBox_filter_choice_stage.currentIndex() == 0:
            num_columns = [0, 2, 3, 4, 7, 9, 10, 11, 13, 14, 16]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'Группа', 'Место гр',
                                              'ПФ', "Группа ПФ", 'Место ПФ', 'Финал', 'Место'])
        elif stage == "Предварительный":
            num_columns = [0, 2, 3, 4, 5, 7, 9]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'Группа', 'Место в гр'])
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            num_columns = [0, 2, 3, 4, 5, 10, 11, 13]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'ПФ', 'Группа ПФ', 'Место ПФ']) 
        else: 
            num_columns = [0, 2, 3, 4, 5, 14, 16]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'Финал', 'Место в финале']) 
    elif tb == 3 or tb == 4 or tb == 5:
        num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        model.setHorizontalHeaderLabels(['id',' Стадия', 'Группа', 'Встреча', '1-й игрок', '2-й игрок', 'Победитель', 'Очки','Общ. счет', 'Счет в партиях']) 
    elif tb == 6:
        model.setHorizontalHeaderLabels(['id',' Место', ' R', 'Фамилия Имя', 'Дата рождения', 'Город', 'Регион']) 
    if tb == 1:
        if my_win.checkBox_15.isChecked():
            my_win.tableView.setSelectionMode(QAbstractItemView.MultiSelection) # выделение несколких строк по клику мышью
        else:
            my_win.tableView.setSelectionMode(QAbstractItemView.SingleSelection) # выделение одной строки по клику мышью
        my_win.tableView.setSelectionBehavior(QAbstractItemView.SelectRows) 
    elif tb == 3 or tb == 4 or tb == 5:
        my_win.tableView.setSelectionMode(QAbstractItemView.SingleSelection) # выделение одной строки по клику мышью
        my_win.tableView.setSelectionBehavior(QAbstractItemView.SelectRows) # 
    else:
        my_win.tableView.setSelectionMode(QAbstractItemView.NoSelection) # нет выделение строк по клику мышью

    if tb == 6:
        if row_count > 0:
            my_win.label_78.setText(f"Поиск спортсмена в рейтинге: найдено всего {row_count} записей(и).")
        else:
            my_win.label_78.setText(f"Поиск спортсмена в рейтинге: не найдено ни одной записи.")
    # создание  data (списка списков)
    if row_count != 0:  # список удаленных игроков пуст если R = 0
        for row in range(row_count):  # добавляет данные из базы в TableWidget
            item_1 = str(list(player_selected[row].values())[num_columns[0]])
            item_2 = str(list(player_selected[row].values())[num_columns[1]])
            item_3 = str(list(player_selected[row].values())[num_columns[2]])
            item_4 = str(list(player_selected[row].values())[num_columns[3]])
            item_5 = str(list(player_selected[row].values())[num_columns[4]])
            item_6 = str(list(player_selected[row].values())[num_columns[5]])
            item_7 = str(list(player_selected[row].values())[num_columns[6]])
            data_table_list = [item_1, item_2, item_3, item_4, item_5, item_6, item_7]
            if tb == 1:
                coach_id = str(list(player_selected[row].values())[num_columns[7]])
                coach = Coach.get(Coach.id == coach_id)
                item_8 = coach.coach
                item_9 = str(list(player_selected[row].values())[num_columns[8]])
                data_table_tmp = [item_8, item_9]
                if my_win.checkBox_6.isChecked():
                    item_10 = str(list(player_selected[row].values())[num_columns[9]])
                    data_table_tmp = [item_8, item_9, item_10]
                data_table_list.extend(data_table_tmp) 
            elif tb ==2:
                if my_win.comboBox_filter_choice_stage.currentIndex() == 0:
                    item_8 = str(list(player_selected[row].values())[num_columns[7]])
                    item_9 = str(list(player_selected[row].values())[num_columns[8]])
                    item_10 = str(list(player_selected[row].values())[num_columns[9]])
                    item_11 = str(list(player_selected[row].values())[num_columns[10]])
                    data_table_tmp = [item_8, item_9, item_10, item_11]
                elif stage == "Предварительный":
                    data_table_tmp = []
                elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
                    item_8 = str(list(player_selected[row].values())[num_columns[7]])
                    data_table_tmp = [item_8]
                else:
                    data_table_tmp = []
                data_table_list.extend(data_table_tmp) 
            elif tb == 3 or tb == 4 or tb == 5:
                item_8 = str(list(player_selected[row].values())[num_columns[7]])
                item_9 = str(list(player_selected[row].values())[num_columns[8]])
                item_10 = str(list(player_selected[row].values())[num_columns[9]])
                data_table_tmp = [item_8, item_9, item_10]
                data_table_list.extend(data_table_tmp)
          

            data.append(data_table_list.copy()) # данные, которые передаются в tableView (список списков)
        my_win.tableView.setModel(model)
        font = my_win.tableView.font()
        font.setPointSize(11)
        my_win.tableView.setFont(font)
        my_win.tableView.horizontalHeader().setFont(QFont("Verdana", 13, QFont.Bold)) # делает заголовки жирный и размер 13
        # y_win.setStyleSheet("#MainWindow{background-color:blue}")
        # my_win.tableView.horizontalHeader().setStyleSheet("#tableView{background-color:blue}")
           
        my_win.tableView.verticalHeader().setDefaultSectionSize(16) # высота строки 20 пикселей
        my_win.tableView.resizeColumnsToContents() # растягивает по содержимому
        my_win.tableView.horizontalHeader().setStretchLastSection(True) # растягивает последнюю колонку до конца
        my_win.tableView.setGridStyle(QtCore.Qt.SolidLine) # вид линии сетки 
    else:
        if tb == 1:
            if my_win.checkBox_15.isChecked() and row_count == 0:
                my_win.statusbar.showMessage(
                "Нет спортсменов из предварительной заявки", 10000)
                my_win.textEdit.setText("Нет спортсменов из предварительной заявки")
            else:
                row = 0
                my_win.statusbar.showMessage(
                    "Нет спортсменов удаленных из списка", 10000)
                my_win.textEdit.setText("Нет спортсменов удаленных из списка")
                my_win.checkBox_6.setChecked(False)
        elif tb == 6:
            row = 0
            my_win.statusbar.showMessage(
                "Такого спортсмена в рейтинг листе нет нет", 10000)
    my_win.tableView.show()
 

def fill_table_R_list():
    """заполняет таблицу списком из текущего рейтинг листа"""
    title = Title.select().where(Title.id == title_id()).get()
    gamer = title.gamer
    if gamer == "Девочки" or gamer == "Девушки" or gamer == "Женщины":
        player_list = R_list_d.select().order_by(R_list_d.r_fname)
    else:
        player_list = R_list_m.select().order_by(R_list_m.r_fname)
    # вставляет в таблицу необходимое кол-во строк
    row_count = len(player_list)
    my_win.label_78.setText(f"Всего {row_count} записей.")
    fill_table(player_list)


def fill_table_R1_list():
    """заполняет таблицу списком из январского рейтинг листа"""
    title = Title.select().where(Title.id == title_id()).get()
    gamer = title.gamer
    if gamer == "Девочки" or gamer == "Девушки" or gamer == "Женщины":
        player_list = R1_list_d.select().order_by(R1_list_d.r1_fname)
    else:
        player_list = R1_list_m.select().order_by(R1_list_m.r1_fname)
    # вставляет в таблицу необходимое кол-во строк
    row_count = len(player_list)
    my_win.label_78.setText(f"Всего {row_count} записей.")
    fill_table(player_list)


def fill_table_results():
    """заполняет таблицу результатов QtableView из db result"""
    system_id_list = []

    system_stage_list = ["Одна таблица", "Предварительный", "1-й полуфинал", "2-й полуфинал"]
    result = Result.select().where(Result.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
    tb = my_win.tabWidget.currentIndex()
    if tb == 3:
        stage = "Предварительный"
        system_id = system.select().where(System.stage == stage).get()
        id_system = system_id.id
        player_list = result.select().where(Result.system_id == id_system)  # проверка есть ли записи в таблице -result
    elif tb == 4:
        player_list = result.select().where((Result.system_stage == "1-й полуфинал") | (Result.system_stage == "2-й полуфинал")) # проверка есть ли записи в таблице -result-
    elif tb == 5:
        for k in system: # заполняе список ид системы финальных этапов
            id_system = k.id
            if k.stage not in system_stage_list:
                system_id_list.append(id_system)
        stage = my_win.comboBox_filter_final.currentText()

        if stage == "все финалы":      
            player_list = result.select().where(Result.system_stage == "Финальный")  # проверка есть ли записи в таблице -result 
        else:
            system_id = system.select().where(System.stage == stage).get()
            id_system = system_id.id
            player_list = result.select().where(Result.system_id == id_system)  # проверка есть ли записи в таблице -result 

    # ==== окрашивает победителя в красный цвет
    # row_num = my_win.tableView.currentIndex().row () # Number of lines
    # row_num = row_num + 1
    # pl_1 = my_win.tableView.model().index(row_num, 4).data()
    # pl_winner = my_win.tableView.model().index(row_num, 6).data()
    # if pl_winner != "None" and pl_winner != "":  # встреча сыграна
    #     if pl_1 == pl_winner:
    #         my_win.tableView.item(row_num, 4).setForeground(
    #             QBrush(QColor(255, 0, 0)))  # окрашивает текст
    #                         # в красный цвет 1-ого игрока
    #     else:
    #         my_win.tableView.model.item(row_num, 5).setForeground(
    #             QBrush(QColor(255, 0, 0)))  # окрашивает текст
    #                         # в красный цвет 2-ого игрока
    # else:
    #     my_win.tableView.item(row_num, 4).setForeground(
    #         QBrush(QColor(0, 0, 0)))  # в черный цвет 1-ого
    #     my_win.tableView.item(row_num, 5).setForeground(
    #         QBrush(QColor(0, 0, 0)))  # в черный цвет 2-ого
                    
    fill_table(player_list)



def fill_table_choice():
    """заполняет таблицу жеребьевки"""
    gamer = my_win.lineEdit_title_gamer.text()
    player_choice = Choice.select().where(Choice.title_id == title_id()).order_by(Choice.rank.desc())
    choice_list = player_choice.dicts().execute()
    row_count = len(choice_list)  # кол-во строк в таблице
    if row_count != 0:
        column_count = len(choice_list[0])  # кол-во столбцов в таблице
        # вставляет в таблицу необходимое кол-во строк
        my_win.tableWidget.setRowCount(row_count)
        for row in range(row_count):  # добавляет данные из базы в TableWidget
            for column in range(column_count):
                item = str(list(choice_list[row].values())[column])
                my_win.tableWidget.setItem(
                    row, column, QTableWidgetItem(str(item)))
        # ставит размер столбцов согласно записям
        my_win.tableWidget.resizeColumnsToContents()
        for i in range(0, row_count):  # отсортировывает номера строк по порядку
            my_win.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))


def fill_table_after_choice():
    """заполняет TableWidget после жеребьевки """
    choice = Choice.select().where(Choice.title_id == title_id())
    pl_choice = choice.select().order_by(Choice.group)
    player_list = pl_choice.select().order_by(Choice.posev_group)
    fill_table(player_list)


def progressbar(count):
    pass
    # progress = QtWidgets.QProgressBar()
    # progress.setValue(100)
    # progress.setMinimum(0)
    # progress.setMaximum(100)
    # m = int(count / 100)
    # for i in range(m, count, m):
    #     progress.setValue(100)
def debtor_R():
    """показывает список должников оплаты рейтинга"""
    player_list = Player.select().where(Player.title_id == title_id())
    if my_win.checkBox_11.isChecked():
        player_debitor_R = player_list.select().where(Player.pay_rejting == "долг")
        my_win.label_dolg_R.setText("Нет спортсменов без лицензии.")
        if len(player_debitor_R) == 0:
            my_win.textEdit.setText("Спортсменов, не оплативших регистрационыый взнос за рейтинг нет.")
    else:
        player_debitor_R = player_list.select().where(Player.pay_rejting == "долг")
        my_win.Button_pay_R.setEnabled(False)
        my_win.textEdit.clear()
    my_win.label_dolg_R.setText(f"Без лицензии: {len(player_debitor_R)} участников.")
    fill_table(player_debitor_R)


def add_player(): 
    """добавляет игрока в список и базу данных"""
    flag = False
    player_list = Player.select().where(Player.title_id == title_id())
    txt = my_win.Button_add_edit_player.text()
    count = len(player_list)
    pl_id = my_win.lineEdit_id.text()
    pl = my_win.lineEdit_Family_name.text()
    bd = my_win.lineEdit_bday.text()
    rn = my_win.lineEdit_R.text()
    ct = my_win.lineEdit_city_list.text()
    rg = my_win.comboBox_region.currentText()
    rz = my_win.comboBox_razryad.currentText()
    ch = my_win.lineEdit_coach.text()
    if pl_id == "": # добавляет нового игрока
        flag = check_repeat_player(pl, bd) # проверка повторного ввода игрока
    else:
        player = Player.select().where(Player.id == pl_id).get()
        pay_R = player.pay_rejting
        comment = player.comment

    num = count + 1
    fn = f"{pl}/{ct}"
    if txt != "Редактировать":
        if flag is True:
            my_win.lineEdit_Family_name.clear()
            my_win.lineEdit_bday.clear()
            my_win.lineEdit_R.clear()
            my_win.lineEdit_city_list.clear()
            my_win.lineEdit_coach.clear()
            return
    add_coach(ch, num)
    txt_edit = my_win.textEdit.toPlainText()
    ms = "" # записвыает место в базу как пустое
    idc = Coach.get(Coach.coach == ch)
    # ==== определяет завявка предварительная или нет
    title = Title.select().where(Title.id == title_id()).get()
    data_start = title.data_start
    date_current = date.today()
    if  date_current < data_start:
        zayavka = "предварительная"
    else:
        zayavka = "основная"
    if my_win.checkBox_6.isChecked():  # если отмечен флажок -удаленные-, то восстанавливает игрока и удаляет из
        # таблицы -удаленные-
        with db:
            player_del = Delete_player.get(Delete_player.id == pl_id)
            player_id = player_del.player_del_id           
            pay_R = player_del.pay_rejting
            comment = player_del.comment
            player_del.delete_instance()
            plr = Player(player_id=player_id, player=pl, bday=bd, rank=rn, city=ct, region=rg,
                         razryad=rz, coach_id=idc, full_name=fn, mesto=ms, title_id=title_id(), pay_rejting=pay_R,
                         comment=comment, coefficient_victories=0, total_game_player=0, total_win_game=0).save()
        my_win.checkBox_6.setChecked(False)  # сбрасывает флажок -удаленные-
    else:  # просто редактирует игрока
        if txt == "Редактировать":
            with db:
                plr =  player_list.select().where(Player.id == pl_id).get()
                plr.player = pl
                plr.bday = bd
                plr.rank = rn
                plr.city = ct
                plr.region = rg
                plr.razryad = rz
                plr.coach_id = idc
                plr.full_name = fn
                plr.pay_rejting = pay_R
                plr.comment = comment
                plr.save()
        elif txt == "Добавить":
            debt = "долг" if txt_edit == "Спортсмену необходимо оплатить рейтинг!" else ""
            with db:
                player = Player(player=pl, bday=bd, rank=rn, city=ct, region=rg, razryad=rz,
                                coach_id=idc, mesto="", full_name=fn, title_id=title_id(), pay_rejting=debt, comment="", 
                                coefficient_victories=0, total_game_player=0, total_win_game=0, application=zayavka).save()
        pl_id = Player.select().order_by(Player.id.desc()).get() # id нового игрока
        player_id = pl_id.id
        # ======== попробовать вставить одну строку в tableView
    player_list = Player.select().where(Player.title_id == title_id())
    fill_table(player_list)
    count = len(player_list)  # подсчитывает новое кол-во игроков
    my_win.label_46.setText(f"Всего: {count} участников")
    list_player_pdf(player_list)
    my_win.lineEdit_Family_name.clear()
    my_win.lineEdit_bday.clear()
    my_win.lineEdit_R.clear()
    my_win.lineEdit_city_list.clear()
    my_win.lineEdit_coach.clear()
    # check_rejting_pay(pl)
    if txt == "Редактировать":
        my_win.Button_add_edit_player.setText("Добавить")
        my_win.Button_del_player.setEnabled(False) 
        my_win.lineEdit_id.clear()       
    my_win.lineEdit_Family_name.setFocus()


def check_rejting_pay(pl):
    """Проверка игрока на оплату рейтинга и запись в базу данных"""
    txt_edit = my_win.textEdit.toPlainText()
    txt_tmp = my_win.label_63.text()
    if txt_tmp == "Поиск в январском рейтинге.":
        b_day = my_win.lineEdit_bday.text()
        year_player = int(b_day[6:])
        date_current = int(datetime.today().strftime("%Y"))
        raznica = date_current - year_player
        if raznica > 11:
            my_win.textEdit.setText("Спортсмену необходимо оплатить рейтинг!")
    elif txt_edit == "Спортсмену необходимо оплатить рейтинг!":
        plr = Player.select().where(Player.title_id == title_id())
        player_id = plr.select().where(Player.player == pl).get()
        with db:
                # player_id = plr.select().where(Player.player == pl).get()
            player_id.pay_rejting = "долг"
            player_id.comment = ""
            player_id.save()


def dclick_in_listwidget():
    """Находит фамилию спортсмена в рейтинге или фамилию тренера и заполняет соответсвующие поля списка"""
    msgBox = QMessageBox
    txt_tmp = my_win.label_63.text()
    text = my_win.listWidget.currentItem().text()
    coach_field = my_win.lineEdit_coach.text()
    if txt_tmp == "Список городов.": # если в listwidget список городов которые есть в базе
        my_win.label_63.setText("")
        my_win.lineEdit_city_list.setText(text)    
        cr = City.get(City.city == text)
        rg = Region.get(Region.id == cr.region_id)
        my_win.comboBox_region.setCurrentText(rg.region)
        my_win.listWidget.clear()   
    elif coach_field == "": # если строка "тренер" пустая значит заполняются поля игрока
        ds = len(text)
        sz = text.index(",")
        sz1 = text.index(",", sz + 1)
        sz2 = text.index(",", sz1 + 1)
        fam_name = text[0:sz]
        znak = fam_name.find(" ")
        fam = fam_name[:znak]
        fam = fam.upper()
        name = fam_name[znak + 1:]
        name = name.capitalize()
        r = text[sz + 2:sz1]
        dr = text[sz1 + 2:sz2]
        znak = dr.find(".")
        # ==== проверка правильность даты для участия в турнире
        title = Title.get(Title.id == title_id())
        vozrast_text = title.vozrast
        if vozrast_text != "": # если играют не мужчины или женщины то проверка на соответсвия возраста
            text_1 = vozrast_text.rfind("моложе")
            text_date = vozrast_text[:2]
            if text_1 == -1 and text_date == "до":
                mark = vozrast_text.find(" ")
                total_old = int(vozrast_text[mark + 1:5])
                year_current = int(datetime.today().strftime("%Y")) # текущий год
                year_bday = year_current - total_old + 1
            elif text_1 > -1: # если возраст г.р и моложе
                year_bday = int(vozrast_text[:4])
                year_current = int(datetime.today().strftime("%Y")) # текущий год
            after_date = date(year_bday, 1, 1)
            if znak != -1:
                date_object = datetime.strptime(dr,"%d.%m.%Y")
            else:                    
                date_object = datetime.strptime(dr,"%Y-%m-%d")
            dr_year = int(date_object.strftime('%Y')) # получаем только год рождения в числовом формате
            current_date = date(dr_year, 1, 1)
            if after_date > current_date: # сравниваем две даты
                result = msgBox.information(my_win, "", "Возраст спортсмена не соответсвует\nвозрастной категории соревнования.\n"
                        "Или возможно в рейтинге указана\nне правильная дата рождения.\nЕсли дата правильная нажмите -ОК-, или -Cancel-",
                                            msgBox.Ok, msgBox.Cancel)
                if result == msgBox.Ok:
                    return
        # ==== переводит строку с датой из базы даннных в строку к обычному виду
        if znak == -1:
            date_object = datetime.strptime(dr,"%Y-%m-%d")
            dr = date_object.strftime('%d.%m.%Y')
        #=====
        ci = text[sz2 + 2:ds] # город
        my_win.lineEdit_Family_name.setText(f"{fam} {name}")
        my_win.lineEdit_bday.setText(dr)
        my_win.lineEdit_R.setText(r)
        my_win.lineEdit_city_list.setText(ci)
         # ======= проверка на рейтинг ====
        if txt_tmp == "Поиск в январском рейтинге.":
            # txt_edit = txt_tmp
            pl = fam_name
            check_rejting_pay(pl)
        c = City.select()  # находит город и соответсвующий ему регион
        c = c.where(City.city ** f'{ci}')  # like
        if (len(c)) == 0:
            my_win.textEdit.setText("Нет такого города в базе, выберите регион где находится населенный пункт.")
            my_win.comboBox_region.setCurrentText("")
        else:  # вставляет регион соответсвующий городу
            cr = City.get(City.city == ci)
            rg = Region.get(Region.id == cr.region_id)
            my_win.comboBox_region.setCurrentText(rg.region)
            my_win.listWidget.clear()
    else:  # идет заполнение поля "тренер" из listWidget
        my_win.lineEdit_coach.setText(text)
        my_win.listWidget.clear()


def load_combobox_filter_final():
    """заполняет комбобокс фильтр финалов для таблицы результаты"""
    my_win.comboBox_filter_final.clear()
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    for sys in system:
        if sys.stage == "Одна таблица":
            if sys.choice_flag is True:
                fin.append(sys.stage)
                break
        else:
            stage = system.select().where(System.stage ** '%-й финал')
            fin = [k.stage for k in stage if k.choice_flag is True]
            fin.insert(0, "все финалы")
            break
    my_win.comboBox_filter_final.addItems(fin)
    my_win.comboBox_filter_choice_stage.addItems(fin)


def load_combobox_filter_group():
    """заполняет комбобокс фильтр групп для таблицы результаты"""
    etap = []
    gr_txt = []
    sender = my_win.menuWidget().sender()
    my_win.comboBox_filter_group.clear()
    my_win.comboBox_filter_choice_stage.clear()

    systems = System.select().where(System.title_id == title_id())  # находит system id последнего
    etap = [i.stage for i in systems ] # все этапы системы

    if etap[0] != "":
        fir_e = "Предварительный"
        flag = fir_e in etap # проверка есть ли в списке этап
        if flag == True:
            sf = systems.select().where(System.stage == fir_e).get()
            kg = int(sf.total_group)  # количество групп
        # if sender == my_win.choice_gr_Action or (my_win.tabWidget.currentIndex() == 2 and my_win.radioButton_group.isChecked()):
        if sender == my_win.choice_gr_Action or (my_win.tabWidget.currentIndex() == 2 and my_win.radioButton_gr_sort.isChecked()):
            gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
            gr_txt.insert(0, "все группы")
            my_win.comboBox_filter_choice_stage.addItems(gr_txt)
        elif my_win.tabWidget.currentIndex() == 3:
            my_win.comboBox_filter_group.addItem("все группы")
            gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
            my_win.comboBox_filter_group.addItems(gr_txt)


def load_combobox_filter_group_semifinal():
    """заполняет комбобокс фильтр групп для таблицы результаты"""
    sf_list = ["-все полуфиналы-"]
    gr_txt = []
    # sender = my_win.menuWidget().sender()
    my_win.comboBox_filter_semifinal.clear()
    # my_win.comboBox_filter_choice.clear()
    my_win.comboBox_filter_group_sf.clear()

    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    systems_sf = system.select().where(System.stage == "1-й полуфинал").get()
    kg = int(systems_sf.total_group)  # количество групп
    system_sf = system.select().where((System.stage == "1-й полуфинал") | (System.stage == "2-й полуфинал"))
    sf_list = [e.stage for e in system_sf]  # получает список этапов на данных соревнованиях
    my_win.comboBox_filter_semifinal.addItems(sf_list)

    my_win.comboBox_filter_group_sf.addItem("все группы")
    # my_win.comboBox_filter_choice.addItem("все группы")
    gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
    my_win.comboBox_filter_group_sf.addItems(gr_txt)
    # my_win.comboBox_filter_choice.addItems(gr_txt)


def load_comboBox_filter_rejting():
    """Загружает комбобоксы вкладки рейтинг"""
    region_rejting = [""]
    city_rejting = [""]
    b_day = ["", "до 12 лет", "до 13 лет", "до 14 лет", "до 16 лет", "до 20 лет", "до 22 лет"]
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]
    id_title = Title.select().where(Title.id == title_id()).get()
    gamer = id_title.gamer
    cur_index = my_win.comboBox_choice_R.currentIndex()

    if cur_index == 0: # если выбран текущий рейтинг
        if gamer == "Девочки" or gamer == "Девушки" or gamer == "Женщины":
            r_data = r_data_w[0]
        else:
            r_data = r_data_m[0] 
    elif cur_index == 1: # если рейтинг за январь
        if gamer == "Девочки" or gamer == "Девушки" or gamer == "Женщины":
            r_data = r_data_w[1]
        else:
           r_data = r_data_m[1]
    player_list = r_data.select()
    for k in player_list:
        region = k.r_region
        city = k.r_city
        if region not in region_rejting:
            region_rejting.append(region)
            region_rejting.sort()
        if city not in city_rejting:
            city_rejting.append(city)
            city_rejting.sort()
    my_win.comboBox_filter_region_in_R.addItems(region_rejting)
    my_win.comboBox_filter_city_in_R.addItems(city_rejting)
    my_win.comboBox_filter_date_in_R.addItems(b_day)


def tab():
    """Изменяет вкладку tabWidget в зависимости от вкладки toolBox"""
    tw = my_win.tabWidget.currentIndex()
    my_win.toolBox.setCurrentIndex(tw)


def tool_page():
    """Изменяет вкладку toolWidget в зависимости от вкладки tabWidget"""
    tb = my_win.toolBox.currentIndex()
    my_win.tabWidget.setCurrentIndex(tb)
    page()


def page():
    """Изменяет вкладку toolBox в зависимости от вкладки tabWidget"""
    msgBox = QMessageBox()
    tb = my_win.toolBox.currentIndex()
    sf = System.select().where(System.title_id == title_id())
    if tb == 0: # -титул-
        my_win.resize(1110, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 275, 841, 496)) # (точка слева, точка сверху, ширина, высота)
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 271))
        my_win.comboBox_referee.setPlaceholderText("Введите фамилию судьи")
        my_win.comboBox_referee.setCurrentIndex(-1)
        my_win.comboBox_referee.setEditable(True)
        my_win.comboBox_secretary.setPlaceholderText("Введите фамилию судьи")
        my_win.comboBox_secretary.setCurrentIndex(-1)
        my_win.comboBox_secretary.setEditable(True)
        db_select_title()
        my_win.tableWidget_GSK.hide()
    elif tb == 1:  # -список участников-
        my_win.resize(1110, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 225, 841, 552))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 221))
        load_comboBox_filter()
        region()
        my_win.Button_app.setEnabled(False)
        my_win.Button_del_player.setEnabled(False)
        my_win.Button_clear_del.setEnabled(False)
        my_win.Button_pay_R.setEnabled(False)
        my_win.Button_add_edit_player.setText("Добавить")
        my_win.statusbar.showMessage("Список участников соревнований", 5000)
        player_list = Player.select().where(Player.title_id == title_id())
        player_debitor_R = Player.select().where((Player.title_id == title_id()) & (Player.pay_rejting == "долг"))
        count_debitor_R = len(player_debitor_R)
        num_debitor_1 = [1]
        num_debitor_2 = [2, 3, 4]
        if count_debitor_R in num_debitor_1:
            end_word = "участник"
        elif count_debitor_R in num_debitor_2:
            end_word = "участника"
        else:
            end_word = "участников"
        fill_table(player_list)  # заполняет TableWidget списком игроков
        count = len(player_list)
        my_win.label_46.setText(f"Всего: {count} участников")
        my_win.label_dolg_R.setText(f"Без лицензии: {count_debitor_R} {end_word}")
        list_player_pdf(player_list)
    elif tb == 2:  # -система-
        my_win.resize(1110, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 308, 841, 471))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 301))
        my_win.tableView.setEditTriggers(QAbstractItemView.NoEditTriggers) # запрет редактирования таблицы
        result = Result.select().where(Result.title_id == title_id())
        result_played = result.select().where(Result.winner != "")
        count_result = len(result_played)
        player_list = Player.select().where(Player.title_id == title_id())
        count = len(player_list)
        my_win.label_8.setText(f"Всего участников: {str(count)} человек")
        my_win.label_52.setText(f"Всего сыграно: {count_result} игр.")
        label_playing_count() # пишет сколько игр сыграно в каждо этапе
        my_win.comboBox_filter_number_group_final.setEnabled(False)
        for k in sf:
            stage = k.stage
            if stage == "Предварительный":
                flag_choice = ready_choice(stage)
                if flag_choice is True:
                    choice_filter_on_system()
                break
        my_win.label_etap_1.hide()
        my_win.label_etap_2.hide()
        my_win.label_etap_3.hide()
        my_win.label_etap_4.hide()
        my_win.label_etap_5.hide()
        my_win.label_etap_6.hide()
        my_win.label_etap_7.hide()
        my_win.label_etap_8.hide()
        my_win.label_etap_9.hide()
        my_win.label_etap_10.hide()
        my_win.label_etap_11.hide()
        my_win.label_101.hide()
        my_win.label_11.hide()
        my_win.label_12.hide()
        my_win.label_19.hide()
        my_win.label_102.hide()
        my_win.label_27.hide()
        my_win.label_30.hide()
        my_win.label_31.hide()
        my_win.label_103.hide()
        my_win.label_104.hide()
        my_win.label_105.hide()
        my_win.label_53.hide()       
        my_win.label_58.hide()
        my_win.label_106.hide()
        my_win.label_107.hide()
        my_win.label_108.hide()
        my_win.label_109.hide()
        my_win.label_110.hide()
        my_win.label_111.hide()
        my_win.label_81.hide()
        my_win.label_82.hide()
        my_win.label_83.hide()
        my_win.label_84.hide()
        my_win.label_85.hide()
        my_win.label_86.hide()

        my_win.comboBox_table_1.hide()
        my_win.comboBox_table_2.hide()
        my_win.comboBox_table_3.hide()
        my_win.comboBox_table_4.hide()
        my_win.comboBox_table_5.hide()
        my_win.comboBox_table_6.hide()
        my_win.comboBox_table_7.hide()
        my_win.comboBox_table_8.hide()
        my_win.comboBox_table_9.hide()
        my_win.comboBox_table_10.hide()
        my_win.comboBox_table_11.hide()

        my_win.spinBox_kol_group.hide()
        # my_win.groupBox_SF.setVisible(False)
        stage = []
        table = []
        game = []
        sum_game = []

        for i in sf:  # цикл по таблице -system-
            total_player = i.total_athletes
            if total_player == 0: # если система только начала создаваться
                return
            stage.append(i.stage)  # добавляет в список этап
            table.append(i.label_string)  # добавляет в список система
            game.append(i.kol_game_string)  # добавляет в список кол-во игр
        count = len(stage)
        for i in range(0, count):  # подсчитывает сумму игр
            txt = game[i]
            t = txt.find(" ")
            txt = int(txt[0:t])
            sum_game.append(txt)
            if i == 0:  # показывает в зависимости от этапов финал, кол-во игр
                my_win.label_101.setText(stage[0])
                my_win.label_19.setText(game[0])
                my_win.label_12.setText(table[0])
                my_win.label_101.show()
                my_win.label_12.show()
                my_win.label_19.show()
            elif i == 1:
                my_win.label_102.setText(stage[1])
                my_win.label_27.setText(game[1])
                my_win.label_etap_2.setText(table[1])
                my_win.label_102.show()
                my_win.label_27.show()
                my_win.label_etap_2.show()
            elif i == 2:
                my_win.label_103.setText(stage[2])
                my_win.label_30.setText(game[2])
                my_win.label_31.setText(table[2])
                my_win.label_30.show()
                my_win.label_31.show()
                my_win.label_103.show()
            elif i == 3:
                my_win.label_104.setText(stage[3])
                my_win.label_53.setText(game[3])
                my_win.label_etap_4.setText(table[3])
                my_win.label_104.show()
                my_win.label_53.show()
                my_win.label_etap_4.show()
            elif i == 4:
                my_win.label_105.setText(stage[4])
                my_win.label_58.setText(game[4])
                my_win.label_etap_5.setText(table[4])
                my_win.label_105.show()
                my_win.label_58.show()
                my_win.label_etap_5.show()
            elif i == 5:
                my_win.label_106.setText(stage[5])
                my_win.label_81.setText(game[5])
                my_win.label_etap_6.setText(table[5])
                my_win.label_106.show()
                my_win.label_81.show()
                my_win.label_etap_6.show()
            elif i == 6:
                my_win.label_107.setText(stage[6])
                my_win.label_82.setText(game[6])
                my_win.label_etap_7.setText(table[6])
                my_win.label_107.show()
                my_win.label_82.show()
                my_win.label_etap_7.show()
            elif i == 7:
                my_win.label_108.setText(stage[7])
                my_win.label_83.setText(game[7])
                my_win.label_etap_8.setText(table[7])
                my_win.label_108.show()
                my_win.label_83.show()
                my_win.label_etap_8.show()
            elif i == 8:
                my_win.label_109.setText(stage[8])
                my_win.label_84.setText(game[8])
                my_win.label_etap_9.setText(table[8])
                my_win.label_109.show()
                my_win.label_84.show()
                my_win.label_etap_9.show()
            elif i == 9:
                my_win.label_110.setText(stage[9])
                my_win.label_85.setText(game[9])
                my_win.label_etap_10.setText(table[9])
                my_win.label_110.show()
                my_win.label_85.show()
                my_win.label_etap_10.show()
            elif i == 10:
                my_win.label_111.setText(stage[10])
                my_win.label_86.setText(game[10])
                my_win.label_etap_11.setText(table[10])
                my_win.label_111.show()
                my_win.label_86.show()
                my_win.label_etap_11.show()

            total_game = sum(sum_game)
            my_win.comboBox_table_1.hide()
            my_win.comboBox_page_vid.setEnabled(False)
            my_win.Button_etap_made.setEnabled(False)
            my_win.Button_system_made.setEnabled(False)
            my_win.label_33.setText(f"Всего {total_game} игр")
            my_win.label_33.show()
            # сделать правильную сортировку по группам
        if my_win.radioButton_gr_sort.isChecked():
            player_list = Choice.select().where(Choice.title_id == title_id()).order_by(Choice.mesto_group, Choice.group)
        elif my_win.radioButton_sf_sort.isChecked():
            player_list = Choice.select().where(Choice.title_id == title_id()).order_by(Choice.mesto_semifinal_group, Choice.sf_group)
        fill_table(player_list)
    elif tb == 3:  # вкладка -группы-
        stage = "Предварительный"
        Button_view_group = QPushButton(my_win.tabWidget) # (в каком виджете размещена)
        Button_view_group.resize(120, 50) # размеры кнопки (длина 120, ширина 50)
        Button_view_group.move(850, 60) # разммещение кнопки (от левого края 850, от верхнего 60) от виджета в котором размещен
        Button_view_group.setText("Просмотр групп")
        Button_view_group.show()
        Button_view_group.clicked.connect(view)

        my_win.resize(1270, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 150, 1000, 626))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 1000, 147))

        system_stage = sf.select().where(System.stage == "Предварительный").get()
        game_visible = system_stage.visible_game
        my_win.checkBox_4.setChecked(game_visible)
        my_win.checkBox_7.setEnabled(False)
        my_win.checkBox_8.setEnabled(False)
        my_win.checkBox_7.setChecked(False)
        my_win.checkBox_8.setChecked(False)

        my_win.Button_Ok_gr.setEnabled(False)
        player_list = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == "Предварительный"))
        fill_table(player_list)
        load_combobox_filter_group()
        load_combo()
        visible_field()
        my_win.label_16.hide()
        my_win.tableView_net.hide() # сетка ручной жеребьевки на 32
    elif tb == 4:  # вкладка -полуфиналы-
        my_win.resize(1270, 825)
        Button_view_semifinal = QPushButton(my_win.tabWidget) # (в каком виджете размещена)
        Button_view_semifinal.resize(120, 50) # размеры кнопки (длина 120, ширина 50)
        Button_view_semifinal.move(850, 60) # разммещение кнопки (от левого края 850, от верхнего 60) от виджета в котором размещен
        Button_view_semifinal.setText("Просмотр полуфиналов")
        Button_view_semifinal.show()
        my_win.tableView.setGeometry(QtCore.QRect(260, 150, 1000, 626))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 1000, 147))
        system_stage = sf.select().where((System.stage == "1-й полуфинал") | (System.stage == "2-й полуфинал")).get()
        game_visible = system_stage.visible_game
        my_win.checkBox_4.setChecked(game_visible)
        my_win.checkBox_7.setEnabled(False)
        my_win.checkBox_8.setEnabled(False)
        my_win.checkBox_7.setChecked(False)
        my_win.checkBox_8.setChecked(False)
        flag = ready_choice(stage="1-й полуфинал")
        if flag is False:
            result = msgBox.information(my_win, "", "Необходимо сделать жеребьевку\nполуфинального этапа.",
                                        msgBox.Ok, msgBox.Cancel)
            if result != msgBox.Ok:
                return
            else:
                my_win.tabWidget.setCurrentIndex(2)
                choice_gr_automat()
                sf.choice_flag = True
                sf.save()
            my_win.tabWidget.setCurrentIndex(3)
        else:  # жеребьевка сделана
            my_win.Button_Ok_pf.setEnabled(False)
            player_list = Result.select().where((Result.system_stage == "1-й полуфинал") | (Result.system_stage == "2-й полуфинал"))
            fill_table(player_list)
            load_combobox_filter_group_semifinal()
            load_combo()
            visible_field()
            my_win.label_16.hide()
    elif tb == 5: # вкладка -финалы-
        my_win.resize(1270, 825)
        Button_view_final = QPushButton(my_win.tabWidget) # (в каком виджете размещена)
        Button_view_final.resize(120, 50) # размеры кнопки (длина 120, ширина 50)
        Button_view_final.move(850, 60) # разммещение кнопки (от левого края 850, от верхнего 60) от виджета в котором размещен
        Button_view_final.setText("Просмотр финалов")
        Button_view_final.show()
        my_win.resize(1270, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 150, 1000, 626))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 1000, 147))
        my_win.checkBox_5.setEnabled(False)
        my_win.checkBox_9.setChecked(False)
        my_win.checkBox_10.setChecked(False)
        my_win.checkBox_9.setEnabled(False)
        my_win.checkBox_10.setEnabled(False)
        my_win.tableView.show()
        my_win.Button_Ok_fin.setEnabled(False)
        my_win.groupBox_kolvo_vstrech_fin.setEnabled(False)
        load_combobox_filter_final()
        player_list = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == "Финальный"))
        fill_table(player_list)
        load_combo()
        visible_field()
        my_win.label_16.hide()
    elif tb == 6: # вкладка -рейтинг-
        my_win.resize(1110, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 75, 841, 702))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 71))
        my_win.comboBox_choice_R.clear()
        my_win.comboBox_filter_date_in_R.clear()
        rejting_month = ["За текуший месяц", "За январь месяц"]
        my_win.comboBox_choice_R.addItems(rejting_month)
        load_comboBox_filter_rejting()
    elif tb == 7: # вкладка -дополнительно-
        my_win.resize(1110, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 250, 841, 525))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 248))
        my_win.Button_made_page_pdf.setEnabled(False)
        my_win.Button_up.setEnabled(False)
        my_win.Button_down.setEnabled(False)
        my_win.Button_made_one_file_pdf.setEnabled(False)
        my_win.Button_print_begunki.setEnabled(False)
        my_win.Button_change_player.setEnabled(False)
        my_win.lineEdit_range_tours.hide()
        my_win.comboBox_first_group.setEnabled(False)
        my_win.comboBox_second_group.setEnabled(False)
        load_combo_etap_begunki()
        # ======
    hide_show_columns(tb)


def label_playing_count():
    """На вкладке -система- пишет сколько игр сыграно в каждом этапе"""
    result = Result.select().where(Result.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
    n = 0
    my_win.label_playing_etap1.hide()
    my_win.label_playing_etap2.hide()
    my_win.label_playing_etap3.hide()
    my_win.label_playing_etap4.hide()
    my_win.label_playing_etap5.hide()
    my_win.label_playing_etap6.hide()
    my_win.label_playing_etap7.hide()
    my_win.label_playing_etap8.hide()
    my_win.label_playing_etap9.hide()
    my_win.label_playing_etap10.hide()
    my_win.label_playing_etap11.hide()
    for k in system:
        n += 1
        system_id = k.id
        result_playing = result.select().where((Result.system_id == system_id) & (Result.winner != ""))
        count_playing = len(result_playing)
        if n == 1:
            my_win.label_playing_etap1.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap1.show()
        elif n == 2:
            my_win.label_playing_etap2.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap2.show()
        elif n == 3:
            my_win.label_playing_etap3.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap3.show()
        elif n == 4:
            my_win.label_playing_etap4.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap4.show()
        elif n == 5:
            my_win.label_playing_etap5.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap5.show()
        elif n == 6:
            my_win.label_playing_etap6.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap6.show()
        elif n == 7:
            my_win.label_playing_etap7.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap7.show()
        elif n == 8:
            my_win.label_playing_etap8.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap8.show()
        elif n == 9:
            my_win.label_playing_etap9.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap9.show()
        elif n == 10:
            my_win.label_playing_etap10.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap10.show()
        elif n == 11:
            my_win.label_playing_etap11.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap11.show()


def add_city():
    """добавляет в таблицу город и соответсвующий ему регион"""
    city_field = my_win.lineEdit_city_list.text()
    city_field = city_field.capitalize()  # Переводит первую букву в заглавную
    index = city_field.find(".")
    if index != -1:
        second_word = city_field[index + 1:]
        second_word = second_word.capitalize()
        city_field = city_field[:index + 1] + second_word
    my_win.lineEdit_city_list.setText(city_field)
    my_win.textEdit.setText("Выберите регион в котором находится населенный пункт.")


def find_coach():
    """поиск тренера в базе"""
    my_win.label_63.setText("Список тренеров.")
    my_win.listWidget.clear()
    list_coach = []
    cp = my_win.lineEdit_coach.text()
    cp = cp.capitalize()  # Переводит первую букву в заглавную
    if my_win.checkBox_find_player.isChecked():
        player = Player.select().where(Player.title_id == title_id())
        coach_list = Coach.select().where(Coach.coach ** f'%{cp}%')  # создает выборку из базы тренеров фамилии,что начинаются на CP
        for pl in coach_list: #походит циклом и создает список с их ID
            c_id = pl.id
            list_coach.append(c_id)

        player_list = player.select().where(Player.coach_id << list_coach) # окончательная выборка со всеми тренерами (id)
        # fill_table(player_list)
    else:
        c = Coach.select()
        c = c.where(Coach.coach ** f'{cp}%')  # like
        tochka = cp.find(".")
        if tochka == -1:
            if (len(c)) != 0:
                for chp in c:
                    full_stroka = chp.coach
                    my_win.listWidget.addItem(full_stroka)


def add_coach(ch, num):
    """Проверяет наличие тренера в базе и если нет, то добавляет"""
    coach = Coach.select()
    count_coach = len(coach)
    if count_coach == 0:  # если первая запись то добавляет без проверки
        with db:
            cch = Coach(coach=ch, player_id=num).save()
        return
    for c in coach:
        coa = Coach.select().where(Coach.coach == ch)
        if bool(coa):
            return
        else:
            cch = Coach(coach=ch, player_id=num).save()


def find_player():
    """Установка курсора в строку поиска спортсмена в загруженном списке"""
    if my_win.checkBox_find_player.isChecked():
        my_win.lineEdit_Family_name.setFocus()
    else:
        pass


def find_player_on_tab_system():
    """выделяет строку в tablewidget при поиске фамилии на вкладке -система_"""
    choice = Choice.select().where(Choice.title_id == title_id())
    txt = my_win.lineEdit_find_player_in_system.text()
    txt = txt.upper()
    player_list = choice.select().where(Choice.family ** f'{txt}%')  # like поиск в текущем рейтинге
    count = len(player_list)
    if count == 1:
        pass
    fill_table(player_list)       


def sort():
    """сортировка таблицы QtableView (по рейтингу или по алфавиту)"""
    sender = my_win.sender()  # сигнал от кнопки
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]
    signal_button_list = [my_win.Button_sort_R, my_win.Button_sort_Name, my_win.Button_sort_mesto]
    id_title = Title.select().where(Title.id == title_id()).get()
    gamer = id_title.gamer
    cur_index = my_win.comboBox_choice_R.currentIndex()
    # if cur_index == 0: # если выбран текущий рейтинг
    #         if gamer == "Девочки" or gamer == "Девушки" or gamer == "Женщины":
    #             r_data = r_data_w[0]
    #         else:
    #             r_data = r_data_m[0]
    #         rejting_name = r_data.r_fname
    #         rejting_list = r_data.r_list
    # elif cur_index == 1: # если рейтинг за январь
    #         if gamer == "Девочки" or gamer == "Девушки" or gamer == "Женщины":
    #             r_data = r_data_w[1]
    #         else:
    #             r_data = r_data_m[1] 
    #         rejting_name = r_data.r1_fname
    #         rejting_list = r_data.r1_list
    if sender == my_win.Button_sort_R:  # в зависимости от сигала кнопки идет сортировка
        player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.rank.desc())  # сортировка по рейтингу
    elif sender == my_win.Button_sort_Name:
        player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.player)  # сортировка по алфавиту
    elif sender == my_win.Button_sort_mesto:
        player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.mesto)  # сортировка по месту
    # elif sender == my_win.Button_sort_alf_R:
    #     player_list = r_data.select().order_by(rejting_name)
    # elif sender == my_win.Button_sort_rejting_in_R:
    #     player_list = r_data.select().order_by(rejting_list.desc())

    fill_table(player_list)
    if sender in signal_button_list:
        list_player_pdf(player_list)


def button_title_made_enable(state):
    """включает кнопку - создание титула - если отмечен чекбокс, защита от случайного нажатия"""
    if state == 2:  # если флажок установлен
        title_str = title_string()
        nm = title_str[0]
        ds = title_str[3]
        de = title_str[4]
        # получение последней записи в таблице
        t = Title.select().order_by(Title.id.desc()).get()
        if t.name == nm and str(t.data_start) == ds and str(t.data_end) == de:
            my_win.Button_title_made.setText("Редактировать")
        else:
            my_win.Button_title_made.setText("Создать")
        my_win.Button_title_made.setEnabled(True)
    else:
        my_win.Button_title_made.setEnabled(False)


def button_system_made_enable(state):
    """включает кнопку - создание системы - если отмечен чекбокс, защита от случайного нажатия"""
    if state == 2:
        my_win.Button_system_made.setEnabled(True)


def list_player_pdf(player_list):
    """создание списка участников в pdf файл"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp
    gamer = tit.gamer
    count = len(player_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in player_list:
        n += 1
        p = l.player
        b = l.bday
        r = l.rank
        c = l.city
        g = l.region
        z = l.razryad
        coach_id = l.coach_id
        t = coach_id.coach
        m = l.mesto
        chop_line(t)
        data = [n, p, b, r, c, g, z, t, m]

        elements.append(data)
    elements.insert(0, ["№", "Фамилия, Имя", "Дата рожд.", "Рейтинг", "Город", "Регион", "Разряд", "Тренер(ы)",
                        "Место"])
    t = Table(elements,
              colWidths=(0.8 * cm, 3.9 * cm, 1.7 * cm, 1.2 * cm, 2.5 * cm, 3.1 * cm, 1.2 * cm, 4.8 * cm, 1.0 * cm),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список участников. {gamer}', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_player_list.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def exit_comp():
    """нажата кнопка -выход-"""
    msgBox = QMessageBox
    result = msgBox.question(my_win, "Выход из программы", "Вы действительно хотите выйти из программы?",
                             msgBox.Ok, msgBox.Cancel)
    if result == msgBox.Ok:
        my_win.close()
        backup()
    else:
        pass


def system_competition():
    """выбор системы проведения при изменении строки в комбобокс этап или мз меню"""
    msgBox = QMessageBox
    sender = my_win.sender()
    tit = Title.get(Title.id == title_id())
    gamer = tit.gamer
    flag_system = ready_system() # False система еще не создана 
    if sender != my_win.comboBox_etap:
        if sender == my_win.system_edit_Action: # редактирование системы из меню
            sb = "Изменение системы проведения соревнования."
            my_win.statusbar.showMessage(sb)
            result = msgBox.question(my_win, "", "Вы хотите изменить систему соревнований?",
                                    msgBox.Ok, msgBox.Cancel)
            if result == msgBox.Ok:
                # очищает таблицы перед новой системой соревнования (system, choice)
                clear_db_before_edit()
                tab_enabled(gamer)  # показывает вкладки по новому
                choice_tbl_made()  # заполняет db жеребьевка
                flag_system = False # ставит флаг, что система еще не создана
                stage = ""
            else:
                return
        elif sender == my_win.system_made_Action: # создание системы из меню
            sb = "Создание системы проведения соревнования."
            my_win.statusbar.showMessage(sb)
            result = msgBox.question(my_win, "", "Вы хотите создать систему соревнований?",
                                msgBox.Ok, msgBox.Cancel)
            if result == msgBox.Ok:
                choice_tbl_made()  # заполняет db жеребьевка
            else:
                return
        my_win.spinBox_kol_group.hide()
        my_win.comboBox_etap.setEnabled(True)
        my_win.label_102.hide()
        my_win.label_27.hide()
        my_win.label_etap_2.hide()
        my_win.label_etap_3.hide()
        my_win.label_etap_4.hide()
        my_win.label_etap_5.hide() 
        my_win.label_etap_6.hide()
        my_win.label_etap_7.hide()
        my_win.label_etap_8.hide() 
        my_win.label_etap_9.hide()
        my_win.label_etap_10.hide()
        my_win.label_etap_11.hide()
        my_win.label_30.hide()
        my_win.label_31.hide()
        my_win.label_103.hide()
        my_win.label_104.hide()
        my_win.label_105.hide()
        my_win.label_106.hide()
        my_win.label_107.hide()
        my_win.label_108.hide()
        my_win.label_109.hide()
        my_win.label_110.hide()
        my_win.label_111.hide()
        my_win.label_53.hide()
        my_win.label_58.hide() 
        my_win.label_81.hide()
        my_win.label_82.hide()
        my_win.label_83.hide()  
        my_win.label_84.hide()
        my_win.label_85.hide()
        my_win.label_86.hide()
 
        my_win.tabWidget.setTabEnabled(2, True)

        if flag_system is True:
            flag_choice = ready_choice(stage)
            if flag_choice is True:
                sb = "Система и жербьевка создана."
            elif flag_choice is False:
                sb = "Система создана, теперь необходимо произвести жеребьевку. " \
                    "Войдите в меню -соревнования- и выберите -жеребьевка-"
            my_win.statusbar.showMessage(sb)
        elif flag_system is False:
            sb = "Выбор системы проведения соревнования."
            my_win.statusbar.showMessage(sb)
# ================
            my_win.spinBox_kol_group.hide()
            my_win.comboBox_etap.clear()
            real_list = ["-выбор этапа-", "Одна таблица", "Предварительный"] # который нужен в комбобокс
            combobox_etap_compare(real_list)
            my_win.comboBox_etap.show()
            my_win.comboBox_table_1.hide()
            my_win.label_10.show()
            my_win.label_10.setText("1-й этап")
            my_win.Button_etap_made.setEnabled(False)
            my_win.comboBox_page_vid.setEnabled(True)
# =======
            player = Player.select().where(Player.title_id == title_id())
            count = len(player)
            if count != 0:
                my_win.tabWidget.setCurrentIndex(2)
            else:
                reply = QMessageBox.information(my_win, 'Уведомление',
                                                "У Вас нет ни одного спортсмена.\nСначала необходимо создать "
                                                "список участников соревнований.\n Перейти к созданию списка?",
                                                msgBox.Ok,
                                                msgBox.Cancel)
                if reply == msgBox.Ok:
                    my_win.tabWidget.setCurrentIndex(1)
                    my_win.lineEdit_Family_name.setFocus()
                else:
                    return        


def one_table(fin, group):
    """система соревнований из одной таблицы запись в System, Game_list, Result"""
    msgBox = QMessageBox()
    system = System.select().where(System.title_id == title_id())
    ch = Choice.select().where(Choice.title_id == title_id())
    count = len(Player.select().where(Player.title_id == title_id()))
    # в зависмости сетка или круг
    cur_index = my_win.comboBox_table.currentIndex()
    if fin == "Одна таблица":
        if cur_index == 1:
            vt = "Сетка (-2) на"
            my_win.comboBox_page_vid.setCurrentText("книжная")
            type_table = "сетка"
        elif cur_index == 2:
            vt = "Сетка (с розыгрышем всех мест) на"
            my_win.comboBox_page_vid.setCurrentText("книжная")
            type_table = "сетка"
        elif cur_index == 3:
            vt = "Сетка (с играми за 1-3 места) на"
            my_win.comboBox_page_vid.setCurrentText("книжная")
            type_table = "сетка"
        elif cur_index == 4:
            vt = "Круговая таблица на"
            my_win.comboBox_page_vid.setCurrentText("альбомная")
            type_table = "круг"

        if type_table == "круг":
            total_athletes = count
        else: # на сколько участников таблица
            if count <= 16:
                total_athletes = 16
            elif count > 16 and count <= 32:
                total_athletes = 32

        flag_ready_system = ready_system()
        if flag_ready_system is False:
            sys_m = System.select().where(System.title_id == title_id()).get()
            total_game = numbers_of_games(cur_index, player_in_final=count)

            sys_m.max_player = count
            sys_m.total_athletes = total_athletes
            sys_m.total_group = group
            sys_m.stage = my_win.comboBox_etap.currentText()
            sys_m.type_table = type_table
            sys_m.page_vid = my_win.comboBox_page_vid.currentText()
            sys_m.label_string = f"{vt} {total_athletes} участников"
            sys_m.kol_game_string =f"{total_game} игр"
            sys_m.save()

            my_win.Button_etap_made.setEnabled(False)
            my_win.comboBox_page_vid.setEnabled(False)

            for k in ch: # записывает в DB после создании системы из одной таблицы basic - Одна таблица
                k.basic = fin
                k.save()
            add_open_tab(tab_page="Система")

            result = msgBox.question(my_win, "", "Система соревнований создана.\n"
                                                 "Теперь необходимо сделать жеребъевку\n"
                                                 "Хотите ее сделать сейчас?",
                                     msgBox.Ok, msgBox.Cancel)
            if result == msgBox.Ok:
                if type_table == "круг":  # функция жеребьевки таблицы по кругу
                    player_in_one_table(fin)
                else:
                    posev_data = player_choice_in_setka(fin)
                    player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                add_open_tab(tab_page="Финалы")
            else:
                return
        sys_m.stage = fin
        sys_m.choice_flag = 1 # запись о том что сделана жеребьевка
        sys_m.save()


def selection_of_the_draw_mode():
    """Выбор режима жеребьевки сетки -автомат- или -ручной-"""
    vid = ["Автоматическая", "Ручная"]
    vid, ok = QInputDialog.getItem(
                    my_win, "Жеребьевка", "Выберите режим жеребьевки сетки.", vid, 0, False)
    if vid == "Автоматическая":
        flag = True
        my_win.tableView_net.hide()
    else:
        flag = False
        my_win.resize(1440, 804)
        my_win.tableView_net.show()
        my_win.tableView_net.setGeometry(QtCore.QRect(1110, 9, 321, 749)) # от лев края, от вверха, ширина и высота)
    return flag
    
              
def kol_player_in_group():
    """подсчет кол-во групп и человек в группах"""
    sender = my_win.sender()  # сигнал от кнопки
    flag_visible = my_win.checkBox_visible_game.isChecked()
    kg = my_win.spinBox_kol_group.text()  # количество групп
    player_list = Player.select().where(Player.title_id == title_id())
    type_table = "группы"
    count = len(player_list)  # количество записей в базе
    # остаток отделения, если 0, то участники равно делится на группы
    e1 = count % int(kg)
    # если количество участников равно делится на группы (кол-во групп)
    p = count // int(kg)
    g1 = int(kg) - e1  # кол-во групп, где наименьшее кол-во спортсменов
    g2 = int(p + 1)  # кол-во человек в группе с наибольшим их количеством
    if e1 == 0:  # то в группах равное количество человек -e1-
        stroka_kol_group = f"{kg} группы по {str(p)} чел."
        skg = int((p * (p - 1) / 2) * int(kg))
        mp = p
    else:
        stroka_kol_group = f"{str(g1)} групп(а) по {str(p)} чел. и {str(e1)} групп(а) по {str(g2)} чел."
        p = int(p)
        skg = int((((p * (p - 1)) / 2 * g1) + ((g2 * (g2 - 1)) / 2 * e1))) # общее количество игр в группах
        mp = g2
    stroka_kol_game = f"{skg} игр"
    my_win.label_11.hide()
    my_win.label_12.setText(stroka_kol_group)
    my_win.label_12.show()
    my_win.label_19.setText(stroka_kol_game)
    my_win.label_19.show()
    my_win.Button_etap_made.setEnabled(True)
    if sender == my_win.Button_etap_made:
        my_win.Button_etap_made.setEnabled(False)
        my_win.comboBox_page_vid.setEnabled(False)
        my_win.spinBox_kol_group.hide()
        # ====== запись в таблицу db -system- первый этап
        s = System.select().where(System.title_id == title_id()).get()
        system = System.get(System.id == s)
        system.max_player = mp
        system.total_athletes = count
        system.total_group = kg
        system.stage = my_win.comboBox_etap.currentText()
        system.type_table = type_table
        system.page_vid = my_win.comboBox_page_vid.currentText()
        system.label_string = stroka_kol_group
        system.kol_game_string = stroka_kol_game
        system.visible_game = flag_visible
        system.save()
    load_combobox_filter_group()


def page_vid():
    """присваивает переменной значение выборат вида страницы"""
    if my_win.comboBox_page_vid.currentText() == "альбомная":
        pv = landscape(A4)
    else:
        pv = A4
    return pv


def view():
    """просмотр PDF файлов средствами OS"""
    from sys import platform
    sender = my_win.sender()
    made_pdf_table_for_view(sender)
    t_id = Title.get(Title.id == title_id())
    tab = my_win.tabWidget.currentIndex()
    short_name = t_id.short_name_comp

    if sender == my_win.view_all_comp_Action: # просмотр полных соревнований в каталоге /competition_pdf
        catalog = 2
        change_dir(catalog)
        view_file = f"{short_name}.pdf"
    else: # просмотр отдельных страниц в каталоге /table_pdf
        catalog = 1
        change_dir(catalog)
        if sender == my_win.view_list_Action:
            view_sort = ["По алфавиту", "По рейтингу", "По месту"]
            view_sort, ok = QInputDialog.getItem(
                        my_win, "Сортировка", "Выберите вид сортировки,\n просмотра списка участников.", view_sort, 0, False)
            if view_sort == "По рейтингу":
                player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.rank.desc())  # сортировка по рейтингу
            elif view_sort == "По алфавиту": 
                player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.player) # сортировка по алфавиту
            elif view_sort == "По месту":
                player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.mesto)  # сортировка по месту
            list_player_pdf(player_list)
            change_dir(catalog)
            view_file =  f"{short_name}_player_list.pdf"
        elif sender == my_win.view_referee_list_Action:
            view_file =  f"{short_name}_referee_list.pdf"
        elif sender == my_win.view_regions_list_Action:
            view_file =  f"{short_name}_regions_list.pdf"
        elif sender == my_win.view_winners_list_Action:
            view_file =  f"{short_name}_winners_list.pdf"    
        elif sender == my_win.view_title_Action:
            view_file = f"{short_name}_title.pdf"
        elif sender == my_win.view_gr_Action or tab == 3:  # вкладка группы
            view_file = f"{short_name}_table_group.pdf"
        elif sender == my_win.view_fin1_Action:
            view_file = f"{short_name}_1-final.pdf"
        elif sender == my_win.view_fin2_Action:
            view_file = f"{short_name}_2-final.pdf"
        elif sender == my_win.view_fin3_Action:
            view_file = f"{short_name}_3-final.pdf"
        elif sender == my_win.view_fin4_Action:
            view_file = f"{short_name}_4-final.pdf"
        elif sender == my_win.view_fin5_Action:
            view_file = f"{short_name}_5-final.pdf"
        elif sender == my_win.view_fin6_Action:
            view_file = f"{short_name}_6-final.pdf"
        elif sender == my_win.view_fin7_Action:
            view_file = f"{short_name}_7-final.pdf"
        elif sender == my_win.view_fin8_Action:
            view_file = f"{short_name}_8-final.pdf"
        elif sender == my_win.view_one_table_Action:
            view_file = f"{short_name}_one_table.pdf"
        elif sender == my_win.view_pf1_Action:
            view_file = f"{short_name}_1-semifinal.pdf"
        elif sender == my_win.view_pf2_Action:
            view_file = f"{short_name}_2-semifinal.pdf"
        elif sender == my_win.clear_s32_Action:
            view_file = "clear_32_net.pdf"
        elif sender == my_win.clear_s16_Action:
            view_file = "clear_16_full_net.pdf"
        elif sender == my_win.clear_s32_full_Action:
            view_file = "clear_32_full_net.pdf"
        elif sender == my_win.clear_s32_2_Action:
            view_file = "clear_32_2_net.pdf"
        elif sender == my_win.clear_s16_2_Action:
            view_file = "clear_16_2_net.pdf"   
        elif sender == my_win.clear_s8_2_Action:
            view_file = "clear_8_2_net.pdf"
        elif sender == my_win.clear_s8_full_Action:
            view_file = "clear_8_full_net.pdf"

    if platform == "linux" or platform == "linux2":  # linux
        pass
    elif platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    os.chdir("..")


def player_in_setka_and_write_Game_list_and_Result(fin, posev_data):
    """заполняет таблицу Game_list данными спортсменами из сетки tds - список списков данных из сетки, а затем
    заполняет таблицу -Result-"""
    system = System.select().where((System.title_id == title_id()) & (System.stage == fin)).get()  # находит system id последнего
    st = "Финальный"
    game = 0
    if fin == "Одна таблица":
        st = "Одна таблица"
    id_system = system.id
    system_table = system.label_string
    mp = system.max_player

    if system_table == "Сетка (с розыгрышем всех мест) на 8 участников":
        game = 12
    elif system_table == "Сетка (-2) на 8 участников":
        game = 14
    elif system_table == "Сетка (с розыгрышем всех мест) на 16 участников":
        game = 32
    elif system_table == "Сетка (-2) на 16 участников":
        game = 38
    elif system_table == "Сетка (с розыгрышем всех мест) на 32 участников":
        game = 80
    elif system_table == "Сетка (-2) на 32 участников":
        game = 94
    elif system_table == "Сетка (1-3 место) на 32 участников":
        game = 32
    # создание сетки со спортсменами согласно жеребьевки
    all_list = setka_data(fin, posev_data)
    tds = all_list[1]
    tds_full_name_city = all_list[3]
    k = 0
    for r in tds:
        if r != "X":
            znak = r.find("/")
            family = r[:znak]
            id_pl = all_list[2][family]
            family_id = f'{family}/{id_pl}'  # фамилия игрока и его id
        else:
            family_id = r
        k += 1
    # записывает в Game_List спортсменов участников сетки и присваивает встречи 1-ого тура и записывает в тбл Results
        with db:
            game_list = Game_list(number_group=fin, rank_num_player=k, player_group=family_id,
                                  system_id=id_system, title_id=title_id()).save()

    for i in range(1, mp // 2 + 1):  # присваивает встречи 1-ого тура и записывает в тбл Results
        num_game = i
        pl1 = tds_full_name_city[i * 2 - 2]
        pl2 = tds_full_name_city[i * 2 - 1]
        if pl1 is not None and pl2 is not None:
            with db:
                results = Result(number_group=fin, system_stage=st, player1=pl1, player2=pl2,
                                 tours=num_game, title_id=title_id(), system_id=id_system).save()
    for i in range(mp // 2 + 1, game + 1):  # дополняет номера будущих встреч
        pl1 = ""
        pl2 = ""
        with db:
            results = Result(number_group=fin, system_stage=st, player1=pl1, player2=pl2,
                             tours=i, title_id=title_id(),system_id=id_system).save()


def player_in_one_table(fin):
    """Соревнования из одной таблицы, создание и заполнение Game_list, Result (создание жеребьевки в круг)"""
    one_table = []
    players = Player.select().where(Player.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
    sys_id = system.select().where(System.stage == fin).get()
    system_id = sys_id.id
    k = 0
    for p in choice:  # цикл заполнения db таблиц -game list-
        k += 1
        player = p.family
        pl_id = p.player_choice_id
        pl_city = players.select().where(Player.id == pl_id).get()
        city = pl_city.city
        player_id = f"{player}/{pl_id}"
        one_table.append(f"{player}/{city}")
        game_list = Game_list(number_group=fin, rank_num_player=k, player_group=player_id, system_id=system_id,
                            title_id=title_id())
        game_list.save()

    tours = tours_list(k - 3)
    round = 0
    for tour in tours: # цикл заполнения db таблиц -Result-
        round += 1
        for match in tour:
            znak = match.find("-")
            first = int(match[:znak])  # игрок под номером в группе
            second = int(match[znak + 1:])  # игрок под номером в группе
            pl1 = one_table[first - 1]
            pl2 = one_table[second - 1]
            results = Result(number_group=fin, system_stage="Одна таблица", player1=pl1, player2=pl2,
                             tours=match, title_id=title_id(), round=round)
            results.save()    


def player_fin_on_circle(fin):
    """заполняет таблицу Game_list данными спортсменами из группы, которые будут играть в финале по кругу
     td - список списков данных из групп"""
    fin_dict = {}
    fin_list = []
 
    players = Player.select().where(Player.title_id == title_id())
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    choice = Choice.select().order_by(Choice.group).where(Choice.title_id == title_id())

    system_id = system.select().where(System.stage == fin).get()
    id_system = system_id.id
    stage_exit = system_id.stage_exit
    st = "Финальный"

    nums = rank_mesto_out_in_group_or_semifinal_to_final(fin) # список мест, выходящих из группы или пф

    if stage_exit == "Предварительный":
        choices_fin = choice.select().where(Choice.mesto_group.in_(nums))
        nt = 1
        for b in nums:
            choices_fin = choice.select().where(Choice.mesto_group == b)
            for n in choices_fin:
                player = n.family
                pl_id = n.player_choice_id
                player_id = f"{player}/{pl_id}"
                fin_dict[nt] = player_id
                nt += 1
    else:
        choices_fin = choice.select().where(Choice.mesto_semi_final.in_(nums))
        nt = 1
        for b in nums:
            choices_fin = choice.select().where(Choice.mesto_semi_final == b)
            for n in choices_fin:
                player = n.family
                pl_id = n.player_choice_id
                player_id = f"{player}/{pl_id}"
                fin_dict[nt] = player_id
                nt += 1

    player_in_final = system_id.max_player # количество игроков в финале
    cp = player_in_final - 3
    tour = tours_list(cp)
    kol_tours = len(tour)  # кол-во туров
    game = len(tour[0])  # кол-во игр в туре
    # ===== получение списка номеров игроков в порядке 1-ого тура
    k = 0
    number_tours = []
    first_tour = tour[0].copy()
    first_tour.sort()
    for n in first_tour:
        z = n.find("-")
        num = int(n[:z])
        number_tours.append(num)
    for n in first_tour:
        z = n.find("-")
        num = int(n[z + 1:])
        number_tours.append(num)

    # for nt in number_tours:
    for nt in range(1, player_in_final + 1):
        fin_list.append(fin_dict[nt]) # список игроков в порядке 1 ого тура
        game_list = Game_list(number_group=fin, rank_num_player=nt, player_group=fin_dict[nt], system_id=system_id,
                            title_id=title_id())
        game_list.save()
  
    # === запись в db игроков которые попали в финал из группы
    ps_final = 1
    for l in fin_list:
        id_pl = int(l[l.find("/") + 1:])
        choices = choice.select().where(Choice.player_choice_id == id_pl).get()
        choices.final = fin
        choices.posev_final = ps_final
        choices.save()
        ps_final += 1
    # исправить если из группы выходят больше 2-ух игроков
    for r in range(0, kol_tours):
        round = r + 1
        tours = tour[r]  # игры тура
        for d in range(0, game):  # цикл по играм тура
            match = tours[d]  # матч в туре
            znak = match.find("-")
            first = int(match[:znak])  # игрок под номером в группе
            # игрок под номером в группе
            second = int(match[znak + 1:])
            pl1_fam_id = fin_list[first - 1] # фамилия первого игрока /id
            z = pl1_fam_id.find("/") # находит черту
            pl1_fam = pl1_fam_id[:z] # отделяет фамилия от ид
            pl1_id = int(pl1_fam_id[z + 1:])
            pl1_city = players.select().where(Player.id == pl1_id).get()
            cit1 = pl1_city.city
            pl2_fam_id = fin_list[second - 1] # фамилия второго игрока
            z = pl2_fam_id.find("/")
            pl2_fam = pl2_fam_id[:z]
            pl2_id = int(pl2_fam_id[z + 1:])
            pl2_city = players.select().where(Player.id == pl2_id).get()
            cit2 = pl2_city.city
            full_pl1 = f"{pl1_fam}/{cit1}"
            full_pl2 = f"{pl2_fam}/{cit2}"
            with db:
                results = Result(number_group=fin, system_stage=st, player1=full_pl1, player2=full_pl2,
                                tours=match, title_id=title_id(), round=round, system_id=id_system).save()
    with db:
        system_id.choice_flag = True
        system_id.save()    
    title = Title.select().where(Title.id == title_id()).get()
    page_title = title.tab_enabled
    if "Финалы" not in page_title:
        page_title = f"{page_title} Финалы"
    gamer = title.gamer
    with db:
        title.tab_enabled = page_title
        title.save()
    tab_enabled(gamer)
    pv = system_id.page_vid
    stage = fin
    table_made(pv, stage)


def player_in_table_group_and_write_Game_list_Result(stage):
    """заполняет таблицу Game_list данными спортсменами из группы td - список списков данных из групп и записывает
    встречи по турам в таблицу -Result- """
    system = System.select().where((System.title_id == title_id()) & (System.stage == stage)).get() # находит system id этапа
    system_id = system.id
    kg = system.total_group
    pv = system.page_vid
    # удаление старых записей в game_list и Result после редактирования жеребьевки групп
    game_list_delete = Game_list.delete().where((Game_list.title_id == title_id()) & (Game_list.system_id == system_id))
    result_delete = Result.delete().where((Result.title_id == title_id()) & (Result.system_id == system_id))
    game_list_delete.execute()
    result_delete.execute()  
    # создание таблиц групп со спортсменами согласно жеребьевки в PDF
    table_made(pv, stage)
    # вызов функции, где получаем список всех участников по группам
    tdt_all = table_data(stage, kg)
    for p in range(0, kg):  # цикл заполнения db таблиц -game list- и  -Results-
        gr = tdt_all[0][p]
        count_player = len(gr) // 2  # максимальное кол-во участников в группе
        number_group = str(p + 1) + ' группа'
        k = 0  # кол-во спортсменов в группе
        for i in range(0, count_player * 2 - 1, 2):
            family_player = gr[i][1]  # фамилия игрока
            posev = int(gr[i][0]) # посев (номер игрока в группе)
            fp = len(family_player) # кол-во знаков фамилии, если 0 значит игрока нет
                # подсчет кол-во знаков в фамилия, если 0 значит игрока нет
            if fp > 0:  # если строка (фамилия игрока) не пустая идет запись в db
                k += 1
                with db:
                    game_list = Game_list(number_group=number_group, rank_num_player=posev, 
                                            player_group=family_player,
                                            system_id=system, title_id=title_id()).save()

        # если 1-я строка (фамилия игрока) пустая выход из группы
        if fp == 0 and k != 0 or k == count_player:
            cp = k - 3
            tour = tours_list(cp)
            kol_tours = len(tour)  # кол-во туров
            game = len(tour[0])  # кол-во игр в туре
            for r in range(0, kol_tours):
                round = r + 1
                tours = tour[r]  # игры тура
                for d in range(0, game):  # цикл по играм тура
                    match = tours[d]  # матч в туре
                    znak = match.find("-")
                    first = int(match[:znak])  # игрок под номером в группе
                    # игрок под номером в группе
                    second = int(match[znak + 1:])
                    pl1_id = gr[first * 2 - 2][1]  # фамилия первого игрока
                    # z = pl1_id.find("/") # находит черту
                    # pl1 = pl1_id[:z] # отделяет фамилия от ид
                    pl2_id = gr[second * 2 - 2][1]  # фамилия второго игрока
                    # z = pl2_id.find("/")
                    # pl2 = pl2_id[:z]
                    cit1 = gr[first * 2 - 1][1] # город 1-ого игрока
                    cit2 = gr[second * 2 - 1][1] # город 2-ого игрока
                    full_pl1 = f"{pl1_id}/{cit1}"
                    full_pl2 = f"{pl2_id}/{cit2}"
                    with db:
                        results = Result(number_group=number_group, system_stage=stage, player1=full_pl1, player2=full_pl2,
                                         tours=match, title_id=title_id(), round=round, system_id=system_id).save()


def chop_line(t, maxline=31):
    """перевод строки если слишком длинный список тренеров"""
    l = len(t)
    if l > maxline:
        s1 = t.find(",", 0, maxline)
        s2 = t.find(",", s1 + 1, maxline)       
        cant = len(t) // maxline
        cant += 1
        strline = ""
        if s2 == -1: # если две фамилии больше 31, перевод после 1-ой фамилии
            for k in range(1, cant):
                index = maxline * k
                strline += "%s\n" % (t[(index - maxline):s1 + 1])
            strline += "%s" % (t[s1 + 1:])
        else:
            for k in range(1, cant):
                index = maxline * k
                strline += "%s\n" % (t[(index - maxline):s2 + 1])
            strline += "%s" % (t[s2 + 1:])
        t = strline
    return t


def chop_line_city(g, maxline=15):
    """перевод строки если слишком длинный список города"""
    l = len(g)
    if l > maxline:
        s1 = g.find(" ", 0, maxline)
        s2 = g.find(" ", s1 + 1, maxline)       
        strline = ""
        if s2 == -1: # если две фамилии больше 31, перевод после 1-ой фамилии
            strline = g[:s1]
        else:
            strline = g[:s2]
        g = strline
    return g


def change_status_visible_and_score_game():
    """изменение статуса колво партий и ввод счета во встречи"""
    sender = my_win.sender()
    system = System.select().where(System.title_id == title_id())
    tab = my_win.tabWidget.currentIndex()
    idx = my_win.tableView.currentIndex() # определиние номера строки
    row_num = idx.row()
    if row_num == -1:
        return
    count = len(system)    
    if tab == 3:
        system_stage = system.select().where(System.stage == "Предварительный").get()
        match_db = system_stage.score_flag
        state_visible = system_stage.visible_game  # флаг, показывающий записывать счет в партиях или нет
        my_win.checkBox_4.setEnabled(state_visible)
        match_current = match_db
        #  ==== изменение состояние =====
        if sender == my_win.checkBox_4:
            for i in my_win.groupBox_kolvo_vstrech_gr.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    match_current = int(i.text())
                    break
            state_visible = my_win.checkBox_4.isChecked()
        elif (sender == my_win.radioButton_match_3 or 
            sender == my_win.radioButton_match_5 or sender == my_win.radioButton_match_7):
            for i in my_win.groupBox_kolvo_vstrech_gr.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    match_current = int(i.text())
                    break
        if match_current == 3:
            my_win.radioButton_match_3.setChecked(True)
            my_win.frame_gr_three.setVisible(True)
            my_win.frame_gr_five.setVisible(False)
            my_win.frame_gr_seven.setVisible(False)
        elif match_current == 5:
            my_win.radioButton_match_5.setChecked(True)
            my_win.frame_gr_three.setVisible(True)
            my_win.frame_gr_five.setVisible(True)
            my_win.frame_gr_seven.setVisible(False)
        elif match_current == 7:
            my_win.radioButton_match_7.setChecked(True)
            my_win.frame_gr_three.setVisible(True)
            my_win.frame_gr_five.setVisible(True)
            my_win.frame_gr_seven.setVisible(True)
        my_win.label_22.setVisible(True)
    elif tab == 4:
        if row_num == -1:
            stage = "1-й полуфинал"
        else:
            id_res = my_win.tableView.model().index(row_num, 0).data() # данные строки tableView
            result = Result.select().where(Result.id == id_res).get()
            stage = result.system_stage
        system_stage = system.select().where(System.stage == stage).get()
        match_db = system_stage.score_flag
        state_visible_db = system_stage.visible_game  # флаг, показывающий записывать счет в партиях или нет
        match_current = match_db
        state_visible = state_visible_db
        my_win.checkBox_14.setEnabled(state_visible)
        #  ==== изменение состояние =====
        if sender == my_win.checkBox_14:
            for i in my_win.groupBox_kolvo_vstrech_pf.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    match_current = int(i.text())
                    break
            state_visible = my_win.checkBox_14.isChecked()
        elif (sender == my_win.radioButton_match_3 or 
            sender == my_win.radioButton_match_5 or sender == my_win.radioButton_match_7):
            for i in my_win.groupBox_kolvo_vstrech_pf.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    match_current = int(i.text())
                    break
            state_visible = state_visible_db

        if match_current == 3:
            my_win.radioButton_match_9.setChecked(True)
            my_win.frame_pf_three.setVisible(True)
            my_win.frame_pf_five.setVisible(False)
            my_win.frame_pf_seven.setVisible(False)
        elif match_current == 5:
            my_win.radioButton_match_10.setChecked(True)
            my_win.frame_pf_three.setVisible(True)
            my_win.frame_pf_five.setVisible(True)
            my_win.frame_pf_seven.setVisible(False)
        elif match_current == 7:
            my_win.radioButton_match_11.setChecked(True)
            my_win.frame_pf_three.setVisible(True)
            my_win.frame_pf_five.setVisible(True)
            my_win.frame_pf_seven.setVisible(True)
        my_win.label_71.setVisible(True)
    else:
        if row_num == -1: # не выбрана ни одна встреча
            system_stage = True
            match_db = 5
            match_current = 5
            state_visible = True
        else:
            if count == 1:
                stage = "Одна таблица"
            else:
                stage = my_win.tableView.model().index(row_num, 2).data() #  данные ячейки (из какого финала играют встречу)
            system_stage = system.select().where(System.stage == stage).get()
            match_db = system_stage.score_flag
            state_visible = system_stage.visible_game  # флаг, показывающий записывать счет в партиях или нет
            # state_visible_db = system_stage.visible_game  # флаг, показывающий записывать счет в партиях или нет
            match_current = match_db    
        #=========

        #  ==== изменение состояние =====
        if sender == my_win.checkBox_5:
            for i in my_win.groupBox_kolvo_vstrech_fin.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    match_current = int(i.text())
                    break
            state_visible = my_win.checkBox_5.isChecked()
        elif (sender == my_win.radioButton_match_4 or 
            sender == my_win.radioButton_match_6 or sender == my_win.radioButton_match_8):
            for i in my_win.groupBox_kolvo_vstrech_fin.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    match_current = int(i.text())
                    break
            state_visible = state_visible_db

        if match_current == 3:
            my_win.radioButton_match_4.setChecked(True)
            my_win.frame_fin_three.setVisible(True)
            my_win.frame_fin_five.setVisible(False)
            my_win.frame_fin_seven.setVisible(False)
        elif match_current == 5:
            my_win.radioButton_match_6.setChecked(True)
            my_win.frame_fin_three.setVisible(True)
            my_win.frame_fin_five.setVisible(True)
            my_win.frame_fin_seven.setVisible(False)
        elif match_current == 7:
            my_win.radioButton_match_8.setChecked(True)
            my_win.frame_fin_three.setVisible(True)
            my_win.frame_fin_five.setVisible(True)
            my_win.frame_fin_seven.setVisible(True)
        my_win.label_40.setVisible(True)
    if state_visible is False:
        frame_gr_list = [my_win.frame_gr_three,  my_win.frame_gr_five, my_win.frame_gr_seven]
        frame_pf_list = [my_win.frame_pf_three,  my_win.frame_pf_five, my_win.frame_pf_seven]
        frame_fin_list = [my_win.frame_fin_three,  my_win.frame_fin_five, my_win.frame_fin_seven]
        if tab == 3:
            for k in frame_gr_list:
                k.setVisible(False)
            my_win.checkBox_4.setChecked(False)
            my_win.lineEdit_pl1_score_total_gr.setFocus(True)
        elif tab == 4:
            for k in frame_pf_list:
                k.setVisible(False)
            my_win.checkBox_14.setChecked(False)
            my_win.lineEdit_pl1_score_total_pf.setFocus(True)
        else:
            for k in frame_fin_list:
                k.setVisible(False)
            my_win.checkBox_5.setChecked(False)
            my_win.lineEdit_pl1_score_total_fin.setFocus(True)
        my_win.label_22.setVisible(False)

    return state_visible


def visible_field():
    """включает или выключает поля для ввода счета, state - игра со счетом, True если включить поля для счета"""
    sender = my_win.sender()
    system = System.select().where(System.title_id == title_id())
    # ==== текущее состояние радиокнопок и чекбокса кол-во партий и ввод счета =====
    tab = my_win.tabWidget.currentIndex()
    idx = my_win.tableView.currentIndex() # номер выделенной строки
    row_num = idx.row()

    if tab == 3:
        stage = "Предварительный"
        system_stage = system.select().where(System.stage == stage).get()
        state_visible = system_stage.visible_game
        my_win.checkBox_4.setChecked(state_visible)
    elif tab == 4:
        my_win.checkBox_14.setChecked(True)
        my_win.radioButton_match_10.setChecked(True)
        state_visible = True
        match_db = 5
    else:
            # устанавливает начальное значение - со счетом ищ 5-ти партий
        if row_num == -1:
            my_win.checkBox_5.setChecked(True)
            my_win.radioButton_match_6.setChecked(True)
            state_visible = True
            match_db = 5
        else:
            stage = my_win.tableWidget.item(row_num , 2).text() # из какого финала играют встречу
            system_stage = system.select().where(System.stage == stage).get()
            match_db = system_stage.score_flag
            state_visible = system_stage.visible_game  # флаг, показывающий записывать счет в партиях или нет
        # ======= записывает изменение в базу данных
    if sender == my_win.checkBox_4 or sender == my_win.checkBox_14 or sender == my_win.checkBox_5: # изменяет состояние чекбокса игра со счетом или нет
        if tab == 3:
            state_visible = my_win.checkBox_4.isChecked()
            if state_visible is True:
                my_win.lineEdit_pl1_s1_gr.setFocus()
            else:
                my_win.lineEdit_pl1_gr_score_total.setFocus()
        elif tab == 4:
            state_visible = my_win.checkBox_14.isChecked()
            if state_visible is True:
                my_win.lineEdit_pl1_s1_pf.setFocus()
            else:
                my_win.lineEdit_pl1_pf_score_total.setFocus()
        else:
            state_visible = my_win.checkBox_5.isChecked()
            if state_visible is True:
                my_win.lineEdit_pl1_s1_fin.setFocus()
            else:
                my_win.lineEdit_pl1_score_total_fin.setFocus()
    change_status_visible_and_score_game()
 
    return state_visible


def select_player_in_list():
    """выводит данные игрока в поля редактирования или удаления"""
    data_list = []
    # row_num = my_win.tableView.currentIndex().row() # определиние номера строки
    # row_num = idx.row()
    # idx = my_win.tableView.currentIndex() # номер выделенной строки
    # row_num = idx.row()
    # pl_id = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    # for idx in my_win.tableView.selectionModel().selectedRows():
    indexes = my_win.tableView.selectionModel().selectedRows()
    for idx in sorted(indexes):

        # print('Row %d is selected' % index.row())
    # for idx in my_win.tableView.selectedIndexes():
    # for idx in my_win.tableView.currentIndex() # определиние номера строки
        row_num = idx.row()
        col_num = idx.column()
        data = my_win.tableView.model().index(row_num, col_num).data()
        data_list.append(data)
# ================================
    my_win.lineEdit_id.setText(data_list[0])
    my_win.lineEdit_id.setEnabled(False)
    my_win.lineEdit_Family_name.setText(data_list[1])
    my_win.lineEdit_bday.setText(data_list[2])
    my_win.lineEdit_R.setText(data_list[3])
    my_win.lineEdit_city_list.setText(data_list[4])
    my_win.comboBox_region.setCurrentText(data_list[5])
    my_win.comboBox_razryad.setCurrentText(data_list[6])
    my_win.lineEdit_coach.setText(data_list[7])

    my_win.Button_add_edit_player.setEnabled(True)
    if my_win.checkBox_6.isChecked():  # отмечен флажок -удаленные-
        my_win.Button_del_player.setEnabled(False)
        my_win.Button_add_edit_player.setText("Восстановить")
    else:
        my_win.Button_del_player.setEnabled(True)
        my_win.Button_add_edit_player.setEnabled(True)
        my_win.Button_add_edit_player.setText("Редактировать")
    if my_win.checkBox_11.isChecked():  # отмечен флажок -оплата R-
        my_win.Button_pay_R.setEnabled(True)
    else:
        my_win.Button_pay_R.setEnabled(False)


def save_in_db_pay_R():
    """запись в базу данных оплату рейтинга"""
    # r = my_win.tableWidget.currentRow()
    # family = my_win.tableWidget.item(r, 1).text()
    idx = my_win.tableView.currentIndex() # номер выделенной строки
    row_num = idx.row()
    pl_id = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    # player = Player.select().where(Player.title_id == title_id())
    # plr = player.select().where(Player.player == family).get()
    # pl_id = plr.id
    comment, ok = QInputDialog.getText(my_win, "Коментарий", "Введите коментарий о месте нахождении квитанции.")
    if ok:
        query = Player.update(pay_rejting="оплачен", comment=comment).where(Player.id == pl_id)
        query.execute()
    else:
        return
    debtor_R()


def check_repeat_player(pl, bd):
    """фукция проверки повтора ввода одно и того же игрока"""
    dr = []
    player_list = Player.select().where(Player.title_id == title_id())
    repeat = player_list.select().where(Player.player == pl) 
    count_family = len(repeat)
    if count_family != 0:
        for l in repeat:
            b_day = l.bday
            dr.append(b_day)
        if bd in dr:
            my_win.textEdit.setText("Такой игрок уже присутствует в списках!")   
            flag = True
        else:
            flag = False
    else:
        flag = False
    return flag


def select_player_in_game():
    """выводит фамилии игроков встречи"""
    tab = my_win.tabWidget.currentIndex()
    if tab == 1:
        select_player_in_list()
    elif tab ==2:
        change_choice_group()
    elif tab == 3:  # вкладка -группы-
        my_win.checkBox_7.setEnabled(True)
        my_win.checkBox_8.setEnabled(True)
        my_win.checkBox_7.setChecked(False)
        my_win.checkBox_8.setChecked(False)
    elif tab == 4:
        my_win.checkBox_12.setEnabled(True)
        my_win.checkBox_13.setEnabled(True)
        my_win.checkBox_12.setChecked(False)
        my_win.checkBox_13.setChecked(False)
    elif tab == 5:  # вкладка -финалы-
        my_win.checkBox_9.setEnabled(True)  # включает чекбоксы неявка
        my_win.checkBox_10.setEnabled(True)
        my_win.checkBox_9.setChecked(False)
        my_win.checkBox_10.setChecked(False)
        # numer_game = my_win.tableView.item(row_num, 3).text()
        my_win.groupBox_match_2.setTitle(f"Встреча №{numer_game}")
    if tab == 3 or tab == 4 or tab == 5:
        my_win.groupBox_kolvo_vstrech_fin.setEnabled(True)
        state_visible = change_status_visible_and_score_game()
        row_num= my_win.tableView.currentIndex().row() # определиние номера строки
        pl1 = my_win.tableView.model().index(row_num, 4).data()
        pl2 = my_win.tableView.model().index(row_num, 5).data()
        pl_win = my_win.tableView.model().index(row_num, 6).data()
        win_pole = my_win.tableView.model().index(row_num, 7).data()
        sc = my_win.tableView.model().index(row_num, 8).data()

        if win_pole != "None" and win_pole != "":  # если встреча сыграна, то заполняет поля общий счет
            if pl1 == pl_win:
                # если в сетке недостающие игроки (bye), то нет счета
                if sc != "":
                    sc1 = sc[0]
                    sc2 = sc[4]
                else:  # оставляет поля общий счет пустыми
                    sc1 = ""
                    sc2 = ""
            else:
                # если в сетке недостающие игроки (bye), то нет счета
                if sc != "":
                    sc1 = sc[4]
                    sc2 = sc[0]
                else:
                    sc1 = ""
                    sc2 = ""
            if tab == 3:
                my_win.lineEdit_pl1_score_total_gr.setText(sc1)
                my_win.lineEdit_pl2_score_total_gr.setText(sc2)
                my_win.lineEdit_player1_gr.setText(pl1)
                my_win.lineEdit_player2_gr.setText(pl2)
                my_win.lineEdit_pl1_s1_gr.setFocus()
            elif tab == 4:
                my_win.lineEdit_pl1_score_total_pf.setText(sc1)
                my_win.lineEdit_pl2_score_total_pf.setText(sc2)
                my_win.lineEdit_player1_pf.setText(pl1)
                my_win.lineEdit_player2_pf.setText(pl2)
                my_win.lineEdit_pl1_s1_pf.setFocus()
            else:
                my_win.lineEdit_pl1_score_total_fin.setText(sc1)
                my_win.lineEdit_pl2_score_total_fin.setText(sc2)
                my_win.lineEdit_player1_fin.setText(pl1)
                my_win.lineEdit_player2_fin.setText(pl2)
                my_win.lineEdit_pl1_s1_fin.setFocus()
        else:
            if tab == 3:
                my_win.checkBox_7.setEnabled(True)
                my_win.checkBox_8.setEnabled(True)
                my_win.lineEdit_player1_gr.setText(pl1)
                my_win.lineEdit_player2_gr.setText(pl2)
                if state_visible is True:
                    my_win.lineEdit_pl1_s1_gr.setFocus()
                else:
                    my_win.lineEdit_pl1_score_total_gr.setFocus()
            elif tab == 4:
                my_win.checkBox_12.setEnabled(True)
                my_win.checkBox_13.setEnabled(True)
                my_win.lineEdit_player1_pf.setText(pl1)
                my_win.lineEdit_player2_pf.setText(pl2)
                if state_visible is True:
                    my_win.lineEdit_pl1_s1_pf.setFocus()
                else:
                    my_win.lineEdit_pl1_score_total_pf.setFocus()
            elif tab == 5:
                my_win.checkBox_5.setEnabled(True)
                my_win.lineEdit_player1_fin.setText(pl1)
                my_win.lineEdit_player2_fin.setText(pl2)
                if pl1 == "X" or pl2 == "X":
                    my_win.Button_Ok_fin.setEnabled(True)
                    my_win.Button_Ok_fin.setFocus()                   
                else:
                    if state_visible is True:
                        my_win.lineEdit_pl1_s1_fin.setFocus()
                    else:
                        my_win.lineEdit_pl1_score_total_fin.setFocus()
        my_win.tableView.selectRow(row_num)


def delete_player():
    """удаляет игрока из списка и заносит его в архив"""
    msgBox = QMessageBox
    # player_current = Player.select().where(Player.title_id == title_id())
    game_list = Game_list.select().where(Game_list.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    idx = my_win.tableView.currentIndex() # определиние номера строки
    row_num = idx.row()

    player_id  = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    player_del  = my_win.tableView.model().index(row_num, 1).data() # данные ячейки tableView
    birthday  = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
    rank   = my_win.tableView.model().index(row_num, 3).data() # данные ячейки tableView
    player_city_del  = my_win.tableView.model().index(row_num, 4).data() # данные ячейки tableView
    region = my_win.tableView.model().index(row_num, 5).data() # данные ячейки tableView
    razryad  = my_win.tableView.model().index(row_num,6).data() # данные ячейки tableView
    coach  = my_win.tableView.model().index(row_num, 7).data() # данные ячейки tableView
    full_name = f"{player_del}/{player_city_del}"
    coach_id = Coach.get(Coach.coach == coach)
    player = Player.select().where(Player.id == player_id).get()
    pay_R = player.pay_rejting
    comment = player.comment

    question = msgBox.question(my_win, "", f"Вы действительно хотите удалить\n"
                                         f" {player_del} город {player_city_del}?",
                             msgBox.Ok, msgBox.Cancel)
    if question == msgBox.Ok:
        systems = system.select().where(System.stage == "Предварительный").get()
        choice_flag = systems.choice_flag
        if choice_flag is True:
            question = msgBox.information(my_win, "", f"Уже была произведена жеребьевка!\n"
                                            f" {player_del} город {player_city_del}\n"
                                            "будет удален(а) из посева.",
                                msgBox.Ok)
            sys = system.select().where(System.stage == "Предварительный").get()
            system_id = sys.id # id системы -Предварительного этапа-
            
            choices = Choice.delete().where(Choice.player_choice_id == player_id)
            choices.execute()
            game_lists = game_list.select().where(Game_list.player_group_id == player_del).get()
            posev = game_lists.rank_num_player
            number_group = game_lists.number_group
            # === изменяет номера посева, если удаляемый игрок не в последний посев ==
            g_list = game_list.select().where((Game_list.system_id == system_id) & 
                                            (Game_list.number_group == number_group))
            for k in g_list:
                gl_id = k.id
                ps = k.rank_num_player # посев игрока
                if posev < ps:
                    rank_in_group = ps - 1
                    gl = Game_list.update(rank_num_player=rank_in_group).where(Game_list.id == gl_id)
                    gl.execute()
                elif posev == ps:
            # === удаляет игрока из Game_list ===
                    gl = Game_list.delete().where(Game_list.id == gl_id)
                    gl.execute()
            # ==== заменяет туры (удаляет встречи с удаленным игроком)
            result_game = result.select().where((Result.system_id == system_id) & 
                                                    (Result.number_group == number_group))
            fam_city_del = f"{player_del}/{player_city_del}"
            for k in result_game:
                pl1 = k.player1
                pl2 = k.player2
                if pl1 == fam_city_del or pl2 == fam_city_del:
                    res = Result.delete().where(Result.id == k)
                    res.execute()
            for k in result_game:
                tour = k.tours
                znak = tour.find("-")
                p1 = int(tour[:znak])  # игрок под номером в группе
                p2 = int(tour[znak + 1:])  # игрок под номером в группе
                if p1 > posev and p2 > posev:
                    p1 -= 1
                    p2 -= 1
                elif p1 > posev:
                    p1 -= 1
                elif p2 > posev:
                    p2 -= 1
                new_tour = f"{p1}-{p2}"
                res = Result.update(tours=new_tour).where(Result.id == k)
                res.execute()
        else: # записывает в таблицу -Удаленные-
            with db: 
                del_player = Delete_player(player_del_id=player_id, bday=birthday, rank=rank, city=player_city_del,
                                            region=region, razryad=razryad, coach_id=coach_id, full_name=full_name,
                                            player=player_del, title_id=title_id(), pay_rejting=pay_R, comment=comment).save()

        player = Player.delete().where(Player.id == player_id)
        player.execute()
        my_win.lineEdit_Family_name.clear()
        my_win.lineEdit_bday.clear()
        my_win.lineEdit_R.clear()
        my_win.lineEdit_city_list.clear()
        my_win.lineEdit_coach.clear()
        player_list = Player.select().where(Player.title_id == title_id())
        count = len(player_list)
        my_win.label_46.setText(f"Всего: {count} участников")
        fill_table(player_list)
    else:
        return


def sortByAlphabet(inputStr):
    inputStr = inputStr.lower()
    return inputStr[0]


def load_comboBox_filter():
    """загрузка комбобокса регионами для фильтрации списка"""
    my_win.comboBox_fltr_region.clear()
    my_win.comboBox_fltr_city.clear()
    reg = []
    gorod = []
    player = Player.select().where(Player.title_id == title_id())
    if my_win.comboBox_fltr_region.count() > 0:  # проверка на заполненность комбобокса данными
        return
    else:
        for r in player:
            reg_n = r.region
            if reg_n not in reg:
                reg.append(reg_n)
        reg.sort(key=sortByAlphabet)
        reg.insert(0, "")
        my_win.comboBox_fltr_region.addItems(reg)
    
    if my_win.comboBox_fltr_city.count() < 0:  # проверка на заполненность комбобокса данными
        for c in player:
            cityes = c.city
            if cityes not in gorod:
                gorod.append(cityes)
        gorod.sort(key=sortByAlphabet)
        gorod.insert(0, "")
        my_win.comboBox_fltr_city.addItems(gorod)


def change_city_from_region():
    """изменяет список городов в комбобоксе фильтра списка в зависимости от региона"""  
    gorod = []
    my_win.comboBox_fltr_city.clear()
    player = Player.select().where(Player.title_id == title_id())
    region = my_win.comboBox_fltr_region.currentText()
    if region == "":
        player_region = player.select()
    else:
        player_region = player.select().where(Player.region == region)
    for pl_reg in player_region:
        if pl_reg.city not in gorod:
            gorod.append(pl_reg.city)
    gorod.sort(key=sortByAlphabet)
    gorod.insert(0, "")
    my_win.comboBox_fltr_city.addItems(gorod)


def filter_player_list(sender):
    """фильтрация списка участников по областям, тренерам, городам"""
    sender = my_win.sender()
    player = Player.select().where(Player.title_id == title_id())
    if sender == my_win.Button_fltr_list:
        region = my_win.comboBox_fltr_region.currentText()
        city = my_win.comboBox_fltr_city.currentText()
        if region != "" and city != "":
            player_list = player.select().where(Player.region == region)
            player_list = player.select().where(Player.city == city)
        elif region == "" and city != "":
            player_list = player.select().where(Player.city == city)
        elif region != "" and city == "":
            player_list = player.select().where(Player.region == region)
    elif sender == my_win.checkBox_15: # отмечен чекбокс предзаявка
        if my_win.checkBox_15.isChecked():
            my_win.Button_app.setEnabled(True)
            player_list = player.select().where(Player.application == "предварительная")
        else:
            my_win.Button_app.setEnabled(False)
            player_list = Player.select().where(Player.title_id == title_id())
    elif sender == my_win.Button_reset_fltr_list:
        player_list = Player.select().where(Player.title_id == title_id())
        my_win.comboBox_fltr_region.setCurrentIndex(0)
        my_win.comboBox_fltr_city.setCurrentIndex(0) 
        my_win.checkBox_15.setChecked(False)      
        load_comboBox_filter()
    fill_table(player_list)


def find_in_player_list():
    """поиск спортсмена или тренера"""
    player = Player.select().where(Player.title_id == title_id())
    txt = my_win.lineEdit_Family_name.text()
    if txt == "":
        my_win.textEdit.clear()
    txt = txt.upper()
    player_list = player.select().where(Player.player ** f'{txt}%')  # like
    if len(player_list) > 0:
        fill_table(player_list)
    else:
        my_win.textEdit.setText("Такого спортсмена нет!")


def find_in_player_rejting_list():
    """поиск спортсмена в рейтинг листе"""
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]
    gamer_w = ["Девочки", "Девушки", "Женщины"]
    id_title = Title.select().where(Title.id == title_id()).get()
    gamer = id_title.gamer
    txt_r = ""
    cur_index = my_win.comboBox_choice_R.currentIndex()
    txt_r = my_win.lineEdit_find_player_in_R.text()
    txt_r = txt_r.capitalize()
    if cur_index == 0: # если выбран текущий рейтинг
        if gamer in gamer_w:
            r_data = r_data_w[0]
        else:
            r_data = r_data_m[0]
        player_list = r_data.select().where(r_data.r_fname ** f'{txt_r}%')   
    elif cur_index == 1: # если рейтинг за январь
        if gamer in gamer_w:
            r_data = r_data_w[1]
        else:
           r_data = r_data_m[1]
        player_list = r_data.select().where(r_data.r1_fname ** f'{txt_r}%')
 
    fill_table(player_list) # заполняет таблицу -tablewidget- списком спортсменов


def filter_rejting_list():
    """Фильтрует вкладку -рейтинг-"""
    sender = my_win.sender()
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]

    gamer_w = ["Девочки", "Девушки", "Женщины"]
    id_title = Title.select().where(Title.id == title_id()).get()
    gamer = id_title.gamer
    cur_index = my_win.comboBox_choice_R.currentIndex()
    region_txt = my_win.comboBox_filter_region_in_R.currentText()
    city_txt = my_win.comboBox_filter_city_in_R.currentText()
    date_txt = my_win.comboBox_filter_date_in_R.currentText()

    if cur_index == 0:
        r_data = r_data_w[0] if gamer in gamer_w else r_data_m[0] # текущий рейтинг
        rejting_name = r_data.r_fname
        rejting_list = r_data.r_list
        rejting_region = r_data.r_region
        rejting_city = r_data.r_city
        rejting_date = r_data.r_bithday
    else:
        r_data = r_data_w[1] if gamer in gamer_w else r_data_m[1] # январский рейтинг 
        rejting_name = r_data.r1_fname
        rejting_list = r_data.r1_list
        rejting_region = r_data.r1_region
        rejting_city = r_data.r1_city
        rejting_date = r_data.r1_bithday

    if sender == my_win.Button_sort_rejting_in_R: 
        if date_txt != "":
            player_list = r_data.select().where(rejting_date > after_date).order_by(rejting_list.desc())       
        elif region_txt == "" and city_txt == "":
            player_list = r_data.select().order_by(rejting_list.desc()) 
        elif region_txt != "" and city_txt == "":
            player_list = r_data.select().where(rejting_region == region_txt).order_by(rejting_list.desc())
        elif region_txt == "" and city_txt != "":
            player_list = r_data.select().where(rejting_city == city_txt).order_by(rejting_list.desc())
    elif sender == my_win.Button_sort_alf_R: 
        if region_txt == "" and city_txt == "":
            player_list = r_data.select().order_by(rejting_name)
        elif region_txt != "" and city_txt == "":
            player_list = r_data.select().where(rejting_region == region_txt).order_by(rejting_name)
        elif region_txt == "" and city_txt != "":
            player_list = r_data.select().where(rejting_city == city_txt).order_by(rejting_name)
    elif date_txt != "":
        znak = date_txt.find(" ")
        year_fltr = int(date_txt[znak: znak + 3])
        year_current = int(datetime.today().strftime("%Y")) # текущий год
        year_bday = year_current - year_fltr + 1
        after_date = date(year_bday, 1, 1)
        player_list = r_data.select().where(rejting_date > after_date)
    else:
        if region_txt == "" and city_txt == "":
            player_list = r_data.select()
        elif region_txt != "" and city_txt == "":
            player_list = r_data.select().where(rejting_region == region_txt)
        elif region_txt == "" and city_txt != "":
            player_list = r_data.select().where(rejting_city == city_txt)

    fill_table(player_list) # заполняет таблицу -tablewidget- списком спортсменов


def enter_total_score():
    """ввод счета во встречи без счета в партиях"""
    msgBox = QMessageBox
    sender = my_win.sender()
    tab = my_win.tabWidget.currentIndex()
    mark = 0
    flag = 0
    mistake = 0
    if sender == my_win.lineEdit_pl1_score_total_gr:
        mark = my_win.lineEdit_pl1_score_total_gr.text()
        flag = 0
    elif sender == my_win.lineEdit_pl2_score_total_gr:
        mark = my_win.lineEdit_pl2_score_total_gr.text()
        flag = 1 
    elif sender == my_win.lineEdit_pl1_score_total_pf:
        mark = my_win.lineEdit_pl1_score_total_pf.text()
        flag = 0
    elif sender == my_win.lineEdit_pl2_score_total_pf:
        mark = my_win.lineEdit_pl2_score_total_pf.text()
        flag = 1  
    elif sender == my_win.lineEdit_pl1_score_total_fin:
        mark = my_win.lineEdit_pl1_score_total_fin.text()
        flag = 0
    elif sender == my_win.lineEdit_pl2_score_total_fin:
        mark = my_win.lineEdit_pl2_score_total_fin.text()
        flag = 1  
    if mark != "":  
        mark = int(mark)
        mistake = check_input_total_score(mark, flag)
        if tab == 3 and flag == 0:
            my_win.lineEdit_pl2_score_total_gr.setFocus() if mistake == 0 else my_win.lineEdit_pl1_score_total_gr.setFocus()
        elif tab == 4 and flag == 0:
            my_win.lineEdit_pl2_score_total_pf.setFocus() if mistake == 0 else my_win.lineEdit_pl1_score_total_pf.setFocus()
        elif tab == 5 and flag == 0:
            my_win.lineEdit_pl2_score_total_fin.setFocus() if mistake == 0 else my_win.lineEdit_pl1_score_total_fin.setFocus()
        elif tab == 3 and flag == 1:
            enter_score(none_player=0) if mistake == 0 else my_win.lineEdit_pl2_score_total_gr.setFocus()
        elif tab == 4 and flag == 1:
            enter_score(none_player=0) if mistake == 0 else my_win.lineEdit_pl2_score_total_pf.setFocus()
        elif tab == 5 and flag == 1:
            enter_score(none_player=0) if mistake == 0 else my_win.lineEdit_pl2_score_total_fin.setFocus()
    else:
        reply = msgBox.information(my_win, 'Уведомление',
                                                "Проверьте правильность ввода счета!",
                                                msgBox.Ok)
        return
    

def check_input_total_score(mark, flag):
    """проверка ввода счета встречи и его правильность"""
    msgBox = QMessageBox
    score_list = []
    tab = my_win.tabWidget.currentIndex() 
    mark_int = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    if tab == 3:
        for i in my_win.groupBox_kolvo_vstrech_gr.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                match_current = int(i.text())
                break
        s1 = my_win.lineEdit_pl1_score_total_gr.text()
        s2 = my_win.lineEdit_pl2_score_total_gr.text()
    elif tab == 4:
        for i in my_win.groupBox_kolvo_vstrech_pf.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                match_current = int(i.text())
                break
        s1 = my_win.lineEdit_pl1_score_total_pf.text()
        s2 = my_win.lineEdit_pl2_score_total_pf.text()
    else:
        for i in my_win.groupBox_kolvo_vstrech_fin.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                match_current = int(i.text())
                break
        s1 = my_win.lineEdit_pl1_score_total_fin.text()
        s2 = my_win.lineEdit_pl2_score_total_fin.text()
    if mark in mark_int:
        if flag == 1:
            score_list.append(int(s1))
            score_list.append(int(s2))
            if match_current // 2 + 1 not in score_list:
                reply = msgBox.information(my_win, 'Уведомление',
                                                "Проверьте правильность ввода счета!\nСчет меньше необходимого.",
                                               msgBox.Ok)
                return
        if match_current // 2 + 1 < mark:
            reply = msgBox.information(my_win, 'Уведомление',
                                                "Проверьте правильность ввода счета!\nЧисло не соответсвует из скольки партий матч.",
                                               msgBox.Ok)
            mistake = 1
        else:
            mistake = 0
            return mistake
    else:
        reply = msgBox.information(my_win, 'Уведомление',
                                                "Вы ввели не правильно символ!",
                                                msgBox.Ok)
        mistake = 1
        return mistake
                 

def focus():
    """переводит фокус на следующую позицию
    sum_total_game список (1-й колво очков которые надо набрать, 2-й сколько уже набрали)"""
    sender = my_win.sender()  # в зависимости от сигала кнопки идет сортировка
    system = System.select().where(System.title_id == title_id())
    tab = my_win.tabWidget.currentIndex()
    stage = my_win.comboBox_filter_final.currentText()
    idx = my_win.tableView.currentIndex() # определиние номера строки
    row_num = idx.row()
    mark_list_gr = [my_win.lineEdit_pl1_s1_gr, my_win.lineEdit_pl2_s1_gr, my_win.lineEdit_pl1_s2_gr, my_win.lineEdit_pl2_s2_gr,
            my_win.lineEdit_pl1_s3_gr, my_win.lineEdit_pl2_s3_gr, my_win.lineEdit_pl1_s4_gr, my_win.lineEdit_pl2_s4_gr,
            my_win.lineEdit_pl1_s5_gr, my_win.lineEdit_pl2_s5_gr, my_win.lineEdit_pl1_s6_gr, my_win.lineEdit_pl2_s6_gr,
            my_win.lineEdit_pl1_s7_gr, my_win.lineEdit_pl2_s7_gr]
    mark_list_sf = [my_win.lineEdit_pl1_s1_pf, my_win.lineEdit_pl2_s1_pf, my_win.lineEdit_pl1_s2_pf, my_win.lineEdit_pl2_s2_pf,
            my_win.lineEdit_pl1_s3_pf, my_win.lineEdit_pl2_s3_pf, my_win.lineEdit_pl1_s4_pf, my_win.lineEdit_pl2_s4_pf,
            my_win.lineEdit_pl1_s5_pf, my_win.lineEdit_pl2_s5_pf, my_win.lineEdit_pl1_s6_pf, my_win.lineEdit_pl2_s6_pf,
            my_win.lineEdit_pl1_s7_pf, my_win.lineEdit_pl2_s7_pf]
    mark_list_fin = [my_win.lineEdit_pl1_s1_fin, my_win.lineEdit_pl2_s1_fin, my_win.lineEdit_pl1_s2_fin, my_win.lineEdit_pl2_s2_fin,
            my_win.lineEdit_pl1_s3_fin, my_win.lineEdit_pl2_s3_fin, my_win.lineEdit_pl1_s4_fin, my_win.lineEdit_pl2_s4_fin,
            my_win.lineEdit_pl1_s5_fin, my_win.lineEdit_pl2_s5_fin, my_win.lineEdit_pl1_s6_fin, my_win.lineEdit_pl2_s6_fin,
            my_win.lineEdit_pl1_s7_fin, my_win.lineEdit_pl2_s7_fin]
    if tab == 3:
        sys = system.select().where(System.stage == "Предварительный").get()
        sf = sys.score_flag  # флаг из скольки партий играется матч
        mark_index = mark_list_gr.index(sender)
        mark = mark_list_gr[mark_index].text()
        flag_mistake = control_mark_in_score(mark)
        if flag_mistake is True:
            return
        if mark_index % 2 == 1:
            if mark_index >= sf:
                sum_total_game = score_in_game()  # подсчет очков в партии
                if len(sum_total_game) == 0: # значит была ошибка в счете и поэтому он вернул пустой список
                    return
                if sum_total_game[0] != sum_total_game[1]:
                    mark_list_gr[mark_index + 1].setFocus()
                else:
                    my_win.Button_Ok_gr.setFocus()
                    return
            mark_list_gr[mark_index + 1].setFocus()    
        else:
            mark_list_gr[mark_index + 1].setFocus()
    elif tab == 4:
        if row_num  == -1:
            stage = "1-й полуфинал"
        else:
            id_res = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
            result = Result.select().where(Result.id == id_res).get()
            stage = result.system_stage
        sys = system.select().where(System.stage == stage).get()
        sf = sys.score_flag  # флаг из скольки партий играется матч
        mark_index = mark_list_sf.index(sender)
        mark = mark_list_sf[mark_index].text()
        control_mark_in_score(mark)
        if mark_index % 2 == 1:
            if mark_index >= sf:
                sum_total_game = score_in_game()  # подсчет очков в партии
                if sum_total_game[0] != sum_total_game[1]:
                    mark_list_sf[mark_index + 1].setFocus()
                else:
                    my_win.Button_Ok_pf.setFocus()
                    return
            mark_list_sf[mark_index + 1].setFocus()    
        else:
            mark_list_sf[mark_index + 1].setFocus()
    elif tab == 5:
        final = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
        if stage == "Одна таблица":
            final = stage
        sys = system.select().where(System.stage == final).get()
        sf = sys.score_flag  # флаг из скольки партий играется матч
        mark_index = mark_list_fin.index(sender)
        mark = mark_list_fin[mark_index].text()
        control_mark_in_score(mark)
        if mark_index % 2 == 1:
            if mark_index >= sf:
                sum_total_game = score_in_game()  # подсчет очков в партии
                if sum_total_game[0] != sum_total_game[1]:
                    mark_list_fin[mark_index + 1].setFocus()
                else:
                    my_win.Button_Ok_fin.setFocus()
                    return
            mark_list_fin[mark_index + 1].setFocus()    
        else:
            mark_list_fin[mark_index + 1].setFocus()


def control_mark_in_score(mark):
    """проверка ввода счета в ячейку """
    msgBox = QMessageBox
    tab = my_win.tabWidget.currentIndex()

    sf = 5
    if tab == 3:
        score_list = [my_win.lineEdit_pl1_score_total_gr.text(), my_win.lineEdit_pl2_score_total_gr.text()] # список общий счет в партии
    elif tab == 4:
        score_list = [my_win.lineEdit_pl1_score_total_pf.text(), my_win.lineEdit_pl2_score_total_pf.text()]
    else:
        score_list = [my_win.lineEdit_pl1_score_total_fin.text(), my_win.lineEdit_pl2_score_total_fin.text()]

    flag = True if str((sf + 1) // 2) in [score_list[0], score_list[1]] else False

    if flag is False:
        mark_number = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
        if mark == "":
            msgBox.critical(my_win, "", "Ошибка при вводе счета!\nвведите счет")
            flag_mistake = True
            return flag_mistake
        else:
            count = len(mark)
            if count > 2:
                msgBox.critical(my_win, "", "Ошибка при вводе счета!\nпроверьте правильность ввода")
                flag_mistake = True
                return flag_mistake
            else:
                for k in range(0, count):
                    mark_zn = mark[k]
                    if mark_zn not in mark_number:
                        msgBox.critical(my_win, "", "Ошибка при вводе счета!\nпроверьте правильность ввода")
                        flag_mistake = True
                        return flag_mistake


def score_in_game():
    """считает общий счет в партиях"""
    msgBox = QMessageBox

    system = System.select().where(System.title_id == title_id())
    stage = my_win.comboBox_filter_final.currentText()
    total_score = []
    ts1 = []
    ts2 = []
    total_game = []
    sum_total_game = []
    row_num = my_win.tableView.currentIndex().row() # определиние номера строки

    tab = my_win.tabWidget.currentIndex()
    s11 = s21 = s12 = s22 = s13 = s23 = s14 = s24 = s15 = s25 = s16 = s26 = s17 = s27 = 0
    # поля ввода счета в партии
    if tab == 3:
        sys = system.select().where(System.stage == "Предварительный").get()
        sf = sys.score_flag  # флаг из скольки партий играется матч
        # ==========
        s11 = my_win.lineEdit_pl1_s1_gr.text()
        s21 = my_win.lineEdit_pl2_s1_gr.text()
        s12 = my_win.lineEdit_pl1_s2_gr.text()
        s22 = my_win.lineEdit_pl2_s2_gr.text()
        s13 = my_win.lineEdit_pl1_s3_gr.text()
        s23 = my_win.lineEdit_pl2_s3_gr.text()
        s14 = my_win.lineEdit_pl1_s4_gr.text()
        s24 = my_win.lineEdit_pl2_s4_gr.text()
        s15 = my_win.lineEdit_pl1_s5_gr.text()
        s25 = my_win.lineEdit_pl2_s5_gr.text()
        s16 = my_win.lineEdit_pl1_s6_gr.text()
        s26 = my_win.lineEdit_pl2_s6_gr.text()
        s17 = my_win.lineEdit_pl1_s7_gr.text()
        s27 = my_win.lineEdit_pl2_s7_gr.text()
    elif tab == 4:
        if row_num == -1:
            stage = "1-й полуфинал"
        else:
            id_res = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
            result = Result.select().where(Result.id == id_res).get()
            stage = result.system_stage
        sys = system.select().where(System.stage == stage).get()
        sf = sys.score_flag  # флаг из скольки партий играется матч
        s11 = my_win.lineEdit_pl1_s1_pf.text()
        s21 = my_win.lineEdit_pl2_s1_pf.text()
        s12 = my_win.lineEdit_pl1_s2_pf.text()
        s22 = my_win.lineEdit_pl2_s2_pf.text()
        s13 = my_win.lineEdit_pl1_s3_pf.text()
        s23 = my_win.lineEdit_pl2_s3_pf.text()
        s14 = my_win.lineEdit_pl1_s4_pf.text()
        s24 = my_win.lineEdit_pl2_s4_pf.text()
        s15 = my_win.lineEdit_pl1_s5_pf.text()
        s25 = my_win.lineEdit_pl2_s5_pf.text()
        s16 = my_win.lineEdit_pl1_s6_pf.text()
        s26 = my_win.lineEdit_pl2_s6_pf.text()
        s17 = my_win.lineEdit_pl1_s7_pf.text()
        s27 = my_win.lineEdit_pl2_s7_pf.text()
    elif tab == 5:
        # из какого финала пара игроков в данный момент
        final = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
        if stage == "Одна таблица":
            final = stage
        sys = system.select().where(System.stage == final).get()
        sf = sys.score_flag  # флаг из скольки партий играется матч
        s11 = my_win.lineEdit_pl1_s1_fin.text()
        s21 = my_win.lineEdit_pl2_s1_fin.text()
        s12 = my_win.lineEdit_pl1_s2_fin.text()
        s22 = my_win.lineEdit_pl2_s2_fin.text()
        s13 = my_win.lineEdit_pl1_s3_fin.text()
        s23 = my_win.lineEdit_pl2_s3_fin.text()
        s14 = my_win.lineEdit_pl1_s4_fin.text()
        s24 = my_win.lineEdit_pl2_s4_fin.text()
        s15 = my_win.lineEdit_pl1_s5_fin.text()
        s25 = my_win.lineEdit_pl2_s5_fin.text()
        s16 = my_win.lineEdit_pl1_s6_fin.text()
        s26 = my_win.lineEdit_pl2_s6_fin.text()
        s17 = my_win.lineEdit_pl1_s7_fin.text()
        s27 = my_win.lineEdit_pl2_s7_fin.text()
    if sf == 3:
        total_score = [s11, s21, s12, s22, s13, s23]
        max_game = 2
    elif sf == 5:
        total_score = [s11, s21, s12, s22, s13, s23, s14, s24, s15, s25]
        max_game = 3
    elif sf == 7:
        total_score = [s11, s21, s12, s22, s13, s23, s14, s24, s15, s25, s16, s26, s17, s27]
        max_game = 4
    point = 0

    n = len(total_score)

    for i in range(0, n, 2):
        if total_score[i] != "":
            sc1 = int(total_score[i])
            sc2 = int(total_score[i + 1])
            
            flag = control_score(sc1, sc2)

            if flag is True:
                if sc1 > sc2:
                    point = 1
                    ts1.append(point)
                else:
                    point = 1
                    ts2.append(point)
                st1 = sum(ts1)
                st2 = sum(ts2)
                # ==============
                if tab == 3:
                    my_win.lineEdit_pl1_score_total_gr.setText(str(st1))
                    my_win.lineEdit_pl2_score_total_gr.setText(str(st2))
                    if st1 == max_game or st2 == max_game:  # сравнивает максимальное число очков и набранные очки одним из игроков
                        # если игрок набрал макс очки активиоует кнопку ОК и переводит на нее фокус
                        my_win.Button_Ok_gr.setEnabled(True)
                        my_win.Button_Ok_gr.setFocus()
                elif tab == 4:
                    my_win.lineEdit_pl1_score_total_pf.setText(str(st1))
                    my_win.lineEdit_pl2_score_total_pf.setText(str(st2))
                    if st1 == max_game or st2 == max_game:  # сравнивает максимальное число очков и набранные очки одним из игроков
                        # если игрок набрал макс очки активиоует кнопку ОК и переводит на нее фокус
                        my_win.Button_Ok_pf.setEnabled(True)
                        my_win.Button_Ok_pf.setFocus()
                elif tab == 5:
                    my_win.lineEdit_pl1_score_total_fin.setText(str(st1))
                    my_win.lineEdit_pl2_score_total_fin.setText(str(st2))
                    if st1 == max_game or st2 == max_game:  # сравнивает максимальное число очков и набранные очки одним из игроков
                        # если игрок набрал макс очки активирует кнопку ОК и переводит на нее фокус
                        my_win.Button_Ok_fin.setEnabled(True)
                        my_win.Button_Ok_fin.setFocus()
                total_game.append(st1)
                total_game.append(st2)
                # находит максимальное число очков из сыгранных партий
                max_score = max(total_game)
                if i == 0:
                    # добавляет в список макс число очков которые надо набрать
                    sum_total_game.append(max_game)
                    # добавляет в список макс число очков которые уже набрал игрок
                    sum_total_game.append(max_score)
                else:
                    sum_total_game[0] = max_game
                    sum_total_game[1] = max_score
            elif flag is False:
                sum_total_game = []
                return sum_total_game
                # желательно сюда ввести чтобы фокус ставился на туже ячейку
    return sum_total_game


def control_score(sc1, sc2):
    """проверка на правильность ввода счета"""
    msgBox = QMessageBox
    if sc1 == 11:
        flag = True if sc2 in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 13] else False
    elif sc1 > 11:
        flag = True if sc2 == sc1 - 2 or sc2 == sc1 + 2 else False
    elif 0 <= sc1 <= 9:
        flag = True if sc2 == 11 else False
    elif sc1 == 10:
        flag = True if sc2 == 12 else False

    if flag == False:
        result = msgBox.information(my_win, "", "Проверьте правильность ввода\n счета в партии!",
                                    msgBox.Ok)
        flag = False
    return flag


def control_winner_player(winner, loser):
    """ Проверка условия победителя матча, различие рейтинга"""
    msgBox = QMessageBox
    flag = True
    player = Player.select().where(Player.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    if winner.rfind("/") == -1: # если спортсмены имеют только имя и фамилию без города (старый вариант)
        f_name_win_id = player.select().where(Player.player == winner).get()
        f_name_los_id = player.select().where(Player.player == loser).get()
        winner = f_name_win_id.full_name
        loser = f_name_los_id.full_name
    
    player_win_id = player.select().where(Player.full_name == winner).get()
    player_los_id = player.select().where(Player.full_name == loser).get()
    r_win = player_win_id.rank
    coefficient_win = player_win_id.coefficient_victories
    total_game_win = player_win_id.total_game_player
    r_los = player_los_id.rank
    coefficient_los = player_los_id.coefficient_victories
    total_game_los = player_los_id.total_game_player
    if total_game_win < 5 or total_game_los < 5: # если еще сыграно мало игр и определяем по разности рейтинга
        if r_win - r_los > 0:
            flag = True
        elif abs(r_los - r_win) < 15:
            flag = True 
        else:
            flag = False      
    else:
        if coefficient_win - coefficient_los > 0:
            flag = True
        elif abs(coefficient_los - coefficient_win) <= 0.3:
            flag = True
        else:
            if (r_win - r_los) < 0 and abs(r_win - r_los) > 30:
                flag = False
    if flag is False:            
        result = msgBox.information(my_win, "", f"Вы уверенны в победе\n {winner} рейтинг{r_win}!\n над {loser} рейтинг {r_los}",
                                    msgBox.Yes, msgBox.No)
        if  result == msgBox.No:
            flag = False
            return flag
    sum_win = player_win_id.total_win_game + 1
    all_game_win = total_game_win + 1
    sum_los = player_los_id.total_win_game
    all_game_los = total_game_los + 1  
    koef_win = sum_win / all_game_win
    koef_los = sum_los / all_game_los
    koef_win = float('{:.3f}'.format(koef_win)) # соотношение выйгранных партий ко всем играм
    koef_los = float('{:.3f}'.format(koef_los)) # соотношение выйгранных партий ко всем играм
 
    with db:
        player_win_id.total_game_player = all_game_win # все игр сыгранных игроком
        player_win_id.total_win_game = sum_win # сколько побед он одержал
        player_win_id.coefficient_victories = koef_win
        player_win_id.save()
        player_los_id.total_game_player = all_game_los
        player_los_id.total_win_game = sum_los
        player_los_id.coefficient_victories = koef_los
        player_los_id.save()


def check_real_player():
    """Изменяет спортсменов по предварительной заявке на реальных"""
    my_win.tabWidget.setCurrentIndex(1)

    indices = my_win.tableView.selectionModel().selectedRows()
    for index in sorted(indices):
        rows = index.row()
        id_pl = my_win.tableView.model().index(rows, 0).data() # данные ячейки tableView
        app = Player.update(application="основная").where(Player.id == id_pl)
        app.execute()
    player_list = Player.select().where((Player.title_id == title_id()) & (Player.application == "предварительная"))
    fill_table(player_list)


def enter_score(none_player=0):
    """заносит в таблицу -результаты- победителя, счет и т.п. sc_total [партии выигранные, проигранные, очки победителя
     очки проигравшего]"""
    tab = my_win.tabWidget.currentIndex()
    row_num = my_win.tableView.currentIndex().row()
    id = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    fin = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
    num_game = my_win.tableView.model().index(row_num, 3).data() # данные ячейки tableView

    sys = System.select().where(System.title_id == title_id())   
    if tab == 3: # группы
        stage = "Предварительный"
    elif tab == 4: # полуфиналы
        if row_num == -1: # не выбрана строка и идет ПФ по умочанию
            stage = "1-й полуфинал"
        else:
            id_res = my_win.tableView.model().index(row_num, 0).data() #  данные ячейки (из какого финала играют встречу)
            result = Result.select().where(Result.id == id_res).get()
            stage = result.system_stage
    else:  # финальный
        if fin == "1 группа":
            stage = "Одна таблица"
        else:
            stage = fin
    # находит system id последнего
    system = sys.select().where(System.stage == stage).get()
    type = system.type_table
    flag = 0
    if stage in ["Предварительный", "1-й полуфинал", "2-й полуфинал"]:
        sc_total = circle_type(none_player, stage)
    elif stage == "Одна таблица":
        if type == "сетка":
            sc_total = setka_type(none_player)
            flag = 1
        else:
            sc_total = circle_type(none_player, stage)
    else:  # финалы
        if type == "сетка":
            sc_total = setka_type(none_player)
            flag = 1
        else:  # по кругу
            sc_total = circle_type(none_player, stage)
    st1 = sc_total[0]  # партия выигранные
    st2 = sc_total[1]  # партии проигранные
    w = sc_total[2]  # очки победителя
    l = sc_total[3]  # очки проигравшего

    if my_win.lineEdit_player1_fin.text() != "X" and my_win.lineEdit_player2_fin.text() != "X":
        if tab == 3:
            pl1 = my_win.lineEdit_player1_gr.text()
            pl2 = my_win.lineEdit_player2_gr.text()
        elif tab == 4:
            pl1 = my_win.lineEdit_player1_pf.text()
            pl2 = my_win.lineEdit_player2_pf.text()
        elif tab == 5:
            pl1 = my_win.lineEdit_player1_fin.text()
            pl2 = my_win.lineEdit_player2_fin.text()
    # ======= 
        if none_player == 0: # встреча состоялась
            if st1 > st2:
                winner = pl1
                loser = pl2
                ts_winner = f"{st1} : {st2}"
                ts_loser = f"{st2} : {st1}"
            else:
                winner = pl2
                loser = pl1
                ts_winner = f"{st2} : {st1}"
                ts_loser = f"{st1} : {st2}"
            winner_string = string_score_game()  # пишет счет в партии
        elif none_player == 1: # не явился 1-й игрок
            winner = pl2
            loser = pl1
            ts_winner = f"{st2} : {st1}"
            ts_loser = f"{st1} : {st2}"
        else:
            winner = pl1
            loser = pl2
            ts_winner = f"{st1} : {st2}"
            ts_loser = f"{st2} : {st1}"
        if none_player != 0: # если победа по неявке
            if type == "сетка":
                winner_string = ""
            elif type == "круг" or type == "группы":
                winner_string = "В : П"
    else: # если нет одного игрока -X-
        if my_win.lineEdit_player1_fin.text() == "X":
            winner = my_win.lineEdit_player2_fin.text()
            loser = my_win.lineEdit_player1_fin.text()
        else:
            winner = my_win.lineEdit_player1_fin.text()
            loser = my_win.lineEdit_player2_fin.text()
        # loser_fam_name = loser # оставляет -X-
        winner_string = ""
        ts_winner = ""
        ts_loser = ""
    loser_fam_name = loser # 
    if none_player == 0: # если победа по неявке, то не проверяет победу
        flag = control_winner_player(winner, loser) # проверка реальности победы игрока (маленький рейтинг над большим)
        if flag is False:
            return

    with db:  # записывает в таблицу -Result- сыгранный матч в группах или финал по кругу
        result = Result.get(Result.id == id)
        result.winner = winner
        result.points_win = w
        result.score_win = winner_string
        result.score_in_game = ts_winner
        result.loser = loser
        result.points_loser = l
        result.score_loser = ts_loser
        result.save()

    if tab == 5:  # записывает в -Result- сыгранный матч со сносками на соответствующие строки победителя и проигравшего
        if type == "сетка":
            vid_setki = system.label_string  # вид сетки и кол-во участников
            # список 1-й номер победителя 2-й проигравшего
            snoska = numer_game(num_game, vid_setki) # snoska список [номер встречи победителя, номер встречи приогравшего, номер в сетке куда сносится проигравший]
            res = Result.select().where(Result.title_id == title_id())
            sys_id = sys.select().where(System.stage == fin).get()
            system_id = sys_id.id
            results = res.select().where(Result.system_id == system_id)
            if snoska[0] != 0:
                with db:  # записывает в db таблицу Result победителя и проигравшего
                    player = winner
                    match_num = result.tours  # номер встречи, в строке
                    game = snoska[2] * -1 # номер встречи число
                    for k in range(0, 2):
                        if int(match_num) == game:
                            res_id = results.select().where(Result.tours == snoska[k]).get() # id встречи, куда попадает победитель и проигравший
                            # =========                            
                            if res_id.player1 == "":
                                res_id.player1 = player
                            else:
                                res_id.player2 = player
                            res_id.save()
                            player = loser_fam_name
        elif type == "круг":
            pass
    fill_table_results()
    if tab == 3:
        line_edit_list = [my_win.lineEdit_pl1_s1_gr, my_win.lineEdit_pl2_s1_gr, my_win.lineEdit_pl1_s2_gr, my_win.lineEdit_pl2_s2_gr,
                          my_win.lineEdit_pl1_s3_gr, my_win.lineEdit_pl2_s3_gr, my_win.lineEdit_pl1_s4_gr, my_win.lineEdit_pl2_s4_gr,
                          my_win.lineEdit_pl1_s5_gr, my_win.lineEdit_pl2_s5_gr, my_win.lineEdit_pl1_s6_gr, my_win.lineEdit_pl2_s6_gr,
                          my_win.lineEdit_pl1_s7_gr, my_win.lineEdit_pl2_s7_gr, my_win.lineEdit_player1_gr,  my_win.lineEdit_player2_gr,
                          my_win.lineEdit_pl1_score_total_gr, my_win.lineEdit_pl2_score_total_gr]
        my_win.checkBox_7.setChecked(False)
        my_win.checkBox_8.setChecked(False)
        filter_gr()
    elif tab == 4:
        line_edit_list = [my_win.lineEdit_pl1_s1_pf, my_win.lineEdit_pl2_s1_pf, my_win.lineEdit_pl1_s2_pf, my_win.lineEdit_pl2_s2_pf,
                          my_win.lineEdit_pl1_s3_pf, my_win.lineEdit_pl2_s3_pf, my_win.lineEdit_pl1_s4_pf, my_win.lineEdit_pl2_s4_pf,
                          my_win.lineEdit_pl1_s5_pf, my_win.lineEdit_pl2_s5_pf, my_win.lineEdit_pl1_s6_pf, my_win.lineEdit_pl2_s6_pf,
                          my_win.lineEdit_pl1_s7_pf, my_win.lineEdit_pl2_s7_pf, my_win.lineEdit_player1_pf,  my_win.lineEdit_player2_pf,
                          my_win.lineEdit_pl1_score_total_pf, my_win.lineEdit_pl2_score_total_pf]
        my_win.checkBox_12.setChecked(False)
        my_win.checkBox_13.setChecked(False)
    elif tab == 5:
        line_edit_list = [my_win.lineEdit_pl1_s1_fin, my_win.lineEdit_pl2_s1_fin, my_win.lineEdit_pl1_s2_fin, my_win.lineEdit_pl2_s2_fin,
                          my_win.lineEdit_pl1_s3_fin, my_win.lineEdit_pl2_s3_fin, my_win.lineEdit_pl1_s4_fin, my_win.lineEdit_pl2_s4_fin,
                          my_win.lineEdit_pl1_s5_fin, my_win.lineEdit_pl2_s5_fin, my_win.lineEdit_pl1_s6_fin, my_win.lineEdit_pl2_s6_fin,
                          my_win.lineEdit_pl1_s7_fin, my_win.lineEdit_pl2_s7_fin, my_win.lineEdit_player1_fin,  my_win.lineEdit_player2_fin,
                          my_win.lineEdit_pl1_score_total_fin, my_win.lineEdit_pl2_score_total_fin]
    for line in line_edit_list:
            line.clear()


def made_pdf_table_for_view(sender):
    """вызов функции заполнения таблицы pdf группы сыгранными играми"""
    tab = my_win.tabWidget.currentIndex()
    sys = System.select().where(System.title_id == title_id())
    if sender == my_win.view_gr_Action or tab == 3:  # вкладка группы
        stage = "Предварительный"
        my_win.tabWidget.setCurrentIndex(3)
    elif sender == my_win.view_fin1_Action:
        stage == "1-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_fin2_Action:
        stage = "2-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_fin3_Action:
        stage = "3-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_fin4_Action:
        stage = "4-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_fin5_Action:
        stage = "5-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_fin6_Action:
        stage = "6-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_fin7_Action:
        stage = "7-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_fin8_Action:
        stage = "8-й финал"
        my_win.tabWidget.setCurrentIndex(5)
        fin = stage
    elif sender == my_win.view_one_table_Action:
        stage = "Одна таблица"
    elif sender == my_win.view_pf1_Action:
        stage = "1-й полуфинал"
        my_win.tabWidget.setCurrentIndex(4)
    elif sender == my_win.view_pf2_Action:
        stage = "2-й полуфинал"
        my_win.tabWidget.setCurrentIndex(4)

    system = sys.select().where(System.stage == stage).get()
    if system.stage == "Предварительный":
        pv = system.page_vid
        table_made(pv, stage)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        pv = system.page_vid
        table_made(pv, stage)
    elif system.stage == "Одна таблица" or system.stage == fin:
        if system.type_table == "круг":
            pv = system.page_vid
            table_made(pv, stage)
        else:
            system_table = system.label_string
            pv = system.page_vid
            if system_table == "Сетка (с розыгрышем всех мест) на 8 участников":
                setka_8_full_made(fin)
            elif system_table == "Сетка (-2) на 8 участников":
                setka_8_2_made(fin)
            elif system_table == "Сетка (с розыгрышем всех мест) на 16 участников":
                setka_16_full_made(fin)
            elif system_table == "Сетка (-2) на 16 участников":
                setka_16_2_made(fin)
            elif system_table == "Сетка (с розыгрышем всех мест) на 32 участников":
                setka_32_full_made(fin)
            elif system_table == "Сетка (-2) на 32 участников":
                setka_32_2_made(fin)
            elif system_table == "Сетка (с розыгрышем всех мест) на 32 участников":
                setka_32_made(fin)    


def setka_type(none_player):
    """сетка ставит очки в зависимости от неявки игрока, встреча состоялась ли пропуск встречи -bye-"""
    sc_total = []
    if my_win.lineEdit_player1_fin.text() == "X" or my_win.lineEdit_player2_fin.text() == "X":
        if my_win.lineEdit_player1_fin.text() != "X":
            winner = my_win.lineEdit_player1_fin.text()
            loser = my_win.lineEdit_player2_fin.text()
        else:
            winner = my_win.lineEdit_player2_fin.text()
            loser = my_win.lineEdit_player1_fin.text()
        w = ""
        l = ""
        st1 = ""
        st2 = ""
    else:
        if none_player == 0:
            st1 = int(my_win.lineEdit_pl1_score_total_fin.text())
            st2 = int(my_win.lineEdit_pl2_score_total_fin.text())
            w = 2
            l = 1
        else:
            if none_player == 1: # не явился 1-й игрок
                st1 = "П"
                st2 = "В"
            elif none_player == 2:
                st1 = "В"
                st2 = "П"
            w = 2
            l = 0
            my_win.lineEdit_pl1_score_total_fin.setText(st1)
            my_win.lineEdit_pl2_score_total_fin.setText(st2)
    sc_total.append(st1)
    sc_total.append(st2)
    sc_total.append(w)
    sc_total.append(l)
    return sc_total


def circle_type(none_player, stage):
    """круговая таблица"""
    sc_total = []
    st1 = ""
    st2 = ""
    w = ""
    l = ""
    if stage == "Предварительный":
        if none_player == 0:
            st1 = int(my_win.lineEdit_pl1_score_total_gr.text())
            st2 = int(my_win.lineEdit_pl2_score_total_gr.text())
            w = 2
            l = 1
        else:
            if none_player == 1:  # не явился 1-й игрок
                st1 = "П"
                st2 = "В"
            elif none_player == 2:  # не явился 2-й игрок
                st1 = "В"
                st2 = "П"
            w = 2
            l = 0
            my_win.lineEdit_pl1_score_total_gr.setText(st1)
            my_win.lineEdit_pl2_score_total_gr.setText(st2)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        if none_player == 0:
            st1 = int(my_win.lineEdit_pl1_score_total_pf.text())
            st2 = int(my_win.lineEdit_pl2_score_total_pf.text())
            w = 2
            l = 1
        else:
            if none_player == 1:  # не явился 1-й игрок
                st1 = "П"
                st2 = "В"
            elif none_player == 2:  # не явился 2-й игрок
                st1 = "В"
                st2 = "П"
            w = 2
            l = 0
            my_win.lineEdit_pl1_score_total_pf.setText(st1)
            my_win.lineEdit_pl2_score_total_pf.setText(st2)    
    else:
        if none_player == 0:
            st1 = int(my_win.lineEdit_pl1_score_total_fin.text())
            st2 = int(my_win.lineEdit_pl2_score_total_fin.text())
            w = 2
            l = 1
        else:
            if none_player == 1:  # не явился 1-й игрок
                st1 = "П"
                st2 = "В"
            elif none_player == 2:  # не явился 2-й игрок
                st1 = "В"
                st2 = "П"
            w = 2
            l = 0
            my_win.lineEdit_pl1_score_total_fin.setText(st1)
            my_win.lineEdit_pl2_score_total_fin.setText(st2)
    sc_total.append(st1)
    sc_total.append(st2)
    sc_total.append(w)
    sc_total.append(l)
    return sc_total


def string_score_game():
    """создает строку со счетом победителя"""
    tab = my_win.tabWidget.currentIndex()
    visible_flag = True
    if tab == 3:
        visible_flag = my_win.checkBox_4.isChecked()
        for i in my_win.groupBox_kolvo_vstrech_gr.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                g = (int(i.text()) + 1) // 2 # число, максимальное кол-во партий для победы
                break
    elif tab == 4:
        visible_flag = my_win.checkBox_14.isChecked()
        for i in my_win.groupBox_kolvo_vstrech_pf.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                g = (int(i.text()) + 1) // 2 # число, максимальное кол-во партий для победы
                break
    else:
        visible_flag = my_win.checkBox_5.isChecked()
        for i in my_win.groupBox_kolvo_vstrech_fin.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                g = (int(i.text()) + 1) // 2
                break
    if tab == 3:
        # поля ввода счета в партии
        st1 = int(my_win.lineEdit_pl1_score_total_gr.text())
        st2 = int(my_win.lineEdit_pl2_score_total_gr.text())
        s11 = my_win.lineEdit_pl1_s1_gr.text()
        s21 = my_win.lineEdit_pl2_s1_gr.text()
        s12 = my_win.lineEdit_pl1_s2_gr.text()
        s22 = my_win.lineEdit_pl2_s2_gr.text()
        s13 = my_win.lineEdit_pl1_s3_gr.text()
        s23 = my_win.lineEdit_pl2_s3_gr.text()
        s14 = my_win.lineEdit_pl1_s4_gr.text()
        s24 = my_win.lineEdit_pl2_s4_gr.text()
        s15 = my_win.lineEdit_pl1_s5_gr.text()
        s25 = my_win.lineEdit_pl2_s5_gr.text()
    elif tab == 4:
        st1 = int(my_win.lineEdit_pl1_score_total_pf.text())
        st2 = int(my_win.lineEdit_pl2_score_total_pf.text())
        s11 = my_win.lineEdit_pl1_s1_pf.text()
        s21 = my_win.lineEdit_pl2_s1_pf.text()
        s12 = my_win.lineEdit_pl1_s2_pf.text()
        s22 = my_win.lineEdit_pl2_s2_pf.text()
        s13 = my_win.lineEdit_pl1_s3_pf.text()
        s23 = my_win.lineEdit_pl2_s3_pf.text()
        s14 = my_win.lineEdit_pl1_s4_pf.text()
        s24 = my_win.lineEdit_pl2_s4_pf.text()
        s15 = my_win.lineEdit_pl1_s5_pf.text()
        s25 = my_win.lineEdit_pl2_s5_pf.text()
    elif tab == 5:
        st1 = int(my_win.lineEdit_pl1_score_total_fin.text())
        st2 = int(my_win.lineEdit_pl2_score_total_fin.text())
        s11 = my_win.lineEdit_pl1_s1_fin.text()
        s21 = my_win.lineEdit_pl2_s1_fin.text()
        s12 = my_win.lineEdit_pl1_s2_fin.text()
        s22 = my_win.lineEdit_pl2_s2_fin.text()
        s13 = my_win.lineEdit_pl1_s3_fin.text()
        s23 = my_win.lineEdit_pl2_s3_fin.text()
        s14 = my_win.lineEdit_pl1_s4_fin.text()
        s24 = my_win.lineEdit_pl2_s4_fin.text()
        s15 = my_win.lineEdit_pl1_s5_fin.text()
        s25 = my_win.lineEdit_pl2_s5_fin.text()
    # создание строки счета победителя
    if st1 > st2:
        if visible_flag is True:
            if int(s11) > int(s21):  # 1-й сет
                n1 = s21
            else:
                n1 = str(f"-{s11}")
            if int(s12) > int(s22):  # 2-й сет
                n2 = s22
            else:
                n2 = str(f"-{s12}")
            if (g == 2 and st1 == 2 and st2 == 0) or (g == 2 and st2 == 0 and st1 == 2):  # из 3-х партий 2-0
                winner_string = f"({n1},{n2})"
                return winner_string
            if int(s13) > int(s23):  # 3-й сет
                n3 = s23
            else:
                n3 = str(f"-{s13}")
            if (g == 2 and st1 == 2 and st2 == 1) or (g == 2 and st2 == 2 and st1 == 1) or \
                    (g == 3 and st1 == 3 and st2 == 0) or (g == 3 and st1 == 0 and st2 == 3):  # из 3-х  2-1 или из 5-и 3-0
                winner_string = f"({n1},{n2},{n3})"
                return winner_string
            if int(s14) > int(s24):  # 4-й сет
                n4 = s24
            else:
                n4 = str(f"-{s14}")
            if (g == 4 and st1 == 4 and st2 == 0) or (g == 4 and st1 == 0 and st2 == 4) or \
                    (g == 3 and st1 == 3 and st2 == 1) or (g == 3 and st1 == 1 and st2 == 3):  # из 5-и 3-1 или из 7-и 4-0
                winner_string = f"({n1},{n2},{n3},{n4})"
                return winner_string
            if int(s15) > int(s25):  # 5-й сет
                n5 = s25
            else:
                n5 = str(f"-{s15}")
            if (g == 4 and st1 == 4 and st2 == 1) or (g == 4 and st1 == 1 and st2 == 4) or \
                    (g == 3 and st1 == 3 and st2 == 2) or (g == 3 and st1 == 2 and st2 == 3):  # из 5-и 3-2 или из 7-и 4-1
                winner_string = f"({n1},{n2},{n3},{n4},{n5})"
        else:
            if visible_flag is True:
                winner_string = f"({st1} : {st2})" 
            else:
                winner_string = f"{st1} : {st2}"      
        return winner_string
    else:
        if visible_flag is True:
            if int(s11) < int(s21):  # 1-й сет
                n1 = s11
            else:
                n1 = str(f"-{s21}")
            if int(s12) < int(s22):  # 2-й сет
                n2 = s12
            else:
                n2 = str(f"-{s22}")
            if (g == 2 and st1 == 2 and st2 == 0) or (g == 2 and st1 == 0 and st2 == 2):  # из 3-х партий 2-0
                winner_string = f"({n1},{n2})"
                return winner_string
            if int(s13) < int(s23):  # 3-й сет
                n3 = s13
            else:
                n3 = str(f"-{s23}")
            if (g == 2 and st1 == 2 and st2 == 1) or (g == 2 and st2 == 2 and st1 == 1) or \
                    (g == 3 and st1 == 3 and st2 == 0) or (g == 3 and st1 == 0 and st2 == 3):  # из 3-х  2-1 или из 5-и 3-0
                winner_string = f"({n1},{n2},{n3})"
                return winner_string
            if int(s14) < int(s24):  # 4-й сет
                n4 = s14
            else:
                n4 = str(f"-{s24}")
            if (g == 4 and st1 == 4 and st2 == 0) or (g == 4 and st1 == 0 and st2 == 4) or \
                    (g == 3 and st1 == 3 and st2 == 1) or (g == 3 and st1 == 1 and st2 == 3):  # из 5-и 3-1 или из 7-и 4-0
                winner_string = f"({n1},{n2},{n3},{n4})"
                return winner_string
            if int(s15) < int(s25):  # 5-й сет
                n5 = s15
            else:
                n5 = str(f"-{s25}")
            if (g == 4 and st1 == 4 and st2 == 1) or (g == 4 and st1 == 1 and st2 == 4) or \
                    (g == 3 and st1 == 3 and st2 == 2) or (g == 3 and st1 == 2 and st2 == 3):  # из 5-и 3-2 или из 7-и 4-1
                winner_string = f"({n1},{n2},{n3},{n4},{n5})"
        else:
            if visible_flag is True:
                winner_string = f"({st2} : {st1})"
            else:
                winner_string = f"{st2} : {st1}"
        return winner_string


def result_filter_name():
    """отсортировывает встречи с участием игрока"""
    cp = my_win.comboBox_find_name.currentText()
    cp = cp.title()  # Переводит первую букву в заглавную
    c = Result.select().where(Result.title_id == title_id())
    c = c.where(Result.player1 ** f'{cp}%')  # like
    result_list = c.dicts().execute()
    row_count = len(result_list)  # кол-во строк в таблице
    column_count = len(result_list[0])  # кол-во столбцов в таблице
    # вставляет в таблицу необходимое кол-во строк
    my_win.tableWidget.setRowCount(row_count)

    for row in range(row_count):  # добавляет данные из базы в TableWidget
        for column in range(column_count):
            item = str(list(result_list[row].values())[column])
            my_win.tableWidget.setItem(
                row, column, QTableWidgetItem(str(item)))


def filter_fin(pl=False):
    """фильтрует таблицу -Result- на вкладке финалы"""
    msgBox = QMessageBox
    data_table_tmp = []
    data_table_list = []
    data = []
    model = MyTableModel(data)
    tb = my_win.tabWidget.currentIndex()
    # player_selected = player_list.dicts().execute()



    sender = my_win.sender()
    num_game_fin = my_win.lineEdit_num_game_fin.text()
    final = my_win.comboBox_filter_final.currentText()
    if final == "":
        final = "все финалы"
    name = my_win.comboBox_find_name_fin.currentText()
    round = my_win.lineEdit_tour.text()
    played = my_win.comboBox_filter_played_fin.currentText()
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    filter = Result.select().where(Result.title_id == title_id())
                                  
    if sender == my_win.lineEdit_num_game_fin:
        fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.tours == num_game_fin))
        count = len(fltr)
        my_win.label_38.setText(f'Всего в {final} {count} игры')
    if final != "все финалы":
        systems = system.select().where(System.stage == final).get()
        id_system = systems.id # получаем ид системы финала
        fltr = filter.select().where(Result.system_id == id_system)
    fin = []
    if final == "Одна таблица":
        if my_win.comboBox_find_name_fin.currentText() != "":
            if pl == False:
                fltr = filter.select().where(Result.player1 == name)
            else:
                fltr = filter.select().where(Result.player2 == name)
            c = len(fltr)
        else:
            if final == "Одна таблица" and played == "все игры" and num_game_fin == "" and round == "":
                fltr = filter.select().where(Result.system_stage == "Одна таблица")
                count = len(fltr)
                my_win.label_38.setText(f'Всего {count} игры')
            elif final == "Одна таблица" and played == "завершенные":
                fl = filter.select().where(Result.system_stage == "Одна таблица")
                fltr = fl.select().where(Result.winner != "")
                count = len(fltr)
                my_win.label_38.setText(f'Сыграно {count} игры')
            elif final == "Одна таблица" and played == "не сыгранные":
                fltr = filter.select().where(Result.system_stage ==
                                             "Одна таблица" and Result.points_win == None)
                count = len(fltr)
                my_win.label_38.setText(f'Не сыграно {count} игры')
            elif final == "Одна таблица" and played == "все игры" and num_game_fin == "" and round != "":
                fl = filter.select().where(Result.system_stage == "Одна таблица")
                fltr = fl.select().where(Result.round == round)
                count = len(fltr)
                my_win.label_38.setText(f'Всего {count} игры')
            elif final == "Одна таблица" and played == "все игры" and num_game_fin != "" and round == "":
                fl = filter.select().where(Result.system_stage == "Одна таблица")
                fltr = fl.select().where(Result.tours == num_game_fin)
    else:
        if final == "все финалы" and played == "все игры" and num_game_fin == "" and round == "":
            fltr = filter.select().where(Result.system_stage == "Финальный")
            if name == "":
                count = len(fltr)
                my_win.label_38.setText(f'Всего в финалах {count} игры')
            else:  # выбор по фамилии спортсмена
                row = 0
                fl = filter.select().where(Result.system_stage == "Финальный")
                fltr = fl.select().where((Result.player1 == name)| (Result.player2 == name)) # объединение запросов (отбор по 2-ум столбцам)
        # один из финалов встречи которые не сыгранные
        elif final != "все финалы" and played == "не сыгранные" and num_game_fin == "" and round == "":
            fl = filter.select().where(Result.system_id == id_system)
            fltr = fl.select().where(Result.points_win == None)
            count = len(fltr)
            my_win.label_38.setText(
                f'Всего в {final} не сыгранно {count} игры')
        elif final != "все финалы" and played == "завершенные" and num_game_fin == "" and round == "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.points_win == 2))
            count_pl = len(fltr)
            my_win.label_38.setText(f'Завершено в {final} {count_pl} игры')
        elif final != "все финалы" and played == "все игры" and num_game_fin == "" and round == "":
            fltr = filter.select().where(Result.system_id == id_system)
            count = len(fltr)
            my_win.label_38.setText(f'Всего в {final} {count} игры')
        elif final != "все финалы" and played == "все игры" and num_game_fin != "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.tours == num_game_fin))
            count = len(fltr)
            my_win.label_38.setText(f'Всего в {final} {count} игры')
        elif final == "все финалы" and played == "все игры" and num_game_fin != "":
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.tours == num_game_fin))
            count = len(fltr)
            my_win.label_38.setText(f'Всего в {final} {count} игры')
        elif final == "все финалы" and played == "все игры" and num_game_fin == "" and round != "":
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.round == int(round)))
            count = len(fltr)
            my_win.label_38.setText(f'Всего в {final} {count} игры')
        elif final != "все финалы" and played == "все игры" and num_game_fin != "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.tours == num_game_fin))
        elif final != "все финалы" and played == "все игры" and num_game_fin == "" and round != "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.round == int(round)))
            count = len(fltr)
            my_win.label_38.setText(f'Всего в {final} {count} игры')  
        elif final != "все финалы" and played == "завершенные" and num_game_fin == "" and round != "":
            fltr_fin = filter.select().where(Result.system_id == id_system)
            fltr = fltr_fin.select().where((Result.round == int(round)) & (Result.points_win == 2))
            count = len(fltr)
            my_win.label_38.setText(f'Всего в {final} {count} игры')   
        elif final == "все финалы" and played == "завершенные" and num_game_fin == "":
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.points_win == 2))
            count_pl = len(fltr)
            my_win.label_38.setText(f' Всего сыграно во всех финалах {count_pl} игры')
        elif final == "все финалы" and played == "не сыгранные" and num_game_fin == "":
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.points_win == None))
            count = len(fltr)
            my_win.label_38.setText(f'Всего в {final} не сыгранно {count} игры')
        else:
            if final != "все финалы" and num_game_fin != "" and round != "":
                fltr = filter.select().where((Result.system_id == id_system) & (Result.round == round))
            else:
                for sys in system:  # отбирает финалы с сеткой
                    if sys.stage != "Предварительный" and sys.stage != "Полуфиналы":
                        txt = sys.label_string
                        txt = txt[:5]
                        if txt == "Сетка":
                            fin.append(sys.stage)
                fin, ok = QInputDialog.getItem(
                    my_win, "Финалы", "Выберите финал, где искать номер встречи.", fin, 0, False)
                fltr = filter.select().where(Result.number_group == fin)
            row = 0
            for result_list in fltr:
                row += 1
                if result_list.tours == num_game_fin:
                    num_game_fin = int(num_game_fin)
                    row_num = num_game_fin - 1
                    my_win.tableView.selectRow(row_num)
                    break

    player_list = fltr
    fill_table(player_list)
    if count == 0: # если в финал по сетке ввели номер тура
        my_win.lineEdit_tour.clear()
        my_win.statusbar.showMessage("Финалы по сетке", 5000)
        return 


def filter_sf():
    """фильтрует таблицу -результаты- на вкладке полуфиналы"""
    data = []
    data_table_tmp = []
    find_player = []
    model = MyTableModel(data)
    sf = ['1-й полуфинал', '2-й полуфинал']
    semifinal = my_win.comboBox_filter_semifinal.currentText()
    group = my_win.comboBox_filter_group_sf.currentText()
    name = my_win.comboBox_find_name_sf.currentText()
    played = my_win.comboBox_filter_played_sf.currentText()
    find_player.append(name)
    fltr_id = Result.select().where(Result.title_id == title_id())
    if group == "все группы" and my_win.comboBox_find_name_sf.currentText() != "":
        if semifinal == "-все полуфиналы-":
            pl1_query = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.player1 == name))
            pl2_query = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.player2 == name))            
        else:
            pl1_query = fltr_id.select().where((Result.system_stage == semifinal) & (Result.player1 == name))
            pl2_query = fltr_id.select().where((Result.system_stage == semifinal) & (Result.player2 == name)) 
        fltr = pl1_query | pl2_query # объдиняет два запроса в один
    elif group == "все группы" and played == "все игры":
        fltr = fltr_id.select().where(Result.system_stage.in_(sf))
    elif group == "все группы" and played == "завершенные":
        if semifinal == "-все полуфиналы-":
            fltr = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.points_win == 2))
        else:
            fltr = fltr_id.select().where((Result.system_stage == semifinal) & (Result.points_win == 2))
    elif group != "все группы" and played == "завершенные":
        if semifinal == "-все полуфиналы-":
            fl = fltr_id.select().where((Result.system_stage.in_(sf)) & (Result.number_group == group))
        else:
            fl = fltr_id.select().where((Result.system_stage == semifinal) & (Result.number_group == group))
        fltr = fl.select().where(Result.points_win == 2)
    elif group != "все группы" and played == "не сыгранные":
        if semifinal == "-все полуфиналы-":
            fl = fltr_id.select().where(Result.system_stage.in_(sf)  & (Result.number_group == group))
        else:
            fl = fltr_id.select().where((Result.system_stage == semifinal) & (Result.number_group == group))
        fltr = fl.select().where(Result.points_win == None)
    elif group == "все группы" and played == "не сыгранные":
        fltr = fltr_id.select().where(Result.points_win != 2 and Result.points_win == None)
    elif group != "все группы" and played == "все игры":
        if semifinal == "-все полуфиналы-":
            fltr = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.number_group == group))
        else:
            fltr = fltr_id.select().where((Result.system_stage == semifinal) & (Result.number_group == group))

    result_list = fltr.dicts().execute()
    row_count = len(result_list)  # кол-во строк в таблице
    if row_count != 0:
        column_count = len(result_list[0])  # кол-во столбцов в таблице
    if played == "завершенные":
        my_win.label_17.setText(f"сыграно {row_count} встреч")
    elif played == "не сыгранные":
        my_win.label_17.setText(f"не сыграно еще {row_count} встреч(а)")
    else:
        my_win.label_17.setText(f"всего {row_count} встреч(а)")
    my_win.label_17.show()
 
    if row_count != 0:
        for row in range(row_count):  # добавляет данные из базы в TableWidget
            for column in range(column_count):
                item = str(list(result_list[row].values())[column])
                data_table_tmp.append(item)
            data.append(data_table_tmp.copy())
            data_table_tmp.clear()
        my_win.tableView.setModel(model)


def filter_gr():
    """фильтрует таблицу -результаты- на вкладке группы"""
    find_player = []

    group = my_win.comboBox_filter_group.currentText()
    name = my_win.comboBox_find_name.currentText()
    played = my_win.comboBox_filter_played.currentText()
    find_player.append(name)
 
    if group == "":
        return
    fltr_id = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == "Предварительный"))
    if group != "все группы":
        player_list = fltr_id.select().where(Result.number_group == group)

    if group == "все группы" and my_win.comboBox_find_name.currentText() != "":
        pl1_query = fltr_id.select().where(Result.player1 == name)
        pl2_query = fltr_id.select().where(Result.player2 == name)
        player_list = pl1_query | pl2_query # объдиняет два запроса в один
    elif group == "все группы" and played == "все игры":
        player_list = fltr_id.select()
    elif group == "все группы" and played == "завершенные":
        player_list = fltr_id.select().where(Result.points_win == 2)
    elif group != "все группы" and played == "завершенные":
        fl = fltr_id.select().where(Result.number_group == group)
        player_list = fl.select().where(Result.points_win == 2)
    elif group != "все группы" and played == "не сыгранные":
        fl = fltr_id.select().where(Result.number_group == group)
        player_list = fl.select().where(Result.points_win != 2 & Result.points_win == None)
    elif group == "все группы" and played == "не сыгранные":
        player_list = fltr_id.select().where((Result.points_win != 2 & Result.points_win == None))

    row_count = len(player_list)  # кол-во строк в таблице

    if played == "завершенные":
        my_win.label_16.setText(f"сыграно {row_count} встреч")
    elif played == "не сыгранные":
        my_win.label_16.setText(f"не сыграно еще {row_count} встреч(а)")
    else:
        my_win.label_16.setText(f"всего {row_count} встреч(а)")
    my_win.label_16.show()

    fill_table(player_list)

 
def load_comboBox_referee():
    """Загружает комбобокс списком судей"""
    msgBox = QMessageBox()
    my_win.comboBox_referee.clear()
    my_win.comboBox_secretary.clear()
    referee_list = []
    referee = Referee.select()
    if len(referee) == 0:
       result = msgBox.information(my_win, "Уведомление", "База данных судей еще пуста!", msgBox.Ok)
       if result == msgBox.Ok:
            return
    else:
        for k in referee:
            family = k.family
            city = k.city
            fam_city = f"{family}/ {city}"
            if fam_city not in referee_list:
                referee_list.append(fam_city)
        my_win.comboBox_referee.addItems(referee_list)
        my_win.comboBox_secretary.addItems(referee_list)      
    return referee_list


def load_combo():
    """загружает комбобокс поиска спортсмена на вкладке группы, пф и финалы фамилиями спортсменов"""
    text = []
    players = Player.select().where(Player.title_id == title_id())
    for i in players:  # цикл по таблице базы данных (I это id строк)
        family = i.player
        city = i.city
        text.append(f"{family}/{city}")
    my_win.comboBox_find_name.addItems(text)
    my_win.comboBox_find_name_sf.addItems(text)
    my_win.comboBox_find_name_fin.addItems(text)
    my_win.comboBox_find_name.setCurrentText("")
    my_win.comboBox_find_name_sf.setCurrentText("")
    my_win.comboBox_find_name_fin.setCurrentText("")


def load_combo_etap_begunki():
    """загружает комбобокс выбора этапов системы на вкладке дополнительно"""
    my_win.comboBox_select_stage_begunki.clear()
    my_win.comboBox_edit_etap1.clear()
    stage_system = ["-Выбор спортсменов-", "Списки участников"]
    results = Result.select().where(Result.title_id == title_id())
    for i in results:
        stage = i.system_stage
        if stage not in stage_system:
            stage_system.append(stage)
    my_win.comboBox_select_stage_begunki.addItems(stage_system)
    my_win.comboBox_edit_etap1.addItems(stage_system)
    my_win.comboBox_edit_etap2.addItems(stage_system)


def reset_filter():
    """сбрасывает критерии фильтрации"""
    sender = my_win.sender()
    if sender == my_win.Button_reset_filter_gr:
        my_win.comboBox_find_name.setCurrentText("")
        my_win.comboBox_filter_played.setCurrentText("все игры")
        my_win.comboBox_filter_group.setCurrentText("все группы")
        filter_gr()
    elif sender == my_win.Button_reset_filter_sf:
        my_win.comboBox_find_name_sf.setCurrentText("")
        my_win.comboBox_filter_played_sf.setCurrentText("все игры")
        my_win.comboBox_filter_group_sf.setCurrentText("все группы")
        filter_sf()
    elif sender == my_win.Button_reset_filter_fin:
        my_win.comboBox_find_name_fin.setCurrentText("")
        my_win.comboBox_filter_played_fin.setCurrentText("все игры")
        my_win.lineEdit_tour.setText("")
        my_win.lineEdit_num_game_fin.setText("")
        if my_win.comboBox_filter_final.currentText() == "Одна таблица":
            my_win.comboBox_filter_final.setCurrentText("Одна таблица")
        else:
            my_win.comboBox_filter_final.setCurrentText("все финалы")
        filter_fin()
    load_combo()


def choice_semifinal_automat(stage):
    """жеребьевка полуфиналов"""
    mesto_first = 0

    system = System.select().where(System.title_id == title_id())
    systems = system.select().where(System.stage == "Предварительный").get()
    total_group = systems.total_group
    system_stage = system.select().where(System.stage == stage).get()
    sys_id = system_stage.id
    mesta_exit = system_stage.mesta_exit

    if stage == "1-й полуфинал":
        mesto_first = 1
    else:
        system_stage = system.select().where(System.stage == "1-й полуфинал").get()
        mesta_exit = system_stage.mesta_exit
        mesto_first = mesta_exit + 1

    for k in range(1, total_group + 1):
        choices = Choice.select().where((Choice.title_id == title_id()) & (Choice.group == f"{k} группа"))
        p = 0 if k <= total_group // 2 else mesta_exit
        n = k if k <= total_group // 2 else total_group - k + 1
        for i in range(mesto_first, mesta_exit + 1):
            p += 1
            choice_mesta = choices.select().where(Choice.mesto_group == i).get()
            with db: # записывает в db номер полуфинала
                choice_mesta.semi_final = stage
                choice_mesta.sf_group = f"{n} группа" # номера группы полуфинала
                choice_mesta.posev_sf = p # номер посева
                choice_mesta.save()
    with db:  # записывает в систему, что произведена жеребъевка
        system = System.get(System.id == sys_id)
        system.choice_flag = True
        system.save()
    player_in_table_group_and_write_Game_list_Result(stage)


def choice_gr_automat():
    "новая система жеребьевки групп"
    " current_region_group - словарь (регион - список номеров групп куда можно сеять)"
    " reg_player - словарь регион ид игрока, player_current - список сеящихся игроков, posev - словарь всего посева"
    posev_tmp = {}
    reg_player = {}
    gr_region = {}
    posev_group = {}
    player_current = []
    pgt = []
    posev = {}
    group_list = []
    start = 0
    end = 1
    step = 0
    stage = "Предварительный"
    sys = System.select().where(System.title_id == title_id())
    sys_id = sys.select().where(System.stage == stage).get()
    group = sys_id.total_group
    max_player = sys_id.max_player  # максимальное число игроков в группе, оно же число посевов
    total_player = sys_id.total_athletes
    for b in range(1,max_player + 1):  # цикл создания словарей (номер посева, списки списков(номер группы и 0 вместо номера регионов))
        for x in range(1, group + 1):
            posev_group[x] = 0
        gr_region = posev_group.copy()
        posev[f"{b}_посев"] = gr_region
        posev_group.clear()
   
    pl_choice = Choice.select().order_by(Choice.rank.desc()).where(Choice.title_id == title_id())
    m = 1  # начальное число посева
    p = 0
    number_poseva = 0  # общий счетчик посева игроков
    reg_list = []
    player_list = []
    for np in pl_choice:
        choice = np.get(Choice.id == np)
        region = choice.region
        pl_id = choice.player_choice_id
        reg = Region.get(Region.region == region)
        region_id = reg.id 
        reg_list.append(region_id)
        player_list.append(pl_id)
    while number_poseva < total_player:
        p += 1
        if number_poseva == 0 or number_poseva % group == 0 :
            group_list = list(range(1, group + 1))  # получение списка групп с помощью функции range

        region_id = reg_list[number_poseva]
        pl_id = player_list[number_poseva]
        posev_tmp = posev[f"{m}_посев"]

        if m == 1:  # 1-й посев       
            posev_tmp[p] = region_id  # создает словарь группа - номер региона
            number_poseva += 1
            player_current.append(pl_id)
            reg_player[pl_id] = number_poseva  # словарь ид игрока его группа при посеве
            if number_poseva == group:  # если доходит окончания данного посева идет запись в db
                choice_save(m, player_current, reg_player)
        else:  # 2-й посев и т.д.
            current_region_group = {}  # словарь регион - список номеров групп куда можно сеять
            key_reg_previous = []
            current = region_player_current(number_poseva, reg_list, group, player_list)  # должен быть получен список текущих регионов посева
            key_reg_current = current[0]  # номера регионов текущего посева
            player_current = current[1]  # номера игроков (id)

            for o in previous_region_group.keys():  # цикл получения списка регионов предыдущих посевов уникальный
                key_reg_previous.append(o)
            pgt.clear()
            remains = total_player - number_poseva  # остаток посева
            finish = 0
            if remains > group: 
                finish = group  # если остаток больше кол-во групп
            else:
                finish = remains            
            for y in range(0, finish):
                group_list_tmp = []  
                z = key_reg_current[y] # список регионов которые уже были посеяны
                pgt.append(y + 1)  # номера групп которые уже посеяны будут удалены из списка

                if z not in key_reg_previous:  # если нет в списке, то добавляет полный список групп
                    current_region_group[z] = group_list
                else:
                    gr_del = previous_region_group[z]  # список групп где уже есть этот регион
                    group_list_tmp = list((Counter(group_list) - Counter(gr_del)).elements()) # удаляет из списка номера групп где уже есть регионы
                    r = len(group_list_tmp)
                    if r == 0:  # если во всех группах уже есть, то начинает опять полный список групп
                        current_region_group[z] = group_list  # получает словарь со списком групп куда сеять
                    else:
                        current_region_group[z] = group_list_tmp  # получает словарь со списком групп куда сеять
                 # система распределения по группам (посев), где m - номер посева начина со 2-ого посева
            sv = add_delete_region_group(key_reg_current, current_region_group, posev_tmp, m, posev, start, end, step, player_current)
            current.clear()
            number_poseva = number_poseva + sv
        if number_poseva != total_player:  # выход из системы жеребьевки при достижении оканчания
            if number_poseva == group * m:  # смена направления посева
                if m % 2 != 0:
                    start = group
                    end = 0
                    step = -1
                else:
                    start = 0
                    end = group
                    step = 1
                m += 1
                previous_region_group = posev_test(posev, group, m)  # возвращает словарь регион  - список номера групп, где он есть
        else:
            fill_table_after_choice()
            with db:  # записывает в систему, что произведена жеребъевка
                system = System.get(System.id == sys_id)
                system.choice_flag = True
                system.save()
            player_in_table_group_and_write_Game_list_Result(stage)
        group_list.clear()


def progress_bar(step):
    """прогресс бар""" 
    msgBox = QMessageBox 
    my_win.progressBar.setValue(step)
    if step >= 99:
       result = msgBox.information(my_win, "Уведомление", "Жеребьевка завершена, проверьте ее результаты!", msgBox.Ok)
       if result == msgBox.Ok:
            my_win.progressBar.setValue(0)
    return step


def check_one_region_in_choice(fin):
    """Проверка на спортсменов одного регионоа в жеребьевке"""
    system = System.select().where(System.title_id == title_id())
    stage_exit = system.stage_exit
    mesta_exit = system.mesta_exit
    choice = Choice.select().where(Choice.stage)


def rank_mesto_out_in_group_or_semifinal_to_final(fin):
    """определение мест, выходящих из группы или полуфинала в финал"""
    systems_stage = System.select().where(System.title_id == title_id())
    system = System.select().where((System.title_id == title_id()) & (System.stage == fin)).get()
    # stage_out = system.stage_exit # откуда вышли в финал
    # how_many_out_in_stage = system.mesta_exit # сколько игроков вышло из этапа в финал
    # == словарь этап число игроков в группе или полуфинале
    player_in_stage = {} # словарь -этап- количество в нем участников
    etap_out_and_player = {}
    num_pl = []
    num_pl_sf1 = []
    num_pl_sf2 = []
    player_out_sf1 = 0
    player_out_sf2 = 0
    for l in systems_stage:
        etap = l.stage
        if etap == "Предварительный":
            num_pl.clear()
            max_pl = l.max_player
            for k in range(1, max_pl +1):
                num_pl.append(k)
            etap_out_and_player[etap] = num_pl
        elif etap == "1-й полуфинал":
            num_pl_sf1.clear()
            max_pl = l.max_player // l.total_group
            for k in range(1, max_pl +1):
                num_pl_sf1.append(k)
            etap_out_and_player[etap] = num_pl_sf1
            player_out_sf1 = l.mesta_exit
            list_mest = etap_out_and_player["Предварительный"]
            del list_mest[:player_out_sf1]
        elif etap == "2-й полуфинал":
            num_pl_sf2.clear()
            max_pl = l.max_player // l.total_group
            for k in range(1, max_pl +1):
                num_pl_sf2.append(k)
            etap_out_and_player[etap] = num_pl_sf2
            player_out_sf2 = l.mesta_exit
            list_mest = etap_out_and_player["Предварительный"]
            del list_mest[:player_out_sf2]
        else:
            system_fin = System.select().where((System.title_id == title_id()) & (System.stage == etap)).get()
            etap_out_fin = system_fin.stage_exit
            pl_out = system_fin.mesta_exit
            list_mest = etap_out_and_player[etap_out_fin]
            if fin != etap:
                del list_mest[:pl_out]
                player_in_stage[etap] = etap_out_and_player[etap_out_fin]
            else:
                del list_mest[pl_out:]
                nums = list_mest
                break
    return nums


def choice_setka_automat(fin, flag, count_exit):
    """автоматическая жеребьевка сетки, fin - финал, count_exit - сколько выходят в финал
    flag - флаг вида жеребьевки ручная или автомат""" 
    msgBox = QMessageBox 
    full_posev = []  # список полного списка участников 1-ого посева
    group_last = []
    number_last = [] # посеянные номера в сетке
    reg_last = []  # посеянные регионы в сетке
    number_posev = []  # список по порядку для посева
    current_region_posev = {} # в текущем посеве список регионов по порядку
    posev_data = {} # окончательные посев номер в сетке - игрок/ город
    num_id_player = {} # словарь номер сетки - id игрока
    possible_number = {}
    #===================================
    system = System.select().where((System.title_id == title_id()) & (System.stage == fin)).get()
    choice = Choice.select().where(Choice.title_id == title_id())
    max_player = system.max_player
  
    posevs = setka_choice_number(fin, count_exit) # выбор номеров сетки для посева
    player_net = posevs[0]
    posev_1 = posevs[1]
    z = len(posevs)

    if z == 3:
        posev_2 = posevs[2]
    elif z == 4:
        posev_2 = posevs[2]
        posev_3 = posevs[3]
    elif z == 5:
        posev_2 = posevs[2]
        posev_3 = posevs[3]
        posev_4 = posevs[4]

    s = 0
    free_seats = 0 # кол-во свободных мест в сетке
    step = 0
    del_num = 0
    free_num = []
    real_all_player_in_final = []

    nums = rank_mesto_out_in_group_or_semifinal_to_final(fin) # получение списка номеров мест выходящих в финал
    stage_exit = system.stage_exit
 
    for n in range (0, count_exit): # начало основного посева
        if system.stage == "Одна таблица":
            real_all_player_in_final.append(len(choice.select().where(Choice.basic == fin)))
        elif fin == "1-й финал":
            if stage_exit == "Предварительный":
                # == реальное число игроков в финале
                real_all_player_in_final = len(choice.select().where(Choice.mesto_group.in_(nums)))
                # == число игроков в конкретном посеве финала
                if n == 0:
                    choice_posev = choice.select().where(Choice.mesto_group == nums[n]).order_by(Choice.group) # если 1-й финал и 1-й посев тос ортирует по группам 
                else:
                    choice_posev = choice.select().where(Choice.mesto_group == nums[n])          
            elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал": # выходят из полуфинала
                real_all_player_in_final = len(choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final.in_(nums))))
                # == число игроков в конкретном посеве финала
                choice_posev = choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final == nums[n])) 
        else:
            num = []
            if stage_exit == "Предварительный":
                for k in range(0, count_exit):
                    num.append(nums[k])
                nums = num.copy()
                choice_posev = choice.select().where(Choice.mesto_group == nums[n])
                real_all_player_in_final = len(choice.select().where(Choice.mesto_group.in_(nums))) # реальное число игроков в сетке
            elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал": # выходят из полуфинала
                choice_posev = choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final == nums[n]))
                real_all_player_in_final = len(choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final.in_(nums))))

        count_player_in_final = len(choice_posev)

        if real_all_player_in_final != max_player and count_exit == 1: # вычеркиваем определенные номера только если одно место выходит из группы
            free_num = free_place_in_setka(max_player, real_all_player_in_final)
            del_num = 1 # флаг, что есть свободные номера
        elif count_player_in_final != max_player // count_exit and count_exit > 1:
            del_num = 1 # флаг, что есть свободные номера
        full_posev.clear()
        for posev in choice_posev: # отбор из базы данных согласно местам в группе для жеребьевки сетки
            psv = []
        
            family = posev.family
            if fin != "Одна таблица":
                if stage_exit == "Предварительный":
                    group = posev.group
                    mesto_group = posev.mesto_group
                else:
                    group = posev.sf_group
                    mesto_group = posev.mesto_semi_final
                ind = group.find(' ')
                group_number = int(group[:ind]) # номер группы
            else:
                group = ""
                group_number = 1
            pl_id = posev.player_choice_id
            region = posev.region
            player = Player.get(Player.id == pl_id)
            city = player.city
            rank = player.rank

            psv = [pl_id, family, region, group_number, group, city, rank, mesto_group]
            full_posev.append(psv)

        if count_exit == 1 or fin == "Одна таблица":
            full_posev.sort(key=lambda k: k[6], reverse=True) # сортировка списка участников по рейтингу
        else:
            full_posev.sort(key=lambda k: k[3]) # сортировка списка участников по группам

        for k in full_posev:
            k.pop(3)
            k.pop(6)
        # ======== начало жеребьевки =========
        end = player_net // count_exit
        for i in range(0, end):
            number_posev.append(i)
        if n == 0:
            posev = posev_1
        elif n == 1:
            posev = posev_2
        elif n == 2:
            posev = posev_3
        elif n == 3:
            posev = posev_4

        count_posev = len(posev)
        for i in range(0, count_posev):  # список посева, разделеный на отдельные посевы
            current_region_posev.clear()
            sev_tmp = posev[i].copy()
            sev = sev_tmp.copy()
            sev_tmp.clear()
            count = len(posev[i]) # всего количество номеров в посеве
            if del_num == 1 and i == count_posev - 1:
                for h in free_num:
                    sev.remove(h)
                free_seats = len(free_num) # сколько свободных мест в сетке
                count = len(posev[i]) - free_seats
                del_num = 0
            for w in range(0, count): # внутренний цикл посева
                l = number_posev[0] # общий список всего посева (порядковый номер посева)
                if i == 0 and n == 0: #  ===== 1-й посев
                    sev = posev[i]  # список номеров посева
                    num_set = sev[w] # номер в сетке на который идет сев
                    count_sev = len(sev) # количество номеров в посеве
                else:
                    num_set = sev[0] # проверить
                    count_sev = len(sev) # конкретное число оставшихся в посеве минус свободных мест(если они есть)
                    if count_sev > 1: # если сеющихся номеров больше одного
                        if w == 0: # 1-й основной посев
                            gr_region_tmp = []
                            for k in range(l, l + count_sev):
                                region = full_posev[k][2]
                                gr = full_posev[k][3]
                                gr_region_tmp.append(region)
                                gr_region_tmp.append(gr)
                                gr_region = gr_region_tmp.copy()
                                current_region_posev[k] = gr_region # словарь регионы, в текущем посеве по порядку
                                gr_region_tmp.clear()
                        number_last.clear()
                        number_last = list(num_id_player.keys()) # список уже посеянных номеров в сетке

                        reg_last.clear()
                        group_last.clear()
                        for v in num_id_player.values():
                            reg_last.append(v[1]) # список уже посеянных регионов
                            group_last.append(v[2]) # список номеров групп уже посеянных
                        if n != 0 or (n == 0 and l > 1):
                # =========== определения кол-во возможны вариантов посева у каждого региона
                            # possible_number = possible_draw_numbers(current_region_posev, reg_last, number_last, group_last, n, sev, num_id_player, player_net)   
                            possible_number = possible_draw_numbers(current_region_posev, reg_last, number_last, group_last, n, sev, num_id_player, player_net, count_exit)                        

                            if i != 0 or n != 0: # отсортирововаем список по увеличению кол-ва возможных вариантов
                                possible_number = {k:v for k,v in sorted(possible_number.items(), key=lambda x:len(x[1]))}
                                num_posev = list(possible_number.keys())   
                            l = list(possible_number.keys())[0]
                            num_set = possible_number[l] # номер куда можно сеять
                            # === выбор ручная или автомат ====
                            if flag is True: # автоматичекая
                                if len(num_set) == 0:
                                    msgBox.information(my_win, "Уведомление", "Автоматическая жеребьевка не получилась, повторите снова.")
                                    sorted_tuple = sorted(num_id_player.items(), key=lambda x: x[0])
                                    dict(sorted_tuple)                                    
                                    player_choice_in_setka(fin)
                                    step = 0
                                elif len(num_set) != 1: # есть выбор из номеров случайно
                                    num_set = random_generator(num_set)
                                elif len(num_set) == 1: # остался только один номер
                                    num_set = num_set[0]
                            else: # manual
                                my_win.tableView.setGeometry(QtCore.QRect(260, 241, 841, 540))
                                player_list = []
                                player_list_tmp = []

                                for j in possible_number.keys():
                                    posev_list = full_posev[j]
                                    pl = posev_list[1] # фамилия
                                    reg = posev_list[2] # регион
                                    pn = possible_number[j] # возможные номера посева
                                    player_list_tmp.append(pl)
                                    player_list_tmp.append(reg)
                                    player_list_tmp.append(pn)   
                                    player_list.append(player_list_tmp.copy())
                                    player_list_tmp.clear()
                                txt_tmp = []
    
                                for g in player_list:
                                    if len(num_id_player) == 2:
                                        fam_city = ""
                                        number_net = ""
                                        view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки
                                    t_str = str(g[2])
                                    txt_str = f"{g[0]} - {g[1]} номера: {t_str}" 
                                    txt_tmp.append(txt_str)
                                text_str = (',\n'.join(txt_tmp))
                                tx = f"Список спортсменов в порядке посева:\n\n{text_str}\n\n" + "Выберите один из номеров и нажмите\n - ОК - если выбрали сами или - Cancel - если хотите выбор случайный"
                                txt = (','.join(list(map(str, num_set))))
                                while True:
                                    try:
                                        number_net, ok = QInputDialog.getText(my_win, f'Возможные номера посева: {txt}', tx)
                                        znak = text_str.find(":")
                                        fam_city = text_str[:znak - 7]
                                        if not ok:
                                            number_net = random.choice(num_set)
                                        msgBox.information(my_win, "Жеребьевка участников", f"{fam_city} идет на номер: {number_net}")
                                        number_net = int(number_net)
                                        view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки
                                    except ValueError:
                                        msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.")
                                        continue
                                    else:
                                        if number_net in num_set:
                                            num_set = number_net
                                            break
                                        else:
                                            msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.") 
                id_player = full_posev[l][0]
                region = full_posev[l][2]
                gr = full_posev[l][3]  
                id_region = []
                id_region.append(id_player)
                id_region.append(region)
                id_region.append(gr)
                num_id_player[num_set] = id_region
            # ======== модуль удаления посеянных номеров =========
                if count_sev > 1:
                    c = len(current_region_posev)
                    if c != 0:
                        del possible_number[l] # удаляет из словаря возможных номеров посеянный порядковый номер
                        del current_region_posev[l] # удаляет из словаря текущий посеянный регион
                        if num_set in sev: # проверяет посеянный номер в посеве
                            sev.remove(num_set)  # удаляет посеянный номер из всех номеров этого посева
                        for z in possible_number.keys():
                            possible_tmp = possible_number[z]
                            #=====
                            if flag is False and len(possible_number) == 1:
                                # number_net = possible_tmp[0]
                                number_net = sev[0]
                                fam_city = f"{pl}/{reg}"
                                view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки 
                            #======
                            if num_set in possible_tmp: # проверяет посеянный номер в возможных номерах
                                possible_tmp.remove(num_set) # удаляет посеянный номер из возможных номеров
                elif count_sev == 1: # удаляет последний номер в посеве
                    sev.clear()
                    possible_number.clear()
                number_posev.remove(l)
                if i != 0:
                    num_posev.remove(l)

                sp = 100 / (real_all_player_in_final)
                step += sp
                # progress_bar(step)
        if step > 99:    
            for i in num_id_player.keys():
                tmp_list = list(num_id_player[i])
                id = tmp_list[0]
                pl_id = Player.get(Player.id == id)
                family_city = pl_id.full_name
                posev_data[i] = family_city
                with db:
                    choice_final = choice.select().where(Choice.player_choice_id == pl_id).get()
                    choice_final.final = fin
                    choice_final.posev_final = i
                    choice_final.save()
            key_set = set(num_id_player.keys()) # получаем сет всех ключей (номеров сетки)
            for j in range(1, player_net + 1):
                free_num.append(j)
            free_number = set((free_num))
            free_number.difference_update(key_set) # вычитаем из всех номеров те которые посеяны и остается номера -X-
            for h in free_number:
                posev_data[h] = "X"
    return posev_data



# def delete_free_seats(current_region_posev, possible_number, num_set, sev, l):
#     """удаляет из возможных посевов номера отсутствующих игроков"""
#     # c = len(current_region_posev)
#     # if c != 0:
#     del possible_number[l] # удаляет из словаря возможных номеров посеянный порядковый номер
#     del current_region_posev[l] # удаляет из словаря текущий посеянный регион
#     if num_set in sev: # проверяет посеянный номер в посеве
#         sev.remove(num_set)  # удаляет посеянный номер из всех номеров этого посева
#     for z in possible_number.keys():
#         possible_tmp = possible_number[z]
#         if num_set in possible_tmp: # проверяет посеянный номер в возможных номерах
#             possible_tmp.remove(num_set) # удаляет посеянный номер из возможных номеров


def view_table_choice(fam_city, number_net, num_id_player):
    """показ таблицы жеребьевки"""
    data = []
    num_fam = []
    manual_choice_dict = {}
    player = Player.select().where(Player.title_id == title_id())
    count_player = max(num_id_player.keys()) # наибольшой ключ в словаре (на сколько сетка)

    manual_choice_dict = num_id_player.copy()

    for r in range(1, count_player + 1):
        manual_choice_dict.setdefault(r, "-")
        list_net = manual_choice_dict[r]
        if r == number_net:
            num_fam_tmp = [r, fam_city, ""]
        elif list_net == "-":
            num_fam_tmp = [r, list_net, ""]
        else:
            id_player = list_net[0]
            group = list_net[2]
            group.replace("группа", "гр.")
            pl_full = player.select().where(Player.id == id_player).get()
            player_full = pl_full.full_name
            num_fam_tmp = [r, player_full, group]
        num_fam = num_fam_tmp.copy()
        num_fam_tmp.clear()
        data.append(num_fam) # список списков
        model = MyTableModel(data)
        my_win.tableView_net.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        my_win.tableView_net.verticalHeader().setDefaultSectionSize(15)
        my_win.tableView_net.setGridStyle(QtCore.Qt.DashDotLine) # вид линии сетки 
        my_win.tableView_net.setModel(model)
        my_win.tableView_net.show()


def setka_choice_number(fin, count_exit):
    """номера сетки при посеве"""
    posevs = []
    posev_1 = []
    posev_2 = []
    posev_3 = []
    posev_4 = []

    system = System.select().where((System.title_id == title_id()) & (System.stage == fin)).get()
 
    type_setka = system.label_string
    if count_exit == 1:
        if type_setka == "Сетка (с розыгрышем всех мест) на 8 участников" or type_setka == "Сетка (-2) на 8 участников":
            posev_1 = [[1, 8], [4, 5], [2, 3, 6, 7]]
            player_net = 8
        elif type_setka == "Сетка (с розыгрышем всех мест) на 16 участников" or type_setka == "Сетка (-2) на 16 участников":
            posev_1 = [[1, 16], [8, 9], [4, 5, 12, 13], [2, 3, 6, 7, 10, 11, 14, 15]]
            player_net = 16
        elif type_setka == "Сетка (с розыгрышем всех мест) на 32 участников" or type_setka == "Сетка (-2) на 32 участников":
            posev_1 = [[1, 32], [16, 17], [8, 9, 24, 25], [4, 5, 12, 13, 20, 21, 28, 29], [2, 3, 6, 7, 10, 11, 14, 15, 18, 19, 22, 23, 26, 27, 30, 31]]
            player_net = 32
    elif count_exit == 2:
        if type_setka == "Сетка (с розыгрышем всех мест) на 8 участников" or type_setka == "Сетка (-2) на 8 участников":
            posev_1 = [[1, 8], [4, 5]]
            posev_2 = [[2, 3, 6, 7]]
            player_net = 8
        elif type_setka == "Сетка (с розыгрышем всех мест) на 16 участников" or type_setka == "Сетка (-2) на 16 участников":
            posev_1 = [[1, 16], [8, 9], [4, 5, 12, 13]]
            posev_2 = [[2, 3, 6, 7, 10, 11, 14, 15]]
            player_net = 16
        elif type_setka == "Сетка (с розыгрышем всех мест) на 32 участников" or type_setka == "Сетка (-2) на 32 участников":
            posev_1 = [[1, 32], [16, 17], [8, 9, 24, 25], [4, 5, 12, 13, 20, 21, 28, 29]]
            posev_2 = [[2, 3, 6, 7, 10, 11, 14, 15, 18, 19, 22, 23, 26, 27, 30, 31]]
            player_net = 32
    elif count_exit == 3:
        pass
    elif count_exit == 4:
        if type_setka == "Сетка (с розыгрышем всех мест) на 8 участников" or type_setka == "Сетка (-2) на 8 участников":
            posev_1 = [[1, 8]]
            posev_2 = [[4, 5]]
            posev_3 = [[3, 6]]
            posev_4 = [[2, 7]]
            player_net = 8
        elif type_setka == "Сетка (с розыгрышем всех мест) на 16 участников" or type_setka == "Сетка (-2) на 16 участников":
            posev_1 = [[1, 16], [8, 9]]
            posev_2 = [[4, 5, 12, 13]]
            posev_3 = [[3, 6, 11, 14]]
            posev_4 = [[2, 7, 10, 15]]
            player_net = 16
        elif type_setka == "Сетка (с розыгрышем всех мест) на 32 участников" or type_setka == "Сетка (-2) на 32 участников":
            posev_1 = [[1, 32], [16, 17], [8, 9, 24, 25]]
            posev_2 = [[4, 5, 12, 13, 20, 21, 28, 29]]
            posev_3 = [[3, 6, 11, 14, 19, 22, 27, 30]]
            posev_4 = [[2, 7, 10, 15, 18, 23, 26, 31]]
            player_net = 32
    posevs.append(player_net)
    if len(posev_1) != 0:
        posevs.append(posev_1)
        if len(posev_2) != 0:
            posevs.append(posev_2)
            if len(posev_3) != 0:
                posevs.append(posev_3)
                if len(posev_4) != 0:
                    posevs.append(posev_4)
    return posevs


def free_place_in_setka(max_player, real_all_player_in_final):
    """вычеркиваем свободные номера в сетке"""
    free_num = []
    free_number_8 = [2, 7, 6, 3]
    free_number_16 = [2, 15, 7, 10, 6, 11, 3, 14]
    free_number_24 = [5, 20, 8, 17, 11, 14, 2, 23]
    free_number_32 = [2, 31, 15, 18, 10, 23, 7, 26, 6, 27, 11, 22, 14, 19, 3, 30]
    count = max_player - real_all_player_in_final# кол-во свободных мест


    if max_player == 8:
        free_number = free_number_8
    if max_player == 16:
        free_number = free_number_16
    elif max_player == 24:
        free_number = free_number_24
    elif max_player == 32:
        free_number = free_number_32

    for i in range (0, count):
        k = free_number[i]
        free_num.append(k)
    return free_num
    

# def possible_draw_numbers(current_region_posev, reg_last, number_last, group_last, n, sev, num_id_player, player_net):
#     """возможные номера посева"""
#     possible_number = {}
#     proba_possible = {} 
#     num_tmp = []
#     reg_tmp = []
# 
#     current_region = list(current_region_posev.values())
#     y = 0
#     for reg in current_region_posev.keys():
#         cur_reg = current_region[y][0] # текущий регион посева
#         cur_gr = current_region[y][1] # номер группы, которая сеятся

#         if n == 0:
#             if cur_reg in reg_last: # если регион который сеятся есть в уже посеянных областях
#                 reg_tuple = tuple(reg_last)
#                 count = reg_tuple.count(cur_reg) # количество регионов уже посеянных 
#                 if count == 1: # значит только один регион в посеве
#                     cur_gr = current_region[y][1]
#                     number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net)
#                     possible_number[reg] = number_posev
#                 else: # если есть уже областей более двух
#                     number_tmp = []
#                     num_tmp.clear()
#                     start = 0
#                     for k in reg_last: # получаем список номеров сетки областей в той половине куда идет сев
#                         if k == cur_reg:
#                             index = reg_last.index(k, start)
#                             set_number = number_last[index] # номер где уже посеянна такая же область
#                             num_tmp.append(set_number)
#                         start += 1
#                     if count % 2 == 0: # если число четное
#                         if count == 2: # посеяны 2 области разводит по четвертям
#                             for h in num_tmp:
#                                 if h <= player_net // 4: # если номер в сетке вверху, то наде сеять вниз
#                                     f = [i for i in sev if i >= player_net // 4 + 1 and i <= player_net // 2] # отсеивает в списке номера 9-16
#                                 elif h > player_net // 4 and h <= player_net // 2: 
#                                     f = [i for i in sev if i <= player_net // 4] # отсеивает в списке номера 1-8
#                                 elif h >= player_net // 2 + 1 and h <= int(player_net * 3 / 4): 
#                                     f = [i for i in sev if i > player_net * 3 / 4] # отсеивает в списке номера 25-32
#                                 elif h > player_net * 3 / 4: 
#                                     f = [i for i in sev if i >= player_net // 2 + 1 and i <= int(player_net * 3 / 4)] # отсеивает в списке номера 17-24
#                                 number_tmp += f
#                         elif count == 4: # посеяны 4 области разводит по восьмушкам
#                             if player_net == 16:
#                                 for h in num_tmp:
#                                     if h <= 2: # если номер в сетке 1-2
#                                         f = [i for i in sev if i >= 3 and i <= 4] # отсеивает в списке номера 3-4 ()
#                                     elif h >= 3 and h <= 4: # если номер в сетке 3-4
#                                         f = [i for i in sev if i < 3] # отсеивает в списке номера 1-2 ()
#                                     elif h >= 5 and h <= 6: # если номер в сетке 5-6
#                                         f = [i for i in sev if i >= 7 and i <= 8] # отсеивает в списке номера 25-32
#                                     elif h >= 7 and h <= 8: # если номер в сетке 7-8
#                                         f = [i for i in sev if i >= 5 and i <= 6] # отсеивает в списке номера 17-24
#                                     elif h >= 9 and h <= 10: # если номер в сетке вверху, то наде сеять вниз
#                                         f = [i for i in sev if i >= 11 and i <= 12] # отсеивает в списке номера 9-16
#                                     elif h >= 11 and h <= 12: 
#                                         f = [i for i in sev if i <= 9 and i <= 10] # отсеивает в списке номера 1-8
#                                     elif h >= 13 and h <= 14: 
#                                         f = [i for i in sev if i > 14] # отсеивает в списке номера 25-32
#                                     elif h > 14: 
#                                         f = [i for i in sev if i >= 12 and i <= 13] # отсеивает в списке номера 17-24    
#                                     number_tmp += f
#                             elif player_net == 32:
#                                 for h in num_tmp:
#                                     if h <= player_net // 8: # если номер в сетке вверху, то наде сеять вниз
#                                         f = [i for i in sev if i >= 5 and i <= 8] # отсеивает в списке номера 3-4 ()
#                                     elif h >= 5 and h <= 8: 
#                                         f = [i for i in sev if i < 5] # отсеивает в списке номера 1-2 ()
#                                     elif h >= 9 and h <= 12: 
#                                         f = [i for i in sev if i >= 13 and i <= 16] # отсеивает в списке номера 25-32
#                                     elif h >= 13 and h <= 16: 
#                                         f = [i for i in sev if i >= 9 and i <= 12] # отсеивает в списке номера 17-24
#                                     elif h >= 17 and h <= 20: # если номер в сетке вверху, то наде сеять вниз
#                                         f = [i for i in sev if i >= 21 and i <= 24] # отсеивает в списке номера 9-16
#                                     elif h >= 21 and h <= 24: 
#                                         f = [i for i in sev if i >= 17 and i <= 20] # отсеивает в списке номера 1-8
#                                     elif h >= 25 and h <= 28: 
#                                         f = [i for i in sev if i >= 29] # отсеивает в списке номера 25-32
#                                     elif h > 28: 
#                                         f = [i for i in sev if i >= 25 and i <= 28] # отсеивает в списке номера 17-24    
#                                     number_tmp += f
#                     elif count > 2:
#                         # number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net, count_exit)
#                         # if 
#                         number_posev = sev
#                         number_tmp = alignment_in_half(player_net, num_tmp, sev, count, number_posev)
                       
#                     number_posev = number_tmp.copy()
#                     possible_number[reg] = number_posev
#             else: # все номера в той части куда можно сеять
#                 possible_number[reg] = sev
#         else: # 2-й посев и последующие 
#             number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net) # возможные номера после ухода от своей группы без учета регионов
#             number_posev_old = number_setka_posev_last(cur_gr, group_last, number_last, n, player_net)
#             reg_tmp.clear()
#             # ======
#             if n > 1:
#                 for k in number_posev_old: # получаем список прошлых посеянных областей в той половине куда идет сев
#                     d = number_last.index(k)
#                     reg_tmp.append(reg_last[d]) # список регионов     
#                 if cur_reg in reg_tmp: # если сеянная область есть в прошлом посеве конкретной половины
#                     num_tmp = [] # список номеров сетки где есть такой же регион (в той половине или четверти с номером который сеятся)
#                     for d in number_posev_old: # номер в сетке в предыдущем посеве
#                         posev_tmp = num_id_player[d]
#                         if cur_reg in posev_tmp:
#                             num_tmp.append(d) # список номеров в сетке, где уже есть такой же регион
#                     count = len(num_tmp) # количество областей в той части сетки, куде сеятся регион
#                 # ======== отбирает номера из -number_posev- , где учитывается регион ======
#                     if count == 1 and n == 1: # есть только одна область в той же половине другой четверти (1 место и 2-е место в группе)
#                         if num_tmp[0] <= player_net // 4: # в первой четверти (1-8)
#                             number_posev = [i for i in number_posev if i > player_net // 4 and i <= player_net // 2] # номера 8-16
#                         elif num_tmp[0] >= (player_net // 4 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (9-16)
#                             number_posev = [i for i in number_posev if i < 9] # номера 1-8
#                         elif num_tmp[0] >= (player_net // 2 + 1) and num_tmp[0] <= player_net // 4 * 3: # в первой четверти (16-24)
#                             number_posev = [i for i in number_posev if i > player_net // 4 * 3] # номера 25-32
#                         elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net: # в первой четверти (25-32)
#                             number_posev = [i for i in number_posev if i > player_net // 2 and i < (player_net // 4 * 3 + 1)] # номера 17-24
#                     elif (count == 1 and n == 2):
#                         if num_tmp[0] <= player_net // 8: # в первой четверти (1-4)
#                             number_posev = [i for i in number_posev if i > player_net // 8 and i <= player_net // 4] # номера 5-8
#                         elif num_tmp[0] >= player_net // 8 + 1 and num_tmp[0] <= player_net // 4: # в первой четверти (5-8)
#                             number_posev = [i for i in number_posev if i < player_net // 8 + 1] # номера 1-4
#                         elif num_tmp[0] >= player_net // 4 + 1 and num_tmp[0] <= player_net // 8 * 3: # в первой четверти (9-12)
#                             number_posev = [i for i in number_posev if i > player_net // 8 * 3 and i <= player_net // 2] # номера 13-16
#                         elif num_tmp[0] >= (player_net // 8 * 3 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (13-16)
#                             number_posev = [i for i in number_posev if i >= player_net // 4 + 1 and i <= player_net // 8 * 3] # номера 9-12
#                         elif num_tmp[0] >= player_net // 2 + 1 and num_tmp[0] <= player_net // 8 * 5: # в первой четверти (17-20)
#                             number_posev = [i for i in number_posev if i > player_net // 8 * 5 and i <= player_net // 4 * 3] # номера 21-24
#                         elif num_tmp[0] >= player_net // 8 * 5 and num_tmp[0] <= (player_net // 4 * 3): # в первой четверти (21-24)
#                             number_posev = [i for i in number_posev if i >(player_net // 2 + 1) and i <= player_net // 8 * 5] # номера 17-20
#                         elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net // 8 * 7: # в первой четверти (25-28)
#                             number_posev = [i for i in number_posev if i > player_net  // 8 * 7 + 1] # номера 29-32
#                         elif num_tmp[0] >= player_net // 8 * 7 + 1: # в первой четверти (29-32)
#                             number_posev = [i for i in number_posev if i >= player_net // 4 * 3 + 1 and i <= player_net  // 8 * 7] # номера 25-28
#                     elif n == 3:
#                         if count == 1 and len(number_posev) != 1:
#                             if num_tmp[0] <= player_net // 8: # в первой четверти (1-4)
#                                 number_posev = [i for i in number_posev if i > player_net // 8 and i <= player_net // 4] # номера 5-8
#                             elif num_tmp[0] >= player_net // 8 + 1 and num_tmp[0] <= player_net // 4: # в первой четверти (5-8)
#                                 number_posev = [i for i in number_posev if i < player_net // 8 + 1] # номера 1-4
#                             elif num_tmp[0] >= player_net // 4 + 1 and num_tmp[0] <= player_net // 8 * 3: # в первой четверти (9-12)
#                                 number_posev = [i for i in number_posev if i > player_net // 8 * 3 and i <= player_net // 2] # номера 13-16
#                             elif num_tmp[0] >= (player_net // 8 * 3 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (13-16)
#                                 number_posev = [i for i in number_posev if i >= player_net // 4 + 1 and i <= player_net // 8 * 3] # номера 9-12
#                             elif num_tmp[0] >= player_net // 2 + 1 and num_tmp[0] <= player_net // 8 * 5: # в первой четверти (17-20)
#                                 number_posev = [i for i in number_posev if i > player_net // 8 * 5 and i <= player_net // 4 * 3] # номера 21-24
#                             elif num_tmp[0] >= player_net // 8 * 5 and num_tmp[0] <= (player_net // 4 * 3): # в первой четверти (21-24)
#                                 number_posev = [i for i in number_posev if i >= (player_net // 2 + 1) and i <= player_net // 8 * 5] # номера 17-20
#                             elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net // 8 * 7: # в первой четверти (25-28)
#                                 number_posev = [i for i in number_posev if i > player_net  // 8 * 7 + 1] # номера 29-32
#                             elif num_tmp[0] >= player_net // 8 * 7 + 1: # в первой четверти (29-32)
#                                 number_posev = [i for i in number_posev if i >= player_net // 4 * 3 + 1 and i <= player_net  // 8 * 7] # номера 25-28
#                     else:  
#                         number_tmp = alignment_in_half(player_net, num_tmp, sev, count, number_posev) # номер (а)куда можно сеять
#                         number_posev.clear()
#                         number_posev = number_tmp.copy()         

#             possible_number[reg] = number_posev
#             proba_possible[cur_gr] = number_posev
#         y += 1
#     return possible_number



def possible_draw_numbers(current_region_posev, reg_last, number_last, group_last, n, sev, num_id_player, player_net, count_exit):
    """возможные номера посева new"""
    possible_number = {}
    proba_possible = {} 
    num_tmp = []
    reg_tmp = []
    # =========
    titles = Title.get(Title.id == title_id())
    multi_reg = titles.multiregion
    #============
    current_region = list(current_region_posev.values())
    y = 0
    for reg in current_region_posev.keys():
        cur_reg = current_region[y][0] # текущий регион посева
        cur_gr = current_region[y][1] # номер группы, которая сеятся
        #=======
        if multi_reg == 0 or (len(num_id_player) >= player_net // 2 and count_exit == 1): # если спортсмены одного региона нет рассеивания
            possible_number[reg] = sev
        else:
            if n == 0:
                if cur_reg in reg_last: # если регион который сеятся есть в уже посеянных областях
                    reg_tuple = tuple(reg_last)
                    count = reg_tuple.count(cur_reg) # количество регионов уже посеянных 
                    if count == 1: # значит только один регион в посеве
                        cur_gr = current_region[y][1]
                        number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net)
                        possible_number[reg] = number_posev
                    else: # если есть уже областей более двух
                        number_tmp = []
                        num_tmp.clear()
                        start = 0
                        for k in reg_last: # получаем список номеров сетки областей в той половине куда идет сев
                            if k == cur_reg:
                                index = reg_last.index(k, start)
                                set_number = number_last[index] # номер где уже посеянна такая же область
                                num_tmp.append(set_number)
                            start += 1
                        if count % 2 == 0: # если число четное
                            if count == 2: # посеяны 2 области разводит по четвертям
                                for h in num_tmp:
                                    if h <= player_net // 4: # если номер в сетке вверху, то наде сеять вниз
                                        f = [i for i in sev if i >= player_net // 4 + 1 and i <= player_net // 2] # отсеивает в списке номера 9-16
                                    elif h > player_net // 4 and h <= player_net // 2: 
                                        f = [i for i in sev if i <= player_net // 4] # отсеивает в списке номера 1-8
                                    elif h >= player_net // 2 + 1 and h <= int(player_net * 3 / 4): 
                                        f = [i for i in sev if i > player_net * 3 / 4] # отсеивает в списке номера 25-32
                                    elif h > player_net * 3 / 4: 
                                        f = [i for i in sev if i >= player_net // 2 + 1 and i <= int(player_net * 3 / 4)] # отсеивает в списке номера 17-24
                                    number_tmp += f
                            elif count == 4: # посеяны 4 области разводит по восьмушкам
                                if player_net == 16:
                                    for h in num_tmp:
                                        if h <= 2: # если номер в сетке 1-2
                                            f = [i for i in sev if i >= 3 and i <= 4] # отсеивает в списке номера 3-4 ()
                                        elif h >= 3 and h <= 4: # если номер в сетке 3-4
                                            f = [i for i in sev if i < 3] # отсеивает в списке номера 1-2 ()
                                        elif h >= 5 and h <= 6: # если номер в сетке 5-6
                                            f = [i for i in sev if i >= 7 and i <= 8] # отсеивает в списке номера 25-32
                                        elif h >= 7 and h <= 8: # если номер в сетке 7-8
                                            f = [i for i in sev if i >= 5 and i <= 6] # отсеивает в списке номера 17-24
                                        elif h >= 9 and h <= 10: # если номер в сетке вверху, то наде сеять вниз
                                            f = [i for i in sev if i >= 11 and i <= 12] # отсеивает в списке номера 9-16
                                        elif h >= 11 and h <= 12: 
                                            f = [i for i in sev if i <= 9 and i <= 10] # отсеивает в списке номера 1-8
                                        elif h >= 13 and h <= 14: 
                                            f = [i for i in sev if i > 14] # отсеивает в списке номера 25-32
                                        elif h > 14: 
                                            f = [i for i in sev if i >= 12 and i <= 13] # отсеивает в списке номера 17-24    
                                        number_tmp += f
                                elif player_net == 32:
                                    for h in num_tmp:
                                        if h <= player_net // 8: # если номер в сетке вверху, то наде сеять вниз
                                            f = [i for i in sev if i >= 5 and i <= 8] # отсеивает в списке номера 3-4 ()
                                        elif h >= 5 and h <= 8: 
                                            f = [i for i in sev if i < 5] # отсеивает в списке номера 1-2 ()
                                        elif h >= 9 and h <= 12: 
                                            f = [i for i in sev if i >= 13 and i <= 16] # отсеивает в списке номера 25-32
                                        elif h >= 13 and h <= 16: 
                                            f = [i for i in sev if i >= 9 and i <= 12] # отсеивает в списке номера 17-24
                                        elif h >= 17 and h <= 20: # если номер в сетке вверху, то наде сеять вниз
                                            f = [i for i in sev if i >= 21 and i <= 24] # отсеивает в списке номера 9-16
                                        elif h >= 21 and h <= 24: 
                                            f = [i for i in sev if i >= 17 and i <= 20] # отсеивает в списке номера 1-8
                                        elif h >= 25 and h <= 28: 
                                            f = [i for i in sev if i >= 29] # отсеивает в списке номера 25-32
                                        elif h > 28: 
                                            f = [i for i in sev if i >= 25 and i <= 28] # отсеивает в списке номера 17-24    
                                        number_tmp += f
                        elif count > 2:
                            # number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net, count_exit)
                            # if 
                            number_posev = sev
                            number_tmp = alignment_in_half(player_net, num_tmp, sev, count, number_posev)
                        
                        number_posev = number_tmp.copy()
                        possible_number[reg] = number_posev
                else: # все номера в той части куда можно сеять
                    possible_number[reg] = sev
            else: # 2-й посев и последующие 
                number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net) # возможные номера после ухода от своей группы без учета регионов
                number_posev_old = number_setka_posev_last(cur_gr, group_last, number_last, n, player_net)
                reg_tmp.clear()
                # ======
                if n > 1:
                    for k in number_posev_old: # получаем список прошлых посеянных областей в той половине куда идет сев
                        d = number_last.index(k)
                        reg_tmp.append(reg_last[d]) # список регионов     
                    if cur_reg in reg_tmp: # если сеянная область есть в прошлом посеве конкретной половины
                        num_tmp = [] # список номеров сетки где есть такой же регион (в той половине или четверти с номером который сеятся)
                        for d in number_posev_old: # номер в сетке в предыдущем посеве
                            posev_tmp = num_id_player[d]
                            if cur_reg in posev_tmp:
                                num_tmp.append(d) # список номеров в сетке, где уже есть такой же регион
                        count = len(num_tmp) # количество областей в той части сетки, куде сеятся регион
                    # ======== отбирает номера из -number_posev- , где учитывается регион ======
                        if count == 1 and n == 1: # есть только одна область в той же половине другой четверти (1 место и 2-е место в группе)
                            if num_tmp[0] <= player_net // 4: # в первой четверти (1-8)
                                number_posev = [i for i in number_posev if i > player_net // 4 and i <= player_net // 2] # номера 8-16
                            elif num_tmp[0] >= (player_net // 4 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (9-16)
                                number_posev = [i for i in number_posev if i < 9] # номера 1-8
                            elif num_tmp[0] >= (player_net // 2 + 1) and num_tmp[0] <= player_net // 4 * 3: # в первой четверти (16-24)
                                number_posev = [i for i in number_posev if i > player_net // 4 * 3] # номера 25-32
                            elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net: # в первой четверти (25-32)
                                number_posev = [i for i in number_posev if i > player_net // 2 and i < (player_net // 4 * 3 + 1)] # номера 17-24
                        elif (count == 1 and n == 2):
                            if num_tmp[0] <= player_net // 8: # в первой четверти (1-4)
                                number_posev = [i for i in number_posev if i > player_net // 8 and i <= player_net // 4] # номера 5-8
                            elif num_tmp[0] >= player_net // 8 + 1 and num_tmp[0] <= player_net // 4: # в первой четверти (5-8)
                                number_posev = [i for i in number_posev if i < player_net // 8 + 1] # номера 1-4
                            elif num_tmp[0] >= player_net // 4 + 1 and num_tmp[0] <= player_net // 8 * 3: # в первой четверти (9-12)
                                number_posev = [i for i in number_posev if i > player_net // 8 * 3 and i <= player_net // 2] # номера 13-16
                            elif num_tmp[0] >= (player_net // 8 * 3 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (13-16)
                                number_posev = [i for i in number_posev if i >= player_net // 4 + 1 and i <= player_net // 8 * 3] # номера 9-12
                            elif num_tmp[0] >= player_net // 2 + 1 and num_tmp[0] <= player_net // 8 * 5: # в первой четверти (17-20)
                                number_posev = [i for i in number_posev if i > player_net // 8 * 5 and i <= player_net // 4 * 3] # номера 21-24
                            elif num_tmp[0] >= player_net // 8 * 5 and num_tmp[0] <= (player_net // 4 * 3): # в первой четверти (21-24)
                                number_posev = [i for i in number_posev if i >(player_net // 2 + 1) and i <= player_net // 8 * 5] # номера 17-20
                            elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net // 8 * 7: # в первой четверти (25-28)
                                number_posev = [i for i in number_posev if i > player_net  // 8 * 7 + 1] # номера 29-32
                            elif num_tmp[0] >= player_net // 8 * 7 + 1: # в первой четверти (29-32)
                                number_posev = [i for i in number_posev if i >= player_net // 4 * 3 + 1 and i <= player_net  // 8 * 7] # номера 25-28
                        elif n == 3:
                            if count == 1 and len(number_posev) != 1:
                                if num_tmp[0] <= player_net // 8: # в первой четверти (1-4)
                                    number_posev = [i for i in number_posev if i > player_net // 8 and i <= player_net // 4] # номера 5-8
                                elif num_tmp[0] >= player_net // 8 + 1 and num_tmp[0] <= player_net // 4: # в первой четверти (5-8)
                                    number_posev = [i for i in number_posev if i < player_net // 8 + 1] # номера 1-4
                                elif num_tmp[0] >= player_net // 4 + 1 and num_tmp[0] <= player_net // 8 * 3: # в первой четверти (9-12)
                                    number_posev = [i for i in number_posev if i > player_net // 8 * 3 and i <= player_net // 2] # номера 13-16
                                elif num_tmp[0] >= (player_net // 8 * 3 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (13-16)
                                    number_posev = [i for i in number_posev if i >= player_net // 4 + 1 and i <= player_net // 8 * 3] # номера 9-12
                                elif num_tmp[0] >= player_net // 2 + 1 and num_tmp[0] <= player_net // 8 * 5: # в первой четверти (17-20)
                                    number_posev = [i for i in number_posev if i > player_net // 8 * 5 and i <= player_net // 4 * 3] # номера 21-24
                                elif num_tmp[0] >= player_net // 8 * 5 and num_tmp[0] <= (player_net // 4 * 3): # в первой четверти (21-24)
                                    number_posev = [i for i in number_posev if i >= (player_net // 2 + 1) and i <= player_net // 8 * 5] # номера 17-20
                                elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net // 8 * 7: # в первой четверти (25-28)
                                    number_posev = [i for i in number_posev if i > player_net  // 8 * 7 + 1] # номера 29-32
                                elif num_tmp[0] >= player_net // 8 * 7 + 1: # в первой четверти (29-32)
                                    number_posev = [i for i in number_posev if i >= player_net // 4 * 3 + 1 and i <= player_net  // 8 * 7] # номера 25-28
                        else:  
                            number_tmp = alignment_in_half(player_net, num_tmp, sev, count, number_posev) # номер (а)куда можно сеять
                            number_posev.clear()
                            number_posev = number_tmp.copy()         

                possible_number[reg] = number_posev
                proba_possible[cur_gr] = number_posev
            y += 1
    return possible_number


def alignment_in_half(player_net, num_tmp, sev, count, number_posev):
    """выравнивание количество областей по половинам 
    -num_tmp- номера где уже есть эта область
    -number_tmp- номера куда можно сеять,
    -number_posev- возможные номера посева
    -count- число регионов посеянных """
    number_tmp = [] 
    upper_half = 0
    quarter_num = -1
    su = 0
    sd = 0
    max_num = max(num_tmp)
    min_num = min(num_tmp)
    if count % 2 != 0: # нечетное число регионов
        upper_half = len([i for i in num_tmp if i <= player_net // 2]) # количество областей в верхней половине сетки 1-16
        if upper_half == count: # все области в верху сетки  1-16
            quarter_num = len([i for i in num_tmp if i <= player_net // 4]) # количество областей в верхней четверти сетки 1-8
            sev_tmp = [i for i in sev if i <= player_net // 2] # оставляет номера нижней половины
        elif upper_half == 0: # все области в низу сетки  17-32:
            quarter_num = len([i for i in num_tmp if i <= player_net * 3 / 4]) # количество областей в верхней четверти сетки 17-24
            sev_tmp = [i for i in sev if i > player_net // 2] # оставляет номера нижней половины
        else: # посеянные области в разных половинах
            for t in num_tmp:
                if t > player_net / 2:
                    sd += 1 # в нижней половине
                else:
                    su += 1 # в вверхней половине
            if sd > su: # больше областей в низу
                sev_tmp = [i for i in sev if i <= player_net // 2] # оставляет номера вверхней половины
                num_tmp = [i for i in num_tmp if i <= player_net // 2] # получает номер, который один в половине сетки 
            else: # больше областей в вверху
                sev_tmp = [i for i in sev if i > player_net // 2] # оставляет номера нижней половины
                num_tmp = [i for i in num_tmp if i > player_net // 2] # получает номер, который один в половине сетки 

            for k in num_tmp:  
                if k <= player_net // 4: # если номер в сетке вверху, то наде сеять вниз 1-8 (1-4)
                    np = [i for i in sev_tmp if i > player_net // 4] # 1-я четверть 32(9-16) 16(5-8)
                elif k > player_net // 4 and k <= player_net // 2: # 9-16 (5-8)
                    np = [i for i in sev_tmp if i <= player_net // 4] # 2-я четверть 32(1-8) 16(1-4)
                elif k > player_net // 2  and k <= player_net * 3 / 4: # 17-24 (9-12)
                    np = [i for i in sev_tmp if i > player_net * 3 / 4] # 3-я четверть 32(17-24) 16(9-12)
                elif k > player_net * 3 / 4: 
                    np = [i for i in sev_tmp if i <= player_net * 3 / 4] # 4-я четверть 32(25-32) 16(13-16)
                number_tmp += np

        if quarter_num == -1:
            return number_tmp
        else:
            if quarter_num == 0:
                pass
            elif quarter_num == 1:
                num_tmp = [min_num]
            elif quarter_num == 2:
                num_tmp = [max_num]
            elif quarter_num == count: # все области с 1 по 8
                pass
        number_posev = sev_tmp

    for k in num_tmp:
        if k <= 4: # в первой четверти (1-4)
            np = [i for i in number_posev if i >= 5 and i <= 8]
        elif k >= 5 and k <= 8: # в первой четверти (5-8)
            np = [i for i in number_posev if i >= 1 and i <= 4]
        elif k >= 9 and k <= 12: # в первой четверти (9-12)
            np = [i for i in number_posev if i >= 13 and i <= 16]
        elif k >= 13 and k <= 16: # в первой четверти (13-16)
            np = [i for i in number_posev if i >= 9 and i <= 12]
        elif k >= 17 and k <= 20: # в первой четверти (17-20)
            np = [i for i in number_posev if i >= 21 and i <= 24]
        elif k >= 21 and k <= 24: # в первой четверти (21-24)
            np = [i for i in number_posev if i >= 17 and i <= 20]
        elif k >= 25 and k <= 28: # в первой четверти (25-28)
            np = [i for i in number_posev if i >= 29 and i <= 32]
        elif k >= 29: # в первой четверти (29-32)
            np = [i for i in number_posev if i >= 25 and i <= 28]
        number_tmp += np
    return number_tmp


def number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net):
    """промежуточные номера для посева в сетке после ухода от своей группы при выоде из группы больше двух"""
    
    if n == 0:
        if cur_reg in reg_last:
            index = reg_last.index(cur_reg)
            set_number = number_last[index] # номер где уже посеянна такая же область 
            if set_number <= player_net // 2: # если номер в сетке вверху, то наде сеять вниз
                number_posev = [i for i in sev if i > player_net // 2] # отсеивает в списке номера больше 16
            else: 
                number_posev = [i for i in sev if i <= player_net // 2] # отсеивает в списке номера больше 16 
    elif n == 1: # уводит 2-е место от 1-ого в другую половину
        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, от которой надо увести 
        if set_number <= player_net // 2: # если номер в сетке вверху, то наде сеять вниз
            number_posev = [i for i in sev if i > player_net // 2] # номера от 17 до 32
        else: 
            number_posev = [i for i in sev if i <= player_net // 2] # номера от 1 до 16 
    elif n > 1: 
        if n == 2: # уводит 3-е место от 2-ого в другую четверть
            group_last = group_last[8:]
            number_last = number_last[8:] # список номеров 2-ого посева
        elif n == 3: # уводит 4-е место от 1-ого в другую четверть
            group_last = group_last[:8] 
            number_last = number_last[:8] # номера 1 мест в группах
        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, во 4-ом посеве от которой надо увести

        if set_number <= player_net // 4: # если номер в сетке вверху, то наде сеять вниз
            number_posev = [i for i in sev if i >= (player_net // 4 + 1) and i < (player_net // 2 + 1)] # номера от 9 до 17
        elif set_number > player_net // 4 and set_number < (player_net // 2 + 1): # если номер в сетке вверху, то наде сеять вниз: 
            number_posev = [i for i in sev if i <= player_net // 4] # номера от 1 до 8 
        elif set_number > player_net // 2 and set_number < (player_net // 4 * 3 + 1): # если номер в сетке вверху, то наде сеять вниз: 
            number_posev = [i for i in sev if i >= (player_net // 4 * 3 + 1)] # номера от 25 до 32   
        elif set_number >= (player_net // 4 * 3 + 1): # если номер в сетке вверху, то наде сеять вниз: 
            number_posev = [i for i in sev if i >= (player_net // 2 + 1) and i < (player_net // 4 * 3 + 1)] # номера от 17 до 24

    return number_posev


def number_setka_posev_last(cur_gr, group_last, number_last, n, player_net):
    """промежуточные номера для посева в сетке
     -number_last- посеянные номера""" 
    if n == 0:
        if cur_gr in group_last:
            index = group_last.index(cur_gr)
            set_number = number_last[index] # номер где уже посеянна такая же область 
            if set_number <= player_net // 2: # если номер в сетке вверху, то надо сеять вниз
                number_posev_old  = [i for i in number_last if i > player_net // 2] # отсеивает в списке номера больше 16
            else: 
                number_posev_old  = [i for i in number_last if i <= player_net // 2] # отсеивает в списке номера больше 16 
    elif n == 1: # уводит 2-е место от 1-ого в другую половину
        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, от которой надо увести 
        if set_number <= player_net // 2: # если номер в сетке вверху, то наде сеять вниз
            number_posev_old  = [i for i in number_last if i > player_net // 2] # номера от 17 до 32
        else: 
            number_posev_old  = [i for i in number_last if i <= player_net // 2] # номера от 1 до 16 
    elif n > 1: 
        if n == 2: # уводит 3-е место от 2-ого в другую четверть
            group_last = group_last[8:] 
            number_last = number_last[8:]  
        elif n == 3: # уводит 4-е место от 1-ого в другую четверть
            group_last = group_last[:8] 
            number_last = number_last[:8]

        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, от которой надо увести 
        if set_number <= 8: # если номер в сетке вверху, то наде сеять вниз
            number_posev_old  = [i for i in number_last if i >= 9 and i < 17] # номера от 9 до 17
        elif set_number > 8 and set_number < 17: # если номер в сетке вверху, то наде сеять вниз: 
            number_posev_old  = [i for i in number_last if i <= 8] # номера от 1 до 8 
        elif set_number > 16 and set_number < 25: # если номер в сетке вверху, то наде сеять вниз: 
            number_posev_old  = [i for i in number_last if i >= 25] # номера от 25 до 32   
        elif set_number >= 25: # если номер в сетке вверху, то наде сеять вниз: 
            number_posev_old  = [i for i in number_last if i >= 17 and i < 25] # номера от 17 до 24

    return number_posev_old


def random_generator(posev_tmp):
    """выдает случайное число из предложенного списка"""
    num_set = random.choice(posev_tmp)
    return num_set


def add_delete_region_group(key_reg_current, current_region_group, posev_tmp, m, posev, start, end, step, player_current):
    """при добавлении в группу региона удалении номера группы из списка сеянных -b- номер группы
    -m- номер посева, kol_group_free - словарь регион и кол-во свободных групп"""
    free_list = []
    reg_list = []
    kol_group_free = {}
    reg_player = dict.fromkeys(player_current, 0)
    player_list = player_current.copy()
    sv = 0
    if start == 0:
        end = len(key_reg_current)
    else:
        start = len(key_reg_current)

    for s in range(start, end, step):
        sv += 1
        for i in key_reg_current:  # получение словаря (регион и кол-во мест (групп) куда можно сеять)
            tmp = current_region_group[i] 
            kol_reg = len(tmp)  # колво регионов (посевов)
            kol_group_free[i] = kol_reg
        # =====
        sorted_tuple = sorted(kol_group_free.items(), key=lambda x: x[1])
        kol_group_free = dict(sorted_tuple)
        # =====
        free_list = list(kol_group_free.values())  # список кол-во свободных групп, куда можно сеять
        reg_list = list(kol_group_free.keys())  # список ключей (регионов)
        last = len(reg_list)  # кол-во остатка посева
        region = reg_list[0]  # номер региона, который сейчас сеется
        # free_gr = kol_group_free[i]  # кол-во групп куда можно сеять
        free_gr = kol_group_free[region]  # кол-во групп куда можно сеять
        # ==== сделать последний посев по наименшему количеству вариантов посева
 
        if 1 in free_list and last > 1 or last == 1 and free_gr == 1 :  # проверка есть ли группа где осталось только одно места для посева
            # сделать посев 1 регион но много групп
            region = reg_list[free_list.index(1)]  # регион если в списке free list есть значение -1-, т.е. осталась одна группа
            u = current_region_group[region][0]  # номер группы 
            values = posev_tmp[u] 
            if values == 0:
                posev_tmp[u] = region  # запись региона в группу (посев)
        else:
            f = current_region_group[region]  # список номеров групп для посева текущего региона
            temp_list = []
            if free_gr != 1:
                # f = current_region_group[region]  # список номеров групп для посева текущего региона
                if len(f) == 0:
                    # temp_list = []
                    for i in range(1, 33):
                        temp_list.append(i)
                    current_region_group[region] = temp_list
                    f = current_region_group[region]  # список номеров групп для посева текущего региона
                if m % 2 != 0:  # в зависимости от четности посева меняет направления посева групп в списке
                     f.sort()
                else:
                    f.sort(reverse = True)
                if s in f:
                    posev_tmp[s] = region
                    u = s #  присваивает переменной u - номер группы, если она идет по порядку
                else:
                    g = f[0] # номер группы
                    posev_tmp[g] = region
                    u = g    # присваивает переменной u - номер группы, если она идет не по порядку
                # else: # во все группах уже есть одинаковое количество регионов и теперь посев снвоа во все группы
                #     pass
            

            # f = current_region_group[region]  # список номеров групп для посева текущего региона
            elif free_gr == 0:
                # temp_list = []                   
                for i in range (1, len(posev) + 1):
                    gr_dict = posev[f"{m}_посев"]
                    gr = gr_dict[i]
                    if gr == 0:
                        temp_list.append(i)
                current_region_group[region] = temp_list
                f = current_region_group[region]                      
            if m % 2 != 0:  # в зависимости от четности посева меняет направления посева групп в списке
                f.sort()
            else:
                f.sort(reverse = True)
            if s in f:
                posev_tmp[s] = region
                u = s #  присваивает переменной u - номер группы, если она идет по порядку
            else:
                g = f[0]
                posev_tmp[g] = region
                u = g    # присваивает переменной u - номер группы, если она идет не по порядку
        # ====не правильное соответствие номера региона и номера группы
        index = key_reg_current.index(region)
        p = player_list[index]
        reg_player[p] = u
        #=====================
        posev[f"{m}_посев"] = posev_tmp
        for d in key_reg_current:  # цикл удаления посеянных групп
            list_group = current_region_group[d]
            if u in list_group:  # находит сеяную группу и удаляет ее из списка групп
                list_group.remove(u)
        player_list.remove(p)       
        key_reg_current.remove(region)  # удаляет регион из списка как посеянный
        count_in_list = key_reg_current.count(region)
        if count_in_list == 0:  # если в посеве больше одного региона, то пропускает удаление из словаря
            del current_region_group[region] 
            del kol_group_free[region]

        if start > end:
            start -= 1
        else:
            start += 1 
    choice_save(m, player_current, reg_player)        
    return sv


def choice_save(m, player_current, reg_player):
    """запись в db результаты жеребьевки конкретного посева"""
    for i in player_current:
        num_group = reg_player[i]
        with db:  # запись в таблицу Choice результата жеребъевки
            choice = Choice.get(Choice.player_choice_id == i)
            choice.group = f"{num_group} группа"
            choice.posev_group = m
            choice.save()


def region_player_current(number_poseva, reg_list, group, player_list):
    """ создание списка номеров регионов в порядке посева для текущего номера посева """
    key_reg_current = []
    key_tmp = []
    player_current = []
    pl_tmp = []
    current = []
    r = 0
    p = 0
    start = number_poseva
    count = len(player_list)  # кол-во игрок
    remains = count - number_poseva  # остаток посева
    if remains > group: 
        end = start + group  # если остаток больше кол-во групп
    else:
        end = start + remains
    for k in range(start, end):
        r = reg_list[k]
        key_tmp.append(r)
        p = player_list[k]
        pl_tmp.append(p)

    key_reg_current = key_tmp.copy()
    player_current = pl_tmp.copy()
    key_tmp.clear()
    pl_tmp.clear()
    current.append(key_reg_current)
    current.append(player_current)
    return current


def posev_test(posev, group, m):
    """возвращает словарь предыдущих посевов регион - группы, где они есть"""
    uniq_region = []  # уникальный список регионов которые уже посеяны
    tmp_posev = {}
    previous_region_group = {} 
    gr = [] 
    gr_tmp = []
    # список регионов данного посева
    for p in range(1, m):
        tmp_posev = posev[f"{p}_посев"]
        for a in range(1, group + 1):
            v = tmp_posev.setdefault(a)
            if v not in uniq_region:
                uniq_region.append(v)
    # уникальный список регионов
    for val in uniq_region:  # цикл получения словаря (номер региона - список групп где они уже есть)
        for d in range(1, m):
            for key, value in posev[f"{d}_посев"].items():
                if val == value:
                    gr_tmp.append(key)
        gr = gr_tmp.copy()
        previous_region_group[val] = gr
        gr_tmp.clear()
    return previous_region_group


def choice_setka(fin):
    """жеребьевки сетки"""
    system = System.select().where((System.title_id == title_id()) & (System.stage == fin)).get()# находит system id последнего

    flag = system.choice_flag
    if flag is True:  # перед повторной жеребьевкой
        del_choice = Game_list.select().where((Game_list.title_id == title_id()) & (Game_list.number_group == fin))
        for i in del_choice:
            i.delete_instance()  # удаляет строки финала (fin) из таблицы -Game_list
        
        del_result = Result.select().where((Result.title_id == title_id()) & (Result.number_group == fin))
        for i in del_result:
            i.delete_instance()  # удаляет строки финала (fin) из таблицы -Result-

        with db:  # запись в таблицу Choice результата жеребъевки
            system.choice_flag = False
            system.save()


def select_stage_for_edit():
    """выбор финалов или номеров групп для редактирования игроков """
    group_list = []
    stage = ""
    title = ""
    sender = my_win.sender()
    etap_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    if sender == my_win.comboBox_edit_etap1:
        my_win.comboBox_first_group.clear()
    else:
        my_win.comboBox_second_group.clear()

    systems = System.select().where(System.title_id == title_id())
    players = Player.select().where(Player.title_id == title_id())
    if sender == my_win.comboBox_edit_etap1:
        index = my_win.comboBox_edit_etap1.currentIndex()
        stage = my_win.comboBox_edit_etap1.currentText()
    elif sender ==  my_win.comboBox_edit_etap2:
        index = my_win.comboBox_edit_etap2.currentIndex()
        stage = my_win.comboBox_edit_etap2.currentText()
    else:
        stage = "-Выбор этапа-"

    if index == 0:
       return
    elif index == 1:
        for k in players:
            f_name = k.full_name
            group_list.append(f_name)
        group_list.sort()
        title = "-Выбор спортсменов-"
    elif stage == "Одна таблица":
        pass
    elif stage in etap_list:
        sys_id = systems.select().where(System.stage == stage).get()
        group = sys_id.total_group
        group_list = [f"{i} группа" for i in range(1, group + 1)] # генератор списка
        title = "-Выбор группы-"
    elif stage not in etap_list:
        etap_fin_list = []
        for s in systems:
            etap = s.stage
            if etap not in etap_list:
                system_choice = systems.select().where(System.stage == etap).get()
                flag = system_choice.choice_flag
                if flag is True:
                    etap_fin_list.append(etap)

        group_list = [i for i in etap_fin_list] # генератор списка
        title = "-Выбор финала-"
    group_list.insert(0, title)
    if sender == my_win.comboBox_edit_etap1:
        my_win.comboBox_first_group.addItems(group_list)
        my_win.comboBox_first_group.setEnabled(True)
    else:
        my_win.comboBox_second_group.addItems(group_list)
        my_win.comboBox_second_group.setEnabled(True)
    if my_win.comboBox_edit_etap1.currentIndex() != 0 and my_win.comboBox_edit_etap2.currentIndex() != 0:
        my_win.Button_change_player.setEnabled(True)


def edit_group_after_draw():
    """редактирование групп после жеребьевки"""
    if my_win.comboBox_edit_etap1.currentText() == "-Выбор этапа-":
        return
    else:
        stage = my_win.comboBox_edit_etap1.currentText()
    if my_win.comboBox_edit_etap2.currentText() == "-Выбор этапа-":
        return
    else:
        stage = my_win.comboBox_edit_etap2.currentText()
    my_win.tableView.setVisible(False)
    my_win.comboBox_first_group.clear()
    my_win.comboBox_second_group.clear()
    system = System.select().where(System.title_id == title_id())
    system_group = system.select().where(System.stage == stage).get()

    players = Player.select().where(Player.title_id == title_id())
    total_gr = system_group.total_group
    group = [f"{i} группа" for i in range(1, total_gr + 1)] # генератор списка
    group.insert(0, "-выберите группу-")   
    my_win.comboBox_first_group.addItems(group)
    my_win.comboBox_second_group.addItems(group)
    player = [k.full_name for k in players]
    player.sort()
    my_win.comboBox_player_group_edit.addItems(player)


def add_item_listwidget():
    """добавление элементов в листвиджет"""
    flag_combo = 0
    sender = my_win.sender()
    choices = Choice.select().where(Choice.title_id == title_id())
    my_win.tableView.setVisible(False)
    coach_list = []
    coach = ""
    if sender == my_win.comboBox_first_group:
        my_win.listWidget_first_group.clear()
        gr = my_win.comboBox_first_group.currentText()
    else:
        my_win.listWidget_second_group.clear()
        gr = my_win.comboBox_second_group.currentText()

    if gr == "":
        return
    else:
        if sender == my_win.comboBox_first_group:
            if my_win.comboBox_edit_etap1.currentText() == "Предварительный":
                group = choices.select().where(Choice.group == gr).order_by(Choice.posev_group)
            elif my_win.comboBox_edit_etap1.currentText() == "1-й полуфинал":
                group = choices.select().where(Choice.sf_group == gr).order_by(Choice.posev_sf)
            else: # финалы
                final = my_win.comboBox_first_group.currentText()
                group = choices.select().where(Choice.final == final)
               
        else:
            if my_win.comboBox_edit_etap2.currentText() == "Предварительный":
                group = choices.select().where(Choice.group == gr).order_by(Choice.posev_group)
            elif my_win.comboBox_edit_etap2.currentText() == "1-й полуфинал":
                group = choices.select().where(Choice.sf_group == gr).order_by(Choice.posev_sf) 
            else: # финалы
                final = my_win.comboBox_second_group.currentText()
                group = choices.select().where(Choice.final == final)
        n = 0
        for k in group:
            item = QListWidgetItem()
            n += 1
            family = k.family
            region = k.region
            coach = k.coach
            text = f"{n}:{family}/{region}/{coach}"
            item.setText(text) 
            if sender == my_win.comboBox_first_group:
                my_win.listWidget_first_group.addItem(item)
            else:
                my_win.listWidget_second_group.addItem(item)
            coach_list.append(coach)
   
        # duplicat = duplicat_coach_in_group(coach_list)
        # if duplicat is not None:
        #     color_coach_in_listwidget(duplicat, flag_combo)
        # color_coach_in_tablewidget(duplicat, coach_list)


def color_coach_in_listwidget(duplicat, flag_combo):
    """отмечает строки с повторяющимися тренерами"""
    if flag_combo == 1:
        item = my_win.listWidget_first_group.item
        count = my_win.listWidget_first_group.count()
    else:
        item = my_win.listWidget_second_group.item
        count = my_win.listWidget_second_group.count()
    for row in range(count):
        find_coach = []
        data_lw = item(row).text()
        mark = data_lw.rfind("/")
        coach_in_row = data_lw[mark + 1:]
        find_mark_1 = coach_in_row.find(",")
        if find_mark_1 != -1:
            coach_first = coach_in_row[:find_mark_1]
            find_mark_2 = coach_in_row.find(",", find_mark_1 + 1)
            find_coach.append(coach_first)
            if find_mark_2 == -1:
                coach_second = coach_in_row[find_mark_1 + 2:]
                find_coach.append(coach_second)
            else:
                coach_second = coach_in_row[find_mark_1 + 2:find_mark_2]
                coach_third = coach_in_row[find_mark_2 + 2:]
                find_coach.append(coach_second)
                find_coach.append(coach_third)
        else:
            find_coach.append(coach_in_row)
        for k in duplicat:
            if k in find_coach:
                item(row).setForeground(QColor(0, 0, 255)) # изменяет весь текст на синий


def list_player_in_group_after_draw():
    """Смена игроков в группах после жеребьевки при отметки в listwidget при редакитровании"""
    sender = my_win.sender()
    if sender == my_win.Button_add_pl1:
        for row in range(my_win.listWidget_first_group.count()):
            select_item = my_win.listWidget_first_group.selectedItems()
        for i in select_item:
            player_first = i.text()
            if my_win.lineEdit_change_pl1.text() == "":
                my_win.lineEdit_change_pl1.setText(player_first)
            else:
                my_win.lineEdit_change_pl1_2.setText(player_first)
    else:
        for row in range(my_win.listWidget_second_group.count()):
            select_item = my_win.listWidget_second_group.selectedItems()
        for i in select_item:
            player_second = i.text()
            if my_win.lineEdit_change_pl2.text() == "":
                my_win.lineEdit_change_pl2.setText(player_second)
            else:
                my_win.lineEdit_change_pl2_2.setText(player_second)


def change_player_between_group_after_draw():
    """Смена игроков в группах после жеребьевки при отметки в listwidget при редакитровании"""
    msgBox = QMessageBox
    player_dict = {}
    game_list = Game_list.select().where(Game_list.title_id == title_id())
    choices = Choice.select().where(Choice.title_id == title_id())
    systems = System.select().where(System.title_id == title_id())
    players = Player.select().where(Player.title_id == title_id())

    etap_1 = my_win.comboBox_edit_etap1.currentText()
    etap_2 = my_win.comboBox_edit_etap2.currentText()
    player1 = my_win.lineEdit_change_pl1.text()
    player2 = my_win.lineEdit_change_pl2.text()
    player1_2 = my_win.lineEdit_change_pl1_2.text() # 2-й игрок из группы для смены в ПФ
    player2_2 = my_win.lineEdit_change_pl2_2.text() # 2-й игрок из группы для смены в ПФ
    gr_pl1 = my_win.comboBox_first_group.currentText() # номер группы
    gr_pl2 = my_win.comboBox_second_group.currentText() # номер группы
    group_list = [gr_pl1, gr_pl2]
    etap_list = [etap_1, etap_2]
    player_list = [player1, player2, player1_2, player2_2]
    #  === получаем full_name для определения id игроков
    full_name_list = []
    for pl in player_list:
        if pl != "":
            znak = pl.find(":")
            znak1 = pl.find("/") 
            znak2 = pl.rfind("/")
            family_name = pl[znak + 1:znak1]
            if znak == -1:
                player_id = players.select().where(Player.full_name == pl).get()               
            else:
                region = pl[znak1 + 1:znak2]
                player_id = players.select().where((Player.player == family_name) & (Player.region == region)).get()
            pl_id = player_id.id
            full_name = player_id.full_name
            full_name_list.append(full_name)
            player_dict[full_name] = pl_id
    # подсчитывает колличество не пустых значений (кол-во участников)
    element_count = len([item for item in player_list if item != ""]) # подсчитывает колличество не пустых значений (кол-во участников) участвующих в редактировании
    # колличество игроков в группе если добавляется игрок из списка (который не прожеребьен)
    if element_count  == 0:
        result = msgBox.information(my_win, "Уведомление", "Вы не выбрали игроков группы!", msgBox.Ok)
        return
    elif element_count == 1: 
        if etap_1 == "Списки участников" or etap_2 == "Списки участников": # добавляет игрока из списка участников в группу
            stage = etap_1 if etap_1 == "Предварительный" else etap_2
            gr = gr_pl1 if etap_1 != "Списки участников" else gr_pl2
            # колличество игроков в группе если добавляется игрок из списка (который не прожеребьен)
            count_in_group = my_win.listWidget_first_group.count() if my_win.listWidget_first_group.count() != 0 else my_win.listWidget_second_group.count()
            # все данные игрока, которого не было в жерербьевке
            for family_pl in full_name_list:
                if family_pl != "":
                    id_pl = player_dict[family_pl]
                    players_data = players.select().where(Player.id == id_pl).get()
                    region = players_data.region
                    coach_id = players_data.coach_id
                    coachs = Coach.select().where(Coach.id == coach_id).get()
                    coach_family = coachs.coach
                    rank = players_data.rank
        else: # перемещает игрока в другую группу
            stage = etap_1 if etap_1 == "Предварительный" else etap_2
            if my_win.listWidget_first_group.count() < my_win.listWidget_second_group.count():
                count_in_group = my_win.listWidget_first_group.count() 
                gr = gr_pl1
            else:
                count_in_group = my_win.listWidget_second_group.count()
                gr = gr_pl2
        system = systems.select().where(System.stage == stage).get()
        system_etap_id = system.id # id этапа

        posev, ok = QInputDialog.getInt(my_win, "Номер посева", "Введите номер посева", min=1, max=(count_in_group + 1))
        if not ok:
            return
        else:
            if posev <= count_in_group: # если пытаются заменить игрока в группе
                result = msgBox.question(my_win, "Уведомление", "Вы хотите заменить игрока группы\n"
                f"{posev} посева?", msgBox.No, msgBox.Ok) 
                if result == msgBox.No:
                    return
            elif posev > count_in_group:  #=== заменяем спортсмена в группу на последний посев и  обновляет Choice (хотя он потом встает в группе по R)  
                query = Choice.update(group=gr, posev_group=posev).where(Choice.player_choice_id == pl_id) # обновляет запись в Choice                 
                query.execute()
            else: # == если добавляют игрока в конец группы                   
                with db:
                    game_list = Game_list(number_group=gr, 
                                        rank_num_player=posev, 
                                        player_group_id=family_name,
                                        system_id=system_etap_id, 
                                        title_id=title_id()
                                        ).save()
                    # если новый игрок, которого не было в жеребьевке
                    choice = Choice(player_choice_id=id_pl,
                                    family=family_name,
                                    region=region,
                                    coach=coach_family,
                                    rank=rank,
                                    group=gr,
                                    posev_group=posev,
                                    title_id=title_id()
                                    ).save()  
    elif element_count == 2: # меняют местами игроков (в группе, ПФ, финале)
        if etap_1 == etap_2: # оба игрока из одного этапа соревнования
            stage = etap_1
            for pl in player_list:
                if pl != "":
                    znak = pl.find(":")
                    posev = int(pl[:znak]) 
                    break
            n = 0      
            for pl in full_name_list:
                pl_id = player_dict[pl]
                gr = group_list[1 - n]
                if stage == "Предварительный":
                    query = Choice.update(group=gr, posev_group=posev).where(Choice.player_choice_id == pl_id) # обновляет запись в Choice  
                elif stage == "1-й полуфинал" or  stage == "2-й полуфинал":
                    query = Choice.update(semi_final=stage, sf_group=gr, posev_sf=posev).where(Choice.player_choice_id == pl_id) 
                elif stage == "Финальный":
                    query = Choice.update(final=gr, posev_final=posev).where(Choice.player_choice_id == pl_id) # обновляет запись в Choice            
                query.execute()
                n += 1
        else: # игроки из разных этапов соревнования для исправления ошибок
            for pl in full_name_list:
                pl_id = player_dict[pl]
                choice = choices.select().where(Choice.player_choice_id == pl_id).get()
                if etap_1 == "Финальный":
                    stage = my_win.comboBox_first_group.currentText()
                    systems_1 = systems.select().where(System.stage == stage).get()
                    system_id_1 = systems_1.id 
                elif etap_2 == "Финальный":
                    stage = my_win.comboBox_second_group.currentText()
                    systems_2 = systems.select().where(System.stage == stage).get()
                    system_id_2 = systems_2.id       
    else:
        # ====== если меняет в полуфинале группы (менять результат) ======
        if element_count == 4: # если присутствуют 2-е игроки для обмена (ПФ смена регионов)
            stage = etap_1
            results = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == stage))
            players = Player.select().where(Player.title_id == title_id())

            for k in range(0, 2): # перезаписывает таблицу Result
                result1 = results.select().where(Result.player1 == full_name_list[k])
                result2 = result1.select().where(Result.player2 == full_name_list[k + 2]).get()
                with db:
                    result2.number_group = group_list[1 - k]
                    result2.save()
            n = 0      
            for pl in full_name_list:
                pl_id = player_dict[pl]
                gr = group_list[1 - n]
                if stage == "1-й полуфинал" or  stage == "2-й полуфинал":
                    query = Choice.update(sf_group=gr).where(Choice.player_choice_id == pl_id) 
                query.execute()
                n += 1
        # =====================
    player_in_table_group_and_write_Game_list_Result(stage) 
    if element_count == 4:   
        load_playing_game_in_table_for_semifinal(stage)  
        # ========
    my_win.lineEdit_change_pl1.clear()
    my_win.lineEdit_change_pl2.clear()
    my_win.lineEdit_change_pl1_2.clear()
    my_win.lineEdit_change_pl2_2.clear()
    my_win.comboBox_first_group.setCurrentText("-выберите группу-")
    my_win.comboBox_second_group.setCurrentText("-выберите группу-")
    my_win.listWidget_first_group.clear()
    my_win.listWidget_second_group.clear()
    my_win.comboBox_edit_etap1.setCurrentIndex(0)
    my_win.comboBox_edit_etap2.setCurrentIndex(0)
    if stage == "Предварительный":
        my_win.tabWidget.setCurrentIndex(3)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
         my_win.tabWidget.setCurrentIndex(4)            
    my_win.tableView.setVisible(True)

# def add_player_to_group():
#     """добавление игрока в группу при редактировании"""
#     player_choice_tmp = []
#     n_group = my_win.comboBox_number_group_edit.currentText()
#     player_gr = my_win.comboBox_player_group_edit.currentText()
#     edit_group_after_draw()


def choice_tbl_made():
    """создание таблицы жеребьевка, заполняет db списком участников для жеребьевки"""
    players = Player.select().order_by(Player.rank.desc()).where(Player.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    if len(choice) != 0:
        for i in choice:
            ch_d = Choice.get(Choice.id == i)
            ch_d.delete_instance()
    for i in players:
        family = i.player
        region = i.region
        rank = i.rank
        coach_id = i.coach_id
        coachs =Coach.select().where(Coach.id == coach_id).get()
        coach = coachs.coach
        chc = Choice(player_choice=i, family=family, region=region, coach=coach, rank=rank,
                    title_id=title_id()).save()


def filter_player_on_system():
    """Фильтрует игроков на вкладке системы по группам ПФ или финалам"""
    choice = Choice.select().where(Choice.title_id == title_id())
    number_group = my_win.comboBox_filter_number_group_final.currentText()
    stage = my_win.comboBox_filter_choice_stage.currentText()
    if number_group == "все группы" or number_group == "все финалы":
        return
    if  stage == "Предварительный":
        player_list = choice.select().where(Choice.group == number_group)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        player_list = choice.select().where((Choice.semi_final == stage) & (Choice.sf_group == number_group)).order_by(Choice.sf_group)
    else:
        player_list = choice.select().where(Choice.final == number_group)
    fill_table(player_list)


def choice_filter_on_system():
    """фильтрует таблицу жеребьевка по группам, полуфиналам или финалам"""
    systems = System.select().where(System.title_id == title_id())
    group_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    stage = my_win.comboBox_filter_choice_stage.currentText()
    number_group = my_win.comboBox_filter_number_group_final.currentText()
    if stage != "-выберите этап-" and stage != "":
        my_win.comboBox_filter_number_group_final.setEnabled(True)
    else:
        my_win.comboBox_filter_number_group_final.setEnabled(False)
    if stage == "-выберите этап-" and number_group == "все группы" and number_group == "все финалы":
        return
    else:
        if stage == "":
            my_win.comboBox_filter_choice_stage.clear()
            etaps_list = [i.stage for i in systems if i.choice_flag is True] # все этапы системы
            etaps_set = set(etaps_list)
            groups_set = set(group_list)
            etaps_set.intersection_update(groups_set)
            etaps = list(etaps_set)
            etaps.sort(reverse = True)
            etaps.append("Финальный")
            etaps.insert(0, "-выберите этап-")
            my_win.comboBox_filter_choice_stage.addItems(etaps)
        elif stage in group_list:
            my_win.comboBox_filter_number_group_final.clear()
            systems_sf = systems.select().where(System.stage == stage).get()
            kg = int(systems_sf.total_group)  # количество групп
            gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
            gr_txt.insert(0, "все группы")
            my_win.comboBox_filter_number_group_final.addItems(gr_txt)
        else:
            my_win.comboBox_filter_number_group_final.clear()
            systems_sf = systems.select().where(System.choice_flag == 1)
            gr_txt = [i.stage for i in systems_sf if i.stage not in group_list]
            gr_txt.insert(0, "все финалы")
            my_win.comboBox_filter_number_group_final.addItems(gr_txt)
           
        player_list = Choice.select().where(Choice.title_id == title_id())
        fill_table(player_list)
 
 
    # duplicat = duplicat_coach_in_group(coach_list)
    # color_coach_in_tablewidget(duplicat, coach_list)


def duplicat_coach_in_group(coach_list):
    """поиск совпадения тренеров в одной группе"""
    tmp_list = []
    count = len(coach_list)
    for i in coach_list:
        znak = i.find(",")
        if znak == -1:
            tmp_list.append(i)
        else:
            coach_1 = i[:znak]
            tmp_list.append(coach_1)
            if i.find(",", znak) == -1:
                znak_1 = i.find(",", znak + 1)
                coach_2 = i[znak: znak_1]
                tmp_list.append(coach_2)
            else:
                coach_2 = i[znak + 2:]
                tmp_list.append(coach_2)
    count_list = len(tmp_list)
    count_uniq = len(set(tmp_list)) 
    if count_list > count_uniq:
        duplicat = [x for i, x in enumerate(tmp_list) if i != tmp_list.index(x)]
        return duplicat


# def color_coach_in_tablewidget(duplicat, coach_list):
#     """окаршиваает в красный цвет повторяющиеся фамилия тренеров"""
#     if duplicat is not None:
#         num_gr = []
#         p = 0
#         for i in coach_list:
#             p += 1
#             for n in duplicat:
#                 if n in i:
#                     if p not in num_gr:
#                         num_gr.append(p) 
#         for k in num_gr:
#             my_win.tableWidget.item(k - 1, 4).setForeground(QBrush(QColor(0, 0, 255)))  # окрашивает текст в красный цвет


def color_region_in_tableWidget(fg):
    """смена цвета шрифта в QtableWidget -fg- номер группы"""
    reg = []
    rid = []

    if fg != "все группы" and fg != "":
        choice = Choice.select().where(Choice.title_id == title_id())
        line = choice.select().where(Choice.group == fg)
        for i in line:
            region = i.region
            region = str(region.rstrip())  # удаляет пробел в конце строки
            reg.append(region)
        if len(reg) != 0:
            for x in reg:
                count_region = reg.count(x)
                if count_region > 1:  # если повторяющихся регионов больше одного
                    p = 0
                    for m in range(len(reg)):
                        ind = reg.index(x, p)
                        p = ind + 1
                        rid.append(ind)                       
                        if m == count_region - 1:
                            break
            rid = list(set(rid))  # получает список индексов повторяющихся регионов
            rows = my_win.tableWidget.rowCount()  # кол-во строк в отсортированной таблице
            if rows != 0:
                for k in rid:
                    my_win.tableWidget.item(k, 3).setForeground(QBrush(QColor(255, 0, 0)))  # окрашивает текст в красный цвет


def hide_show_columns(tb):
    """скрывает или показывает столбцы TableView"""
    for k in range(0, 19):
        my_win.tableView.hideColumn(k)
    if tb == 0: # титул
        pass
    elif tb == 1: # участники
        my_win.tableView.showColumn(1) # фамилия имя
        my_win.tableView.showColumn(2) # др
        my_win.tableView.showColumn(3) # рейтинг
        my_win.tableView.showColumn(4) # город
        my_win.tableView.showColumn(5) # регион
        my_win.tableView.showColumn(6) # разряд
        my_win.tableView.showColumn(7) # тренеры
        my_win.tableView.showColumn(8) # место
        # my_win.tableView.showColumn(16) # заявка
    elif tb == 2: # система
        my_win.tableView.showColumn(1) # фамилия имя
        my_win.tableView.showColumn(2) # регион
        my_win.tableView.showColumn(3) # предварительный
        my_win.tableView.showColumn(4) # место группы
        my_win.tableView.showColumn(5) # пф
        my_win.tableView.showColumn(6) # тренеры
        my_win.tableView.showColumn(7) # место
        my_win.tableView.showColumn(8) # место в группе

        my_win.tableView.showColumn(9) # тренеры
        my_win.tableView.showColumn(10) # место
        my_win.tableView.showColumn(11) # место в группе
    elif tb == 3 or tb == 4 or tb == 5:
        my_win.tableView.showColumn(2)
        my_win.tableView.showColumn(3) # регион
        my_win.tableView.showColumn(4) # игрок 1
        my_win.tableView.showColumn(5) # игрок 2
        my_win.tableView.showColumn(6) # победитель        
        my_win.tableView.showColumn(8) # общий счет
        my_win.tableView.showColumn(9) # счет в партиях
    elif tb == 6:
        # my_win.tableView.showColumn(0)
        my_win.tableView.showColumn(1)
        my_win.tableView.showColumn(2)
        my_win.tableView.showColumn(3)
        my_win.tableView.showColumn(4)
        my_win.tableView.showColumn(5)
        my_win.tableView.showColumn(6)
    # elif tb == 7:
        # my_win.tableWidget.showColumn(0)
        # my_win.tableWidget.showColumn(1)



def etap_made():
    """создание этапов соревнований"""
    system = System.select().where(System.title_id == title_id())
    sum_game = []
    etap = my_win.comboBox_etap.currentText()
    if etap == "Одна таблица":
        fin = my_win.comboBox_etap.currentText()
        one_table(fin, group=1)
        gamer = my_win.lineEdit_title_gamer.text()
        tab_enabled(gamer)
        return
    if etap == "Предварительный":    
        kol_player_in_group() # кол-во участников в группах
    elif etap == "Финальный":
        total_game_table(exit_stage="", kpt=0, fin="", pv="") # сколько игр в финале или пф 
        systems = system.select().order_by(System.id.desc()).get()
        state_visible = my_win.checkBox_5.isChecked() # записывает в DB измененный статус видимости
        with db:
            systems.visible_game = state_visible
            systems.save()
        # суммирует все игры этапов    
    for k in system:
        kol_game_str = k.kol_game_string
        zn = kol_game_str.find(" ")
        number = int(kol_game_str[:zn])
        sum_game.append(number)
    all_sum_game = sum(sum_game)
    my_win.label_33.setText(f"Всего:{all_sum_game} игр.")
    my_win.checkBox_visible_game.setChecked(True)
    flag = control_all_player_in_final(etap) # проверяет все ли игроки распределены по финалам
    if flag is True: # продолжает выбор этапа
        made_system_load_combobox_etap()
    my_win.Button_etap_made.setEnabled(False)


def total_game_table(exit_stage, kpt, fin, pv):
    """количество участников и кол-во игр"""
    sender = my_win.sender()
    sum_player = [0]
    etap_text = my_win.comboBox_etap.currentText()
    flag_visible = my_win.checkBox_visible_game.isChecked()
    system = System.select().where(System.title_id == title_id()) # находит system id последнего
    systems = system.select().where(System.stage == "Предварительный").get()
    total_athletes = systems.total_athletes
    total_gr = systems.total_group

    for sys in system:
        fin_type = sys.type_table
        if fin_type == "круг" or fin_type == "сетка":
            fin_player = sys.max_player
            sum_player.append(fin_player)

        sum_pl = sum(sum_player)
    if kpt != 0:  # подсчет кол-во игр из выбора кол-ва игроков вышедших из группы и системы финала
        if etap_text == "Полуфиналы":
            vt = "группы"
            type_table = "группы"
            gr_pf = total_gr // 2
            player_in_final = gr_pf * kpt * 2 # колво участников в полуфинале
            cur_index = 0
        elif etap_text == "Финальный" or etap_text == "Суперфинал":
            cur_index = current_index_combobox_table(sender)
            if cur_index == 1:
                vt = "Сетка (-2) на"
                my_win.comboBox_page_vid.setCurrentText("книжная")
                type_table = "сетка"
            elif cur_index == 2:
                vt = "Сетка (с розыгрышем всех мест) на"
                my_win.comboBox_page_vid.setCurrentText("книжная")
                type_table = "сетка"
            elif cur_index == 3:
                vt = "Сетка (с играми за 1-3 места) на"
                my_win.comboBox_page_vid.setCurrentText("книжная")
                type_table = "сетка"
            elif cur_index == 4:
                vt = "Круговая таблица на"
                type_table = "круг"

            if exit_stage == "1-й полуфинал" or exit_stage == "2-й полуфинал":
                system_exit = system.select().where(System.stage == exit_stage).get()
                total_gr = system_exit.total_group 
            player_in_final_full = total_gr * kpt # колво участников в конкретном финале, если в группах полный состав
            player_in_final_current = total_athletes - sum_pl # кол-во участников в последнем финале (разница всех игроков минус уже разведенных по финалам)
            player_in_final = player_in_final_current if player_in_final_current <  player_in_final_full else player_in_final_full

            if etap_text == "Суперфинал":
                player_in_final = kpt # колво участников в конкретном финале, если в группах полный состав

        total_games = numbers_of_games(cur_index, player_in_final, kpt) # подсчет кол-во игр

        if etap_text == "Полуфиналы":
            gr_pf = total_gr // 2
            str_setka = f"{gr_pf} {vt} по {kpt * 2} участника"
            total_gr = gr_pf
        else:
            str_setka = f"{vt} {player_in_final} участников"
            total_gr = 0
 
        stroka_kol_game = f"{total_games} игр"

        system = System(title_id=title_id(), total_athletes=total_athletes, total_group=total_gr, kol_game_string=stroka_kol_game,
                        max_player=player_in_final, stage=fin, type_table=type_table, page_vid=pv, label_string=str_setka,
                        choice_flag=0, score_flag=5, visible_game=flag_visible, stage_exit=exit_stage, mesta_exit=kpt).save()    
        
        return [str_setka, player_in_final, total_athletes, stroka_kol_game]


def current_index_combobox_table(sender):
    """определяет индекс значения комбобокса"""
    if sender == my_win.comboBox_table_2:
        cur_index = my_win.comboBox_table_2.currentIndex()
    elif sender == my_win.comboBox_table_3:
        cur_index = my_win.comboBox_table_3.currentIndex()
    elif sender == my_win.comboBox_table_4:
        cur_index = my_win.comboBox_table_4.currentIndex()
    elif sender == my_win.comboBox_table_5:
        cur_index = my_win.comboBox_table_5.currentIndex()
    elif sender == my_win.comboBox_table_6:
        cur_index = my_win.comboBox_table_6.currentIndex()
    elif sender == my_win.comboBox_table_7:
        cur_index = my_win.comboBox_table_7.currentIndex() 
    elif sender == my_win.comboBox_table_8:
        cur_index = my_win.comboBox_table_8.currentIndex()  
    elif sender == my_win.comboBox_table_9:
        cur_index = my_win.comboBox_table_9.currentIndex()
    elif sender == my_win.comboBox_table_10:
        cur_index = my_win.comboBox_table_10.currentIndex() 
    elif sender == my_win.comboBox_table_11:
        cur_index = my_win.comboBox_table_11.currentIndex()      
    return cur_index


def control_all_player_in_final(etap):
    """проверка все ли игроки распределены по финалам и дает сигнал об окончании создании системы"""
     # титул id и стадия содержит слово финал (1 и 2 заменяет %)
    msgBox = QMessageBox
    gamer = my_win.lineEdit_title_gamer.text()
    system = System.select().order_by(System.id).where(System.title_id == title_id())
    system_stage = system.select().where(System.stage == "Предварительный").get()
    total_player = system_stage.total_athletes
    system_id = system.select().where(System.stage ** '% финал')
    tot_fin = len(system_id) # если 0, значит финалы еще не созданы или этап -одна таблица-
    sum_final = []

    for i in system_id:
        if i.stage != "Предварительный" and i.stage != "1-й полуфинал" and i.stage != "2-й полуфинал":
            player_in_etap = i.max_player
            sum_final.append(player_in_etap)

    total_final = sum(sum_final)
    t = total_player - total_final # оставшиеся не распределенные участники по финалам
    txt = ""
    if total_final == total_player or t <= 2: # все игроки посеяны по финалам или остался 1 или 2 игрока окончание создание системы
        if t == 1 and etap != "Суперфинал":     
            txt = "Остался 1 участник, не вошедший в финальную часть"
            msgBox.information(my_win, "Уведомление", txt)
        elif t == 2 and etap != "Суперфинал":
            txt = "Остались 2 игрока, они могут сыграть за место между собой"
            msgBox.information(my_win, "Уведомление", txt)
                    # ====== вставить вопрос о суперфинале и игры за 3 место если финал сетка
        if etap != "Суперфинал":           
            result = msgBox.question(my_win, "", "Будет ли суперфинал?",
                                    msgBox.Yes, msgBox.No) 
            if result == msgBox.Yes:
                flag = True
            # else:
            #     flag = False
            # return flag   
            else:        # ========= 
                add_open_tab(tab_page="Система")
                result = msgBox.question(my_win, "", "Система соревнований создана.\n"
                                                            "Теперь необходимо сделать жеребъевку\n"
                                                            "предварительного этапа.\n"
                                                            "Хотите ее сделать сейчас?",
                                            msgBox.Ok, msgBox.Cancel)
                if result == msgBox.Ok:
                    choice_gr_automat()
                    add_open_tab(tab_page="Группы")
                    tab_enabled(gamer)
                    with db:
                        system_stage.choice_flag = True
                        system_stage.save()
                        flag = True
                else:
                    return    
    elif t >= 3: # продолжает создание системы
        flag = True
    elif t == 0:
        flag = False
    return flag


def combobox_etap_compare(real_list):
    """сравнение и изменение значение комбокса в зависиости от выбора этапа
    -real_list- список значений для будущего этапа"""
    count_items = my_win.comboBox_etap.count()
    item_list = [] # текущий набор комбобокса
    for i in range(0, count_items):
        cur_item = my_win.comboBox_etap.itemText(i)
        item_list.append(cur_item) # значение которые есть в комбобокс после добавления отсутствующих
    res = [x for x in real_list + item_list if x not in real_list or x not in item_list] # список, который надо убрать из комбобокса

    if len(res) != 0:
        my_win.comboBox_etap.clear()
        my_win.comboBox_etap.addItems(real_list)


def made_system_load_combobox_etap():
    """подготовка комбобокса для дальнейшего создания системы"""
    sender = my_win.sender()
    ct = my_win.comboBox_etap.currentText()
    label_etap_list = [my_win.label_101, my_win.label_102, my_win.label_103, my_win.label_104, my_win.label_105, my_win.label_106,
                       my_win.label_107, my_win.label_108, my_win.label_109, my_win.label_110,  my_win.label_111]
    combobox_etap_list = [my_win.comboBox_table_1, my_win.comboBox_table_2, my_win.comboBox_table_3, my_win.comboBox_table_4, 
                          my_win.comboBox_table_5, my_win.comboBox_table_6, my_win.comboBox_table_7, my_win.comboBox_table_8,
                          my_win.comboBox_table_9, my_win.comboBox_table_10, my_win.comboBox_table_11]
    label_text = my_win.label_10.text() # номер текущего этапа
    real_list = []
 # после выбора из комбобокса соответственно этапу включает label   
    if ct == "" or ct == "-выбор этапа-" or sender == my_win.Button_etap_made: # после нажатия кнопки создать этап готовит следующий этап
        if ct == "-выбор этапа-" or ct == "":
            return   
        if label_text == "1-й этап": # пишет следующий этап
            if ct == "Одна таблица":
                my_win.comboBox_table_1.show()
                my_win.spinBox_kol_group.hide()
                my_win.label_11.hide()
                my_win.label_101.hide()
            elif ct == "Предварительный":
                my_win.spinBox_kol_group.hide()
                my_win.comboBox_table_1.hide()
                my_win.label_11.hide()
                my_win.label_12.show()
                real_list = ["-выбор этапа-", "Полуфиналы", "Финальный"] # который нужен в комбобокс
                combobox_etap_compare(real_list)
                my_win.label_10.setText("2-й этап")
        elif  label_text == "2-й этап":  
            # ==== поиск всех занчений комбобокса
            if ct == "Полуфиналы":
                real_list = ["-выбор этапа-", "Полуфиналы", "Финальный"] # который нужен в комбобокс
            elif ct == "Финальный": 
                real_list = ["-выбор этапа-", "Финальный", "Суперфинал"] 
            combobox_etap_compare(real_list)
            my_win.label_10.setText("3-й этап")
        elif  label_text == "3-й этап": # текущий этап
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_104.show()
            my_win.label_10.setText("4-й этап")
        elif  label_text == "4-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_105.show()
            my_win.label_10.setText("5-й этап")
        elif  label_text == "5-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_106.show()
            my_win.label_10.setText("6-й этап")
        elif  label_text == "6-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_107.show()
            my_win.label_10.setText("7-й этап")
        elif  label_text == "7-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_108.show()
            my_win.label_10.setText("8-й этап")
        elif  label_text == "8-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_109.show()
            my_win.label_10.setText("9-й этап")
        elif  label_text == "9-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_110.show()
            my_win.label_10.setText("10-й этап")
        elif  label_text == "10-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_111.show()
            my_win.label_10.setText("11-й этап")
        my_win.comboBox_etap.setCurrentText("-выбор этапа-")   
    else:   # выбор значения из комбобокса создания этапов
        if ct == "Одна таблица":
            my_win.comboBox_table_1.show()
            my_win.spinBox_kol_group.hide()
            my_win.label_11.hide()
            my_win.label_101.hide()
        elif ct == "Предварительный":
            my_win.spinBox_kol_group.show()
            my_win.comboBox_table_1.hide()
            my_win.label_101.show()
            my_win.label_101.setText("Предварительный этап")
            my_win.label_11.show()
            my_win.label_12.hide()
        elif ct == "Полуфиналы":
            mark = label_text.find("-")
            number_etap = int(label_text[:mark])
            label_etap_list[number_etap - 1].show()
            label_etap_list[number_etap - 1].setText(f"{number_etap - 1}-й полуфинал")
            kol_player_in_final()
        elif ct == "Финальный":
            if label_text == "2-й этап":
                my_win.label_102.show()
                my_win.label_102.setText("1-й финал")
                my_win.comboBox_table_2.show()
            elif label_text == "3-й этап":
                last_etap = my_win.label_102.text() 
                my_win.label_103.show()               
                if last_etap == "1-й полуфинал":
                    my_win.label_103.setText("1-й финал")
                else:
                    my_win.label_103.setText("2-й финал")
                my_win.comboBox_table_3.show()               
            elif label_text == "4-й этап":
                last_etap = my_win.label_103.text() 
                if last_etap == "2-й полуфинал":
                    my_win.label_104.setText("1-й финал")
                else: 
                    txt = my_win.label_103.text()
                    znak = txt.find("-") 
                    fin = int(txt[:znak])
                    final = f"{fin + 1}-й финал"    
                    my_win.label_104.setText(final)
                my_win.comboBox_table_4.show()
            elif label_text == "5-й этап":
                txt = my_win.label_104.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_105.setText(final)
                my_win.comboBox_table_5.show()
            elif label_text == "6-й этап":
                txt = my_win.label_105.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_106.setText(final)
                my_win.comboBox_table_6.show()
            elif label_text == "7-й этап":
                txt = my_win.label_106.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_107.setText(final)
                my_win.comboBox_table_7.show()
            elif label_text == "8-й этап":
                txt = my_win.label_107.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_108.setText(final)
                my_win.comboBox_table_8.show()
            elif label_text == "9-й этап":
                txt = my_win.label_108.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_109.setText(final)
                my_win.comboBox_table_9.show()
            elif label_text == "10-й этап":
                txt = my_win.label_109.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_110.setText(final)
                my_win.comboBox_table_10.show()
            elif label_text == "11-й этап":
                txt = my_win.label_110.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_111.setText(final)
                my_win.comboBox_table_11.show()
            else:
                mark = label_text.find("-")
                fin = int(txt[:mark])
                final = f"{fin + 1}-й финал"  
                label_etap_list[number_etap - 1].setText(final) 
                combobox_etap_list[number_etap - 1].show() 
        elif ct == "Суперфинал":
            mark = label_text.find("-")
            number_etap = int(label_text[:mark])
            label_etap_list[number_etap - 1].show()
            label_etap_list[number_etap - 1].setText("Суперфинал")
            combobox_etap_list[number_etap - 1].show()


def total_games_in_final_without_group_games(player_in_final, total_gr, kpt):
    """всего игр в финале без учета сыгранных игр в предварительном этапе"""
    # остаток отделения, если 0, то участники равно делится на группы
    remains = player_in_final % int(total_gr)
    if remains == 0:  # если в группах равное количество человек
        playing_game = (kpt * (kpt - 1)) // 2 * total_gr
    else:
        full_group = player_in_final // kpt # кол-во групп с полным количеством участников
        no_full_group = total_gr - remains
        playing_game_in_full_group = (kpt * (kpt - 1)) // 2 * full_group
        kpt_min = kpt - 1
        playing_game_in_no_full_group = (kpt_min * (kpt_min - 1)) // 2 * no_full_group
        playing_game = playing_game_in_full_group + playing_game_in_no_full_group
    total_games = (player_in_final * (player_in_final - 1)) // 2 - playing_game
    return total_games


def total_games_in_final_with_group_games(player_in_final, gr_pf, kpt):
    """всего игр в полуфинале с учетом сыгранных игр в предварительном этапе"""
    # остаток отделения, если 0, то участники равно делится на группы
    remains = player_in_final % int(gr_pf)
    if remains == 0:  # если в группах равное количество человек
        playing_game_in_group = (kpt * (kpt - 1)) # кол-во игр, сыгранных в группе
        total_games = (((kpt * 2 * (kpt * 2 - 1)) // 2) - playing_game_in_group) * gr_pf # всего игр в пф
    else:
        full_group = player_in_final // kpt # кол-во групп с полным количеством участников
        no_full_group = gr_pf - remains
        playing_game_in_group = (kpt * (kpt - 1)) # кол-во игр, сыгранных в группе
        playing_game_in_full_group = (((kpt * (kpt - 1)) // 2) - playing_game_in_group) * full_group
        kpt_min = kpt - 1
        playing_game_in_no_full_group = ((kpt_min * (kpt_min - 1)) // 2 - playing_game_in_group) * no_full_group
        total_games = playing_game_in_full_group + playing_game_in_no_full_group
    return total_games


def numbers_of_games(cur_index, player_in_final, kpt):
    """подсчет количество игр в зависимости от системы (пока сетки на 16)"""
    systems = System.select().where(System.title_id == title_id())
    system = systems.select().where(System.stage == "Предварительный").get()
    system_etap = my_win.comboBox_etap.currentText() #
    gr = system.total_group
    if system_etap == "Полуфиналы":
        gr_pf = gr // 2
        total_games = total_games_in_final_with_group_games(player_in_final, gr_pf, kpt)
    else:
        if cur_index == 1:  # сетка - 2
            if player_in_final == 8:
                total_games = 14
            elif player_in_final > 4 and player_in_final < 8: # если игроков не полная сетка
                tours = 3
                free = 8 - player_in_final
                if free == 1:
                    total_games = 14 - free * tours
                elif free > 1:
                    total_games = 14 - (free * tours - 1)
            elif player_in_final == 16:
                total_games = 38
            elif player_in_final > 8 and player_in_final < 16: # если игроков не полная сетка
                tours = 4
                free = 16 - player_in_final
                if free == 1:
                    total_games = 38 - free * tours
                elif free > 1:
                    total_games = 38 - (free * tours - 1)
            elif player_in_final == 32:
                total_games = 94
        elif cur_index == 2:  # прогрессивная сетка
            if player_in_final == 8:
                total_games = 12
            elif player_in_final == 16:
                total_games = 32
            elif player_in_final == 32:
                total_games = 80
            else:
                total_games= 80
        elif cur_index == 3:  # сетка с розыгрышем призовых мест
            if player_in_final == 8:
                total_games = 4
            elif player_in_final == 16:
                total_games = 8
            elif player_in_final == 32:
                total_games = 32
        elif cur_index == 4: # игры в круг
            total_games = total_games_in_final_without_group_games(player_in_final, gr, kpt)

    return total_games


def clear_db_before_edit():
    """очищает таблицы при повторном создании системы"""
    system = System.select().where(System.title_id == title_id())
    title = Title.select().where(Title.id == title_id()).get()
    for i in system:  # удаляет все записи
        i.delete_instance()
    sys = System(title_id=title_id(), total_athletes=0, total_group=0, max_player=0, stage="", type_table="", page_vid="",
                 label_string="", kol_game_string="", choice_flag=False, score_flag=5, visible_game=True,
                 stage_exit="", mesta_exit="").save()
    with db:
        # записывает в таблицу -Title- новые открытые вкладки
        title.tab_enabled = "Титул Участники"
        title.save()

    gl = Game_list.select().where(Game_list.title_id == title_id())
    for i in gl:
        gl_d = Game_list.get(Game_list.id == i)
        gl_d.delete_instance()
    chc = Choice.select().where(Choice.title_id == title_id())
    for i in chc:
        ch_d = Choice.get(Choice.id == i)
        ch_d.delete_instance()
    rs = Result.select().where(Result.title_id == title_id())
    for i in rs:
        r_d = Result.get(Result.id == i)
        r_d.delete_instance()


def clear_db_before_choice(stage):
    """очищает систему перед повторной жеребьевкой и изменяет кол-во участников если они изменились"""
    msgBox = QMessageBox
    sys = System.select().where(System.title_id == title_id())
    player = Player.select().where(Player.title_id == title_id())
    system = sys.select().where(System.stage == "Предварительный").get()
    sys_id = system.id
    tg = system.total_group
    total_player = system.total_athletes
    max_pl = system.max_player
    new_total_player = len(player)
    if total_player != new_total_player: #  если изменилось число участников
        result = msgBox.question(my_win, "Список участников", "Был изменено число участников.\n"
        "вы хотите изменить систему соревнований?",
                                    msgBox.Ok, msgBox.Cancel)
        if result == msgBox.Ok:
             # очищает таблицы перед новой системой соревнования (system, choice)
            clear_db_before_edit()
            choice_tbl_made()  # заполняет db жеребьевка
        else:
            e1 = new_total_player % tg  # остаток до полного посева групп, где спортсменов на одного больше
                # если количество участников равно делится на группы (кол-во групп)
            p_min = new_total_player // tg  # минимальное число спортсменов в группах
            g1 = int(tg) - e1  # кол-во групп, где наименьшее кол-во спортсменов
            p_max= p_min + 1  # кол-во человек в группе с наибольшим их количеством
            if e1 == 0:  # то в группах равное количество человек -e1-
                stroka_kol_group = f"{tg} группы по {str(p_min)} чел."
                skg = int((p_min * (p_min - 1) / 2) * int(tg))
                max_pl = p_min
            else:
                stroka_kol_group = f"{str(g1)} групп(а) по {str(p_min)} чел. и {str(e1)} групп(а) по {str(p_max)} чел."
                skg = int((((p_min * (p_min - 1)) / 2 * g1) + ((p_max * (p_max - 1)) / 2 * e1)))
                max_pl = p_max
            kgs = f"{skg} игр"
            sys_t = System.select().where(System.id == sys_id).get()
            sys_t.max_player = max_pl
            sys_t.label_string = stroka_kol_group
            sys_t.kol_game_string = kgs
            sys_t.save()

            for x in sys:
                x.total_athletes = new_total_player
                x.save()
    # else:  # если число спортсменов не изменилось (просто смена участников)
    gl = Game_list.select().where(Game_list.title_id == title_id())
    system_id = sys.select().where(System.stage == stage).get()
    gamelists = gl.select().where(Game_list.system_id == system_id)
    for i in gamelists:
        gl_d = Game_list.get(Game_list.id == i)
        gl_d.delete_instance()
    choices = Choice.select().where(Choice.title_id == title_id())
    if stage == "Предварительный":
        for i in choices:
            ch_d = Choice.get(Choice.id == i)
            ch_d.delete_instance()
    results = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == stage))
    for i in results:
        r_d = Result.get(Result.id == i)
        r_d.delete_instance()
    if stage == "Предварительный":
        choice_tbl_made()


def clear_db_before_choice_final(fin):
    """очищает базу данных -Game_list- и -Result- перед повторной жеребьевкой финалов"""
    gamelist = Game_list.select().where(Game_list.title_id == title_id())
    gl = gamelist.select().where(Game_list.number_group == fin)
    for i in gl:
        gl_d = Game_list.get(Game_list.id == i)
        gl_d.delete_instance()
    results = Result.select().where(Result.title_id == title_id())
    rs = results.select().where(Result.number_group == fin)
    for i in rs:
        r_d = Result.get(Result.id == i)
        r_d.delete_instance()
    choice = Choice.select().where(Choice.title_id == title_id())
    ch = choice.select().where(Choice.final == fin)
    for i in ch:
        ch_d = Choice.get(Choice.id == i)
        ch_d.posev_final = ""
        ch_d.save()


def clear_db_before_choice_semifinal(stage):
    """очищает базу данных -Game_list- и -Result- перед повторной жеребьевкой полуфиналов"""
    system = System.select().where(System.title_id == title_id()) 
    system_id = system.select().where(System.stage == stage).get()
    gamelist = Game_list.select().where(Game_list.title_id == title_id())
    gl = gamelist.select().where(Game_list.system_id == system_id)
    for i in gl:
        gl_d = Game_list.get(Game_list.id == i)
        gl_d.delete_instance()
    results = Result.select().where(Result.title_id == title_id())
    rs = results.select().where(Result.system_stage == stage)
    for i in rs:
        r_d = Result.get(Result.id == i)
        r_d.delete_instance()


def ready_system():
    """проверка на готовность системы"""
    all_player_in_final = []
    system = System.select().where(System.title_id == title_id())  # находит system id первого
    count = len(system)
    if count == 1:
        for k in system:
            stage = k.stage
        if stage == "Одна таблица":
            my_win.statusbar.showMessage("Система соревнований создана", 5000)
            flag = True
        else:
            my_win.statusbar.showMessage("Необходимо создать систему соревнований", 500)
            flag = False
    elif count > 1:
        sys = system.select().order_by(System.id.desc()).where(System.title_id == title_id()).get()
        stage_system = sys.stage
        if stage_system == "Предварительный" or stage_system == "1-й полуфинал" or stage_system == "2-й полуфинал":
            my_win.statusbar.showMessage("Необходимо создать систему соревнований", 500)
            flag = False
        else:
            sys_min = system.select().order_by(System.id).where(System.title_id == title_id()).get()
            total_player = sys_min.total_athletes
            system_id = system.select().where(System.stage ** '%финал')
            for k in system_id:
                tot_player = k.max_player
                all_player_in_final.append(tot_player)
            all_player = sum(all_player_in_final)
            if all_player > total_player:
                my_win.statusbar.showMessage("Система соревнований создана", 5000)
                flag = True
            else:
                my_win.statusbar.showMessage("Необходимо создать систему соревнований", 500)
                flag = False
    return flag


def ready_choice(stage):
    """проверка на готовность жеребьевки групп"""
    sys = System.select().where(System.title_id == title_id())
    greb_flag = False
    if stage != "":
        system = sys.select().where(System.stage == stage).get()
        greb_flag = system.choice_flag
    
    if greb_flag is True:
        my_win.statusbar.showMessage("Жеребьевка сделана", 5000)
        flag = True
    else:
        my_win.statusbar.showMessage("Жеребьевка групп еще не выполнена", 5000)
        flag = False
    return flag


def select_choice_final():
    """выбор жеребьевки финала"""
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    fin = []
    for sys in system:
        if sys.stage != "Предварительный" and sys.stage != "1-й полуфинал" and sys.stage != "2-й полуфинал":
            fin.append(sys.stage)
    fin, ok = QInputDialog.getItem(my_win, "Выбор финала", "Выберите финал для жеребъевки", fin, 0, False)
    if ok:
        return fin
    else:
        fin = None
        return fin


def select_choice_semifinal():
    """выбор жеребьевки полуфинала"""
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    semifinal = []
    for sys in system:
        if sys.stage == "1-й полуфинал" or sys.stage == "2-й полуфинал":
            semifinal.append(sys.stage)
    semifinal, ok = QInputDialog.getItem(my_win, "Выбор полуфинала", "Выберите полуфинал для жеребъевки", semifinal, 0, False)
    if ok:
        return semifinal
    else:
        semifinal = None
        return semifinal


def manual_choice_setka(fin, count_exit, mesto_first_poseva):
    """Ручная жеребьевка сетки"""
    choice = Choice.select().where(Choice.title_id == title_id())
    posevs = setka_choice_number(fin, count_exit)
    player_net = posevs[0]
    posev_1 = posevs[1]
    z = len(posevs)

    if z == 3:
        posev_2 = posevs[2]
    elif z == 4:
        posev_2 = posevs[2]
        posev_3 = posevs[3]
    elif z == 5:
        posev_2 = posevs[2]
        posev_3 = posevs[3]
        posev_4 = posevs[4]
    for n in range (0, count_exit): # начало основного посева
        if fin == "1-й финал":
            choice_posev = choice.select().where(Choice.mesto_group == mesto_first_poseva + n)
        else:
            choice_posev = choice.select().order_by(Choice.rank).where(Choice.mesto_group == mesto_first_poseva + n)


def check_choice(fin):
    """Проверяет перед жеребьевкой финалов, была ли она произведена ранее или еще нет"""
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    system_final = system.select().where(System.stage == fin).get() # получаем запись конкретного финала
    check_flag = system_final.choice_flag
 
    return check_flag


def checking_possibility_choice(fin):
    """Проверяет перед жеребьевкой финалов, сыграны ли все партиии в группах или полуфиналах"""
    msg = QMessageBox
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    system_final = system.select().where(System.stage == fin).get() # получаем запись конкретного финала
    if fin == "Одна таблица":
        check_flag = system_final.choice_flag
    else:    
        exit = system_final.stage_exit  # запись откуда идет выход в финал
        if exit == "1-й полуфинал" or exit == "2-й полуфинал":
            exit_str = f"{exit}е"
        else:
            exit_str = "предварительном этапе"
        gr = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == exit)) # отбираем записи с выходом в финал

        for i in gr:
            game = i.points_win 
            check_flag = True      
            if game is None:
                result = msg.information(my_win, "Предварительный этап", "Еще не все встречи сыграны в" f"{exit_str}.",
                                        msg.Ok)
                check_flag = False
                break                        
    return check_flag


def del_player_table():
    """таблица удаленных игроков на данных соревнованиях"""
    if my_win.checkBox_6.isChecked():
        my_win.Button_clear_del.setEnabled(True)
        player_list = Delete_player.select().where(Delete_player.title_id == title_id())
        count = len(player_list)
        if count == 0:
            my_win.statusbar.showMessage(
                "Удаленных участников соревнований нет", 10000)
            # return
            fill_table(player_list)
        else:
            my_win.tableView.hideColumn(8)
            # my_win.tableView.hideColumn(9)
            my_win.tableView.hideColumn(10)
            my_win.tableView.hideColumn(11)
            my_win.tableView.hideColumn(12)
            my_win.tableView.hideColumn(13)
            fill_table(player_list)
            my_win.statusbar.showMessage(
                "Список удаленных участников соревнований", 10000)
            if my_win.lineEdit_Family_name.text() != "":
                my_win.Button_add_edit_player.setText("Восстановить")
                my_win.Button_add_edit_player.setEnabled(True)
            else:
                my_win.Button_add_edit_player.setEnabled(False)
    else:
        player_list = Player.select().where(Player.title_id == title_id())
        fill_table(player_list)
        my_win.tableView.showColumn(8)
        my_win.Button_add_edit_player.setText("Добавить")
        my_win.Button_add_edit_player.setEnabled(True)
        my_win.Button_clear_del.setEnabled(False)
        my_win.statusbar.showMessage("Список участников соревнований", 10000)


def clear_del_player():
    """Очистка базы данных удаленных игроков"""
    msgBox = QMessageBox
    del_player = Delete_player.select().where(Delete_player.title_id == title_id())
    result = msgBox.question(my_win, "Участники", "Вы действительно хотите очистить список\n"
                                "удаленных игроков?",
                                        msgBox.Ok, msgBox.Cancel)
    if result == msgBox.Ok:
        for i in del_player:
            i.delete_instance()
        my_win.Button_clear_del.setEnabled(False)  
        my_win.checkBox_6.setChecked(False)      
    else:
        return


def remains_in_group(etap_system, etap_system_dict):
    """подсчет игроков в группе и полуфиналов после создания финалов"""
    stage_dict = {} # словарь (этап: кол0во игроков)
    number_player_gr = 0
    number_player_pf1 = 0
    number_player_pf2 = 0
    out_pf1 = 0
    out_pf2 = 0
    out_f = 0
    system = System.select().where(System.title_id == title_id())
    for m in range(0, 2):
        for k  in system:
            etap_system = k.stage
            if etap_system == "Предварительный":
                number_player_gr = k.max_player
                stage_dict[etap_system] = number_player_gr if m == 0 else number_player_gr - out_pf1 - out_pf2
            elif etap_system == "1-й полуфинал":
                number_player_pf1 = k.max_player // k.total_group # кол-во игрок в 1-ом пф
                out_pf1 = etap_system_dict[etap_system] # сколько вышло из группы 1-й пф
                stage_dict[etap_system] = number_player_pf1
            elif etap_system == "2-й полуфинал":
                number_player_pf2 = k.max_player // k.total_group
                out_pf2 = etap_system_dict[etap_system] # сколько вышло из группы 1-й пф
                stage_dict[etap_system] = number_player_pf2
            elif (etap_system == "1-й финал" or etap_system == "2-й финал" or etap_system == "3-й финал" or
                etap_system == "4-й финал" or etap_system == "5-й финал" or etap_system == "6-й финал" or
                etap_system == "7-й финал" or etap_system == "8-й финал" or etap_system == "9-й финал"):
                if m == 1:
                    systems = system.select().where(System.stage == etap_system).get()
                    exit_stage = systems.stage_exit # откуда выходят в финал
                    out_f = etap_system_dict[etap_system] # сколько вышло из группы 1-й пф
                    stage_dict[exit_stage] = stage_dict[exit_stage ] - out_f  # сколько вышло из 1-й пф в 1-й финал
    return stage_dict


def max_player_and_exit_stage(etap):
    """определяет максимальное число спортсменов в комбобоксе и стадию откуда выход в финал
    etap - текущий этап, stage - предыдущий этап, label_text - номер этапа, mx_pl - максимальное число в комбобоксе
    # etap_list список [этап, кол-во игроков, из какого этапа вышли"""
    exit_player_stage = []
    etap_list = []
    etap_list_tmp = []
    total_stage = []
    etap_dict = {}
    etap_system_dict = {}
    system = System.select().where(System.title_id == title_id())
    i = 0
    for k in system: # получение словаря этапов
        i += 1
        etap_system = k.stage
        mesta_exit = k.mesta_exit
        stage_exit = k.stage_exit
        etap_list_tmp.append(etap_system)
        etap_list_tmp.append(mesta_exit)
        etap_list_tmp.append(stage_exit)
        total_stage.append(etap_system)
        etap_list = etap_list_tmp.copy()
        etap_list_tmp.clear()
        etap_dict[i] = etap_list
        etap_system_dict[etap_system] = mesta_exit
    number_etap = i + 1
    dict_etap = remains_in_group(etap_system, etap_system_dict)

    listing_etap = etap_dict[i] # список этапа (название, выход)
    last_etap = listing_etap[0] 
    system_last = system.select().where(System.stage == last_etap).get()
    mesta_exit = listing_etap[1]
    stage_exit = listing_etap[2]

    if number_etap == 2:
        if etap == "Полуфиналы":
            fin = "1-й полуфинал"
        elif etap == "Финальный":
            fin = "1-й финал"
        exit_stage = "Предварительный" # откуда попадают в полуфинал игроки
        max_pl = system_last.max_player # максимальное допустимое число игроков при выборе в комбобоксе
    elif number_etap == 3:
        if etap == "Полуфиналы":
            fin = "2-й полуфинал"
            exit_stage = "Предварительный"
        elif etap == "Финальный":
            fin = "1-й финал" if "1-й полуфинал" in total_stage else "2-й финал"
            exit_stage = "1-й полуфинал" if "1-й полуфинал" in total_stage else "Предварительный"
        max_pl = dict_etap[exit_stage]
    elif (number_etap == 4 or number_etap == 5 or number_etap == 6 or number_etap == 7
        or number_etap == 8 or number_etap == 9 or number_etap == 10 or number_etap == 11):
        fin = number_final(last_etap) # текущий этап
        if "2-й полуфинал" in total_stage:
            if dict_etap["1-й полуфинал"] == 0 and dict_etap["2-й полуфинал"] != 0:
                exit_stage = "2-й полуфинал"
            elif dict_etap["1-й полуфинал"] != 0 and dict_etap["2-й полуфинал"] != 0:
                exit_stage = "1-й полуфинал"
            elif dict_etap["1-й полуфинал"] == 0 and dict_etap["2-й полуфинал"] == 0:
                exit_stage = "Предварительный"
        elif "1-й полуфинал" in total_stage:
            if dict_etap["1-й полуфинал"] == 0:
                exit_stage = "Предварительный"
            elif dict_etap["1-й полуфинал"] != 0:
                exit_stage = "1-й полуфинал"
        else:
            exit_stage = "Предварительный"
        max_pl = dict_etap[exit_stage]

    exit_player_stage.append(max_pl)
    exit_player_stage.append(exit_stage)
    exit_player_stage.append(fin)

    return exit_player_stage


def number_final(last_etap):
    """определяет номер финала исходя из предыдущего"""
    if last_etap == "2-й полуфинал":
        fin = "1-й финал"
    else:
        znak = last_etap.find("-") 
        fin_num = int(last_etap[:znak])
        fin = f"{fin_num + 1}-й финал"
    return fin


def kol_player_in_final():
    """после выбора из комбобокс системы финала подсчитывает сколько игр в финале"""
    sender = my_win.sender()
    pv = my_win.comboBox_page_vid.currentText()
    etap = my_win.comboBox_etap.currentText()
    player = Player.select().where(Player.title_id == title_id())
    count = len(player)
    fin = ""
    exit_stage = ""
    label_text = my_win.label_10.text()
    if etap != "Суперфинал":
        if my_win.comboBox_etap.currentText() == "Одна таблица":
            if my_win.comboBox_table_1.currentText() == "Круговая система":
                kol_game = count * (count - 1) // 2
                my_win.label_etap_1.show()
                my_win.label_19.show()
                my_win.label_101.show()
                my_win.label_101.setText(my_win.comboBox_etap_1.currentText())
                my_win.label_19.setText(f"{kol_game} игр.")
                my_win.label_33.setText(f"Всего: {kol_game} игр.")
                my_win.label_etap_1.setText(f"{count} человек по круговой системе.")
                my_win.comboBox_etap.hide()
                my_win.comboBox_table_1.hide()
                my_win.comboBox_page_vid.setCurrentText("альбомная")
            else: # система из одной таблицы по олимпийской системе
                my_win.comboBox_page_vid.setCurrentText("книжная")
                cur_index = my_win.comboBox_table_1.currentIndex()
                total_game = 0
                if cur_index != 0:
                    player_in_final = count
                    total_game = numbers_of_games(cur_index, player_in_final)
                my_win.label_etap_1.show()
                my_win.label_19.show()
                my_win.label_19.setText(f"{total_game} игр.")
                my_win.label_33.setText(f"Всего: {total_game} игр.")
                my_win.label_etap_1.setText(f"{count} человек в сетке.")
                my_win.comboBox_table_1.hide()
        else:
            etap = my_win.comboBox_etap.currentText()
            exit_player_stage = max_player_and_exit_stage(etap)
            max_exit_group = exit_player_stage[0]
            exit_stage = exit_player_stage[1]
            fin = exit_player_stage[2]
        # изменение падежа этапов в комбобоксе
        if exit_stage == "Предварительный":
            exit_stroka = "Предварительного этапа"
        elif exit_stage == "1-й полуфинал":
            exit_stroka = "1-ого полуфинала"
        elif exit_stage == "2-й полуфинал":
            exit_stroka = "2-ого полуфинала" 
    else:
        # systems = System.select().where(System.title_id == title_id())
 
        fin = "Суперфинал"
        exit_stage = "1-й финал"
        exit_stroka = "1-ого финала"
        max_exit_group = 12
    
    kpt, ok = QInputDialog.getInt(my_win, "Число участников", "Введите число участников, выходящих\n "
                                                                    f"из {exit_stroka} в {fin}", min=1, max=max_exit_group)
                    
            # возвращает из функции несколько значения в списке
    list_pl_final = total_game_table(exit_stage, kpt, fin, pv)

    if ok is True: # заполняет этапы значениями (label)
        if label_text == "1-й этап":
            my_win.label_19.show()
            my_win.label_19.setText(list_pl_final[3])
            my_win.label_etap_1.show()
            my_win.label_etap_1.setText(list_pl_final[0])
        elif label_text == "2-й этап":
            my_win.label_27.setText(list_pl_final[3])
            my_win.label_27.show()
            my_win.label_etap_2.setText(list_pl_final[0])
            my_win.label_etap_2.show()
            my_win.comboBox_table_2.hide()
        elif label_text == "3-й этап":
            my_win.label_30.setText(list_pl_final[3])
            my_win.label_30.show()
            my_win.label_etap_3.setText(list_pl_final[0])
            my_win.label_etap_3.show()
            my_win.comboBox_table_3.hide()
        elif label_text == "4-й этап":
            my_win.label_53.setText(list_pl_final[3])
            my_win.label_53.show()
            my_win.label_etap_4.setText(list_pl_final[0])
            my_win.label_etap_4.show()
            my_win.comboBox_table_4.hide()
        elif label_text == "5-й этап":
            my_win.label_58.setText(list_pl_final[3])
            my_win.label_58.show()
            my_win.label_etap_5.setText(list_pl_final[0])
            my_win.label_etap_5.show()
            my_win.comboBox_table_5.hide()
        elif label_text == "6-й этап":
            my_win.label_81.setText(list_pl_final[3])
            my_win.label_81.show()
            my_win.label_etap_6.setText(list_pl_final[0])
            my_win.label_etap_6.show()
            my_win.comboBox_table_6.hide()
        elif label_text == "7-й этап":
            my_win.label_82.setText(list_pl_final[3])
            my_win.label_82.show()
            my_win.label_etap_7.setText(list_pl_final[0])
            my_win.label_etap_7.show()
            my_win.comboBox_table_7.hide()
        elif label_text == "8-й этап":
            my_win.label_83.setText(list_pl_final[3])
            my_win.label_83.show()
            my_win.label_etap_8.setText(list_pl_final[0])
            my_win.label_etap_8.show()
            my_win.comboBox_table_8.hide()
        elif label_text == "9-й этап":
            my_win.label_84.setText(list_pl_final[3])
            my_win.label_84.show()
            my_win.label_etap_9.setText(list_pl_final[0])
            my_win.label_etap_9.show()
            my_win.comboBox_table_9.hide()
        elif label_text == "10-й этап":
            my_win.label_85.setText(list_pl_final[3])
            my_win.label_85.show()
            my_win.label_etap_10.setText(list_pl_final[0])
            my_win.label_etap_10.show()
            my_win.comboBox_table_10.hide()
        elif label_text == "11-й этап":
            my_win.label_86.setText(list_pl_final[3])
            my_win.label_86.show()
            my_win.label_etap_11.setText(list_pl_final[0])
            my_win.label_etap_11.show()
            my_win.comboBox_table_11.hide()
        my_win.Button_etap_made.setEnabled(True)
        my_win.comboBox_page_vid.setEnabled(True)
        my_win.Button_etap_made.setFocus(True)


def max_exit_player_out_in_group(exit_stage):
    """максимальное число игроков для комбобокса"""
    system = System.select().where(System.title_id == title_id())
    systems = system.select().where(System.stage == exit_stage).get()
    stroka = systems.label_string
    ind = stroka.find("по")
    max_exit_group = int(stroka[ind + 3:ind + 5]) # получаем число игроков в полуфинале
    return max_exit_group


def no_play():
    """победа по неявке соперника"""
    tb = my_win.tabWidget.currentIndex()
    check_gr_pl1 = my_win.checkBox_7.isChecked() 
    check_gr_pl2 = my_win.checkBox_8.isChecked()
    check_sf_pl1 = my_win.checkBox_9.isChecked() 
    check_sf_pl2 = my_win.checkBox_10.isChecked()
    check_fin_pl1 = my_win.checkBox_11.isChecked() 
    check_fin_pl2 = my_win.checkBox_12.isChecked()
    if tb == 3:
        if check_gr_pl1 is False and check_gr_pl2 is False:
            return
    elif tb == 4:
        if check_sf_pl1 is False and check_sf_pl2 is False:
            return
    elif tb == 5:
        if check_fin_pl1 is False and check_fin_pl2 is False:
            return
    none_player = 1 if check_gr_pl1 is True or check_sf_pl1 is True or check_fin_pl1 is True else 2
    enter_score(none_player)


def backup():
    """резервное копирование базы данных"""
    # time_date = date.today()
    try:
        db = sqlite3.connect('comp_db.db')
        db_backup = sqlite3.connect(f'comp_db_backup.db')
        with db_backup:
            db.backup(db_backup, pages=3, progress=None)
        # показывает статус бар на 5 секунд
        my_win.statusbar.showMessage(
            "Резервное копирование базы данных завершено успешно", 5000)
    except sqlite3.Error as error:
        # показывает статус бар на 5 секунд
        my_win.statusbar.showMessage(
            "Ошибка при копировании базы данных", 5000)
    finally:
        if (db_backup):
            db_backup.close()
            db.close()
            my_win.close()


def title_id():
    """возвращает title id в зависимости от соревнования"""
    name = my_win.lineEdit_title_nazvanie.text()  # определяет название соревнований из титула
    if name != "":       
        data = my_win.dateEdit_start.text()
        gamer = my_win.lineEdit_title_gamer.text()
        titles_data = Title.select().where((Title.name == name) & (Title.gamer == gamer)) # получает эту строку в db
        titles = titles_data.select().where(Title.data_start == data).get()
        title_id = titles.id
    else:
        # получение последней записи в таблице
        t_id = Title.select().order_by(Title.id.desc()).get()
        title_id = t_id.id
    return title_id


def func_zagolovok(canvas, doc):
    """создание заголовка страниц"""
    pagesizeW = doc.width
    pagesizeH = doc.height
    current_button = "0"
    if pagesizeH > pagesizeW:
        pv = A4
    else:
        pv = landscape(A4)
    (width, height) = pv

    title = Title.get(Title.id == title_id())

    nz = title.name
    ms = title.mesto
    sr = f"среди {title.sredi} {title.vozrast}"
    data_comp = data_title_string()
    # === определяет если список судей то подпись ставит выше
    tb = my_win.tabWidget.currentIndex()
    if tb == 7:
        for i in my_win.tabWidget.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    current_button = i.text()
                    break
    # =========   
    canvas.saveState()
    canvas.setFont("DejaVuSerif-Italic", 14)
    # центральный текст титула
    canvas.drawCentredString(width / 2.0, height - 1.1 * cm, nz)
    canvas.setFont("DejaVuSerif-Italic", 11)
    # текста титула по основным
    canvas.drawCentredString(width / 2.0, height - 1.5 * cm, sr)
    canvas.drawRightString(width - 1 * cm, height - 1.6 * cm, f"г. {ms}")  # город
    canvas.drawString(0.8 * cm, height - 1.6 * cm, data_comp)  # дата начала

    canvas.setFont("DejaVuSerif-Italic", 11)
    canvas.setFillColor(blue)  # меняет цвет шрифта списка судейской коллеги
    if pv == landscape(A4):
        main_referee_collegia = f"Гл. судья: судья {title.kat_ref}______________ {title.referee}   " \
                                f"Гл. секретарь: судья {title.kat_sec} ______________{title.secretary}"
        if current_button == "3":
            # текста титула по основным
            canvas.drawCentredString(width / 2.0, height - 15 * cm, main_referee_collegia)
        else:
            # текста титула по основным
            canvas.drawCentredString(width / 2.0, height - 20 * cm, main_referee_collegia)
    else:
        main_referee = f"Гл. судья: судья {title.kat_ref} ______________{title.referee}"
        if current_button == "1":
            # подпись главного судьи
            canvas.drawString(2 * cm, 20 * cm, main_referee)
        elif current_button == "2":
            regions_list = [] # сдвигает подпись судьи относительно кол-во регионов
            regions = Player.select().where(Player.title_id == title_id())
            for k in regions:
                region = k.region
                regions_list.append(region)
            regions_set = set(regions_list)
            count = len(regions_set)
            # подпись главного судьи
            canvas.drawString(2 * cm, ((29 - count ) * cm), main_referee)
        else:
            main_secretary = f"Гл. секретарь: судья {title.kat_sec} ______________{title.secretary} "
            # подпись главного судьи
            canvas.drawString(2 * cm, 2 * cm, main_referee)
            # подпись главного секретаря
            canvas.drawString(2 * cm, 1 * cm, main_secretary)
    canvas.restoreState()
    return func_zagolovok


def tbl(stage, kg, ts, zagolovok, cW, rH):
    """данные таблицы и применение стиля и добавления заголовка столбцов
    tdt_new - [[[участник],[регион счет в партиях]]]"""
    from reportlab.platypus import Table
    group_dict = {}
    cell_list_tmp = []
    cell_list = [] # список номеров ячеек где большие фамилиии
    tdt_temp = []
    tdt_new_tmp = []
    dict_tbl = {}
    tdt_all = table_data(stage, kg)  # данные результатов в группах
    # данные результатов победителей в группах для окрашивания очков в красный цвет
    tdt_new = tdt_all[0]
    # убирает id от фамилии и перезаписывает tdt_new
    g = 0
    for group in tdt_new:
        l = 0
        for z in group:
            if l % 2 == 0:
                fam_id = z[1]
                znak = fam_id.find("/")
                if znak != -1:
                    family = fam_id[:znak]
                else:
                    family = fam_id
                z[1] = family 
            # === определяет большие фамилии
                length_of_last_name = len(z[1])
                if length_of_last_name > 17:
                    cell_list_tmp.append(l)
            l += 1
        cell_list = cell_list_tmp.copy()
        group_dict[g] = cell_list
        cell_list_tmp.clear()
        g += 1

    for k in tdt_new:
        tdt_temp = k.copy()
        k.clear()
        tdt_new_temp = tdt_temp.copy()
        tdt_new_tmp.append(tdt_new_temp)
        tdt_temp.clear()
    tdt_new.clear()
    tdt_new = tdt_new_tmp.copy()
    # ===========================
    for i in range(0, kg):
        tdt_new[i].insert(0, zagolovok)       
        dict_tbl[i] = Table(tdt_new[i], colWidths=cW, rowHeights=rH)
        list_cells = group_dict[i]
        ts.add('TEXTCOLOR', (0, 0), (-1, -1), colors.darkblue)
        for m in list_cells:
            ts.add('FONTSIZE', (1, m + 1), (1, m + 1), 5)
            # ts.add('TEXTCOLOR', (1, m + 1), (1, m + 1), colors.green)
        # ставит всю таблицу в синий цвет
        for k in tdt_all[1][i]:
            col = k[0]  # столбец очков победителя
            row = k[1]  # ряд очков победителя
            ts.add('TEXTCOLOR', (col, row + 1), (col, row + 1), colors.red)  # красный цвет очков победителя
        dict_tbl[i].setStyle(ts)  # применяет стиль к таблице данных
    return dict_tbl


def tbl_begunki(ts, stage, number_group, tours, list_tours):
    """данные таблицы и применение стиля и добавления заголовка столбцов
    tdt_new - [[[участник],[регион счет в партиях]]]"""
    stiker = []
    final_type = "круг"
    from reportlab.platypus import Table
    systems = System.select().where(System.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    if stage != "Финальный":
        system = systems.select().where(System.stage == stage).get()
        # total_group = system.total_group
    else:
        system = systems.select().where(System.stage == number_group).get()
        final_type = system.type_table
     # # кол-во столбцов в таблице и их ширина
    cW = (1.6 * cm)
    rH = (0.6 * cm, 0.9 * cm, 1 * cm, 0.6 * cm, 0.6 * cm, 0.6 * cm, 0.6 * cm, 0.6 * cm,
           0.5 * cm, 0.5 * cm)
    dict_tbl = {}
    tdt_new_tmp = []

    if final_type == "сетка":
        result_setka = result.select().where(Result.number_group == number_group)
        result_all = result_setka.select().where((Result.player1 != "") & (Result.player2 != ""))
        result_group = result_all.select().where(Result.winner.is_null())
    else:    
        if number_group == "все" and tours == "все":
            result_group = result.select().where(Result.system_stage == stage)
        elif number_group == "все" and tours == "диапазон":
            result_group = result.select().where((Result.system_stage == stage) & (Result.round.in_(list_tours)))
        elif number_group != "все" and tours == "все":
            if stage == "1-й полуфинал" or stage == "2-й полуфинал":
                result_group = result.select().where((Result.system_stage == stage) & (Result.number_group == number_group))
            else:
                result_group = result.select().where(Result.number_group == number_group)
        elif number_group != "все" and tours == "диапазон":
            result_group = result.select().where((Result.number_group == number_group) & (Result.round.in_(list_tours)))
 
    shot_stage = ""

    for res_id in result_group:
        tours = res_id.tours # номера игроков в туре
        pl1 = res_id.player1 # 1-й игроков и его город в туре
        pl2 = res_id.player2 # 2-й игроков и его город в туре
        st = res_id.number_group # этап
        n_gr = ""
        if stage == "Предварительный":
            shot_stage = "ПР"
            mark = st.find(" ")
            gr = st[:mark]
            sys_stage = f"{shot_stage}"
            n_gr = f"{gr}гр"
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            shot_stage = "ПФ"
            mark = st.find(" ")
            gr = st[:mark]
            sys_stage = f"{shot_stage}"
            n_gr = f"{gr}гр"
        elif stage == "Финальный":
            shot_stage = "Ф"
            mark = st.find("-")
            sys_stage = f"{st[:mark]}{shot_stage}"

        round = res_id.round # раунд
        s1 = pl1.find("/")  
        s2 = pl2.find("/")   
        player1 = pl1[:s1]
        city1 = pl1[s1 + 1:]
        player2 = pl2[:s2]
        city2 = pl2[s2 + 1:]
        pl1 = f"{player1}\n{city1}" # делает фамилия и город на разнызх строчках
        pl2 = f"{player2}\n{city2}"
            # список строк бегунка
        d_tmp = [[n_gr, 'тур', 'вст', 'стол'],
                [sys_stage, round, tours, ''],
                [pl1, '', pl2, ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['общ счет:', '', '', ''],
                ['Победитель', '', '', '']]
        tdt_temp = d_tmp.copy()
        d_tmp.clear()
        tdt_new_temp = tdt_temp.copy()
        tdt_new_tmp.append(tdt_new_temp)
        tdt_temp.clear()
    game = len(tdt_new_tmp)
        # ===========================
    for i in range(0, game):      
        dict_tbl[i] = Table(tdt_new_tmp[i], colWidths=cW, rowHeights=rH)
        dict_tbl[i].setStyle(ts)  # применяет стиль к таблице данных
    stiker.append(dict_tbl)
    stiker.append(game)
    return stiker


def begunki_made():
    """создание бегунков"""
    from sys import platform
    from reportlab.platypus import Table
    msgBox = QMessageBox
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    result = Result.select().where(Result.title_id == title_id())
    number_group = my_win.comboBox_select_group_begunki.currentText()
    stage = my_win.comboBox_select_stage_begunki.currentText()
    tours = my_win.comboBox_select_tours.currentText()
    elements = []
    ts = []
    tblstyle = []
    for p in range(0, 8):
        fn = ('SPAN',(0, 2 + p), (1, 2 + p))
        tblstyle.append(fn)
        fn = ('SPAN',(2, 2 + p), (3, 2 + p))
        tblstyle.append(fn)

    ts.append(tblstyle)
    # span (0,2), (1,2) - объединяет 0 и 1 столбец и строки 2 (0-столбец, 2-строка), (1-столбец, 2-строка)
    ts = TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
                        ('BOX', (0,0), (-1,-1), 1, colors.black)]
                        + tblstyle +
                        [('FONTSIZE', (0, 1), (0, 1), 20),
                        ('VALIGN', (0, 1), (0, 1), 'TOP'),
                        ('ALIGN',(0, 1), (0, 1),'CENTER'),
                        ('FONTSIZE', (0, 2), (3, 2), 7), 
                        ('VALIGN', (1, 0), (3, 0), 'MIDDLE'),
                        ('FONTSIZE', (1, 1), (3, 1), 12), 
                        ('VALIGN', (1, 1), (3, 1), 'MIDDLE'),
                        ('ALIGN',(1, 1), (3, 1),'CENTER'),
                        ('FONTSIZE', (0, 0), (0, 0), 12), 
                        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                        ('ALIGN',(0, 0), (0, 0),'CENTER')])
    #  ========= формирование диапазона печати бегунков ==========
    if stage == "Финальный":
        sys = system.select().where(System.stage == number_group).get()
    elif stage == "Предварительный":
        sys = system.select().where(System.stage == stage).get()

    final_type = sys.type_table
    list_tours = []
    if final_type == "сетка":
        list_tours.append("несыгранные")
        result_setka = result.select().where(Result.number_group == number_group)
        result_all = result_setka.select().where((Result.player1 != "") & (Result.player2 != ""))
        result_fin = result_all.select().where(Result.winner.is_null())
        if len(result_fin) == 0:
            msgBox.information(my_win, "Уведомление", "Все встречи сыграны,\n в печати бегунков нет необходимости.")
            return
    elif final_type == "круг" or final_type == "группы":
        if tours != "все":
            range_tours_str = my_win.lineEdit_range_tours.text()
            txt = range_tours_str.replace(" ", "")
            range_tours_list = list(txt)
            if "-" in range_tours_list:
                range_tours_list.remove("-")
                result_int = [int(item) for item in range_tours_list] # преобразовывает список строковых данных в числовой тип
                for b in range (result_int[0], result_int[1] + 1):
                    b = int(b)
                    list_tours.append(b)
            else:
                tours_list = range_tours_list
                for b in tours_list:
                    if b != ",":
                        b = int(b)
                        list_tours.append(b)
        else:
            if number_group != "все":
                if stage == "1-й полуфинал" or stage == "2-й полуфинал":
                    result_group = result.select().where((Result.system_stage == stage) & (Result.number_group == number_group))
                else:   
                    result_group = result.select().where(Result.number_group == number_group)
                for i in result_group:
                    r = int(i.round)
                    if r not in list_tours:
                        list_tours.append(r)
        
    stiker = tbl_begunki(ts, stage, number_group, tours, list_tours) # здесь надо менять данные бегунков
    dict_table = stiker[0]
    game = stiker[1]

    data_tmp = []
    data_temp = []
    tmp = []
    temp = []
    data = []
    celoe = game // 3
    ostatok = game % 3
    end = 0
    row = 3
    if ostatok == 0:
        end = celoe + 1
    else:
        end = celoe + 2
    a = 0
    for k in range(1, end):
        if ostatok !=0 and k == end - 1:
            row = ostatok
        for i in range(0, row): # кол-во бегунков в 
            data_tmp.append(dict_table[a])
            a += 1
        tmp = data_tmp.copy()
        data_temp.append(tmp) 
        temp = data_temp.copy()
        data.append(temp)
        data_tmp.clear()
        data_temp.clear()
    shell_table = []
    count_data = len(data)
    s_tmp = []
    for l in range(0, count_data): 
        shell_tmp = Table(data[l], colWidths=["*"])
        s_tmp.append(shell_tmp)
        tmp_copy = s_tmp.copy()
        shell_table.append(tmp_copy)
        s_tmp.clear()
        elements.append(shell_table[l][0])
 
    name_table = "begunki.pdf"
    # устанавливает поля на странице pdf
    doc = SimpleDocTemplate(name_table, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements)
    my_win.lineEdit_range_tours.clear()
    my_win.lineEdit_range_tours.hide()
    view_file = name_table
    if platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    os.chdir("..") # возврат на предыдущий уровень


def select_stage_for_begunki():
    """выбор финалов или номеров групп для печати бегунков"""
    my_win.comboBox_select_group_begunki.clear()
    if my_win.comboBox_select_stage_begunki.currentIndex() != 0:
        my_win.Button_print_begunki.setEnabled(True)
    systems = System.select().where(System.title_id == title_id())
    group_list = ["все"]
    stage = my_win.comboBox_select_stage_begunki.currentText()
    if stage == "-Выбор этапа-":
        pass
    elif stage == "Предварительный" or stage == "1-й полуфинал" or stage == "2-й полуфинал":
        sys_id = systems.select().where(System.stage == stage).get()
        group = sys_id.total_group
        group_list = [f"{i} группа" for i in range(1, group + 1)] # генератор списка
        group_list.insert(0, "все")
    elif stage == "Одна таблица":
        pass
    else:
        for k in systems:
            if k.stage == "Предварительный":
                pass
            elif k.stage == "Полуфинал":
                pass
            else:
                group_list.append(k.stage)
    my_win.comboBox_select_group_begunki.addItems(group_list)

        
def select_tour_for_begunki():
    """выбор номеров тура или диапазона туров""" 
    my_win.comboBox_select_tours.clear()
    tour_list = ["все", "диапазон"]
    my_win.comboBox_select_tours.addItems(tour_list)
    index = my_win.comboBox_select_tours.currentIndex()
    if index != 0:
        my_win.lineEdit_range_tours.show()


def select_diapazon():
    """показывает поле для ввода дмапазона туров"""
    my_win.lineEdit_range_tours.clear()
    index = my_win.comboBox_select_tours.currentIndex()
    if index != 0:
        my_win.lineEdit_range_tours.show()
        my_win.lineEdit_range_tours.setFocus()
    else:
        my_win.lineEdit_range_tours.hide()


def enter_print_begunki():
    """Печать бегунков при нажатии энтер на поле диапазона"""
    sender = my_win.sender()
    if sender == my_win.lineEdit_range_tours:
        begunki_made()


def merdge_pdf_files():
    """Слияние все таблиц соревнований в один файл"""
    pdf_merger = PdfMerger()

    title = Title.get(Title.id == title_id())
    pdf_files_list = []
    short_name = title.short_name_comp
 
    count = my_win.tableWidget.rowCount()
    for k in range(0, count):
        name_files = my_win.tableWidget.item(k, 1).text()
        pdf_files_list.append(name_files)

    my_win.tableWidget.setColumnCount(2) # устанавливает колво столбцов
    my_win.tableWidget.setRowCount(count)
    column_label = ["№", "Этапы"]

    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
    my_win.tableWidget.setDragDropOverwriteMode(True)
    my_win.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
    my_win.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
    my_win.tableWidget.show()
    catalog = 1 # переходит в каталог /table_pdf и считывает все файлы этого соревнования
    change_dir(catalog)
    with contextlib.ExitStack() as stack:
        files = [stack.enter_context(open(pdf, 'rb')) for pdf in pdf_files_list]
        for f in files:
            pdf_merger.append(f)
        os.chdir("..")
        catalog = 2
        change_dir(catalog)
        with open(f'{short_name}.pdf', 'wb') as f:
            pdf_merger.write(f)
            title.pdf_comp = f'{short_name}.pdf'
            title.save()
    my_win.tableWidget.show()


def load_name_net_after_choice_for_wiev(fin):
    """загружает список сетки после жеребьевки для ее просмотра"""
    system = System.select().where(System.title_id == title_id())
    for k in system:
        stage = k.stage
        if stage == fin:
            vid_net = k.label_string
            break
    if vid_net == 'Сетка (с розыгрышем всех мест) на 8 участников':
        setka_8_full_made(fin)
    elif vid_net == 'Сетка (-2) на 8 участников':
        setka_8_2_made(fin)
    elif vid_net == 'Сетка (с розыгрышем всех мест) на 16 участников':
        setka_16_full_made(fin)
    elif vid_net == 'Сетка (-2) на 16 участников':
        setka_16_2_made(fin)
    elif vid_net == 'Сетка (с розыгрышем всех мест) на 32 участников':
        setka_32_full_made(fin)
    elif vid_net == 'Сетка (-2) на 32 участников':
        setka_32_2_made(fin)
    elif vid_net == 'Сетка (1-3 место) на 32 участников':
        setka_32_made(fin)


def table_made(pv, stage):
    """создание таблиц kg - количество групп(таблиц), g2 - наибольшое кол-во участников в группе
     pv - ориентация страницы, е - если участников четно группам, т - их количество"""
    # start_time = time.time()
    from reportlab.platypus import Table
    system = System.select().where((System.title_id == title_id()) & (System.stage == stage)).get()  # находит system id последнего
    type_tbl = system.type_table
    titles = Title.select().where(Title.id == title_id()).get()
    sex = titles.gamer

    # =========
    # styles = getSampleStyleSheet()
    # styleN = styles['Normal']
    # styleH = styles['Heading1']
    # story = []
    #     #add some flowables
    # story.append(Paragraph("This is a Heading",styleH))
    # story.append(Paragraph("This is a paragraph in <i>Normal</i> style.",
    #        styleN))
    # canvas  = Canvas('mydoc.pdf', pagesize=landscape(A4))
    # f = Frame(5* cm, 3 * cm, 6 * cm, 25 * cm, showBoundary=0) # высота прямоугольника  6 Х 25, showBoundary = 1, рамка 0- нет
    # f.addFromList(story,c)
    # c.save()

    # ==========
 
    if (stage == "Одна таблица" and type_tbl == "круг") or (stage != "Одна таблица" and type_tbl == "круг"):
        kg = 1
        max_pl = system.max_player
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        kg = system.total_group  # кол-во групп
        max_pl = system.max_player // kg
    else:  # групповые игры
        kg = system.total_group  # кол-во групп
        max_pl = system.max_player
        
    family_col = 3.2
    if pv == "альбомная":  # альбомная ориентация стр
        pv = landscape(A4)
        center_stage = 300 # откуда начинается надпись -предварительный этап-
        if kg == 1 or max_pl in [10, 11, 12, 13, 14, 15, 16]:
            # ширина столбцов таблицы в зависимости от кол-во чел (1 таблица)
            wcells = 21.4 / max_pl
        else:
            # ширина столбцов таблицы в зависимости от кол-во чел (2-ух в ряд)
            wcells = 7.4 / max_pl
    else:  # книжная ориентация стр
        pv = A4
        center_stage = 150 # откуда начинается надпись -предварительный этап-
        if max_pl < 7:
            family_col = 4.0
            wcells = 12.0 / max_pl  # ширина столбцов таблицы в зависимости от кол-во чел
        else:
            family_col = 3.8
            wcells = 12.8 / max_pl  # ширина столбцов таблицы в зависимости от кол-во чел
    col = ((wcells * cm,) * max_pl)
    elements = []

    # кол-во столбцов в таблице и их ширина
    cW = ((0.4 * cm, family_col * cm) + col + (1 * cm, 1 * cm, 1 * cm))
    if kg == 1:
        rH = (0.45 * cm)  # высота строки
    else:
        if max_pl < 5:
            rH = (0.34 * cm)  # высота строки
        else:
            rH = (0.32 * cm)  # высота строки
    num_columns = []  # заголовки столбцов и их нумерация в зависимости от кол-во участников

    for i in range(max_pl):
        i += 1
        i = str(i)
        num_columns.append(i)
    zagolovok = (['№', 'Участники/ Город'] + num_columns + ['Очки', 'Соот', 'Место'])

    tblstyle = []
    # =========  цикл создания стиля таблицы ================
    for q in range(1, max_pl + 1):  # город участника делает курсивом
        # город участника делает курсивом
        fn = ('FONTNAME', (1, q * 2), (1, q * 2), "DejaVuSerif-Italic")
        tblstyle.append(fn)
        fn = ('FONTNAME', (1, q * 2 - 1), (1, q * 2 - 1),
              "DejaVuSerif-Bold")  # участника делает жирным шрифтом
        tblstyle.append(fn)
        # центрирование текста в ячейках)
        fn = ('ALIGN', (1, q * 2 - 1), (1, q * 2 - 1), 'LEFT')
        tblstyle.append(fn)
        # объединяет 1-2, 3-4, 5-6, 7-8 ячейки 1 столбца
        fn = ('SPAN', (0, q * 2 - 1), (0, q * 2))
        tblstyle.append(fn)
        # объединяет клетки очки
        fn = ('SPAN', (max_pl + 2, q * 2 - 1), (max_pl + 2, q * 2))
        tblstyle.append(fn)
        # объединяет клетки соот
        fn = ('SPAN', (max_pl + 3, q * 2 - 1), (max_pl + 3, q * 2))
        tblstyle.append(fn)
        # объединяет клетки  место
        fn = ('SPAN', (max_pl + 4, q * 2 - 1), (max_pl + 4, q * 2))
        tblstyle.append(fn)
        # объединяет диагональные клетки
        fn = ('SPAN', (q + 1, q * 2 - 1), (q + 1, q * 2))
        tblstyle.append(fn)
        fn = ('BACKGROUND', (q + 1, q * 2 - 1), (q + 1, q * 2),
              colors.lightgreen)  # заливает диагональные клетки
        tblstyle.append(fn)

    ts = []
    ts.append(tblstyle)
    # ============= полный стиль таблицы ======================
    ts = TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                     ('FONTSIZE', (0, 0), (-1, -1), 6),
                     # вставить размер шрифта конкретной ячей под длинную фамилию
                     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                     ('FONTNAME', (0, 0), (max_pl + 5, 0), "DejaVuSerif-Bold"),
                     ('VALIGN', (0, 0), (max_pl + 5, 0), 'MIDDLE')]  # центрирование текста в ячейках вертикальное
                    + tblstyle +
                    [('BACKGROUND', (0, 0), (max_pl + 5, 0), colors.yellow),
                     # цвет шрифта в ячейках
                     ('TEXTCOLOR', (0, 0), (-1, -1), colors.darkblue),
                     ('LINEABOVE', (0, 0), (-1, 1), 1,
                      colors.black),  # цвет линий нижней
                     # цвет и толщину внутренних линий
                     ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                     ('BOX', (0, 0), (-1, -1), 2, colors.black)])  # внешние границы таблицы

    #  ============ создание таблиц и вставка данных =================
    h1 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic",
            leftIndent=center_stage, spacebefore=10, textColor="green")  # стиль параграфа ()

    # styles = getSampleStyleSheet()
    # title_style = styles['Title']
    # title_style.textColor = colors.red
    # title_style.fontSize = 24
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=300, spacebefore=20, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
            #========
    dict_table = tbl(stage, kg, ts, zagolovok, cW, rH)
    if kg == 1:  # одна таблицу
        data = [[dict_table[0]]]
        shell_table = Table(data, colWidths=["*"])
        if stage == "Одна таблица":
            text = ""
        else:
            text = stage
        elements.append(Paragraph(text, h2))
        elements.append(shell_table)
    else:
        data_tmp = []
        data_temp = []
        tmp = []
        temp = []
        data = []
        if pv == landscape(A4):  # страница альбомная, то таблицы размещаются обе в ряд
            for k in range(1, kg // 2 + 1):
                for i in range(0, 2):
                    data_tmp.append(dict_table[(k * 2 - 2) + i])  
                tmp = data_tmp.copy()
                data_temp.append(tmp) 
                temp = data_temp.copy()
                data.append(temp)
                data_tmp.clear()
                data_temp.clear()
            shell_table = []
            s_tmp = []
            for l in range(0, kg // 2): 
                shell_tmp = Table(data[l], colWidths=["*"])
                s_tmp.append(shell_tmp)
                tmp_copy = s_tmp.copy()
                shell_table.append(tmp_copy)
                s_tmp.clear()
                text = f'группа {l * 2 + 1} группа {l * 2 + 2}'
                elements.append(Paragraph(text, h2))
                elements.append(shell_table[l][0])
                # =======
        else:  # страница книжная, то таблицы размещаются обе в столбец
            for k in range(1, kg // 2 + 1):
                for i in range(0, kg):
                    data_tmp.append(dict_table[i])  
                    tmp = data_tmp.copy()
                    data_temp.append(tmp) 
                    temp = data_temp.copy()
                    data.append(temp)
                    data_tmp.clear()
                    data_temp.clear()
            shell_table = []
            s_tmp = []
            for l in range(0, kg): 
                shell_tmp = Table(data[l], colWidths=["*"])
                s_tmp.append(shell_tmp)
                tmp_copy = s_tmp.copy()
                shell_table.append(tmp_copy)
                s_tmp.clear()
                elements.append(Paragraph(f'группа {l + 1}', h2))
                elements.append(shell_table[l][0])

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    short_name = t_id.short_name_comp

    if stage == "Одна таблица":
        name_table = f"{short_name}_one_table.pdf"
    elif stage == "Предварительный":
        title = "Предварительный этап"
        name_table = f"{short_name}_table_group.pdf"
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        txt = stage.rfind("-")
        number_fin = stage[:txt]
        title = stage
        name_table = f"{short_name}_{number_fin}-semifinal.pdf"
    else:
        txt = stage.rfind("-")
        number_fin = stage[:txt]
        title = stage
        name_table = f"{short_name}_{number_fin}-final.pdf"
    doc = SimpleDocTemplate(name_table, pagesize=pv)
    catalog = 1
    change_dir(catalog)
    doc.topMargin = 1.8 * cm # высота отступа от верха листа pdf
    doc.bottomMargin = 1.6 * cm
    elements.insert(0, (Paragraph(f"{title}. {sex}", h1)))
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")

    # end_time = time.time()
    # execution_time = end_time - start_time
    # print(f"Время выполнения: {execution_time} секунд")
 # ========
# def create_report():
#     doc = SimpleDocTemplate("sample_report.pdf")
     
#     styles = getSampleStyleSheet()
#     title_style = styles['Title']
#     title_style.textColor = colors.red
#     title_style.fontSize = 24
     
#     story = []
#     title = Paragraph("Пример отчета", styles['Title'])
#     story.append(title)
     
#     doc.build(story)
# # ===========


def list_regions_pdf():
    """список субъектов РФ"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    region_list = []
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp
    regions = Player.select().where(Player.title_id == title_id())

    for k in regions:
        reg = k.region
        if reg not in region_list:
            region_list.append(reg)

    kp = len(region_list)
    region_list.sort()
    n = 0
    for reg in region_list:
        n += 1
        num = n
        data = [num, reg]
        elements.append(data)
    elements.insert(0, ["№", "Субъекты РФ"])
    t = Table(elements,
              colWidths=(1.0 * cm, 10.0 * cm),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 10),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, -1), 'CENTER'), # (1-я ячейка столб, ряд)  (2-я ячейка столб, ряд)
                        #    ('ALIGN', (0, 0), (0, kp), 'CENTER'), # (1-я ячейка столб, ряд)  (2-я ячейка столб, ряд)
                           ('ALIGN', (1, 1), (1, kp), 'LEFT'),
                           ('BACKGROUND', (0, 0), (1, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (1, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.8, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150, textColor="green",
            firstLineIndent=-20)  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список субъектов РФ', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_regions_list.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def list_winners_pdf():
    """список призеров"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp
   
    count_Row = my_win.tableWidget.rowCount()
    kp = count_Row + 1
    n = 0
    for l in range(0, count_Row):
        mesto = my_win.tableWidget.item(l, 0).text()
        player = my_win.tableWidget.item(l, 1).text()
        bday = my_win.tableWidget.item(l, 2).text()
        rank = my_win.tableWidget.item(l, 3).text()
        city = my_win.tableWidget.item(l, 4).text()
        region = my_win.tableWidget.item(l, 5).text()
        razryad = my_win.tableWidget.item(l, 6).text()
        coach = my_win.tableWidget.item(l, 7).text()
        n += 1

        data = [mesto, player, bday, rank, city, region, razryad, coach]
        elements.append(data)
    elements.insert(0, ["Место", "Фамилия, Имя", "Дата рождения", "Рейтинг", "Город", "Регион", "Разряд", "Тренеры"])
    t = Table(elements,
              colWidths=(2.3 * cm, 5.6 * cm, 3.0 * cm, 2.0 * cm, 3.0 * cm, 4.0 * cm, 2.0 * cm, 5.5 * cm,),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 9),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (0, 7), 'CENTER'),
                           ('ALIGN', (1, 0), (1, kp), 'LEFT'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.8, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=200, textColor="green",
            firstLineIndent=-20)  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список победителей и призеров', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_winners_list.pdf", pagesize=landscape(A4))
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def list_referee_pdf():
    """список судейской коллегии"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp

    count_Row = my_win.tableWidget.rowCount()
    kp = count_Row + 1
    n = 0
    for l in range(0, count_Row):
        if n < 2: 
            num = my_win.tableWidget.item(l, 0).text()
            post = my_win.tableWidget.item(l, 1).text()
            fam_city = my_win.tableWidget.item(l, 2).text()
            category = my_win.tableWidget.item(l, 3).text()
        else:
            post_combobox = my_win.tableWidget.cellWidget(l, 1)
            post = post_combobox.currentText()
            fam_city_comboBox = my_win.tableWidget.cellWidget(l, 2)
            fam_city = fam_city_comboBox.currentText()
            category_combobox = my_win.tableWidget.cellWidget(l, 3)
            category = category_combobox.currentText()
        num = my_win.tableWidget.item(l, 0).text()
        n += 1
        data = [num, post, fam_city, category]
        elements.append(data)
    elements.insert(0, ["№", "Должность", "Фамилия Имя/ город", "Категория"])
    t = Table(elements,
              colWidths=(0.6 * cm, 7.0 * cm, 7.7 * cm, 3.2 * cm),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 10),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.9, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150, textColor="green",
            firstLineIndent=-20)  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список главной судейской коллегии', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_referee_list.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def setka_8_full_made(fin):
    """сетка на 8 в pdf"""
    sender = my_win.sender()
    from reportlab.platypus import Table
    table = "setka_8_full"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 10
    # добавить в аргументы функции
    final = fin
    if sender != my_win.clear_s8_full_Action:
        first_mesto = mesto_in_final(fin)
    else:
        first_mesto = 1  # временный финал для чистой сетки
    for i in range(0, 40):
        # column_count[9] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= места ==========
    y = 0
    for i in range(0, 16, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=1, row_step=2, col_n=2, number_of_columns=3, number_of_game=1, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=16, row_step=2, col_n=6, number_of_columns=2, number_of_game=8, player=2, data=data) # рисует номера встреч 1-32
    draw_num(row_n=20, row_step=2, col_n=4, number_of_columns=2, number_of_game=9, player=4, data=data) # рисует номера встреч 1-32
    draw_num_lost(row_n=16, row_step=2, col_n=4, number_of_game=5, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=20, row_step=2, col_n=2, number_of_game=1, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=28, row_step=2, col_n=4, number_of_game=9, player=2, data=data) # номера минус проигравшие встречи -1 -16
   
    data[28][6] = str(12)  # создание номеров встреч 15
    data[13][6] = str(-7)
    data[18][6] = str(-8)
    data[25][6] = str(-11)
    data[30][6] = str(-12)
    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 4.8 * cm, 1.5 * cm, 0.4 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    color_mesta(data, first_mesto, table) # раскрашивает места участников красным цветом
    t = Table(data, cw, 40 * [0.6 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 1, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_setka(3, 20, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 28, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 16, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    # ======= встречи за места =====
    for q in range(0, 7, 6):
        fn = ('LINEABOVE', (7, q + 8), (8, q + 8),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 3, 2):
        fn = ('LINEABOVE', (7, q + 17), (8, q + 17),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
        fn = ('LINEABOVE', (7, q + 29), (8, q + 29),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (7, q + 23), (8, q + 23),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)

    for i in range(1, 6, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 39), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 39), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 39), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 39), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)
    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 16), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 16), 7),
                           # 10 столбец с 0 по 68 ряд (цвет места)
                           ('TEXTCOLOR', (8, 0), (8, 39), colors.red),
                           ('ALIGN', (8, 0), (8, 39), 'RIGHT'),
                           ('ALIGN', (7, 0), (7, 39), 'LEFT'),
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 39), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                           ] + ts))
# === надпись финала
    h2 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic",
            leftIndent=200, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(fin, h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        else:
            name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_8_full_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok)
    os.chdir("..") # переходит на один уровень на верх
    return tds


def setka_8_2_made(fin):
    """сетка на 8 минус 2 в pdf"""
    sender = my_win.sender()
    from reportlab.platypus import Table
    table = "setka_8_2"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 10
    # добавить в аргументы функции
    final = fin
    if sender != my_win.clear_s8_2_Action:
        first_mesto = mesto_in_final(fin)
    else:
        first_mesto = 1  # временный финал для чистой сетки
    for i in range(0, 40):
        # column_count[9] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= места ==========
    y = 0
    for i in range(0, 16, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=1, row_step=2, col_n=2, number_of_columns=3, number_of_game=1, player=8, data=data) # рисует номера встреч 1-32
    draw_num_2(row_n=17, row_step=2, col_n=2, number_of_columns=2, number_of_game=8, player=4, data=data) # рисует номера встреч 33-47 
    draw_num_lost_2(row_n=15, row_step=2, col_n=2, revers_number=1, number_of_game=5, player=2, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost(row_n=17, row_step=2, col_n=0, number_of_game=1, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=25, row_step=2, col_n=4, number_of_game=10, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=31, row_step=2, col_n=4, number_of_game=8, player=2, data=data) # номера минус проигравшие встречи -1 -16
    
    data[13][6] = str(-7)
    data[17][6] = str(12)  # создание номеров встреч 12
    data[25][6] = str(13)
    data[31][6] = str(14)
    data[22][6] = str(-12)
    data[28][6] = str(-13)
    data[34][6] = str(-14)  # создание номеров встреч 27
    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 4.8 * cm, 1.0 * cm, 0.2 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    color_mesta(data, first_mesto, table) # раскрашивает места участников красным цветом
    t = Table(data, cw, 40 * [0.6 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 1, 8, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka_2(1, 17, 4, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka(5, 25, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_setka(5, 31, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    # ======= встречи за места =====
    for q in range(0, 7, 6):
        fn = ('LINEABOVE', (7, q + 8), (8, q + 8),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 5, 4):
        fn = ('LINEABOVE', (7, q + 19), (8, q + 19),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (7, q + 26), (8, q + 26),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (7, q + 32), (8, q + 32),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)

    for i in range(1, 6, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 39), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 39), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 39), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 39), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)
    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 15), "DejaVuSerif-Bold"),
                        #    ('FONTSIZE', (1, 0), (1, 15), 7),
                           # 10 столбец с 0 по 68 ряд (цвет места)
                           ('TEXTCOLOR', (8, 0), (8, 39), colors.red),
                           # столбец с фамилиями за места выравнивает слева
                           ('ALIGN', (7, 0), (7, 39), 'LEFT'), 
                           # столбец с местами выравнивает справа
                           ('ALIGN', (8, 0), (8, 39), 'RIGHT'),
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 39), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                           ] + ts))
# === надпись финала
    h2 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic",
            leftIndent=200, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(fin, h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        else:
            name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_8_2_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok)
    os.chdir("..")
    return tds


def setka_16_full_made(fin):
    """сетка на 16 в pdf"""
    sender = my_win.sender()
    from reportlab.platypus import Table
    table = "setka_16_full"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 11
    # добавить в аргументы функции
    final = fin
    if sender != my_win.clear_s16_Action:
        first_mesto = mesto_in_final(fin)
    else:
        first_mesto = 1  # временный финал для чистой сетки
    for i in range(0, 69):
        # column_count[10] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp) # пустая основа сетки
    # ========= места ==========
    y = 0
    for i in range(0, 32, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=1, row_step=2, col_n=2, number_of_columns=4, number_of_game=1, player=16, data=data) # рисует номера встреч 1-32
    draw_num(row_n=32, row_step=2, col_n=6, number_of_columns=2, number_of_game=17, player=4, data=data) # рисует номера встреч 1-32
    draw_num(row_n=41, row_step=2, col_n=4, number_of_columns=3, number_of_game=21, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=58, row_step=2, col_n=6, number_of_columns=2, number_of_game=29, player=4, data=data) # рисует номера встреч 1-32
    draw_num_lost(row_n=29, row_step=2, col_n=6, number_of_game=13, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=32, row_step=2, col_n=4, number_of_game=9, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=39, row_step=2, col_n=6, number_of_game=17, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=41, row_step=2, col_n=2, number_of_game=1, player=8, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=58, row_step=2, col_n=4, number_of_game=21, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=55, row_step=2, col_n=6, number_of_game=25, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=65, row_step=2, col_n=6, number_of_game=29, player=2, data=data) # номера минус проигравшие встречи -1 -16
   
    data[8][8] = str(15)  # создание номеров встреч 15
    data[25][8] = str(-15)
    data[29][8] = str(16)  # создание номеров встреч 16
    data[31][8] = str(-16)
    data[37][8] = str(-19)
    data[39][8] = str(20)
    data[41][8] = str(-20)
    data[44][8] = str(27)  # создание номеров встреч 27
    data[52][8] = str(-27)
    data[55][8] = str(28)  # создание номеров встреч 28
    data[57][8] = str(-28)
    data[63][8] = str(-31)
    data[65][8] = str(32)  # создание номеров встреч 32
    data[67][8] = str(-32)

    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm,
           0.4 * cm, 4.4 * cm, 1.3 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    color_mesta(data, first_mesto, table) # раскрашивает места участников красным цветом
    t = Table(data, cw, 69 * [0.35 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 1, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka(7, 29, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 32, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 39, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(3, 41, 8, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 55, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 58, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 65, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    # ======= встречи за места =====
    for q in range(0, 11, 10):
        fn = ('LINEABOVE', (9, q + 16), (10, q + 16),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 3, 2):
        fn = ('LINEABOVE', (9, q + 30), (10, q + 30),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 40), (10, q + 40),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 56), (10, q + 56),
              1, colors.darkblue)  # за 11-12 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 66), (10, q + 66),
              1, colors.darkblue)  # за 15-16 место
        style.append(fn)
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (9, q + 35), (10, q + 35),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 61), (10, q + 61),
              1, colors.darkblue)  # за 13-14 место
        style.append(fn)
    for q in range(0, 6, 5):
        fn = ('LINEABOVE', (9, q + 48), (10, q + 48),
              1, colors.darkblue)  # за 9-10 место
        style.append(fn)

    for i in range(1, 8, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 68), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 68), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 68), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 68), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)
    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 32), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 32), 7),
                           # 10 столбец с 0 по 68 ряд (цвет места)
                           ('TEXTCOLOR', (10, 0), (10, 68), colors.red),
                        #    ('ALIGN', (10, 0), (10, 68), 'RIGHT'),
                           ('ALIGN', (9, 0), (9, 68), 'LEFT'),
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                           ] + ts))
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=200, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(fin, h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        else:
            name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_16_full_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=2.2*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_16_2_made(fin):
    """сетка на 16_2 в pdf"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_16_2"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 11
    # добавить в аргументы функции
    final = fin
    if sender != my_win.clear_s16_2_Action:
        first_mesto = mesto_in_final(fin)
    else:
        first_mesto = 1  # временный финал для чистой сетки
    for i in range(0, 86):
        # column_count[10] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= места ==========
    y = 0
    for i in range(2, 34, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=4, number_of_game=1, player=16, data=data) # рисует номера встреч 1-32
    draw_num_lost_2(row_n=45, row_step=1, col_n=0, revers_number=0, number_of_game=1, player=8, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost_2(row_n=44, row_step=2, col_n=2, revers_number=1, number_of_game=9, player=4, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost_2(row_n=43, row_step=4, col_n=6, revers_number=0, number_of_game=13, player=2, data=data) # номера минус проигравшие встречи -17-24
    draw_num(row_n=62, row_step=2, col_n=2, number_of_columns=2, number_of_game=31, player=4, data=data) # рисует номера встреч 1-32
    draw_num(row_n=74, row_step=2, col_n=2, number_of_columns=2, number_of_game=35, player=4, data=data) # рисует номера встреч 1-32

    draw_num(row_n=46, row_step=2, col_n=2, number_of_columns=1, number_of_game=16, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=45, row_step=2, col_n=4, number_of_columns=2, number_of_game=20, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=44, row_step=4, col_n=8, number_of_columns=1, number_of_game=26, player=4, data=data) # рисует номера встреч 1-32
    draw_num_lost(row_n=62, row_step=2, col_n=0, number_of_game=20, player=4, data=data) # номера минус проигравшие встречи -20 -23
    draw_num_lost(row_n=74, row_step=2, col_n=0, number_of_game=16, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=61, row_step=2, col_n=6, number_of_game=26, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=67, row_step=2, col_n=6, number_of_game=24, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=73, row_step=2, col_n=6, number_of_game=31, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=79, row_step=2, col_n=6, number_of_game=35, player=2, data=data) # номера минус проигравшие встречи -1 -16
   
    data[46][10] = str(28)  # создание номеров встреч 15
    data[34][8] = str(-15)
    data[57][8] = str(-28)
    data[70][4] = str(-33)
    data[82][4] = str(-37)
    data[61][8] = str(29)  # создание номеров встреч 27
    data[64][8] = str(-29)
    data[67][8] = str(30)  # создание номеров встреч 28
    data[70][8] = str(-30)
    data[73][8] = str(34)  # создание номеров встреч 32
    data[76][8] = str(-34)
    data[79][8] = str(38)  # создание номеров встреч 32
    data[82][8] = str(-38)


    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm,
           0.4 * cm, 4.4 * cm, 0.4 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table) # раскрашивает места участников красным цветом
    t = Table(data, cw, 86 * [0.55 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 3, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka(1, 62, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(1, 74, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka_2(1, 46, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_setka_made(9, 46, 2, 8, 1, style) # рисует кусок сетки(номер столбца, колво уч, шаг между линиями)
    style = draw_setka(7, 61, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 67, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 73, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 79, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    # ======= встречи за места =====
    for q in range(0, 18, 17):
        fn = ('LINEABOVE', (9, q + 18), (10, q + 18),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 9, 8):
        fn = ('LINEABOVE', (9, q + 50), (10, q + 50),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
 
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (9, q + 62), (10, q + 62),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 68), (10, q + 68),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 74), (10, q + 74),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 80), (10, q + 80),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)

    for q in range(0, 7, 6):
        fn = ('LINEABOVE', (5, q + 65), (6, q + 65),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)
        fn = ('LINEABOVE', (5, q + 77), (6, q + 77),
              1, colors.darkblue)  # за 9-10 место
        style.append(fn)

    for i in range(1, 10, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 85), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 85), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 85), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 85), 'CENTER')
        style.append(fn)
   
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)

    for b in style_color: # цикл окрашивания мест красным цветом
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 32), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 32), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 40), colors.blue),
                           ('TEXTCOLOR', (0, 41), (0, 85), colors.green),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=200, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(fin, h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        else:
            name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_16_2_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_32_made(fin):
    """сетка на 32 с розыгрышем 1-3 места"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_32"
    elements = []
    style = []
    data = []
    column = ['']
    column_count = column * 13
    final = fin
    if sender != my_win.clear_s32_Action:
        first_mesto = mesto_in_final(fin)
    else:
        first_mesto = 1
    strok = 69
    for i in range(0, strok):
        # column_count[12] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= нумерация встреч сетки ==========
    y = 0
    for i in range(1, 65, 2):
        y += 1
        data[i + 1][0] = str(y)  # рисует начальные номера таблицы 1-32
    number_of_game = draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=5, number_of_game=1, player=32, data=data) # рисует номера встреч 1-32
    data[60][8] = str((number_of_game - 3) * -1)  # номера проигравших 29
    data[62][8] = str((number_of_game - 2) * -1)  # номера проигравших 30
    data[18][10] = str(number_of_game - 1)  # создание номеров встреч (31)
    data[55][10] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-31)
    data[61][10] = str(number_of_game)  # создание номеров встреч 32
    data[66][10] = str((number_of_game) * -1)  # номер проигравшего финал (-32)

    # ============= данные игроков и встреч и размещение по сетке =============
    # ======= создать словарь  ключ - номер встречи, значение - номер ряда
    dict_num_game = {}
    for d in range(2, 11, 2):
        for r in range(0, 69):
            key = data[r][d]
            if key != "":
                dict_num_game[key] = r
 
    tds = write_in_setka(data, fin, first_mesto, table)
    cw = ((0.2 * cm, 3.8 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm,
        2.5 * cm, 0.35 * cm, 3.0 * cm, 0.3 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table) # раскрашивает места участников красным цветом
    t = Table(data, cw, strok * [0.35 * cm])
    # =========  цикл создания стиля таблицы =======
    # ========= 1 страница =========
    style = draw_setka(1, 3, 32, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
   
    for l in range(34, 57, 22):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
        style.append(fn)
    for l in range(62, 68, 5):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 3-4 места
        style.append(fn)
    for l in range(61, 64, 2):
        fn = ('LINEABOVE', (9, l), (10, l), 1, colors.darkblue)  # рисует линии встреч за -29 -30
        style.append(fn)
    fn = ('BOX', (10, 61), (10, 62), 1, colors.darkblue)
    style.append(fn) 
    fn = ('SPAN', (10, 61), (10, 62))  # встреча 32
    style.append(fn)       
    fn = ('BACKGROUND', (10, 61), (10, 62), colors.lightyellow)  # встречи 32 за 3-4 место
    style.append(fn)

    for i in range(0, 11, 2):
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, strok), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i, 0), (i, strok), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i + 1, 0), (i + 1, strok), 'LEFT')
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i, 0), (i, strok), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)
    ts = style   # стиль таблицы (список оформления строк и шрифта)
    for b in style_color:
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 32), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 32), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')])) 
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=200, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(fin, h2))
# ====                         
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_32_net"
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_32_full_made(fin):
    """сетка на 32 с розыгрышем всех мест"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_32_full"
    elements = []
    style = []
    data = []
    column = ['']
    column_count = column * 13
    final = fin
    #===== выбор чистая
    if sender != my_win.clear_s32_full_Action:
        first_mesto = mesto_in_final(fin)
    else:
        first_mesto = 1
    strok = 207
    for i in range(0, strok):
        # column_count[12] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= нумерация встреч сетки ==========
    y = 0
    for i in range(1, 65, 2):
        y += 1
        data[i + 1][0] = str(y)  # рисует начальные номера таблицы 1-32
    number_of_game = draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=5, number_of_game=1, player=32, data=data) # рисует номера встреч 1-32
    data[60][8] = str((number_of_game - 3) * -1)  # номера проигравших 29
    data[62][8] = str((number_of_game - 2) * -1)  # номера проигравших 30
    data[18][10] = str(number_of_game - 1)  # создание номеров встреч (31)
    data[55][10] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-31)
    data[61][10] = str(number_of_game)  # создание номеров встреч 32
    data[66][10] = str((number_of_game) * -1)  # номер проигравшего финал (-32)
    #===== 2-я страница =========
    y = 0
    for i in range(78, 89, 10):
        y += 1
        data[i][8] = str((y + 34) * -1)  # номер проигравшего финал (-35, -36)
    y = 0
    for i in range(101, 112, 9):
        y += 1
        data[i][8] = str((y + 42) * -1)  # номер проигравшего финал (-43, -44)
    y = 0
    for i in range(120, 131, 10):
        y += 1
        data[i][8] = str((y + 46) * -1)  # номер проигравшего финал (-47, -48)
    y = 0
    for i in range(172, 187, 14):
        y += 4
        data[i][10] = str((y + 60) * -1)  # номер проигравшего финал (-64, -68)
    y = 0
    for i in range(198, 206, 7):
        y += 4
        data[i][10] = str((y + 72) * -1)  # номер проигравшего финал (-76, -80)
    data[178][8] = str(-67)  # номер проигравшего финал (-67)
    data[191][6] = str(-75)  # номер проигравшего финал (-75)
    data[164][8] = str(-63)  # номер проигравшего финал (-63)
    data[203][4] = str(-79)  # номер проигравшего финал (-79)

    number_of_game = draw_num(row_n=72, row_step=2, col_n=6, number_of_columns=2, number_of_game=33, player=4, data=data) # рисует номера встреч 1-32
    data[84][8] = str(number_of_game)  # создание номеров встреч 36
    number_of_game = draw_num(row_n=89, row_step=2, col_n=4, number_of_columns=3, number_of_game=37, player=8, data=data) # рисует номера встреч 1-32
    data[106][8] = str(number_of_game)
    number_of_game = draw_num(row_n=114, row_step=2, col_n=6, number_of_columns=2, number_of_game=45, player=4, data=data) # рисует номера встреч 1-32
    data[126][8] = str(number_of_game)  # создание номеров встреч 48
    draw_num_lost(row_n=72, row_step=2, col_n=4, number_of_game=25, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=84, row_step=2, col_n=6, number_of_game=33, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=89, row_step=2, col_n=2, number_of_game=17, player=8, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=106, row_step=2, col_n=6, number_of_game=41, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=114, row_step=2, col_n=4, number_of_game=37, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=126, row_step=2, col_n=6, number_of_game=45, player=2, data=data) # номера минус проигравшие встречи -1 -16
#========== 3-я страница ==============
    number_of_game = draw_num(row_n=140, row_step=2, col_n=2, number_of_columns=4, number_of_game=49, player=16, data=data) # рисует номера встреч 1-32
    data[168][10] = str(number_of_game)  # создание номеров встреч 64
    number_of_game = draw_num(row_n=172, row_step=2, col_n=6, number_of_columns=2, number_of_game=65, player=4, data=data) # рисует номера встреч 1-32
    data[182][10] = str(number_of_game)  # создание номеров встреч 68
    number_of_game = draw_num(row_n=179, row_step=2, col_n=2, number_of_columns=3, number_of_game=69, player=8, data=data) # рисует номера встреч 1-32
    data[194][10] = str(number_of_game)
    number_of_game = draw_num(row_n=197, row_step=2, col_n=2, number_of_columns=2, number_of_game=77, player=4, data=data) # рисует номера встреч 1-32
    data[201][10] = str(number_of_game)  # создание номеров встреч 68
    draw_num_lost(row_n=140, row_step=2, col_n=0, number_of_game=1, player=16, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=172, row_step=2, col_n=4, number_of_game=57, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=179, row_step=2, col_n=0, number_of_game=49, player=8, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=197, row_step=2, col_n=0, number_of_game=69, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=168, row_step=2, col_n=8, number_of_game=61, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=182, row_step=2, col_n=8, number_of_game=65, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=194, row_step=2, col_n=8, number_of_game=73, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=201, row_step=2, col_n=8, number_of_game=77, player=2, data=data) # номера минус проигравшие встречи -1 -16
    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.2 * cm, 3.8 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm,
        2.5 * cm, 0.35 * cm, 3.0 * cm, 0.3 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table) # раскрашивает места участников красным цветом
    t = Table(data, cw, strok * [0.35 * cm])
    # =========  цикл создания стиля таблицы =======
    # ========= 1 страница =========
    style = draw_setka(1, 3, 32, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
   
    for l in range(34, 57, 22):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
        style.append(fn)
    for l in range(62, 68, 5):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 3-4 места
        style.append(fn)
    for l in range(61, 64, 2):
        fn = ('LINEABOVE', (9, l), (10, l), 1, colors.darkblue)  # рисует линии встреч за -29 -30
        style.append(fn)

    fn = ('BOX', (10, 61), (10, 62), 1, colors.darkblue)
    style.append(fn) 
    fn = ('SPAN', (10, 61), (10, 62))  # встреча 32
    style.append(fn)       
    fn = ('BACKGROUND', (10, 61), (10, 62), colors.lightyellow)  # встречи 32 за 3-4 место
    style.append(fn)
# =========== 2 страница ===================
    # ======= встречи (33-35) за 5-6 место =====
    style = draw_setka(5, 72, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=75, col=9, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (36) за 7-8 место =====
    style = draw_setka(7, 84, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=85, col=9, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (37-43) за 9-10 место =====
    style = draw_setka(3, 89, 8, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=96, col=9, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (44) за 11-12 место =====
    style = draw_setka(7, 106, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=107, col=9, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (45-47) за 13-14 место =====
    style = draw_setka(5, 114, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=117, col=9, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (48) за 15-16 место =====
    style = draw_setka(7, 126, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=127, col=9, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
# =========== 3 страница ==================
    # ======= встречи (49-56) за 17-18 место =====
    style = draw_setka(1, 140, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16человека)
    style = draw_mesta(row=155, col=9, player=16, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (64) за 19-20 место =====
    style = draw_setka(9, 168, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=169, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (33-35) за 21-24 место =====
    style = draw_setka(5, 172, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=175, col=8, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (68) за 25-26 место =====
    style = draw_setka(9, 182, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=183, col=11, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (69 - 75) за 25-26 место =====
    style = draw_setka(1, 179, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_mesta(row=186, col=7, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 194, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=195, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(1, 197, 4, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=200, col=5, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 201, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=202, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    
# =========================================
    for i in range(0, 11, 2):
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 206), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i, 0), (i, 206), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i + 1, 0), (i + 1, 206), 'LEFT')
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i, 0), (i, 206), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)
    ts = style   # стиль таблицы (список оформления строк и шрифта)
    for b in style_color:
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 64), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 64), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
                           
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=200, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(fin, h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_32_full_net"
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3.5*cm, bottomMargin=1.0*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_32_2_made(fin):
    """сетка на 32 (-2) с розыгрышем всех мест"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_32_2"
    elements = []
    style = []
    data = []
    column = ['']
    column_count = column * 15
    final = fin
    if sender != my_win.clear_s32_2_Action:
        first_mesto = mesto_in_final(fin)
    else:
        first_mesto = 1
    strok = 207
    for i in range(0, strok):
        # column_count[14] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= нумерация встреч сетки ==========
    y = 0
    for i in range(1, 64, 2):
        y += 1
        data[i + 1][0] = str(y)  # рисует начальные номера таблицы 1-32
    number_of_game = draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=5, number_of_game=1, player=32, data=data) # рисует номера встреч 1-32 
    data[18][10] = str(number_of_game - 1)  # создание номеров встреч (31)
    data[55][10] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-31)
 # ======= 2-я страница ===========
    draw_num_lost(row_n=74, row_step=2, col_n=0, number_of_game=1, player=16, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=102, row_step=2, col_n=10, number_of_game=58, player=2, data=data) # номера минус проигравшие встречи -58-59
    draw_num_lost(row_n=110, row_step=2, col_n=10, number_of_game=56, player=2, data=data) # номера минус проигравшие встречи -56-57
    draw_num_lost(row_n=112, row_step=2, col_n=0, number_of_game=52, player=4, data=data) # номера минус проигравшие встречи -52-55
    draw_num_lost(row_n=124, row_step=2, col_n=0, number_of_game=48, player=4, data=data) # номера минус проигравшие встречи -48-51
    draw_num_lost(row_n=120, row_step=2, col_n=8, number_of_game=63, player=2, data=data) # номера минус проигравшие встречи -63-64
    draw_num_lost(row_n=128, row_step=2, col_n=8, number_of_game=67, player=2, data=data) # номера минус проигравшие встречи -67-68
    draw_num_lost_2(row_n=72, row_step=2, col_n=2, revers_number=1, number_of_game=17, player=8, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost_2(row_n=71, row_step=4, col_n=6, revers_number=1, number_of_game=25, player=2, data=data) # номера минус проигравшие встречи -25-26
    draw_num_lost_2(row_n=87, row_step=4, col_n=6, revers_number=1, number_of_game=27, player=2, data=data) # номера минус проигравшие встречи -27-28
    draw_num_lost_2(row_n=71, row_step=8, col_n=10, revers_number=1, number_of_game=29, player=2, data=data) # номера минус проигравшие встречи -29-30

    number_of_game = draw_num_2(row_n=74, row_step=2, col_n=2, number_of_columns=2, number_of_game=32, player=16, data=data) # рисует номера встреч 33-47 
    number_of_game = draw_num_2(row_n=74, row_step=4, col_n=6, number_of_columns=2, number_of_game=48, player=16, data=data) # рисует номера встреч 48-55
    number_of_game = draw_num_2(row_n=74, row_step=8, col_n=10, number_of_columns=1, number_of_game=56, player=16, data=data) # рисует номера встреч 56-57
    number_of_game = draw_num_2(row_n=72, row_step=8, col_n=12, number_of_columns=1, number_of_game=58, player=16, data=data) # рисует номера встреч 58-59
    number_of_game = draw_num(row_n=112, row_step=2, col_n=2, number_of_columns=2, number_of_game=63, player=4, data=data) # рисует номера встреч 63-65
    number_of_game = draw_num(row_n=124, row_step=2, col_n=2, number_of_columns=2, number_of_game=67, player=4, data=data) # рисует номера встреч 67-69

    data[75][14] = str(number_of_game - 10)  # создание номеров встреч (60)
    data[98][12] = str((number_of_game - 10) * -1)  # номер проигравшего финал (-60)
    data[102][12] = str(number_of_game - 9)  # создание номеров встреч (61)
    data[106][12] = str((number_of_game - 9) * -1)  # номер проигравшего финал (-61)
    data[110][12] = str(number_of_game - 8)  # создание номеров встреч (62)
    data[114][12] = str((number_of_game - 8) * -1)  # номер проигравшего финал (-62)
    data[120][10] = str(number_of_game - 4)  # создание номеров встреч (66)
    data[124][10] = str((number_of_game - 4) * -1)  # номер проигравшего финал (-66)
    data[128][10] = str(number_of_game)  # создание номеров встреч (70)
    data[132][10] = str((number_of_game) * -1)  # номер проигравшего финал (-70)
    data[118][4] = str((number_of_game - 5) * -1)  # номер проигравшего финал (-65)
    data[130][4] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-69)
# ======= 3-я страница ===========
    draw_num_lost(row_n=141, row_step=2, col_n=0, number_of_game=40, player=8, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=156, row_step=2, col_n=8, number_of_game=75, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=171, row_step=2, col_n=8, number_of_game=79, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=186, row_step=2, col_n=8, number_of_game=87, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=201, row_step=2, col_n=8, number_of_game=91, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=160, row_step=2, col_n=2, number_of_game=71, player=4, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=171, row_step=2, col_n=0, number_of_game=32, player=8, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=192, row_step=2, col_n=2, number_of_game=83, player=4, data=data) # номера минус проигравшие встречи

    number_of_game = draw_num(row_n=141, row_step=2, col_n=2, number_of_columns=3, number_of_game=71, player=8, data=data) # рисует номера встреч 49
    data[153][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[156][10] = str(number_of_game)  # создание номеров встреч 78
    data[160][10] = str(number_of_game * -1)  # создание номеров встреч -78
    number_of_game = draw_num(row_n=160, row_step=2, col_n=4, number_of_columns=2, number_of_game=79, player=4, data=data) # рисует номера встреч 49
    data[166][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[171][10] = str(number_of_game)  # создание номеров встреч 68
    data[175][10] = str(number_of_game * -1)  # создание номеров встреч 68
    number_of_game = draw_num(row_n=171, row_step=2, col_n=2, number_of_columns=3, number_of_game=83, player=8, data=data) # рисует номера встреч 49
    data[183][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[186][10] = str(number_of_game)  # создание номеров встреч 90
    data[190][10] = str(number_of_game * -1)  # создание номеров встреч -90
    number_of_game = draw_num(row_n=192, row_step=2, col_n=4, number_of_columns=2, number_of_game=91, player=4, data=data) # рисует номера встреч 49
    data[198][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[201][10] = str(number_of_game)  # создание номеров встреч 94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94

    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94

    # ============= данные игроков и встреч и размещение по сетке =============
    # ======= создать словарь  ключ - номер встречи, значение - номер ряда
    dict_num_game = {}
    for d in range(2, 15, 2):
        for r in range(0, 69):
            key = data[r][d]
            if key != "":
                dict_num_game[key] = r
    # ===== добавить данные игроков и счета в data ==================
    tds = write_in_setka(data, fin, first_mesto, table)
    # ==============
    cw = ((0.2 * cm, 3.5 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm,
        2.4 * cm, 0.35 * cm, 2.6 * cm, 0.35 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table) # раскрашивает места участников красным цветом
    t = Table(data, cw, strok * [0.35 * cm])
    # =========  цикл создания стиля таблицы =======
    # ========= 1 страница =========
    style = draw_setka(1, 3, 32, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
   
    for l in range(34, 57, 22):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
        style.append(fn)
# =========== 2 страница ===================
    # # ======= встречи (33-35) за 3-4 место =====
    style = draw_setka_2(1, 74, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    for k in range(0, 7, 6):
        for l in range(72 + k, 89 + k, 16):
            fn = ('LINEABOVE', (11, l), (12, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
            style.append(fn)   
    for k in range(0, 17, 16):
        fn = ('BOX', (12, 72 + k), (12, 77 + k), 1, colors.darkblue)
        style.append(fn) 
        fn = ('SPAN', (12, 72 + k), (12, 77 + k))  # встреча 32
        style.append(fn)       
        fn = ('BACKGROUND', (12, 72 + k), (12, 77 + k), colors.lightyellow)  # встречи 32 за 3-4 место
        style.append(fn) 
           
    for l in range(75, 101, 8):
        fn = ('LINEABOVE', (13, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 3-4 места
        style.append(fn)
    fn = ('BOX', (14, 75), (14, 90), 1, colors.darkblue)
    style.append(fn) 
    fn = ('SPAN', (14, 75), (14, 90))  # встреча 32
    style.append(fn)       
    fn = ('BACKGROUND', (14, 75), (14, 90), colors.lightyellow)  # встречи 32 за 3-4 место
    style.append(fn)  
    # # ======= встречи (61) за 5-6 место =====
    style = draw_setka(11, 102, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=103, col=13, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (36) за 7-8 место =====
    style = draw_setka(11, 110, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=111, col=13, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (37-43) за 9-10 место =====
    style = draw_setka(1, 112, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=115, col=5, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (44) за 11-12 место =====
    style = draw_setka(9, 120, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=121, col=11, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (37-43) за 13-14 место =====
    style = draw_setka(1, 124, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=127, col=5, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (44) за 15-16 место =====
    style = draw_setka(9, 128, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=129, col=11, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
# =========== 3 страница ==================
    # ======= встречи (49-56) за 17-18 место =====
    style = draw_setka(1, 141, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_mesta(row=148, col=7, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
   # ======= встречи (64) за 19-20 место =====
    style = draw_setka(9, 156, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=157, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (33-35) за 21-24 место =====
    style = draw_setka(3, 160, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=163, col=6, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (68) за 25-26 место =====
    style = draw_setka(9, 171, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=172, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (69 - 75) за 25-26 место =====
    style = draw_setka(1, 171, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_mesta(row=178, col=7, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 186, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=187, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(3, 192, 4, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=195, col=6, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 201, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=202, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
   
# =========================================
    for i in range(0, 15, 2):
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 206), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i, 0), (i, 206), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i + 1, 0), (i + 1, 206), 'LEFT')
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i, 0), (i, 206), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)
    ts = style   # стиль таблицы (список оформления строк и шрифта)
    for b in style_color:
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 64), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 64), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=200, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(fin, h2))
# ====                       
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_32_2_net"
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3.4*cm, bottomMargin=1.0*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def mesto_in_final(fin):
    """с какого номера расставляются места в финале, в зависимости от его номера и кол-во участников fin - финал"""
    final = []
    mesto = {}

    system = System.select().where(System.title_id == title_id()) # находит system id последнего
    system_id = system.select().where(System.stage == "1-й финал").get()
    id_system = system_id.id
    count = len(system)
    k = 0
    if fin == "Одна таблица" or fin == "1-й финал":
       mesto[fin] = 1 
    else:
        for k in range(id_system, id_system + count):
            sys = system.select().where(System.id == k).get()
            max_player = sys.max_player
            stage = sys.stage
            if stage == fin:
                final.append(max_player)
                break
        mesto[fin] = sum(final) + 1
    first_mesto = mesto[fin]

    return first_mesto


def write_in_setka(data, fin, first_mesto, table):
    """функция заполнения сетки результатами встреч data поступает чистая только номера в сетке, дальше идет заполнение игроками и счетом"""
    "row_num_win - словарь, ключ - номер игры, значение - список(номер строки 1-ого игрока, номер строки 2-ого игрока) и записвает итоговые места в db"
    sender = my_win.sender()
    player = Player.select().where(Player.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    row_num_los = {}
    row_end = 0  # кол-во строк для начальной расстоновки игроков в зависимости от таблицы
    flag_clear = False
    # уточнить кол-во столбцов
    if table == "setka_8_full":
        row_last = 33
        column_last = 8
        row_end = 15
        row_num_win = {1: [1], 2: [5], 3: [9], 4: [13], 5: [3], 6: [11], 7: [7], 8: [16], 9: [20], 10: [24],
                        11: [22], 12: [28]}
                 # ======= list mest
        mesta_dict = {7: 7, 8: 16, 11: 22, 12: 28}
    elif table == "setka_8_2":
        # это вариант при сетке минус 2
        # если встреча верху четная, то на встречу куда идет победитель ( список наоборот) 12: [20, 16]
        row_last = 39
        column_last = 9
        row_end = 15
        row_num_win = {1: [1], 2: [5], 3: [9], 4: [13], 5: [3], 6: [11], 7: [7], 8: [17], 9: [21],
                        10: [16], 11: [20], 12: [18], 13: [25], 14: [31]}
                 # ======= list mest
        mesta_dict = {7: 7, 12: 18, 13: 25, 14: 31}
    elif table == "setka_16_full":
        row_last = 69
        column_last = 11
        row_end = 31
        row_num_win = {1: [1], 2: [5], 3: [9], 4: [13], 5: [17], 6: [21], 7: [25], 8: [29], 9: [3], 10: [11], 11: [19], 12: [27], 13: [7], 14: [23], 
                       15: [15], 16: [29], 17: [32], 18: [36], 19: [34], 20: [39], 21: [41], 22: [45], 23: [49], 24: [53], 25: [43], 26: [51], 27: [47],
                       28: [55], 29: [58], 30: [62], 31: [60], 32: [65]}
                 # ======= list mest
        mesta_dict = {15: 15, 16: 29, 19: 34, 20: 39, 27: 47, 28: 55, 31: 60, 32: 65}
    elif table == "setka_16_2": # встречи, где играют победители и проигравший из основного тура  например 22: [54, 54] в списке одинаковые строки
        row_last = 85
        column_last = 11
        row_end = 33
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [5], 10: [13], 11: [21], 12: [29], 13: [9], 14: [25], 15: [17], 
                       16: [46], 17: [50], 18: [54], 19: [58], 20: [46], 21: [50], 22: [54], 23: [58], 24: [47], 25: [55], 26: [45], 27: [53], 28: [49], 29: [61],
                        30: [67],  31: [62], 32: [66], 33: [64], 34: [73], 35: [74], 36: [78], 37: [76], 38: [79]} 
                 # ======= list mest
        mesta_dict = {15: 17, 28: 49, 29: 61, 33: 64, 30: 67, 34: 73, 37: 76, 38: 79} # номер встречи - номер строки
    elif table == "setka_32":
        row_last = 69
        column_last = 11
        row_end = 65
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [35], 10: [39], 11: [43], 12: [47],
        13: [51], 14: [55], 15: [59], 16: [63], 17: [5], 18: [13], 19: [21], 20: [29], 21: [37], 22: [45], 23: [53], 24: [61],
        25: [9], 26: [25], 27: [41], 28: [57], 29: [17], 30:[49], 31: [33], 32: [61]}
        mesta_dict = {31: 33, 32: 61}
    elif table == "setka_32_2":
        # встреч, которые попадают на сноски (в сетке за 3 место) должно быть в row_num_win а список состоит из одного номера встречи куда идет победитель
        row_last = 207
        column_last = 15
        row_end = 65
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [35], 10: [39], 11: [43], 12: [47],
        13: [51], 14: [55], 15: [59], 16: [63], 17: [5], 18: [13], 19: [21], 20: [29], 21: [37], 22: [45], 23: [53], 24: [61],
        25: [9], 26: [25], 27: [41], 28: [57], 29: [17], 30:[49], 31: [33], 32: [74], 33: [78], 34: [82], 35: [86], 36: [90],
        37: [94], 38: [98], 39: [102], 40: [73], 41: [77], 42: [81], 43: [85], 44:[89], 45: [93], 46: [97], 47: [101], 
        48: [75], 49: [83], 50: [91],  51: [99], 52: [73], 53: [81], 54: [89], 55: [97], 56: [77], 57: [93], 58: [74], 59: [90], 
        60: [81], 63: [112], 64: [116],  67: [124], 68: [128], 71: [141], 72: [145], 73: [149], 74: [153], 75: [143], 76: [151], 
        79: [160], 80: [164], 83: [171], 84: [175], 85: [179], 86: [183], 87: [173], 88: [181], 91: [192], 92: [196]}
                 # ======= dict mest
        mesta_dict = {31: 33, 60: 82, 61: 102, 62: 110, 65: 114, 66: 120, 69: 126, 70: 128, 77: 147,
                        78: 156, 81: 162, 82: 171, 89: 177, 90: 186, 93: 194, 94: 201}
    elif table == "setka_32_full":
        row_last = 207
        column_last = 11
        row_first = 0
        row_end = 65
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [35], 10: [39], 11: [43], 12: [47],
        13: [51], 14: [55], 15: [59], 16: [63], 17: [5], 18: [13], 19: [21], 20: [29], 21: [37], 22: [45], 23: [53], 24: [61],
        25: [9], 26: [25], 27: [41], 28: [57], 29: [17], 30:[49], 31: [33], 32: [61], 33: [72], 34: [76], 35: [74], 36: [84], 37: [89],
        38: [93], 39: [97], 40: [101], 41: [91], 42: [99], 43: [95], 44: [106], 45: [114], 46: [118], 47: [116], 48: [126],  49: [140],
        50: [144], 51: [148], 52: [152], 53: [156], 54: [158], 55: [162], 56: [166], 57: [142], 58: [150], 59: [158], 60: [166], 61: [146],
        62: [162], 63: [154], 64: [168], 65: [172], 66: [176], 67: [174], 68: [182], 69: [179], 70: [183], 71: [187], 72: [191], 73: [181],
        74: [189], 75: [185], 76: [194], 77: [197], 78: [201], 79: [199]}
                 # ======= dict mest
        mesta_dict = {31: 33, 32: 61, 35: 74, 36: 84, 43: 95, 44: 106, 47: 116, 48: 126, 63: 154,
                        64: 168, 67: 174, 68: 182, 75: 185, 76: 194, 79: 199, 80: 201}
    
    if sender == my_win.clear_s32_Action or sender == my_win.clear_s32_full_Action or sender == my_win.clear_s32_2_Action:
        all_list = setka_data_clear(fin, table)  # печать чистой сетки
        col_first = 0
        row_first = 2
        flag_clear = True
    elif sender == my_win.clear_s16_Action:
        all_list = setka_data_clear(fin, table)  # печать чистой сетки
        col_first = 2
        row_first = 0
        flag_clear = True
    elif sender == my_win.clear_s16_2_Action:
        all_list = setka_data_clear(fin, table)  # печать чистой сетки
        col_first = 2
        row_first = 2
        flag_clear = True
    elif sender == my_win.clear_s8_full_Action:
        all_list = setka_data_clear(fin, table)  # печать чистой сетки
        col_first = 2
        row_first = 2
        flag_clear = True
    elif sender == my_win.clear_s8_2_Action:
        all_list = setka_data_clear(fin, table)  # печать чистой сетки
        col_first = 2
        row_first = 2
        flag_clear = True
    else:
        sys = System.select().where(System.title_id == title_id())
        system = sys.select().where(System.stage == fin).get()
        setka_string = system.label_string
        if setka_string == "Сетка (с розыгрышем всех мест) на 8 участников":
            col_first = 0
            row_first = 0
        elif setka_string == "Сетка (-2) на 8 участников":
            col_first = 0
            row_first = 0
        elif setka_string == "Сетка (с розыгрышем всех мест) на 16 участников":
            col_first = 2
            row_first = 0
        elif setka_string == "Сетка (-2) на 16 участников":
            col_first = 0
            row_first = 2
        elif setka_string == "Сетка (с розыгрышем всех мест) на 32 участников":
            col_first = 0
            row_first = 2
        elif setka_string == "Сетка (-2) на 32 участников":
            col_first = 0
            row_first = 2
        posev_data = setka_player_after_choice(fin) # игроки 1-ого посева
        all_list = setka_data(fin, posev_data)
        id_sh_name = all_list[2] # словарь {Фамилия Имя: id}
    tds = []
    tds.append(all_list[0]) # список фамилия/ город 1-ого посева
    # ======
    if flag_clear is False:
        tds.append(id_sh_name)
 
    for d in range(col_first, column_last, 2):
        for r in range(row_first, row_last):
            key = data[r][d]
            if key != "":
                k = int(key)
            if key != "" and k < 0:
                row_num_los[key] = r # словарь номер игры, сноски - номер строки

    n = 0
    for t in range(row_first, row_end, 2):  # цикл расстановки игроков по своим номерам в 1-ом посеве (фамилия инциалы имени/ город)
        data[t][1] = tds[0][n]
        n += 1
    # ==============
    if flag_clear is False:
        # функция расстановки счетов и сносок игроков
        dict_setka = score_in_setka(fin)
        key_list = []
        mesta_list = []
        for k in dict_setka.keys():
            key_list.append(k)
        for v in mesta_dict.keys():
            mesta_list.append(v)
        # ============
        for i in key_list: # спиисок встреч которые сыграны
            match = dict_setka[i]
            pl_win = match[1]
            pl_los = match[4]
            if pl_win != "X":
                id_win = id_sh_name[pl_win]
            if pl_los != "X":
                id_los = id_sh_name[pl_los]
            else:
                id_los = ""
            r = str(match[3]) # сноска проигравшего
            # ===== определение итоговых мест и запись в db
            if i in mesta_list: # i - номер данной встречи
                index = mesta_list.index(i)
                mesto = first_mesto + (index * 2)
                # записывает места в таблицу -Player-
                player = Player.get(Player.id == id_win)
                win = f"{player.player}/{player.city}"
                player.mesto = mesto
                player.save()
                m = 0
                for n in [id_win, id_los]: # записывает место в сетке в таблицу -choice-
                    choice_pl = Choice.get(Choice.player_choice_id == n)
                    choice_pl.mesto_final = mesto + m
                    choice_pl.save()
                    m += 1
                if id_los != "":
                    player = Player.get(Player.id == id_los)
                    los = f"{player.player}/{player.city}"
                    player.mesto = mesto + 1
                    player.save()
                else:
                    los = "X"
            c = match[0] # номер встречи, куда попадают победитель данной встречи (i)
            # ========== расстановка для сетки на 16
            if c != 0: #  номер встречи в сетке куда попадает победиель (кроме встреч за места)
                row_win_list = row_num_win[i]  # номера строк данной встречи в сетке
                row_win = row_win_list[0]
                c1 = []
                c1_tmp = []
                c = str(c)
                win = match[1]
                los = match[4]
            elif c == 0:  # встречи за места
                row_win = mesta_dict[i]
                c = str(i)
           # цикл создания списков номеров встреч по столбцам новый
            column_dict = {}
            for cd in range(2, column_last, 2):
                c1_tmp.clear()
                for rd in range(0, row_last):
                    d1 = data[rd][cd]
                    if d1 != "" and type(d1) == str and int(d1) > 0:
                        c1_tmp.append(d1)
                        c1 = c1_tmp.copy()
                column_dict[cd] = c1    # ключ -номер столбца, значение - список номеров встреч   
            # =======
            for k in column_dict.keys():
                num_game_list = column_dict[k]  
                if str(i) in num_game_list:
                    if k == column_last - 1:
                        col_win = k - 1
                    else:
                        col_win = k + 1
                    break
            row_los = row_num_los[r]  # строка проигравшего
            score = match[2]  # счет во встречи
            row_list_los = data[row_los]  # получаем список строки, где ищет номер куда сносится проигравший
            col_los = row_list_los.index(r) # номер столбца проигравшего            
            data[row_win][col_win] = win
            data[row_win + 1][col_win] = score
            data[row_los][col_los + 1] = los
        return tds


def setka_data_clear(fin, table):
    """заполняет сетку для просмотра пустыми фамилиями"""
    all_list = []
    tmp = [""]
    if table == "setka_8_full" or table == "setka_8_2":
        max_pl = 8
    elif table == "setka_16_full" or table == "setka_16_2":
        max_pl = 16
    elif table == "setka_32" or table == "setka_32_full" or table == "setka_32_2":
        max_pl = 32
    tds = tmp * max_pl
    all_list.append(tds)
    return all_list
    

def kol_player(stage):
    """выводит максимальное количество человек в группе t если все группы равны, а g2 если разное количество"""
    system = System.select().where((System.title_id == title_id()) & (System.stage == stage)).get()
    if stage == "Предварительный":
        all_players = system.total_athletes
        all_group = system.total_group
        flat = all_players % all_group  # если количество участников равно делится на группы
    # если количество участников не равно делится на группы, g2 наибольшое кол-во человек в группе
        player_flat = all_players // all_group
        if flat == 0:
            max_gamer = player_flat
        else:
            max_gamer = player_flat + 1
    else:
        max_gamer = system.max_player // system.total_group
    return max_gamer


def  table_data(stage, kg):
    """циклом создаем список участников каждой группы или финалов если играют по кругу"""
    tdt_all = []  # список списков [tdt_new] и [tdt_color]
    tdt_color = []
    tdt_new = []
    result = Result.select().where(Result.title_id == title_id())  # находит system id последнего
    system = System.select().where(System.title_id == title_id())
    system_id = system.select().where(System.stage == stage).get()
    id_system = system_id.id
    if kg == 1:  # система одна таблица круг или финалу по кругу
        # список словарей участник и его регион
        result_fin = result.select().where(Result.system_id == id_system)
        tr = len(result_fin)  # общее кол-во игр в финалах или одной таблице
        posev_data = player_choice_one_table(stage) # posev_data (фамилия/ id)
        count_player_group = len(posev_data)
        max_gamer = count_player_group
        num_gr = stage
        tdt_tmp = tdt_news(max_gamer, posev_data, count_player_group, tr, num_gr)
        tdt_new.append(tdt_tmp[0])
        tdt_color.append(tdt_tmp[1])
        tdt_all.append(tdt_new)
        tdt_all.append(tdt_color)
    else:
        max_gamer = kol_player(stage)
        result_stage = result.select().where(Result.system_id == id_system)
        tr = len(result_stage)  # общее кол-во игр в группах
        for p in range(0, kg):
            num_gr = f"{p + 1} группа"
            if stage == "Предварительный":
                posev_data = player_choice_in_group(num_gr) # словарь фамилия:игрок/id регион: область
            else:
                posev_data = player_choice_semifinal(stage, num_gr)
            count_player_group = len(posev_data)
            tdt_tmp = tdt_news(max_gamer, posev_data, count_player_group, tr, num_gr)
            tdt_new.append(tdt_tmp[0])
            tdt_color.append(tdt_tmp[1])
            tdt_all.append(tdt_new)
            tdt_all.append(tdt_color)
    return tdt_all


def tdt_news(max_gamer, posev_data, count_player_group, tr, num_gr):
    tdt_tmp = []
    tbl_tmp = []  # временный список группы tbl
    # цикл нумерации строк (по 2-е строки на каждого участника)
    for k in range(1, max_gamer * 2 + 1):
        st = ['']
        # получаем пустой список (номер, фамилия и регион, клетки (кол-во уч), оч, соот, место)
        s = (st * (max_gamer + 4))
        s.insert(0, str((k + 1) // 2))  # получаем нумерацию строк по порядку
        tbl_tmp.append(s)
    for i in range(1, count_player_group * 2 + 1, 2):
        posev = posev_data[((i + 1) // 2) - 1]
        fam_id = posev["фамилия"]
        znak = fam_id.find("/")
        if znak != -1:
            tbl_tmp[i - 1][1] = fam_id[:znak]
        else:
            tbl_tmp[i - 1][1] = posev["фамилия"]
        tbl_tmp[i][1] = posev["регион"]
 
    td = tbl_tmp.copy()  # cписок (номер, фамилия, город и пустые ячейки очков)
    td_color = []

    if tr != 0:  # если еще не была жеребьевка, то пропуск счета в группе
        # список очки победителя красные (ряд, столбец) без заголовка
        td_color = score_in_table(td, num_gr)

    tdt_new = td
    tdt_tmp.append(tdt_new)
    tdt_tmp.append(td_color)

    return tdt_tmp


def setka_player_after_choice(fin):
    """список игроков сетки после жеребьевки"""
    p_data = {}
    posev_data = []
    player = Player.select().where(Player.title_id == title_id())
    game_list = Game_list.select().where(Game_list.title_id == title_id())
    pl_list = game_list.select().where(Game_list.number_group == fin)
    for i in pl_list:
        p_data['посев'] = i.rank_num_player
        txt = i.player_group_id
        if txt != "X":
            line = txt.find("/")  # находит черту отделяющий имя от города
            id_pl = int(txt[line + 1:])
            pl = player.select().where(Player.id == id_pl).get()
            p_data['фамилия'] = pl.full_name
        else:
            p_data['фамилия'] = "X"
        tmp = p_data.copy()
        posev_data.append(tmp)
        p_data.clear()
    return posev_data


def setka_data(fin, posev_data):
    """данные сетки"""
    id_ful_name = {}
    id_name = {}
    tds = []
    fam_name_city = []
    all_list = []

    system = System.select().where((System.title_id == title_id()) & (System.stage == fin)).get()  # находит system id последнего

    mp = system.max_player
    for i in range(1, mp * 2 + 1, 2):
        posev = posev_data[((i + 1) // 2) - 1]
        family = posev['фамилия'] # фамилия имя / город
        # id_f_name = full_player_id(family) # словарь {name: фамилия/город, id: номер игрока}, {name: фамилия, id: номер мгрока}
        name_list = full_player_id(family) # словарь {name: фамилия/город, id: номер игрока}, {name: фамилия, id: номер мгрока}
        id_f_n = name_list[0] # словарь name: фамилия/город, id: номер игрока
        id_s_n = name_list[1] # {name: фамилия, id: номер игрока}
            # словарь ключ - полное имя/ город, значение - id
        id_ful_name[id_f_n["name"]] = id_f_n["id"]
        id_name[id_s_n["name"]] = id_s_n["id"]
            # =================
        if family != "X":
            # находит пробел отделяющий имя от фамилии
            space = family.find(" ")
            line = family.find("/")  # находит черту отделяющий имя от города
            city_slice = family[line:]  # получает отдельно город
            # получает отдельно фамилия и первую букву имени
            family_slice = family[:space + 2]
            family_city = f'{family_slice}.{city_slice}'   # все это соединяет
            tds.append(family_city)
            fam_name_city.append(family)
        else:
            tds.append(family)
            fam_name_city.append(family)
    all_list.append(tds)
    all_list.append(id_ful_name)
    all_list.append(id_name)
    all_list.append(fam_name_city)
    return all_list


def full_player_id(family):
    """получает словарь -фамилия игрока и его город и соответствующий ему id в таблице Players"""
    full_name = {}
    short_name = {}
    
    player = Player.select().where(Player.title_id == title_id())
    pl_id = player.select().where(Player.full_name == family).get()
    player_id = pl_id.id # ид игрока
    if family != "X":
        # space_mark = family.find("/")  # находит косую черту отделяющий город и игрока
        # player_in_net = family[:space_mark]
        # plr = player.select().where(Player.player == player_in_net).get()

        # id_player = plr.id
        # city = plr.city
        # name = plr.player
        f_name = pl_id.full_name
        s_name = pl_id.player
        full_name["name"] = f_name
        full_name["id"] = player_id 
        short_name["name"] = s_name
        short_name["id"] = player_id
        # ====  вариант фамилия и имя
        # space = name.find(" ")  # находит пробел отделяющий имя от фамилии
        # full_name["name"] = f"{name}/{city}"
        # full_name["id"] = id_player
        # short_name["name"] = f"{name}"
        # short_name["id"] = id_player
        # =======
    else:
        full_name["name"] = "X"
        full_name["id"] = 0
        short_name["name"] = "X"
        short_name["id"] = 0
    name_list = []
    name_list.append(full_name)
    name_list.append(short_name)

    return name_list


def score_in_table(td, num_gr):
    """заносит счет и места в таблицу группы или таблицу по кругу pdf
    -td- список строки таблицы, куда пишут счет"""
    td_color = []
    total_score = {}  # словарь, где ключ - номер участника группы, а значение - очки
    tab = my_win.tabWidget.currentIndex()
    system = System.select().where(System.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    gamelist = Game_list.select().where(Game_list.title_id == title_id())
    if tab == 7: # открыта вкладка для редактирования групп
        stage = my_win.comboBox_edit_etap1.currentText()
        system_id = system.select().where(System.stage == stage).get()
        mp = len(gamelist.select().where((Game_list.system_id == system_id) & (Game_list.number_group == num_gr)))
        r = result.select().where((Result.system_stage == stage) & (Result.number_group == num_gr))
        ch = choice.select().where((Choice.semi_final == stage) & (Choice.sf_group == num_gr))  # фильтрует по группе
    elif tab == 3:
        ta = system.select().where(System.stage == "Предварительный").get()  # находит system id последнего
        r = result.select().where((Result.system_stage == "Предварительный") & (Result.number_group == num_gr))
        ch = choice.select().where(Choice.group == num_gr)  # фильтрует по группе
        mp = ta.max_player
        stage = ta.stage
    elif tab == 4:
        ta = system.select().where((System.stage == "1-й полуфинал") | (System.stage == "2-й полуфинал")).get()  # находит system id последнего
        stage = ta.stage
        mp = len(gamelist.select().where((Game_list.system_id == ta) & (Game_list.number_group == num_gr)))
        r = result.select().where((Result.system_stage == stage) & (Result.number_group == num_gr))
        ch = choice.select().where((Choice.semi_final == stage) & (Choice.sf_group == num_gr))  # фильтрует по группе
    elif tab == 5:
        system_id = system.select().where(System.stage == num_gr).get()
        id_system = system_id.id
        r = result.select().where(Result.system_id == id_system)
        if num_gr == "Одна таблица":
            stage = "Одна таблица"
            ch = choice.select().where(Choice.basic == "Одна таблица")  # фильтрует по одной таблице
        else:
            stage = "Финальный"
            ch = choice.select().where(Choice.final == num_gr)
        mp = len(gamelist.select().where(Game_list.system_id == id_system))

    count = len(r)  # сколько игр в группе
    count_player = len(ch)  # определяет сколько игроков в группе
    result_list = r.dicts().execute()
    for s in range(1, count_player + 1):
        total_score[s] = 0
    for i in range(0, count):
        sc_game = str(list(result_list[i].values())[9])  # счет в партиях
        if sc_game != "" or sc_game != "None":
            scg = 9
        else:  # номер столбца
            scg = 8
        tours = str(list(result_list[i].values())[3])  # номера игроков в туре
        znak = tours.find("-")
        p1 = int(tours[:znak])  # игрок под номером в группе
        p2 = int(tours[znak + 1:])  # игрок под номером в группе

        win = str(list(result_list[i].values())[6])
        player1 = str(list(result_list[i].values())[4])
        # ==== убираю город из фамилии, чтоб сравнивать игроков 
        # znak_player1 = player1.find("/") # если игрок с городом, то удаляет название города
        # if znak_player1 != -1:
        #     player1 = player1[:znak_player1]
        # ==============
        if win != "" and win != "None":  # если нет сыгранной встречи данного тура
            if win == player1:  # если победитель игрок под первым номером в туре
                # очки 1-ого игрока
                td[p1 * 2 - 2][p2 + 1] = str(list(result_list[i].values())[7])  # ячейка в таблице очки
                # счет 1-ого игрока
                td[p1 * 2 - 1][p2 + 1] = str(list(result_list[i].values())[scg])   # ячейка в таблице счет впартии
                # очки 2-ого игрока
                td[p2 * 2 - 2][p1 + 1] = str(list(result_list[i].values())[11])  # ячейка в таблице очки
                # счет 2-ого игрока
                td[p2 * 2 - 1][p1 + 1] = str(list(result_list[i].values())[12])  # ячейка в таблице счет впартии
                # очки 1-ого игрока
                tp1 = str(list(result_list[i].values())[7])
                # очки 2-ого игрока
                tp2 = str(list(result_list[i].values())[11])
                tp1 = 0 if tp1 == "" else str(list(result_list[i].values())[7])
                tp2 = 0 if tp2 == "" else str(list(result_list[i].values())[11])
                # считывает из словаря 1-ого игрока всего очков
                plr1 = total_score[p1]
                # считывает из словаря 2-ого игрока всего очков
                plr2 = total_score[p2]
                plr1 = plr1 + int(tp1)  # прибавляет очки 1-ого игрока
                plr2 = plr2 + int(tp2)  # прибавляет очки 2-ого игрока
                total_score[p1] = plr1  # записывает сумму очков 1-му игроку
                total_score[p2] = plr2  # записывает сумму очков 2-му игроку
                col = p1 * 2 - 2
                row = p2 + 1
            else:  # если победитель игрок под вторым номером в туре
                # очки 1-ого игрока
                td[p1 * 2 - 2][p2 + 1] = str(list(result_list[i].values())[11])
                # счет 1-ого игрока
                td[p1 * 2 - 1][p2 + 1] = str(list(result_list[i].values())[12])
                # очки 2-ого игрока
                td[p2 * 2 - 2][p1 + 1] = str(list(result_list[i].values())[7])
                # счет 2-ого игрока
                td[p2 * 2 - 1][p1 + 1] = str(list(result_list[i].values())[scg])
                # очки 1-ого игрока
                tp1 = str(list(result_list[i].values())[11])
                # очки 2-ого игрока
                tp2 = str(list(result_list[i].values())[7])
                tp1 = 0 if tp1 == "" else str(list(result_list[i].values())[11])
                tp2 = 0 if tp2 == "" else str(list(result_list[i].values())[7])
                # считывает из словаря 1-ого игрока очки
                plr1 = total_score[p1]
                # считывает из словаря 2-ого игрока очки
                plr2 = total_score[p2]
                plr1 = plr1 + int(tp1)  # прибавляет очки 1-ого игрока
                plr2 = plr2 + int(tp2)  # прибавляет очки 2-ого игрока
                total_score[p1] = plr1  # записывает сумму очков 1-му игроку
                total_score[p2] = plr2  # записывает сумму очков 2-му игроку
                col = p2 * 2 - 2
                row = p1 + 1
            # список ряд столбец, где очки надо красить в красный
            td_tmp = [row, col]
            td_color.append(td_tmp)
    for t in range(0, count_player):  # записывает очки в зависимости от кол-во игроков в группе
        # записывает каждому игроку сумму очков
        td[t * 2][mp + 2] = total_score[t + 1]
    # ===== если сыграны все игры группе то выставляет места =========
    count_game = (count_player * (count_player - 1)) // 2
    if num_gr == "Одна таблица":
        results = result.select().where(Result.system_stage == num_gr)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        results = r.select().where(Result.number_group == num_gr)
    else:
        results = result.select().where((Result.system_stage == stage) & (Result.number_group == num_gr))

    results_playing = results.select().where(Result.points_win == 2)
    a = len(results_playing) # кол-во сыгранных игр

    if a == count_game:
        rank_in_group(total_score, td, num_gr, stage)  # определяет места в группе
    return td_color


def numer_game(num_game, vid_setki):
    """определяет куда записывать победителя и проигравшего по сноске в сетке, номера встреч"""
    snoska = []
    num_game = int(num_game)
    if vid_setki == 'Сетка (с розыгрышем всех мест) на 8 участников':
        dict_winner = {1:5, 2:5, 3:6, 4:6, 5:7, 6:7, 9:11, 10:11}
        dict_loser = {1:9, 2:9, 3:10, 4:10, 5:8, 6:8, 9:12, 10:12}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12}
        dict_mesta = [7, 8, 11, 12]
    elif vid_setki == 'Сетка (-2) на 8 участников':
        dict_winner = {1:5, 2:5, 3:6, 4:6, 5:7, 6:7, 8:10, 9:11, 10:12, 11:12}
        dict_loser = {1:8, 2:8, 3:9, 4:9, 5:11, 6:10, 10:13, 11:13, 8:14, 9:14}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                      14: -14}
        dict_mesta = [7, 12, 13, 14]
    elif vid_setki == 'Сетка (с розыгрышем всех мест) на 16 участников':
        dict_winner = {1: 9, 2: 9, 3: 10, 4: 10, 5: 11, 6: 11, 7: 12, 8: 12, 9: 13, 10: 13, 11: 14, 12: 14, 13: 15, 14: 15,
                   17: 19, 18: 19, 21: 25, 22: 25, 23: 26, 24: 26, 25: 27, 26: 27, 29: 31, 30: 31}
        dict_loser = {1: 21, 2: 21, 3: 22, 4: 22, 5: 23, 6: 23, 7: 24, 8: 24, 9: 17, 10: 17, 11: 18, 12: 18, 13: 16, 14: 16,
                  17: 20, 18: 20, 21: 29, 22: 29, 23: 30, 24: 30, 25: 28, 26: 28, 29: 32, 30: 32}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                      14: -14, 17: -17, 18: -18, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 29: -29, 30: -30}
        dict_mesta = [15, 16, 19, 20, 27, 28, 31, 32]
    elif vid_setki == 'Сетка (-2) на 16 участников':
        dict_winner = {1:9, 2:9, 3:10, 4:10, 5:11, 6:11, 7:12, 8:12, 9:13, 10:13, 11:14, 12:14, 13:15, 14:15,
                   16:20, 17:21, 18:22, 19:23, 20:24, 21:24, 22:25, 23:25, 24:26, 25:27, 26:28, 27:28, 31:33, 32:33, 35:37, 36:37}
        dict_loser = {1:16, 2:16, 3:17, 4:17, 5:18, 6:18, 7:19, 8:19, 9:23, 10:22, 11:21, 12:20, 13:26, 14:27,
                  16:35, 17:35, 18:36, 19:36, 20:31, 21:31, 22:32, 23:32, 24:30, 25:30, 26:29, 27:29, 31:34, 32:34, 35:38, 36:38}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                      14: -14, 16: -16, 17: -17, 18: -18, 19: -19, 20: -20, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 
                      27: -27, 31: -31, 32: -32, 35: -35, 36: -36}
        dict_mesta = [15, 28, 29, 30, 33, 34, 37, 38]
    elif vid_setki == 'Сетка (с розыгрышем всех мест) на 32 участников':
        dict_winner = {1: 17, 2: 17, 3: 18, 4: 18, 5: 19, 6: 19, 7: 20, 8: 20, 9: 21, 10: 21, 11: 22, 12: 22, 13: 23, 14: 23,
                   15: 24, 16: 24, 17: 25, 18: 25, 19: 26, 20: 26, 21: 27, 22: 27, 23: 28, 24: 28, 25: 29, 26: 29, 27: 30, 28: 30, 
                   29: 31, 30: 31, 33: 35, 34: 35, 37: 41, 38: 41, 39: 42, 40: 42, 41: 43, 42: 43, 45: 47, 46: 47, 49: 57, 50: 57,
                   51: 58, 52: 58, 53: 59, 54: 59, 55: 60, 56: 60, 57: 61, 58: 61, 59: 62, 60: 62, 61: 63, 62: 63, 65: 67, 66: 67,
                   69: 73, 70: 73, 71: 74, 72: 74, 73: 75, 74: 75, 77: 79, 78: 79}
        dict_loser = {1: 49, 2: 49, 3: 50, 4: 50, 5: 51, 6: 51, 7: 52, 8: 52, 9: 53, 10: 53, 11: 54, 12: 54, 13: 55, 14: 55, 15: 56, 16: 56,
                  17: 37, 18: 37, 19: 38, 20: 38, 21: 39, 22: 39, 23: 40, 24: 40, 25: 33, 26: 33, 27: 34, 28: 34, 29: 32, 30: 32,
                  33: 36, 34: 36, 37: 45, 38: 45, 39: 46, 40: 46, 41: 44, 42: 44, 45: 48, 46: 48, 49: 69, 50: 69, 51: 70, 52: 70, 53: 71, 54: 71, 55: 72, 56: 72,
                  57: 65, 58: 65, 59: 66, 60: 66, 61: 64, 62: 64, 65: 68, 66: 68, 69: 77, 70: 77, 71: 78, 72: 78, 73: 76, 74: 76, 77: 80, 78: 80}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                    14: -14, 15: -15, 16: -16, 17: -17, 18: -18, 19: -19, 20: -20, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 
                    27: -27, 28: -28, 29: -29, 30: -30, 33: -33, 34: -34, 37: -37, 38: -38, 39: -39, 40: -40, 41: -41, 42: -42, 45: -45, 46: -46,
                    49: -49, 50: -50, 51: -51, 52: -52, 53: -53, 54: -54, 55: -55, 56: -56, 57: -57, 58: -58, 59: -59, 60: -60, 61: -61, 62: -62, 65: -65, 66: -66,
                    69: -69, 70: -70, 71: -71, 72: -72, 73: -73, 74: -74, 75: -75, 77: -77, 78: -78, 79: -79}
        dict_mesta = [31, 32, 35, 36, 43, 44, 47, 48, 63, 64, 67, 68, 75, 76, 79, 80]
    elif vid_setki == 'Сетка (-2) на 32 участников':
        dict_winner = {1: 17, 2: 17, 3: 18, 4: 18, 5: 19, 6: 19, 7: 20, 8: 20, 9: 21, 10: 21, 11: 22, 12: 22, 13: 23, 14: 23,
                   15: 24, 16: 24, 17: 25, 18: 25, 19: 26, 20: 26, 21: 27, 22: 27, 23: 28, 24: 28, 25: 29, 26: 29, 27: 30, 28: 30, 
                   29: 31, 30: 31, 32: 40, 33: 41, 34: 42, 35: 43, 36: 44, 37: 45, 38: 46, 39: 47, 40: 48, 41: 48, 42: 49, 43: 49, 44: 50, 45: 50, 46: 51,
                   47: 51, 48: 52, 49: 53, 50: 54, 51: 55, 52: 56, 53: 56, 54: 57, 55: 57, 56: 58, 57: 59, 58: 60, 59: 60, 63: 65, 64: 65, 67: 69, 68: 69, 71: 75, 72: 75, 73: 76, 74: 76, 
                   75: 77, 76: 77, 79: 81, 80: 81, 83: 87, 84: 87, 85: 88, 86: 88, 87: 89, 88: 89, 91: 93, 92: 93}
        
        dict_loser = {1: 32, 2: 32, 3: 33, 4: 33, 5: 34, 6: 34, 7: 35, 8: 35, 9: 36, 10: 36, 11: 37, 12: 37, 13: 38, 14: 38, 15: 39, 16: 39,
                  17: 47, 18: 46, 19: 45, 20: 44, 21: 43, 22: 42, 23: 41, 24: 40, 25: 53, 26: 52, 27: 55, 28: 54, 29: 59, 30: 58,
                  32: 83, 33: 83, 34: 84, 35: 84, 36: 85, 37: 85, 38: 86, 39: 86, 40: 71, 41: 71, 42: 72, 43: 72, 44: 73, 45: 73, 46: 74, 47: 74, 48: 67, 49: 67, 
                  50: 68, 51: 68, 52: 63, 53: 63, 54: 64, 55: 64, 56: 62, 57: 62, 58: 61, 59: 61, 63: 66, 64: 66, 67: 70, 68: 70, 71: 79, 72: 79, 73: 80, 74: 80, 
                  79: 82, 80: 82, 75: 78, 76: 78, 83: 91, 84: 91, 85: 92, 86: 92, 87: 90, 88: 90, 91: 94, 92: 94}
        
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                    14: -14, 15: -15, 16: -16, 17: -17, 18: -18, 19: -19, 20: -20, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 
                    27: -27, 28: -28, 29: -29, 30: -30, 32: -32, 33: -33, 34: -34, 35: -35, 36: -36, 37: -37, 38: -38, 39: -39, 40: -40, 
                    41: -41, 42: -42, 43: -43, 44: -44, 45: -45, 46: -46, 47: -47, 48: -48, 49: -49, 50: -50, 51: -51, 52: -52, 53: -53, 54: -54, 
                    55: -55, 56: -56, 57: -57, 58: -58, 59: -59, 63: -63, 64: -64, 67: -67, 68: -68,
                    71: -71, 72: -72, 73: -73, 74: -74, 75: -75, 76: -76, 79: -79, 80: -80, 83: -83, 84: -84, 85: -85, 86: -86, 87: -87,
                    88: -88, 91: -91, 92: -92}
        dict_mesta = [31, 60, 61, 62, 65, 66, 69, 70, 77, 78, 81, 82, 89, 90, 93, 94]
    elif vid_setki == 'Сетка (1-3 место) на 32 участников':
        pass

    if num_game in dict_mesta: # если встреча за места
        index = dict_mesta.index(num_game)
        snoska = [0, 0]
        # для отображения в pdf (встречи с минусом)
        game_loser = dict_mesta[index] * -1
        snoska.append(game_loser)
    else:
        game_winner = dict_winner[num_game]  # номер игры победителя
        snoska.append(game_winner)
        game_loser = dict_loser[num_game]  # номер игры проигравшего
        snoska.append(game_loser)
        # для отображения в pdf (встречи с минусом)
        game_loser = dict_loser_pdf[num_game]
        snoska.append(game_loser) # список: номер встречи победителя, номер - проигравшего и куда снести проигравшего
    return snoska


def score_in_setka(fin):
    """ выставляет счет победителя и сносит на свои места в сетке"""
    dict_setka = {}
    match = []
    tmp_match = []
    sys = System.select().where(System.title_id == title_id())
    system = sys.select().where(System.stage == fin).get()
    system_id = system.id
    vid_setki = system.label_string
    # получение id последнего соревнования
    player = Player.select().where(Player.title_id == title_id())
    result = Result.select().where(Result.system_id == system_id)
 
    for res in result:
        num_game = int(res.tours)

        if res.winner is not None and res.winner != "":
            id_pl_win = player.select().where(Player.full_name == res.winner).get()
            id_pl_los = player.select().where(Player.full_name == res.loser).get()
            short_name_win = id_pl_win.player
            short_name_los = id_pl_los.player

            snoska = numer_game(num_game, vid_setki)
            tmp_match.append(snoska[0]) # номер на сетке куда идет победитель
            tmp_match.append(short_name_win)
            tmp_match.append(f'{res.score_in_game} {res.score_win}')
            tmp_match.append(snoska[2])
            tmp_match.append(short_name_los)
            match = tmp_match.copy()
            tmp_match.clear()
            dict_setka[num_game] = match
    return dict_setka


def result_rank_group_in_choice(num_gr, player_rank_group, stage):
    """записывает места из группы в таблицу -Choice-, а если одна таблица в финале по кругу то в список
    player_rank_group список списков 1-е число номер игрок в группе, 2-е его место"""
    tab = my_win.tabWidget.currentIndex()
    chc = Choice.select().where(Choice.title_id == title_id())
    if len(player_rank_group) > 0:
        if tab == 3:
            choice = chc.select().where(Choice.group == num_gr)
        elif tab == 4:
            choice = chc.select().where((Choice.semi_final == stage) & (Choice.sf_group == num_gr))
        else:
            if num_gr == "Одна таблица":
                choice = chc.select().where(Choice.basic == "Одна таблица")
            else:
                choice = chc.select().where(Choice.final == num_gr)
        count = len(choice)
        n = 0
        for ch in choice:
            if tab == 3:
                for i in range(0, count):  # цикл по номерам посева в группе
                    # если есть совпадение, то место в списке
                    if ch.posev_group == player_rank_group[i][0]:
                        with db:
                            # записывает в таблицу -Choice- места в группе
                            ch.mesto_group = player_rank_group[i][1]
                            ch.save()
            elif tab == 4:
                for i in range(0, count):  # цикл по номерам посева в группе
                    # если есть совпадение, то место в списке
                    if ch.posev_sf == player_rank_group[i][0]:
                        with db:
                            # записывает в таблицу -Choice- места в группе
                            ch.mesto_semi_final = player_rank_group[i][1]
                            ch.save()
            else:
                player_rank_group.sort()
                ch.mesto_final = player_rank_group[n][1]
                player_id = ch.player_choice_id
                ch.save()
                player = Player.get(Player.id == player_id)
                player.mesto = player_rank_group[n][1]
                player.save()
                n += 1


def rank_in_group(total_score, td, num_gr, stage):
    """выставляет места в группах соответственно очкам 
    -men_of_circle - кол-во человек в крутиловке
    -player_rank_group - список списков номер игрока - место 
    -num_player -player_rank - список списков участник - место
    -player_group - кол-во участников в группе"""
    tr_all = []
    pps = []
 
    pp = {}  # ключ - игрок, значение его очки
    pg_win = {}
    pg_los = {}
    tr = []
    player_rank_tmp = []
    # player_rank = []
    rev_dict = {}  # словарь, где в качестве ключа очки, а значения - номера групп
    player_rank_group = []    
    sys = System.select().where(System.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    game_list = Game_list.select().where(Game_list.title_id == title_id())

    if stage == "Предварительный":
        system = sys.select().where(System.stage == stage).get()
        max_person = system.max_player
        game_max = result.select().where((Result.system_stage == stage) & (Result.number_group == num_gr))
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        system = sys.select().where(System.stage == stage).get()
        system_id = system.id
        game_list_group = game_list.select().where((Game_list.system_id == system_id) & (Game_list.number_group == num_gr))
        game_max = result.select().where((Result.system_stage == stage) & (Result.number_group == num_gr))  # сколько всего игр в группе
        max_person = len(game_list_group)
    else:
        system = sys.select().where(System.stage == num_gr).get()
        id_system = system.id
        game_max = result.select().where(Result.system_id == id_system)  # сколько всего игр в группе
        game_list_group = game_list.select().where(Game_list.system_id == id_system)
        max_person = len(game_list_group)
    # elif num_gr == "Одна таблица":
    #     system = sys.select().where(System.stage == stage).get()
    #     id_system = system.id
    #     game_max = result.select().where(Result.system_id == id_system)  # сколько всего игр в группе

    # else:
    #     system = sys.select().where(System.stage == num_gr).get()
    #     id_system = system.id
    #     game_list_group = game_list.select().where(Game_list.system_id == id_system)
    #     game_max = result.select().where(Result.system_id == system_id)  # сколько всего игр в финале по кругу
    #     max_person = len(game_list_group)
    # ======== проверка на неявку ======
    fio_no_player = []
    game_not_player = game_max.select().where(Result.points_loser == 0)
    count_not_player = len(game_not_player)
    if count_not_player != 0:
        for k in game_not_player:
            pl = k.points_loser
            if pl == 0:
                player_no = k.loser
                if player_no not in fio_no_player:
                    fio_no_player.append(player_no)
                    
        for fio in fio_no_player:
            fio_loser = game_not_player.select().where(Result.loser == fio)
            count_fio_loser = len(fio_loser)

            game_one_person = max_person // 2
            if count_fio_loser >= game_one_person: # игры по неявке более 50%
                game_id_not_player = game_not_player.select().where(Result.loser == fio)
                for game_id in game_id_not_player:
                    game_id.points_win = ""
                    game_id.points_loser = ""
                    game_id.save()
     # ===========================================
    # 1-й запрос на выборку с группой
    game_played = game_max.select().where((Result.winner is None) | (Result.winner != ""))  # 2-й запрос на выборку
    # с победителями из 1-ого запроса
    kol_tours_played = len(game_played)  # сколько игр сыгранных
    kol_tours_in_group = len(game_max)  # кол-во всего игр в группе

    for key, value in total_score.items():
        rev_dict.setdefault(value, set()).add(key) # словарь (число очков, номера участников группы у которых они есть)
    res = [key for key, values in rev_dict.items() if len(values) > 1] # список очки, которых более одного

    # отдельно составляет список ключей (номера участников группы)
    val_list = list(total_score.values())  # отдельно составляет список значений (очки каждого игрока)
    # ======== новый вариант =========
    # получает словарь(ключ - номер участника, значение - очки)
    ds = {index: value for index, value in enumerate(val_list)}  
    # сортирует словарь по убыванию соот
    sorted_tuple = {k: ds[k] for k in sorted(ds, key=ds.get, reverse=True)}
    valuesList = list(sorted_tuple.values())  # список очков по убыванию
    unique_numbers = list(set(valuesList))  # множество уникальных очков
    unique_numbers.sort(reverse=True)  # список уникальных очков по убыванию
    mesto = 1

    for f in unique_numbers:  # проходим циклом по уник. значениям
        num_player = rev_dict.get(f)
        for x in num_player:
            tr.append(str(x))  # создает список (встречи игроков)
        m_new = valuesList.count(f)  # подсчитываем сколько раз оно встречается

        if m_new == 1:  # если кол-во очков у одного спортсмена
            p1 = x
            # записывает место победителю
            td[p1 * 2 - 2][max_person + 4] = mesto
            player_rank_tmp.append([p1, mesto])
        elif m_new == 2:  # если кол-во очков у двух спортсмена (определение мест по игре между собой)
            player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr)
        elif m_new == 3: # если кол-во очков у трех спортсмена
            men_of_circle = m_new
            # получает список 1-й уникальные
            u = summa_points_person(men_of_circle, tr, tr_all, pp, pg_win, pg_los, num_gr, stage)
            # значения очков и список значения очков и у скольких спортсменов они есть
            z = u[1]  # список списков кол-во очков и у сколько игроков они есть
            points_person = z[0] # список [колво очко, у скольки игроков они есть]
            player_rank_tmp = circle_3_player(men_of_circle, points_person, tr, td, max_person, mesto, player_rank_tmp, num_gr, tr_all,
                    pg_win, pg_los, pp, pps, stage)
        elif m_new > 3:  # если кол-во очков у более трех спортсменов (крутиловка)
            m_circle = m_new
            men_of_circle = m_new
            player_rank_tmp = circle(men_of_circle, tr, num_gr, td, max_person, mesto, m_circle, stage)
        tr.clear()

        for i in player_rank_tmp:
            # список участников в группе и его место
            player_rank_group.append(i)

        mesto = mesto + m_new
        player_rank_tmp.clear()
    if kol_tours_played == kol_tours_in_group:  # когда все встречи сыграны
        # функция простановки мест из группы в -Choice-
        result_rank_group_in_choice(num_gr, player_rank_group, stage)


def get_unique_numbers(pp_all):
    """получение списка уникальных значений"""
    unique = []
    for number in pp_all:
        if number not in unique:
            unique.append(number)
    return unique


def circle(men_of_circle, tr, num_gr, td, max_person, mesto, m_circle, stage):
    """выставляет места в крутиловке -tour- встречи игроков, p1, p2 фамилии, num_gr номер группы
    -tr- список всех туров (номеров) участников в крутиловке men_of_circle кол-во игроков с одинаковым кол-вом очков,
    max_person общее кол-во игроков в группе player_rank - список (номер игроков и их места)"""
    pl_rank_tmp = []  # список списков (игрок и его место)
    player_rank_tmp = []
    tr_all = []
    ps = []
    pps = []
    rev_dict = {}  # словарь, где в качестве ключа очки, а значения - номера групп
    pp = {}  # ключ - игрок, значение его очки
    pg_win = {}
    pg_los = {}

    # получает список 1-й уникальные
    u = summa_points_person(tr, tr_all, pp, pg_win, pg_los, num_gr, stage)
    # значения очков и список значения очков и у скольких спортсменов они есть
    unique_numbers = u[0]
    tr.clear()
    sort_tuple = {k: pp[k] for k in sorted(pp, key=pp.get, reverse=True)}
    for key, value in sort_tuple.items():
        rev_dict.setdefault(value, set()).add(key)

    for f in unique_numbers:  # проходим циклом по уник. значениям, очки в крутиловке
        m_new = 0
        num_player = rev_dict.get(f)
        count_point = len(num_player)

        if count_point == 1:
            for x in num_player:
                p1 = x
            # записывает место победителю
            td[p1 * 2 - 2][max_person + 4] = mesto
            td[p1 * 2 - 2][max_person + 3] = f  # записывает место победителю
            player_rank_tmp.append([p1, mesto])
            m_new += 1
        elif count_point == 2:
            for x in num_player:
                tr.append(str(x))  # создает список (встречи игроков)
                m_new += 1
            player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr)
        else:
            point = 0
            for x in num_player:
                tr.append(str(x))  # создает список (встречи игроков)
                m_new += 1
            player_rank_tmp = circle_in_circle(m_new, td, max_person, mesto, tr, num_gr, point,
                                               player_rank_tmp, tr_all, pp, pg_win, pg_los, x, pps, ps)
        mesto = mesto + m_new
        tr.clear()

        # заменяет список (места еще не проставлены) на новый с правильными местами
        for i in player_rank_tmp:
            pl_rank_tmp.append(i)
        player_rank_tmp.clear()
    player_rank_tmp = pl_rank_tmp
    return player_rank_tmp


def circle_in_circle(m_new, td, max_person, mesto, tr, num_gr, point, player_rank_tmp,
                     tr_all, pp, pg_win, pg_los, x, pps, ps, stage):
    """крутиловка в крутиловке"""
    num_player = []
    if m_new == 1:
        p1 = x
        td[p1 * 2 - 2][max_person + 4] = mesto  # записывает место победителю
        td[p1 * 2 - 2][max_person + 3] = point  # очки во встрече победителя
        player_rank_tmp.append([p1, mesto])
    elif m_new == 2:
        player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr)
    elif m_new == 3:
        # men_of_circle = m_new
        # получает список 1-й уникальные
        u = summa_points_person(tr, tr_all, pp, pg_win, pg_los, num_gr, stage)
        # значения очков и список значения очков и у скольких спортсменов они есть
        z = u[1]
        points_person = z[0]
        player_rank_tmp = circle_3_player(points_person, tr, td, max_person, mesto, player_rank_tmp, num_gr, ps,
                                          tr_all, pg_win, pg_los, pp, pps, stage)
    elif m_new > 3:
        dict_ratio = {}
        for k in range(1, m_new + 1):
            pg_win[k] = sum(pg_win[k])  # сумма выигранных партий
            pg_los[k] = sum(pg_los[k])  # сумма проигранных партий
            x = pg_win[k] / pg_los[k]
            x = float('{:.3f}'.format(x)) # соотношение выйгранных партий к проигранным
            dict_ratio[k] = x
        sorted_ratio = {k: dict_ratio[k] for k in
                            sorted(dict_ratio, key=dict_ratio.get, reverse=True)}  # сортирует словарь по убыванию соот 
        k_list = list(sorted_ratio.keys())  # отдельно составляет список ключей (группы)
        v_list = list(sorted_ratio.values())  # отдельно составляет список значений (соотношение)
        ratio_person = get_unique_numbers(v_list)  
        list_uniq = []  # список списков соотношение (выигранный партии к проигравшем) и кол-во игроков их имеющих
        list_tmp = []
        u = []

        for p in ratio_person:
            a = v_list.count(p)
            list_tmp.append(p)
            list_tmp.append(a)
            # список (очки и скольких игроков они встречаются)
            list_uniq.append(list_tmp.copy())
            list_tmp.clear()

        for k in list_uniq:
            point = k[0] # соотношение
            total_uniq = k[1] # сколько раз встречается
            index = v_list.index(k[0])
            p1 = k_list[index]
            if total_uniq == 1:
                td[p1 * 2 - 2][max_person + 4] = mesto  # записывает место победителю
                td[p1 * 2 - 2][max_person + 3] = point  # очки во встрече победителя
                player_rank_tmp.append([p1, mesto])   
                mesto += 1
            elif total_uniq == 2:
                for i in range(len(v_list)):
                    if v_list[i] == point:
                        num_pl = k_list[i]
                        num_player.append(num_pl)
                tr.clear()
                for x in num_player:
                    tr.append(str(x))  # создает список (встречи игроков)
                    m_new += 1
                player_rank_temp = circle_2_player(tr, td, max_person, mesto, num_gr)
                player_rank_tmp = player_rank_tmp + player_rank_temp
            elif total_uniq == 3:
                pass

    tr_all.clear()
    tr.clear()
    return player_rank_tmp


def tour_circle(pp, per_circ, circ):
    tr_new = []
    k_list = list(pp.keys())  # отдельно составляет список ключей (группы)
    v_list = list(pp.values())  # отдельно составляет список значений (очки)
    y = 0
    for s in range(0, circ):
        index = v_list.index(per_circ, y)
        per = str(k_list[index])
        y = index + 1
        tr_new.append(per)
    return tr_new


def summa_points_person(men_od_circle, tr, tr_all, pp, pg_win, pg_los, num_gr, stage):
    """подсчитывает сумму очков у спортсменов в крутиловке 
    -tr- номера игроков в группе, у которых крутиловка
    -tr_all- все варианты встреч в крутиловке
    -pg_los- словарь (номер игрока: список (кол-во проигранных партий)
    -pg_win- словарь (номер игрока: список (кол-во выйгранных партий)"""
    pp_all = []
    u = []
    tr_all.clear()
    pg_win.clear()
    pg_los.clear()
    pp.clear()
    for r in tr:
        r = int(r)
        pp[r] = []  # словарь (игрок - сумма очков)
        pg_win[r] = []
        pg_los[r] = []

    for i in combinations(tr, 2):  # получает список с парами игроков в крутиловке
        i = list(i)
        tr_all.append(i)
    count_game_circle = len(tr_all)

    for n in range(0, count_game_circle):
        tour = "-".join(tr_all[n])  # получает строку встреча в туре
        ki1 = int(tr_all[n][0])  # 1-й игрок в туре
        ki2 = int(tr_all[n][1])  # 2-й игрок в туре

        sum_points_circle(num_gr, tour, ki1, ki2, pg_win, pg_los, pp, stage)  # сумма очков игрока

    for i in tr:  # суммирует очки каждого игрока
        i = int(i)
        s = sum(pp[i])
        pp[i] = s  # словарь (участник - его очки)
        pp_all.append(s)

    list_uniq = []  # список списков сумма очков и кол-во игроков их имеющих
    list_tmp = []
    points_person = get_unique_numbers(pp_all)
    points_person.sort(reverse=True)

    for p in points_person:
        a = pp_all.count(p)
        list_tmp.append(p)
        list_tmp.append(a)
        # список (очки и скольких игроков они встречаются)
        list_uniq.append(list_tmp.copy())
        list_tmp.clear()
    u.append(points_person)
    u.append(list_uniq)
    return u # список списков 1-й кол-во игроков 2-й очки выйигранные и проигранные


def circle_2_player(tr, td, max_person, mesto, num_gr):
    """крутиловка из 2-ух человек"""
    result = Result.select().where(Result.title_id == title_id())
    player_rank_tmp = []
    tour = "-".join(tr)  # делает строку встреча в туре
    # =====приводит туры к читаемому виду (1-й игрок меньше 2-ого)
    znak = tour.find("-")
    p1 = int(tour[:znak])  # игрок под номером в группе
    p2 = int(tour[znak + 1:])  # игрок под номером в группе
    if p1 > p2:  # меняет последовательность игроков в туре на обратную, чтоб у 1-ого игрока был номер меньше
        p_tmp = p1
        p1 = p2
        p2 = p_tmp
        tour = f"{p1}-{p2}"
    if num_gr != "Одна таблица":
        c = result.select().where((Result.number_group == num_gr) &
                                  (Result.tours == tour)).get()  # ищет в базе
    # строчку номер группы и тур по двум столбцам
    else:
        c = result.select().where((Result.system_stage == num_gr) &
                                  (Result.tours == tour)).get()  # ищет в базе
        # строчку номер группы и тур по двум столбцам
    if c.winner == c.player1:
        points_p1 = c.points_win  # очки во встрече победителя
        points_p2 = c.points_loser  # очки во встрече проигравшего
        td[p1 * 2 - 2][max_person + 4] = mesto  # записывает место победителю
        td[p2 * 2 - 2][max_person + 4] = mesto + 1  # записывает место проигравшему
        player_rank_tmp.append([p1, mesto])
        player_rank_tmp.append([p2, mesto + 1])
    else:
        points_p1 = c.points_loser
        points_p2 = c.points_win
        td[p1 * 2 - 2][max_person + 4] = mesto + 1  # записывает место победителю
        td[p2 * 2 - 2][max_person + 4] = mesto  # записывает место проигравшему
        player_rank_tmp.append([p1, mesto + 1])
        player_rank_tmp.append([p2, mesto])
    td[p1 * 2 - 2][max_person + 3] = points_p1  # очки во встрече победителя
    td[p2 * 2 - 2][max_person + 3] = points_p2  # очки во встрече проигравшего

    return player_rank_tmp


def circle_3_player(men_of_circle, points_person, tr, td, max_person, mesto, player_rank_tmp, num_gr, tr_all,
                    pg_win, pg_los, pp, pps, stage):
    """в крутиловке 3-и спортсмена
    -pp- словарь (номер игрока, очки)
    -ps- список коэфициентов
    -points_person - список [1-е значение колво очков, 2-е у скольки участников оное есть"""
    ps = []
    if points_person[1] == 3:  # у всех трех участников равное кол-во очков   
        for k in tr:  # суммирует выигранные и проигранные партии каждого игрока
            k = int(k)
            pg_win[k] = sum(pg_win[k])  # сумма выигранных партий
            pg_los[k] = sum(pg_los[k])  # сумма проигранных партий
            x = pg_win[k] / pg_los[k]
            x = float('{:.3f}'.format(x)) # соотношение выйгранных партий к проигранным
            ps.append(x) # коэфициент
            pps.append(pp[k])
        # получает словарь(ключ, номер участника)
        d = {index: value for index, value in enumerate(tr)}
        # получает словарь(ключ, соотношение)
        ds = {index: value for index, value in enumerate(ps)}
        # сортирует словарь по убываню соот
        sorted_tuple = {k: ds[k] for k in sorted(ds, key=ds.get, reverse=True)}
        key_l = list(sorted_tuple.keys())
        val_l = list(sorted_tuple.values())
        vls = set(val_l)  # группирует разные значения
        vl = len(vls)  # подсчитывает их количество
        m = 0
        if vl == 1:  # подсчитывает соотношения выигранных и проигранных мячей в партиях
            plr_ratio = score_in_circle(tr_all, men_of_circle, num_gr, tr, stage)
            sorted_ratio = {k: plr_ratio[k] for k in
                            sorted(plr_ratio, key=plr_ratio.get, reverse=True)}  # сортирует словарь по убыванию соот
            # получает список ключей отсортированного словаря
            key_ratio = list(sorted_ratio.keys())
            r = 0
            for i in key_ratio:
                ratio = sorted_ratio[i]  # соотношение в крутиловке
                person = int(d[i])  # номер игрока
                # записывает соотношение
                td[person * 2 - 2][max_person + 3] = str(ratio)
                td[person * 2 - 2][max_person + 4] = str(mesto + r)  # записывает место
                # добавляет в список группа, место, чтоб занести в таблицу Choice
                player_rank_tmp.append([person, mesto + r])
                r += 1
        else:
            for i in val_l:
                # получает ключ, по которому в списке ищет игрока
                w = key_l[val_l.index(i)]
                # получает номер участника, соответствующий
                wq = int(d.setdefault(w))
                # записывает соотношения игроку
                td[wq * 2 - 2][max_person + 3] = str(i)
                # записывает место
                td[wq * 2 - 2][max_person + 4] = str(m + mesto)
                # добавляет в список группа, место, чтоб занести в таблицу Choice
                player_rank_tmp.append([wq, m + mesto])
                m += 1
    elif points_person[1] == 2 or points_person[1] == 1:
        # получает словарь(ключ, номер участника)
        d = {index: value for index, value in enumerate(tr)}
        # сортирует словарь по убыванию соот
        sorted_tuple = {k: pp[k] for k in sorted(pp, key=pp.get, reverse=True)}
        key_l = list(sorted_tuple.keys()) # номера игроков по убыванию очков
        val_l = list(sorted_tuple.values()) # очки игроков по убыванию
        m = 0
        # вставить если в крутиловке игра по неявке
        if points_person[1] == 2:
            if val_l[0] == val_l[1]:
                tr = [str(key_l[0]), str(key_l[1])] 
                player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr)
                player_rank_tmp.append([key_l[2], mesto + 2])
            else:
                player_rank_tmp = ([key_l[0], mesto])
                tr = [str(key_l[1]), str(key_l[2])] 
                player_rank_temp = circle_2_player(tr, td, max_person, mesto, num_gr)
                player_rank_tmp.extend(player_rank_temp)
            for k in player_rank_tmp:
                td[k[0] * 2 - 2][max_person + 4] = str(k[1])  # записывает место
        else:
            for i in val_l:
                q = val_l.index(i) # индекс в списке
                wq = key_l[q] # получает номер участника группы, соответствующий
                # записывает соотношения игроку
                td[wq * 2 - 2][max_person + 3] = str(i)
                td[wq * 2 - 2][max_person + 4] = str(m + mesto)  # записывает место
                # добавляет в список группа, место, чтоб занести в таблицу Choice
                player_rank_tmp.append([wq, m + mesto])
                m += 1
    return player_rank_tmp


def sum_points_circle(num_gr, tour, ki1, ki2, pg_win, pg_los, pp, stage):
    """сумма очков каждого игрока в крутиловке"""
    
    # # =====приводит туры к читаемому виду (1-й игрок меньше 2-ого)
    znak = tour.find("-")
    p1 = int(tour[:znak])  # игрок под номером в группе
    p2 = int(tour[znak + 1:])  # игрок под номером в группе
    if p1 > p2:  # меняет последовательность игроков в туре на обратную, чтоб у 1-ого игрока был номер меньше
        # уточнить смену тура при p1>p2
        tour = f"{p2}-{p1}"
        ki1 = p2
        ki2 = p1
    result = Result.select().where(Result.title_id == title_id())
    if num_gr == "Одна таблица":
        res = result.select().where(Result.system_stage == num_gr)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        res = result.select().where((Result.system_stage == stage) & (Result.number_group == num_gr))
    else:
        res = result.select().where(Result.number_group == num_gr)
    c = res.select().where(Result.tours == tour).get()  # ищет в базе  данную встречу c - id - встречи в таблице Result
 
    if c.winner == c.player1:  # победил 1-й игрок
        points_p1 = c.points_win  # очки победителя
        points_p2 = c.points_loser  # очки проигравшего
        game_p1 = c.score_in_game  # счет во встречи (выигранные и проигранные партии) победителя
        game_p2 = c.score_loser # счет во встречи (выигранные и проигранные партии) проигравшего
        if game_p1 == "В : П":
            p1_game_win = 0
            p1_game_los = 0
            p2_game_win = 0
            p2_game_los = 0
        else:
            p1_game_win = int(game_p1[0]) # кол-во выигранных партий 1 игрока
            p1_game_los = int(game_p1[4])
            p2_game_win = int(game_p2[0])
            p2_game_los = int(game_p2[4])
    else: # победил 2-й игрок
        points_p1 = c.points_loser # очки 1-ого игрока проигранные
        points_p2 = c.points_win # очки 2-ого игрока выигранные
        game_p1 = c.score_loser # счет во встречи 1-ого игрока
        game_p2 = c.score_in_game # счет во встречи 2-ого игрока
        # ======= если победа по неявке исправить
        if game_p1 == "П : В":
            p1_game_win = 0
            p1_game_los = 0
            p2_game_win = 0
            p2_game_los = 0
        else:
            p1_game_win = int(game_p1[0]) # кол-во выигранных партий 1 игрока
            p1_game_los = int(game_p1[4])
            p2_game_win = int(game_p2[0])
            p2_game_los = int(game_p2[4])
    pp[ki1].append(points_p1)  # добавляет очки 1-ому игроку встречи
    pp[ki2].append(points_p2)  # добавляет очки 2-ому игроку встречи
# записывает в словарь счет во встречи 1-ого игрока
    pg_win[ki1].append(p1_game_win)
    # записывает в словарь счет во встречи 1-ого игрока
    pg_los[ki1].append(p1_game_los)
    # записывает в словарь счет во встречи 2-ого игрока
    pg_win[ki2].append(p2_game_win)
    # записывает в словарь счет во встречи 2-ого игрока
    pg_los[ki2].append(p2_game_los)


def score_in_circle(tr_all, men_of_circle, num_gr, tr, stage):
    """подсчитывает счет по партиям в крутиловке"""
    result = Result.select().where(Result.title_id == title_id())
    plr_win = {0: [], 1: [], 2: []}
    plr_los = {0: [], 1: [], 2: []}
    plr_ratio = {0: [], 1: [], 2: []}
    for n in range(0, men_of_circle):
        tour = "-".join(tr_all[n])  # получает строку встреча в туре
        znak = tour.find("-")
        p1 = int(tour[:znak])  # игрок под номером в группе
        p2 = int(tour[znak + 1:])  # игрок под номером в группе
        if p1 > p2:  # меняет последовательность игроков в туре на обратную, чтоб у 1-ого игрока был номер меньше
            tour = f"{p2}-{p1}"
        c_res = result.select().where((Result.system_stage == stage) & (Result.number_group == num_gr))
        c = c_res.select().where(Result.tours == tour).get()
        k1 = tr_all[n][0]  # 1-й игрок в туре
        k2 = tr_all[n][1]  # 2-й игрок в туре
        ki1 = tr.index(k1)  # получение индекса 1-й игрока
        ki2 = tr.index(k2)
        g = c.score_win
        g_len = len(g)
        g = g[1:g_len - 1]
        sc_game = g.split(",")
 
        if c.winner == c.player1:  # победил 1-й игрок
            for i in sc_game:
                i = int(i)
                if i < 0:
                    plr_win[ki1].append(abs(i))
                    plr_los[ki2].append(abs(i))
                    if abs(i) < 10:
                        plr_los[ki1].append(11)
                        plr_win[ki2].append(11)
                    else:
                        plr_los[ki1].append(abs(i) + 2)
                        plr_win[ki2].append(abs(i) + 2)
                elif 0 <= i < 10:
                    plr_win[ki1].append(11)
                    plr_los[ki1].append(i)
                    plr_win[ki2].append(i)
                    plr_los[ki2].append(11)
                elif i >= 10:
                    plr_win[ki1].append(i + 2)
                    plr_los[ki1].append(i)
                    plr_win[ki2].append(i)
                    plr_los[ki2].append(i + 2)
        else:  # если победил 2-й игрок
            for i in sc_game:
                i = int(i)
                if i < 0:  # партию проиграл
                    plr_win[ki2].append(abs(i))
                    plr_los[ki1].append(abs(i))
                    if abs(i) < 10:
                        plr_los[ki2].append(11)
                        plr_win[ki1].append(11)
                    else:
                        plr_los[ki2].append(abs(i) + 2)
                        plr_win[ki1].append(abs(i) + 2)
                elif 0 <= i < 10:  # выиграл партию
                    plr_win[ki2].append(11)
                    plr_los[ki2].append(i)
                    plr_win[ki1].append(i)
                    plr_los[ki1].append(11)
                elif i >= 10:  # выиграл партию на больше меньше
                    plr_win[ki2].append(i + 2)
                    plr_los[ki2].append(i)
                    plr_win[ki1].append(i)
                    plr_los[ki1].append(i + 2)
    for n in range(0, men_of_circle):
        plr_win[n] = sum(plr_win[n])
        plr_los[n] = sum(plr_los[n])
        x = plr_win[n] / plr_los[n]
        x = float('{:.3f}'.format(x))
        plr_ratio[n] = x
    return plr_ratio


def player_choice_in_group(num_gr):
    """распределяет спортсменов по группам согласно жеребьевке"""
    posev_data = []
    choice_group = Choice.select().where((Choice.title_id == title_id()) & (Choice.group == num_gr))
    players = Player.select().where(Player.title_id == title_id())
    for posev in choice_group:
        pl = players.select().where(Player.id == posev.player_choice_id).get()
        city = pl.city
        id_pl = posev.player_choice_id
        posev_data.append({
            'фамилия': f"{posev.family}/{id_pl}",
            'регион': city,
        })
    return posev_data


def player_choice_one_table(stage):
    """список спортсменов одной таблицы"""
    posev_data = []
    choices = Choice.select().where(Choice.title_id == title_id())
    players = Player.select().where(Player.title_id == title_id())
    if stage == "Одна таблица":
        choice = choices.select().where(Choice.basic == "Одна таблица")
    else:
        choice = choices.select().order_by(Choice.posev_final).where(Choice.final == stage)
    for posev in choice:
        pl = players.select().where(Player.id == posev.player_choice_id).get()
        city = pl.city
        posev_data.append({
            'фамилия': posev.family,
            'регион': city,
        })
    return posev_data


def player_choice_semifinal(stage, num_gr):
    """список спортсменов полуфиналов"""
    posev_data = []
    choice = Choice.select().where(Choice.title_id == title_id())
    choice_pf = choice.select().where(Choice.semi_final == stage)
    choice_group_pf = choice_pf.select().order_by(Choice.posev_sf).where(Choice.sf_group == num_gr)
    players = Player.select().where(Player.title_id == title_id())
    for posev in choice_group_pf:
        pl = players.select().where(Player.id == posev.player_choice_id).get()
        city = pl.city
        id_pl = posev.player_choice_id
        posev_data.append({
            'фамилия': f"{posev.family}/{id_pl}",
            'регион': city,
        })
    return posev_data


def player_choice_in_setka(fin):
    """распределяет спортсменов в сетке согласно жеребьевке"""
    system = System.select().where(System.title_id == title_id())
    system_id = system.select().where(System.stage == fin).get()
    count_exit = system_id.mesta_exit # сколько игроков выходят в финал
 
    flag = selection_of_the_draw_mode() # выбор ручная или автоматическая жеребьевка
    posev = choice_setka_automat(fin, flag, count_exit)

    posev_data = []
    for key in posev.keys():
        posev_data.append({'посев': key, 'фамилия': posev[key]})  
    # сортировка (списка словарей) по ключу словаря -посев-
    posev_data = sorted(posev_data, key=lambda i: i['посев'])
    sys_tem = system.select().where(System.stage == fin).get()
    with db:  # записывает в db, что жеребьевка произведена
        sys_tem.choice_flag = True
        sys_tem.save()
    return posev_data


def change_choice_group():
    """Смена жеребьевки групп если в группе 2 и более одинаковых регион чтоб развести тренеров"""
    msg = QMessageBox
    sender = my_win.sender()
    if my_win.radioButton_repeat_regions.isChecked():
        reg = []
        reg_d = []
        gr_key = []
        reg_tmp = []
        double_reg = {}
        fg = my_win.comboBox_filter_choice_stage.currentText()
        choice = Choice.select().where(Choice.title_id == title_id())
        system = System.select().where(System.title_id == title_id())
        sys = system.select().where(System.stage == "Предварительный").get()
        total_gr = sys.total_group
        for i in range(1, total_gr + 1):
            m = 0
            group = choice.select().where(Choice.group == f"{i} группа")
            for k in group:
                m += 1
                reg_n = k.region
                if reg_n not in reg:
                    reg.append(reg_n)
                else:
                    reg_tmp.append(reg_n)
            reg_d = reg_tmp.copy()
            count =len(reg_d)
            if count > 0:
                double_reg[f"{i} группа"] = reg_d
            reg_tmp.clear()
            reg.clear()
        dr_count = len(double_reg)
        if dr_count != 0:
            for key in double_reg.keys():
                gr_key.append(key)     
            my_win.comboBox_filter_choice_stage.clear()
            my_win.comboBox_filter_choice_stage.addItems(gr_key)
        else:
            msg.information(my_win, "Уведомление", "Нет групп с повторяющимися регионами.")
    else:
        return


def change_page_vid():
    """Смена вида страницы с таблицами""" 
    msgBox = QMessageBox
    sys = []
    sys.append("")
    system = System.select().where(System.title_id == title_id()) 
    for i in system:
        stage = i.stage
        sys.append(stage)
    stage, ok = QInputDialog.getItem(my_win, "Таблицы", "Выберите таблицы из списка для\n"
                                        "смены ориентации страницы", sys)
    if ok:                                   
        sys = system.select().where(System.stage == stage).get()
        vid = sys.page_vid
        vid_ed = "альбомная"
        if vid == "альбомная":
            vid_ed = "книжная"
        else:
            vid_ed = "альбомная"
        ok = msgBox.question(my_win, "Таблицы", "Текущая ориентация страницы\n"
                                            f"-{stage}-: {vid},\n"
                                            "Хотите ее изменить на:" f"{vid_ed}?", msgBox.Ok, msgBox.Cancel)
        if ok:
            sys.page_vid = vid_ed
            sys.save()
        else:
            return
    else:
        return


def change_dir(catalog):
    """смена директории, чтоб все pdf фалы сохранялися в папке table_pdf"""

    dir_path = pathlib.Path.cwd()
    parent_dir = str(dir_path)
    f1 = parent_dir.rfind("table_pdf")
    f2 = parent_dir.rfind("competition_pdf")
    if catalog == 1:
        if f1 == -1 :
            os.chdir("table_pdf") # переходит в каталог pdf       
    else:
        if f2 == -1 :
            os.chdir("competition_pdf") # переходит в каталог pdf       
 

def draw_setka_made(col, row, num, step, tur, style):
    """рисование сетки встреч игроков
    col - начальный столбец, row - начальный ряд, num - кол-во игроков"""
    style_set = []  
   
    col_fin = (col + 1) + (2 * (tur - 1)) # последний столбец
    row_fin = row + (num - 1) * step # последняя строка 
    for i in range (col, col_fin + 1, 2): # номер столбца 
        for k in range(row, row_fin + 1, step): # номер строки
            fn = ('LINEABOVE', (i, k), (i + 1, k), 1, colors.darkblue)  # рисует линии встреч
            style_set.append(fn)  
    for m in range(col + 1, col_fin + 1, 2):
        for q in range(row, row_fin, step):  # встречи 33-34
            fn = ('SPAN', (m, q), (m, q + step - 1 ))             
            style_set.append(fn)
            fn = ('BACKGROUND', (m, q), (m, q + step - 1 ), colors.lightyellow)  
            style_set.append(fn) 
            fn = ('BOX', (m, q), (m, q + step - 1), 1, colors.darkblue)
            style_set.append(fn) 
    for fn in style_set:
        style.append(fn)
    return style


def draw_setka(col, row, num, style):
    """рисование сетки встреч игроков
    col - начальный столбец, row - начальный ряд, num - кол-во игроков"""
    style_set = []  
    s = 1
    cf = 0  # кол-во туров
    if num == 2:  # кол-во игроков
        cf = 1
    elif num == 4:
        cf = 2
    elif num == 8:
        cf = 3
    elif num == 16:
        cf = 4
    elif num == 32:
        cf = 5
    row_b = row
    col_fin = col + cf * 2 # последний столбец
    row_fin = row + num * 2 - 1 # последняя строка 
    for i in range (col, col_fin, 2): # номер столбца 
        s *= 2
        for k in range(row, row_fin, s): # номер строки
            fn = ('LINEABOVE', (i, k), (i + 1, k), 1, colors.darkblue)  # рисует линии встреч
            style_set.append(fn)  
        row = row + s // 2
    s = 1
    for m in range(col + 1, col_fin + 1, 2):
        s *= 2
        for q in range(row_b, row_fin, s * 2):  # встречи 33-34
            fn = ('SPAN', (m, q), (m, q + s - 1 ))             
            style_set.append(fn)
            fn = ('BACKGROUND', (m, q), (m, q + s - 1 ), colors.lightyellow)  
            style_set.append(fn) 
            fn = ('BOX', (m, q), (m, q + s - 1), 1, colors.darkblue)
            style_set.append(fn)
        row_b = row_b + s // 2   
    for fn in style_set:
        style.append(fn)
    return style


def draw_setka_2(col, row, num, style):
    """рисование сетки встреч игроков
    col - начальный столбец, row - начальный ряд, num - кол-во игроков"""
    style_set = []  
    s = 1
    cf = 0  # кол-во туров
    if num == 2:  # кол-во игроков
        cf = 1
    elif num == 4:
        cf = 3
    elif num == 8:
        cf = 4
    elif num == 16:
        cf = 5
    
    row_b = row
    col_fin = col + cf * 2 # последний столбец
    row_fin = row + num * 2 - 1 # последняя строка 
    for i in range (col, col_fin, 2): # номер столбца 
        s *= 2
        for k in range(row, row_fin, s): # номер строки
            fn = ('LINEABOVE', (i, k), (i + 1, k), 1, colors.darkblue)  # рисует линии встреч
            style_set.append(fn)  
        if i == 1:
            row -= 1
            s = 1
        elif i == 3:
            row += 1
            s = 2
        elif i == 5:
            row -= 2
            row_fin -= 1
            s = 2
        elif i == 7:
            row += 2
            s = 4
 
    s = 1
    row_fin = row_b + num * 2 - 2 # последняя строка 
    for m in range(col + 1, col_fin + 1, 2):
        s *= 2
        for q in range(row_b, row_fin, s * 2):  # встречи 33-34
            fn = ('SPAN', (m, q), (m, q + s - 1 ))             
            style_set.append(fn)
            fn = ('BACKGROUND', (m, q), (m, q + s - 1 ), colors.lightyellow)  
            style_set.append(fn) 
            fn = ('BOX', (m, q), (m, q + s - 1), 1, colors.darkblue)
            style_set.append(fn)
        if m == 2:
            row_b -= 1
            s = 1
        elif m == 4:
            row_b += 1
            s = 2
        elif m == 6:
            row_b -= 2
            s = 2
        elif m == 8:
            row_b += 2
            s = 4

    for fn in style_set:
        style.append(fn)
    return style


def draw_mesta(row, col, player, style):
    """рисует линии встреч за место"""
    p = 0
    if player == 2:
        p = 4
    elif player == 4:
        p = 4
    elif player == 8:
        p = 6
    elif player == 16:
        p = 10

    col_f = 11

    if col == 9:
        col_f = col + 2
    else:
        col_f = col + 1

    for l in range(row, row + p + 1, p):
        fn = ('LINEABOVE', (col, l), (col_f, l), 1, colors.darkblue)  # рисует линии мест 5-6 места (4 чел)
        style.append(fn)
    return style


def draw_num(row_n, row_step, col_n, number_of_columns, number_of_game, player, data):
    """рисует номера встреч, row_n - начальный ряд, col_n - начальный столбец, 
    number_of_game - начальный номер встречи, player - кол-во участников, number_of_columns - кол-во столбцов """
    s = 1
    col_f = col_n + number_of_columns * 2 - 1
    row_f = row_n + (player - 2) * row_step 
    for k in range(col_n, col_f, 2):
        step = row_step * 2
        for i in range (row_n, row_f + 1, step):
            data[i][k] = str(number_of_game)
            number_of_game += 1
        row_step *= 2
        s *= 2
        row_n = row_n + s // 2
    return number_of_game


def draw_num_2(row_n, row_step, col_n, number_of_columns, number_of_game, data, player):
    """рисует номера встреч, row_n - начальный ряд, col_n - начальный столбец, 
    number_of_game - начальный номер встречи, player - кол-во участников, 
    number_of_columns - число столбцов"""
  
    col_f = col_n + number_of_columns * 2 - 1
    row_f = row_n + (player * 2 - 1)
    for k in range(col_n, col_f, 2): 
        for i in range (row_n, row_f, row_step * 2):
            data[i][k] = str(number_of_game)
            number_of_game += 1
        
        row_n -= int(row_step / 2)
        row_f -= int(row_step / 2)

    return number_of_game


def draw_num_lost(row_n, row_step, col_n, number_of_game, player, data):
    """нумерация встреч проигранных"""
    row_f = row_n + (player - 1) * 2 + 1
    for d in range(row_n, row_f, row_step):
            data[d - 1][col_n] = str(number_of_game * -1)
            number_of_game += 1


def draw_num_lost_2(row_n, row_step, col_n, revers_number, number_of_game, player, data):
    """нумерация встреч проигранных"""
    if revers_number == 0:
        row_n = row_n
        row_f = row_n + (player - 1) * row_step * 2 + 1
        step = row_step * 2
    else:
        row_n = row_n + (player - 1) * row_step * 2
        row_f = row_n - (player - 1) * row_step * 2 - 1
        step = row_step * 2 * -1

    for d in range(row_n, row_f, step):
            data[d][col_n] = str(number_of_game * -1)
            number_of_game += 1


def color_mesta(data, first_mesto, table):
    """окрашивает места в красный цвет"""
    b = 0
    style_color = []
    ml = [] # столбец, ряд -1 ого места, ряд 2-ого места + 1, шаг между местами
    f = 0 # количество столбцов
    if table == "setka_32":
        f = 2
    elif table == "setka_32_full":
        f = 13
    elif table == "setka_32_2":
        f = 16
    elif table == "setka_16_full":
        f = 8
    elif table == "setka_16_2":
        f= 8
    elif table == "setka_8_full":
        f= 4
    elif table == "setka_8_2":
        f= 4

    for c in range(0, f):
        if c == 0: # 1-2 место
            if table == "setka_32_2":
                ml = [13, 31, 54, 22] 
            elif table == "setka_16_full":
                ml = [10, 15, 26, 10] 
            elif table == "setka_16_2":
                ml = [9, 15, 33, 17]
            elif table == "setka_8_full":
                ml = [8, 7, 14, 6]
            elif table == "setka_8_2":
                ml = [8, 7, 14, 6]
            elif table == "setka_32_full":
                ml = [11, 31, 54, 22] 
            elif table == "setka_32":
                ml = [11, 31, 54, 22]
        elif c == 1: # 3-4 место
            if table == "setka_32_2":
                ml = [13, 80, 97, 16]  
            elif table == "setka_16_full":
                ml = [10, 29, 32, 2] 
            elif table == "setka_16_2":
                ml = [9, 48, 56, 7] 
            elif table == "setka_8_full":
                ml = [8, 16, 19, 2]
            elif table == "setka_8_2": 
                ml = [8, 18, 23, 4]
            elif table == "setka_32_full":               
                ml = [11, 59, 65, 5] 
            elif table == "setka_32":               
                ml = [11, 59, 65, 5]    
        elif c == 2: # 5-6 место
            if table == "setka_32_2":
                ml = [13, 101, 106, 4]  
            elif table == "setka_16_full":
                ml = [10, 34, 38, 3] 
            elif table == "setka_16_2":
                ml = [9, 60, 64, 3] 
            elif table == "setka_8_full":
                ml = [8, 22, 26, 3]
            elif table == "setka_8_2": 
                ml = [8, 25, 29, 3] 
            else:
                ml = [11, 72, 92, 5]
        elif c == 3: # 7-8 место
            if table == "setka_32_2":
                ml = [13, 109, 114, 4]  
            elif table == "setka_16_full":
                ml = [10, 39, 42, 2] 
            elif table == "setka_16_2":
                ml = [9, 66, 70, 3] 
            elif table == "setka_8_full":
                ml = [8, 28, 31, 2] 
            elif table == "setka_8_2": 
                ml = [8, 31, 35, 3]
            else:
                ml = [11, 94, 95, 1]
        elif c == 4: # 9-10 место
            if table == "setka_32_2":
                ml = [5, 113, 118, 4]  
            elif table == "setka_16_full":
                ml = [10, 47, 53, 5] 
            elif table == "setka_16_2":
                ml = [5, 63, 70, 6] 
            else:
                ml = [11, 99, 133, 5]
        elif c == 5: # 11-12 место
            if table == "setka_32_2":
                ml = [11, 119, 124, 4]  
            elif table == "setka_16_full":
                ml = [10, 55, 58, 2] 
            elif table == "setka_16_2":
                ml = [9, 72, 76, 3] 
            else:
                ml = [11, 152, 163, 10]
        elif c == 6: # 13-14 место
            if table == "setka_32_2":
                ml = [5, 125, 130, 4]  
            elif table == "setka_16_full":
                ml = [10, 60, 64, 3]
            elif table == "setka_16_2":
                ml = [5, 75, 82, 6]  
            else:
                ml = [11, 167, 172, 4]
        elif c == 7: # 15-16 место
            if table == "setka_32_2":
                ml = [11, 127, 132, 4] 
            elif table == "setka_16_full":
                ml = [10, 65, 68, 2]
            elif table == "setka_16_2":
                ml = [9, 78, 82, 3]  
            else:
                ml = [9, 173, 178, 4]
        elif c == 8: # 17-18 место
            if table == "setka_32_2":
                ml = [7, 145, 153, 7]  
            else:
                ml = [11, 180, 186, 5]
        elif c == 9: # 19-20 место
            if table == "setka_32_2":
                ml = [11, 154, 160, 5]  
            else:
                ml = [7, 184, 191, 6]
        elif c == 10: # 21-22 место
            if table == "setka_32_2":
                ml = [7, 161, 166, 4]  
            else:
                ml = [11, 192, 198, 5]
        elif c == 11: # 23-24 место
            if table == "setka_32_2":
                ml = [11, 169, 175, 5]  
            else:
                ml = [5, 198, 203, 4]
        elif c == 12: # 25-26 место
            if table == "setka_32_2":
                ml = [7, 175, 183, 7] 
            else:
                ml = [11, 200, 205, 4]
        elif c == 13: # 27-28 место
            ml = [11, 184, 190, 5]  
        elif c == 14: # 29-30 место 
            ml = [7, 193, 198, 4] 
        elif c == 15: # 31-32 место
            ml = [11, 199, 205, 5]
            
        for i in range(ml[1], ml[2], ml[3]):
            data[i][ml[0]] = str(first_mesto + b) + " Место"
            fn = (('TEXTCOLOR', (ml[0], i), (ml[0], i), colors.red))
            style_color.append(fn)
            fn =  ('ALIGN', (ml[0], i), (ml[0], i), 'CENTER')
            style_color.append(fn)
            b += 1    
    return style_color   


def last_competition():
    """заполняе меню -последние- прошедшими соревнованиями 5 штук"""
    title = Title.select().order_by(Title.data_start.desc())
    i = 0
    for t in title:
        full_name = t.full_name_comp
        if i > 5:
            break
        if i == 0: 
            if full_name != "":
                my_win.first_comp_Action.setText(full_name)
            else:
                my_win.first_comp_Action.setText("Пусто")
        elif i == 1: 
            if full_name != "":
                my_win.second_comp_Action.setText(full_name)
            else:
                my_win.second_comp_Action.setText("Пусто")
        elif i == 2: 
            if full_name != "":
                my_win.third_comp_Action.setText(full_name)
            else:
                my_win.third_comp_Action.setText("Пусто")
        elif i == 3: 
            if full_name != "":
                my_win.fourth_comp_Action.setText(full_name)
            else:
                my_win.fourth_comp_Action.setText("Пусто")
        elif i == 4: 
            if full_name != "":
                my_win.fifth_comp_Action.setText(full_name)
            else:
                my_win.fifth_comp_Action.setText("Пусто")
        i += 1


def tours_list(cp):
    """туры таблиц по кругу в зависимости от кол-во участников (-cp- + 3) кол-во участников"""
    tour_list = []
    tr = [[['1-3'], ['1-2'], ['2-3']],
          [['1-3', '2-4'], ['1-2', '3-4'], ['2-3', '1-4']],
          [['2-4', '1-5'], ['1-4', '3-5'], ['1-3', '2-5'], ['2-3', '4-5'], ['1-2', '3-4']],
          [['2-4', '1-5', '3-6'], ['1-4', '2-6', '3-5'], ['1-3', '2-5', '4-6'], ['2-3', '1-6', '4-5'],
            ['1-2', '3-4', '5-6']],
          [['2-6', '3-5', '1-7'], ['2-5', '1-6', '4-7'], ['1-5', '4-6', '3-7'], ['4-5', '2-7', '3-6'],
            ['1-3', '2-4', '5-7'], ['1-4', '2-3', '6-7'], ['1-2', '3-4', '5-6']],
          [['2-6', '3-5', '1-7', '4-8'], ['2-5', '1-6', '3-8', '4-7'], ['1-5', '2-8', '4-6', '3-7'],
            ['1-8', '4-5', '2-7', '3-6'], ['1-3', '2-4', '5-7', '6-8'], ['1-4', '2-3', '6-7', '5-8'],
            ['1-2', '3-4', '5-6', '7-8']],
          [['1-9', '2-8', '3-7', '4-6'], ['5-9', '1-8', '2-7', '3-6'], ['4-9', '5-8', '1-7', '2-6'],
            ['3-9', '4-8', '5-7', '1-6'], ['2-4', '1-5', '3-8', '7-9'], ['4-1', '5-3', '9-2', '8-6'],
            ['1-3', '2-5', '4-7', '6-9'], ['3-2', '5-4', '8-9', '7-6'], ['1-2', '3-4', '5-6', '7-8']],
          [['1-9', '2-8', '3-7', '4-6', '5-10'], ['5-9', '1-8', '2-7', '3-6', '4-10'], ['4-9', '5-8', '1-7', '2-6', '3-10'],
            ['3-9', '4-8', '5-7', '1-6', '2-10'], ['2-4', '1-5', '3-8', '7-9', '6-10'], ['4-1', '5-3', '9-2', '8-6', '7-10'],
            ['1-3', '2-5', '4-7', '6-9', '8-10'], ['3-2', '5-4', '8-9', '7-6', '1-10'], ['1-2', '3-4', '5-6', '7-8', '9-10']],
          [['1-11', '2-10', '3-9', '4-8', '5-7'], ['6-11', '1-10', '2-9', '3-8', '4-7'], ['5-11', '6-10', '1-9', '2-8', '3-7'],
            ['4-11', '5-10', '6-9', '1-8', '2-7'], ['3-11', '4-10', '5-9', '6-8', '1-7'], ['2-11', '3-10', '4-9', '5-8', '6-7'],
            ['2-4', '1-5', '3-6', '7-10', '9-11'], ['1-4', '2-6', '3-5', '8-10', '7-11'], ['1-3', '2-5', '4-6', '7-9', '8-11'],
            ['2-3', '1-6', '4-5', '8-9', '10-11'], ['1-2', '3-4', '5-6', '7-8', '9-10']],
          [['1-11', '2-10', '3-9', '4-8', '5-7', '6-12'], ['6-11', '1-10', '2-9', '3-8', '4-7', '5-12'],
            ['5-11', '6-10', '1-9', '2-8', '3-7', '4-12'], ['4-11', '5-10', '6-9', '1-8', '2-7', '3-12'],
            ['3-11', '4-10', '5-9', '6-8', '1-7', '2-12'], ['2-11', '3-10', '4-9', '5-8', '6-7', '1-12'],
            ['2-4', '1-5', '3-6', '7-10', '9-11', '8-12'], ['1-4', '2-6', '3-5', '8-10', '7-11', '9-12'],
            ['1-3', '2-5', '4-6', '7-9', '8-11', '10-12'], ['2-3', '1-6', '4-5', '8-9', '10-11', '7-12'],
            ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12']],
          [['1-13', '2-12', '3-11', '4-10', '5-9', '6-8'], ['7-13', '1-12', '2-11', '3-10', '4-9', '5-8'],
            ['6-13', '7-12', '1-11', '2-10', '3-9', '4-8'], ['5-13', '6-12', '7-11', '1-10', '2-9', '3-8'],
            ['4-13', '5-12', '6-11', '7-10', '1-9', '2-8'], ['3-13', '4-12', '5-11', '6-10', '7-9', '1-8'],
            ['1-7', '2-6', '3-5', '4-11', '9-13', '10-12'], ['1-6', '2-5', '4-7', '3-12', '8-11', '10-13'],
            ['1-4', '2-7', '3-6', '5-10', '8-13', '9-12'], ['1-5', '3-7', '4-6', '2-13', '8-12', '9-11'],
            ['1-3', '2-4', '5-7', '6-9', '8-10', '11-13'], ['2-3', '4-5', '6-7', '8-9', '10-11', '12-13'],
            ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12']],
          [['1-13', '2-12', '3-11', '4-10', '5-9', '6-8', '7-14'], ['7-13', '1-12', '2-11', '3-10', '4-9', '5-8', '6-14'],
          ['6-13', '7-12', '1-11', '2-10', '3-9', '4-8', '5-14'], ['5-13',
                                                                   '6-12', '7-11', '1-10', '2-9', '3-8', '4-14'],
          ['4-13', '5-12', '6-11', '7-10', '1-9', '2-8', '3-14'], ['3-13',
                                                                   '4-12', '5-11', '6-10', '7-9', '1-8', '2-14'],
          ['1-7', '2-6', '3-5', '4-11', '9-13', '10-12', '8-14'], ['1-6',
                                                                   '2-5', '4-7', '3-12', '8-11', '10-13', '9-14'],
          ['1-4', '2-7', '3-6', '5-10', '8-13', '9-12', '11-14'], ['1-5',
                                                                   '3-7', '4-6', '2-13', '8-12', '9-11', '10-14'],
          ['1-3', '2-4', '5-7', '6-9', '8-10', '11-13', '12-14'], ['2-3',
                                                                   '4-5', '6-7', '8-9', '10-11', '12-13', '1-14'],
          ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14']],
          [['1-15', '2-14', '3-13', '4-12', '5-11', '6-10', '7-9'], ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9'],
          ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9'], ['7-15',
                                                                    '8-14', '1-13', '2-12', '3-11', '4-10', '5-9'],
          ['6-15', '7-14', '8-13', '1-12', '2-11', '3-10', '4-9'], ['5-15',
                                                                    '6-14', '7-13', '8-12', '1-11', '2-10', '3-9'],
          ['4-15', '5-14', '6-13', '7-12', '8-11', '1-10', '2-9'], ['3-15',
                                                                    '4-14', '5-13', '6-12', '7-11', '8-10', '1-9'],
          ['2-15', '3-14', '4-13', '5-12', '6-11', '7-10', '8-9'], ['1-7',
                                                                    '2-6', '3-5', '4-8', '9-13', '12-14', '11-15'],
          ['1-6', '2-5', '3-8', '4-7', '9-14', '10-13', '12-15'], ['1-5',
                                                                   '2-8', '3-7', '4-6', '9-15', '10-14', '11-13'],
          ['1-4', '2-7', '3-6', '5-8', '9-12', '10-15', '11-14'], ['1-3',
                                                                   '2-4', '5-7', '6-8', '9-11', '10-12', '13-15'],
          ['1-8', '2-3', '4-5', '6-7', '10-11', '12-13', '14-15'], ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14']],
          [['1-15', '2-14', '3-13', '4-12', '5-11', '6-10', '7-9', '8-16'],
          ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9', '7-16'],
          ['7-15', '8-14', '1-13', '2-12', '3-11', '4-10', '5-9', '6-16'],
          ['6-15', '7-14', '8-13', '1-12', '2-11', '3-10', '4-9', '5-16'],
          ['5-15', '6-14', '7-13', '8-12', '1-11', '2-10', '3-9', '4-16'],
          ['4-15', '5-14', '6-13', '7-12', '8-11', '1-10', '2-9', '3-16'],
          ['3-15', '4-14', '5-13', '6-12', '7-11', '8-10', '1-9', '2-16'],
          ['2-15', '3-14', '4-13', '5-12', '6-11', '7-10', '8-9', '1-16'],
          ['1-7', '2-6', '3-5', '4-8', '9-13', '12-14', '11-15', '10-16'],
          ['1-6', '2-5', '3-8', '4-7', '9-14', '10-13', '12-15', '11-16'],
          ['1-5', '2-8', '3-7', '4-6', '9-15', '10-14', '11-13', '12-16'],
          ['1-4', '2-7', '3-6', '5-8', '9-12', '10-15', '11-14', '13-16'],
          ['1-3', '2-4', '5-7', '6-8', '9-11', '10-12', '13-15', '14-16'],
          ['1-8', '2-3', '4-5', '6-7', '10-11', '12-13', '14-15', '9-16'],
          ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14', '15-16']]]

    tour_list = tr[cp]
    return tour_list


# ====
# def load_playing_game_in_table_for_semifinal(stage):
#     """растановка в полуфинале игроков со встречей сыгранной в группе"""
#     id_player_exit_out_gr = [] # список ид игроков попадающих в финал из группы в порядке занятых место по возрастанию
#     posev_player_exit_out_gr = []
#     player_exit = []    
#     mesto_rank = 1 # начальное место с которого вышли в финал
#     system = System.select().where(System.title_id == title_id())
#     choice = Choice.select().where(Choice.title_id == title_id())
#     results = Result.select().where(Result.title_id == title_id())
#     sys = system.select().where(System.stage == "Предварительный").get()
#     sys_semifin = system.select().where(System.stage == stage).get()
#     kol_gr = sys.total_group
#     if stage == "1-й полуфинал":
#         mesto_rank = 1
#     else:
#         sys_fin_last = system.select().where(System.stage == stage).get()
#         mesto_rank = sys_fin_last.mesta_exit + 1 # место, попадающих в финал из группы начало
#     how_many_mest_exit = sys_semifin.mesta_exit # количество мест попадающих из предварительного этапа
#     for i in range(1, kol_gr + 1): # цикл по группам
#         posev_player_exit_out_gr.clear()
#         id_player_exit_out_gr.clear()
#         choice_group = choice.select().where(Choice.group == f"{i} группа") 
#         kol_player = len(choice_group) # число участников в группе
#         if mesto_rank + how_many_mest_exit <= kol_player:
#             mesto_rank_end = mesto_rank + how_many_mest_exit
#         else:
#             mesto_rank_end = kol_player + 1
#         n = 0
#         for k in range(mesto_rank, mesto_rank_end): # цикл в группе начиная с места с которого выходят в финал (зависит скольк игроков выходят из группы)
#             ch_mesto_exit = choice_group.select().where(Choice.mesto_group == k).get()
#             pl_id = ch_mesto_exit.player_choice_id # id игрока, занявшего данное место
#             pl_posev = ch_mesto_exit.posev_group
#             id_player_exit_out_gr.append(pl_id)
#             posev_player_exit_out_gr.append(pl_posev) # номера игроков в группе вышедших в финал
#             n += 1

#         posev_pl = []
#         temp = []
#         posev_id_pl = []
#         all_posev_id_pl = []
#         if n > 1:
#             # получаем все варианты встреч, сыгранных в группе игроков которые попали в финал
#             for i in combinations(posev_player_exit_out_gr, 2):
#                 posev_player_exit = list(i)
#                 for v in posev_player_exit:
#                     ind = posev_player_exit_out_gr.index(v)
#                     id_player = id_player_exit_out_gr[ind]
#                     temp.append(id_player)
#                     posev_id_pl = temp.copy()
#                 temp.clear()
#                 posev_pl.append(posev_player_exit)
#                 all_posev_id_pl.append(posev_id_pl)

#             result_pre = results.select().where(Result.system_stage == "Предварительный") # изменить откуда выходят из группы или пф
#             for d in range(0, len(posev_pl)):
#                 posev_exit = posev_pl[d]
#                 id_player_exit = all_posev_id_pl[d]
#                 if posev_exit[0] > posev_exit[1]: # если спортсмены заняли места не по расстановки в табл меняем на номера встречи в правильном порядке по возр
#                     id_player_exit.reverse()
                    
#                 player_exit.clear()
#                 posev_exit.clear()
#                 for l in id_player_exit:
#                     players = Player.select().where(Player.id == l).get()
#                     family_city = players.full_name
#                     player_exit.append(family_city)  
#                     # номер ид в таблице -Result- встречи игроков, попавших в полуфинал идущих по расстоновке в таблице   
#                 result_gr = result_pre.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get() 

#                 result_pre_fin = results.select().where(Result.system_stage == stage)
#                 result_semifin_player1 = result_pre_fin.select().where(Result.player1.in_(player_exit))
#                 result_semifin = result_semifin_player1.select().where(Result.player2.in_(player_exit)).get()

#                 with db:
#                     result_semifin.winner = result_gr.winner
#                     result_semifin.points_win = result_gr.points_win
#                     result_semifin.score_in_game = result_gr.score_in_game
#                     result_semifin.score_win = result_gr.score_win
#                     result_semifin.loser = result_gr.loser
#                     result_semifin.points_loser = result_gr.points_loser
#                     result_semifin.score_loser = result_gr.score_loser
#                     result_semifin.save()
#     pv = sys_semifin.page_vid
#     my_win.tabWidget.setCurrentIndex(4)
#     table_made(pv, stage)


# ====





def load_playing_game_in_table_for_semifinal(stage):
    """растановка в полуфинале игроков со встречей сыгранной в группе"""
    id_player_exit_out_gr = [] # список ид игроков попадающих в финал из группы в порядке занятых место по возрастанию
    posev_player_exit_out_gr = []
    player_exit = []    
    mesto_rank = 1 # начальное место с которого вышли в финал
    system = System.select().where(System.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    results = Result.select().where(Result.title_id == title_id())
    sys = system.select().where(System.stage == "Предварительный").get()
    sys_semifin = system.select().where(System.stage == stage).get()
    kol_gr = sys.total_group
    if stage == "1-й полуфинал":
        mesto_rank = 1
    else:
        sys_fin_last = system.select().where(System.stage == stage).get()
        mesto_rank = sys_fin_last.mesta_exit + 1 # место, попадающих в финал из группы начало
    how_many_mest_exit = sys_semifin.mesta_exit # количество мест попадающих из предварительного этапа
    for i in range(1, kol_gr + 1): # цикл по группам
        posev_player_exit_out_gr.clear()
        id_player_exit_out_gr.clear()
        choice_group = choice.select().where(Choice.group == f"{i} группа") 
        kol_player = len(choice_group) # число участников в группе
        if mesto_rank + how_many_mest_exit <= kol_player:
            mesto_rank_end = mesto_rank + how_many_mest_exit
        else:
            mesto_rank_end = kol_player + 1
        n = 0
        for k in range(mesto_rank, mesto_rank_end): # цикл в группе начиная с места с которого выходят в финал (зависит скольк игроков выходят из группы)
            ch_mesto_exit = choice_group.select().where(Choice.mesto_group == k).get()
            pl_id = ch_mesto_exit.player_choice_id # id игрока, занявшего данное место
            pl_posev = ch_mesto_exit.posev_group
            id_player_exit_out_gr.append(pl_id)
            posev_player_exit_out_gr.append(pl_posev) # номера игроков в группе вышедших в финал
            n += 1

        posev_pl = []
        temp = []
        posev_id_pl = []
        all_posev_id_pl = []
        if n > 1:
            # получаем все варианты встреч, сыгранных в группе игроков которые попали в полуфинал
            for i in combinations(posev_player_exit_out_gr, 2):
                posev_player_exit = list(i)
                for v in posev_player_exit:
                    ind = posev_player_exit_out_gr.index(v)
                    id_player = id_player_exit_out_gr[ind]
                    temp.append(id_player)
                    posev_id_pl = temp.copy()
                temp.clear()
                posev_pl.append(posev_player_exit)
                all_posev_id_pl.append(posev_id_pl)

            result_pre = results.select().where(Result.system_stage == "Предварительный") # изменить откуда выходят из группы или пф
            for d in range(0, len(posev_pl)):
                posev_exit = posev_pl[d]
                id_player_exit = all_posev_id_pl[d]
                if posev_exit[0] > posev_exit[1]: # если спортсмены заняли места не по расстановки в табл меняем на номера встречи в правильном порядке по возр
                    id_player_exit.reverse()
                    
                player_exit.clear()
                posev_exit.clear()
                for l in id_player_exit:
                    players = Player.select().where(Player.id == l).get()
                    family_city = players.full_name
                    player_exit.append(family_city)  
                    # номер ид в таблице -Result- встречи игроков, попавших в полуфинал идущих по расстоновке в таблице   
                result_gr = result_pre.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get() 

                result_pre_fin = results.select().where(Result.system_stage == stage)
                result_semifin_player1 = result_pre_fin.select().where(Result.player1.in_(player_exit))
                result_semifin = result_semifin_player1.select().where(Result.player2.in_(player_exit)).get()

                with db:
                    result_semifin.winner = result_gr.winner
                    result_semifin.points_win = result_gr.points_win
                    result_semifin.score_in_game = result_gr.score_in_game
                    result_semifin.score_win = result_gr.score_win
                    result_semifin.loser = result_gr.loser
                    result_semifin.points_loser = result_gr.points_loser
                    result_semifin.score_loser = result_gr.score_loser
                    result_semifin.save()
    pv = sys_semifin.page_vid
    my_win.tabWidget.setCurrentIndex(4)
    # table_made(pv, stage)


def load_playing_game_in_table_for_final(fin):
    """растановка в финале игроков со встречей сыгранной в группе"""
    id_player_exit_out_gr = [] # список ид игроков попадающих в финал из группы в порядке занятых место по возрастанию
    posev_player_exit_out_gr = []
    player_exit = []
    mesto_rank = 1 # начальное место с которого вышли в финал
    system = System.select().where(System.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    results = Result.select().where(Result.title_id == title_id())
    sys = system.select().where(System.stage == "Предварительный").get()
    sys_fin = system.select().where(System.stage == fin).get()
    kol_gr = sys.total_group
    if fin == "1-й финал":
        mesto_rank = 1
    else:
        sum_player = []
        etap_exit = sys_fin.stage_exit
        for m in system:
            if m.stage == fin:
                break
            else:
                if m.stage_exit == etap_exit:
                    total_player_exit = m.mesta_exit
                    sum_player.append(total_player_exit)
        sum_player_exit = sum(sum_player) # сумма игроков вышедших в финал   
        mesto_rank = sum_player_exit + 1
    how_many_mest_exit = sys_fin.mesta_exit # количество мест попадающих из предварительного этапа
    for i in range(1, kol_gr + 1): # цикл по группам
        posev_player_exit_out_gr.clear()
        id_player_exit_out_gr.clear()
        choice_group = choice.select().where(Choice.group == f"{i} группа") 
        kol_player = len(choice_group) # число участников в группе
        if mesto_rank + how_many_mest_exit <= kol_player:
            mesto_rank_end = mesto_rank + how_many_mest_exit
        else:
            mesto_rank_end = kol_player + 1
        n = 0
        for k in range(mesto_rank, mesto_rank_end): # цикл в группе начиная с места с которого выходят в финал (зависит скольк игроков выходят из группы)
            ch_mesto_exit = choice_group.select().where(Choice.mesto_group == k).get()
            pl_id = ch_mesto_exit.player_choice_id # id игрока, занявшего данное место
            pl_posev = ch_mesto_exit.posev_group
            id_player_exit_out_gr.append(pl_id)
            posev_player_exit_out_gr.append(pl_posev) # номера игроков в группе вышедших в финал
            n += 1

        posev_pl = []
        temp = []
        posev_id_pl = []
        all_posev_id_pl = []
        if n > 1:
            # получаем все варианты встреч, сыгранных в группе игроков которые попали в финал
            for i in combinations(posev_player_exit_out_gr, 2):
                posev_player_exit = list(i)
                for v in posev_player_exit:
                    ind = posev_player_exit_out_gr.index(v)
                    id_player = id_player_exit_out_gr[ind]
                    temp.append(id_player)
                    posev_id_pl = temp.copy()
                temp.clear()
                posev_pl.append(posev_player_exit)
                all_posev_id_pl.append(posev_id_pl)

            result_pre = results.select().where(Result.system_stage == "Предварительный") # изменить откуда выходят из группы или пф
            for d in range(0, len(posev_pl)):
                posev_exit = posev_pl[d]
                id_player_exit = all_posev_id_pl[d]
                if posev_exit[0] > posev_exit[1]: # если спортсмены заняли места не по расстановки в табл меняем на номера встречи в правильном порядке по возр
                    id_player_exit.reverse()
                    
                player_exit.clear()
                posev_exit.clear()
                for l in id_player_exit:
                    players = Player.select().where(Player.id == l).get()
                    family_city = players.full_name
                    player_exit.append(family_city)  
                    # номер ид в таблице -Result- встречи игроков, попавших в финал идущих по расстоновке в таблице   
                result_gr = result_pre.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get() 

                result_pre_fin = results.select().where(Result.number_group == fin)

                result_fin_1 = result_pre_fin.select().where((Result.player1 == player_exit[0]))
                result_fin = result_fin_1.select().where(Result.player2 == player_exit[1])
                count = len(result_fin)

                if count != 1:
                    result_fin = result_pre_fin.select().where((Result.player1 == player_exit[1]) & (Result.player2 == player_exit[0])).get()
                else:
                    result_fin = result_pre_fin.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get()

                with db:
                    result_fin.winner = result_gr.winner
                    result_fin.points_win = result_gr.points_win
                    result_fin.score_in_game = result_gr.score_in_game
                    result_fin.score_win = result_gr.score_win
                    result_fin.loser = result_gr.loser
                    result_fin.points_loser = result_gr.points_loser
                    result_fin.score_loser = result_gr.score_loser
                    result_fin.save()
    stage = fin
    pv = sys_fin.page_vid
    table_made(pv, stage)


def made_file_excel_for_rejting():
    """создание файла Excel для обсчета рейтинга"""
    result = Result.select().where(Result.title_id == title_id())
    player_result = result.select().where(Result.points_loser != 0).order_by(Result.winner)
    book = op.Workbook()
    worksheet = book.active
    names_headers = ["Победитель", "Проигравший", "Счет"]
    for m in range(1, 4):
        c =  worksheet.cell(row = 1, column = m)
        c.value = names_headers[m - 1]
    k = 2
    for l in player_result:
        pl_win = l.winner
        pl_los = l.loser
        score = l.score_in_game
        c1 = worksheet.cell(row = k, column = 1)
        c1.value = pl_win
        c2 = worksheet.cell(row = k, column = 2)
        c2.value = pl_los
        c3 = worksheet.cell(row = k, column = 3)
        c3.value = score
        k += 1

    t_id = Title.get(Title.id == title_id())
    short_name = t_id.short_name_comp 
    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['b'].width = 40
    book.save(f"{short_name}_report.xlsx")
    

def button_move_enabled():
    """включает или выключает кнопки перемещения по таблице в зависимости от выделенной строки"""
    count = my_win.tableWidget.rowCount()
    row = my_win.tableWidget.currentRow()
    if row == 0:
        my_win.Button_down.setEnabled(True)
        my_win.Button_up.setEnabled(False)
    elif row == count - 1:
        my_win.Button_up.setEnabled(True)
        my_win.Button_down.setEnabled(False)
    else:
        my_win.Button_up.setEnabled(True)
        my_win.Button_down.setEnabled(True)


def move_row_in_tablewidget():
    """перемещяет выделенную строку по таблице вверх или вниз"""
    sender = my_win.sender()
    row_count = my_win.tableWidget.rowCount()
    row_cur = my_win.tableWidget.currentRow()
    if row_cur == 1:
        my_win.Button_up.setEnabled(False)
    if row_cur == row_count:
        my_win.Button_down.setEnabled(False)
    item_cur = my_win.tableWidget.item(row_cur, 1).text()
    item_cur_name = my_win.tableWidget.item(row_cur, 2).text()
    if sender == my_win.Button_down:
        item_tmp = my_win.tableWidget.item(row_cur + 1, 1).text()
        item_temp = my_win.tableWidget.item(row_cur + 1, 2).text()
        my_win.tableWidget.selectRow(row_cur + 1)
        my_win.tableWidget.setItem(row_cur + 1, 1, QTableWidgetItem(str(item_cur)))
        my_win.tableWidget.setItem(row_cur, 1, QTableWidgetItem(str(item_tmp)))
        my_win.tableWidget.setItem(row_cur + 1, 2, QTableWidgetItem(str(item_cur_name)))
        my_win.tableWidget.setItem(row_cur, 2, QTableWidgetItem(str(item_temp)))
    else:
        item_tmp = my_win.tableWidget.item(row_cur - 1, 1).text()
        item_temp = my_win.tableWidget.item(row_cur - 1, 2).text()
        my_win.tableWidget.selectRow(row_cur - 1)
        my_win.tableWidget.setItem(row_cur - 1, 1, QTableWidgetItem(str(item_cur)))
        my_win.tableWidget.setItem(row_cur, 1, QTableWidgetItem(str(item_tmp)))
        my_win.tableWidget.setItem(row_cur - 1, 2, QTableWidgetItem(str(item_cur_name)))
        my_win.tableWidget.setItem(row_cur, 2, QTableWidgetItem(str(item_temp)))


def made_list_referee():
    """создание списка судейской коллегии"""
    my_win.radioButton_GSK.setChecked(True)
    my_win.Button_made_page_pdf.setEnabled(True)
    num_columns = [0, 1, 2, 3]
    number_of_referee, ok = QInputDialog.getInt(my_win, "Главная судейская коллегия", "Укажите число судей списка\n главной cудейской коллегии.", 4, 3, 10)
  
    if ok:
        title = Title.get(Title.id == title_id())
        referee = title.referee
        kat_referee = title.kat_ref
        secretary = title.secretary
        kat_secretary = title.kat_sec
        list_referee = [referee, secretary]
        list_kategory = [kat_referee, kat_secretary]

        ref_list = Referee.select()
        referee_selected = ref_list.dicts().execute()
        row_count = len(referee_selected)  # кол-во строк в таблице

        title_competition = Title.get(Title.id == title_id())
        referee = title_competition.referee
        kat_ref = title_competition.kat_ref
        secretary = title_competition.secretary
        kat_sec = title_competition.kat_sec

        data = [["1", "Главный судья", referee, kat_ref], ["2", "Главный секретарь", secretary, kat_sec]]
        data_extend = ["", "", "", ""]
        data_tmp = []

        for k in range(number_of_referee - 2):
            data_extend[0] = str(k + 3)
            data_tmp = data_extend.copy()      
            data.append(data_tmp.copy())
            data_tmp.clear()

        model = MyTableModel(data)
        model.setHorizontalHeaderLabels(["№", "Должность", "Фамилия Имя Отчество/ Город", "Категория"])
        my_win.tableView.setModel(model)
        #======== ++++
        post_list = ["-выберите должность-", "Зам. Главного судьи", "Зам. Главного секретаря", "Ведущий судья"]
        combo = QComboBox()
        # lineEd = QLineEdit()
        combo.addItems(post_list)
        for i in range(2, number_of_referee):
            my_win.tableView.setIndexWidget(my_win.tableView.model().index(i, 1), combo) # вставляет в строку i, 1 столбец combobox
        #     my_win.tableView.setIndexWidget(my_win.tableView.model().index(i, 2), lineEd) # вставляет в строку i, 1 столбец combobox
        # ======= ++++++++
        delegate = LineDelegate(my_win.tableView)
        my_win.tableView.setItemDelegateForColumn(2, delegate)

        # delegate = ComboBoxDelegate(post_list)
        # my_win.tableView.setItemDelegateForColumn(1, delegate)
#  ======== +++++++ ==========
        # # delegate = ComboBoxDelegate(my_win.tableView)
        # # my_win.tableView.setIndexWidget(my_win.tableView.model().index(3, 1), delegate)
        # my_win.tableView.setItemDelegateForColumn(2, ComboBoxDelegate(my_win.tableView))
        # delegate = MyComboDelegate(my_win.view)
        # my_win.view.setItemDelegateForColumn(3, delegate)
        # my_win.view.show()

        # category_list = ["-выберите категорию-", "ССВК", "1-й кат.", "2-й кат."]
        # post_list = ["-выберите должность-", "Зам. Главного судьи", "Зам. Главного секретаря", "Ведущий судья"]


 
        # -  my_win.tableView.setModel(model)
       
        # color_delegate = MyColorDelegate(my_win.tableView)
        # delegate = MyComboDelegate(my_win.tableView)
        # delegate = ComboDelegate(my_win.tableView)
        # my_win.tableView.setItemDelegateForColumn(0, color_delegate)
        # my_win.tableView.setItemDelegateForColumn(2, delegate)
        my_win.tableView.resizeColumnsToContents() # растягивает по содержимому

        my_win.tableView.show()
  

    else:
        return
def made_list_GSK():
    """создание списка судейской коллегии"""
    # Dialog = QInputDialg()
    my_win.tableWidget_GSK.clear()
    my_win.tableWidget_GSK.show()
    # my_win.radioButton_GSK.setChecked(True)
    # my_win.Button_made_page_pdf.setEnabled(True)
    number_of_referee, ok = QInputDialog.getInt(my_win, "Главная судейская коллегия", "Укажите число судей списка\n главной cудейской коллегии.", 4, 3, 10)
    for l in range(0, number_of_referee):
        my_win.tableWidget_GSK.setItem(l, 0, QTableWidgetItem(str(l + 1)))
    if ok:
        title = Title.get(Title.id == title_id())
        referee = title.referee
        kat_referee = title.kat_ref
        secretary = title.secretary
        kat_secretary = title.kat_sec
        list_referee = [referee, secretary]
        list_kategory = [kat_referee, kat_secretary]

        my_win.tableWidget_GSK.setColumnCount(4) # устанавливает колво столбцов
        my_win.tableWidget_GSK.setRowCount(number_of_referee)
        column_label = ["№", "Должность", "Фамилия Имя Отчество/ Город", "Категория"]
        my_win.tableWidget_GSK.setColumnWidth(2, 10000)
        for i in range(0, 4):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
            my_win.tableWidget_GSK.showColumn(i)
            item = QtWidgets.QTableWidgetItem()
            brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
            brush.setStyle(QtCore.Qt.SolidPattern)
            item.setForeground(brush)
            my_win.tableWidget_GSK.setHorizontalHeaderItem(i, item)
        my_win.tableWidget_GSK.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
        referee_list = []
        post_list = ["", "ССВК", "1-й кат.", "2-й кат."]
        category_list = ["","Зам. Главного судьи", "Зам. Главного секретаря", "Ведущий судья"]
        my_win.tableWidget_GSK.setItem(0, 1, QTableWidgetItem("Гл. судья"))
        my_win.tableWidget_GSK.setItem(1, 1, QTableWidgetItem("Гл. секретарь"))
    else:
        return
    for k in range(0, 2):
        my_win.tableWidget_GSK.setItem(k, 2, QTableWidgetItem(str(list_referee[k])))
        my_win.tableWidget_GSK.setItem(k, 3, QTableWidgetItem(str(list_kategory[k])))
    for n in range(2, int(number_of_referee)): 
        comboBox_list_post = QComboBox()
        comboBox_list_category = QComboBox()  
        comboBox_family_city = QComboBox()
        referee_list = load_comboBox_referee()

        comboBox_family_city.setPlaceholderText("Введите фамилию судьи")
        comboBox_family_city.setCurrentIndex(-1)
        comboBox_family_city.setEditable(True)
        comboBox_list_category.addItems(category_list)
        comboBox_list_post.addItems(post_list) 
        comboBox_family_city.addItems(referee_list)

        my_win.tableWidget_GSK.setCellWidget(n, 1, comboBox_list_category)
        my_win.tableWidget_GSK.setCellWidget(n, 2, comboBox_family_city)
        my_win.tableWidget_GSK.setCellWidget(n, 3, comboBox_list_post)   

        # my_win.tableWidget_GSK.itemChanged.connect(change_on_comboBox_referee)
        comboBox_family_city.currentTextChanged.connect(change_on_comboBox_referee)   

def change_on_comboBox_referee(comboBox_family_city):
    """добавляет в базу данных судей если их там нет"""
    row_cur = my_win.tableWidget_GSK.currentRow()
    mark = comboBox_family_city.find("/") # если еще нет фамилии и города
    if mark != 0 and mark != -1:
        add_referee_to_db()
        family_referee = comboBox_family_city[:mark]
        family_referee = family_referee.title()
        referees = Referee.select().where(Referee.family == family_referee)
        if len(referees) > 0:
            for ref in referees:
                kategor = ref.category
                kat = my_win.tableWidget_GSK.cellWidget(row_cur, 3)
                kat.setCurrentText(kategor)
                return
    else:
        kat = my_win.tableWidget_GSK.cellWidget(row_cur, 3)
        kat.setCurrentText("")


def add_referee_to_db():
    """добавляет в базу данных новых судей"""
    sender = my_win.sender()
    # count = my_win.tableWidget.rowCount()
    if sender == my_win.comboBox_kategor_ref:
        kat = my_win.comboBox_kategor_ref.currentText()
        item = my_win.comboBox_referee.currentText()
        whrite_referee_to_db(kat, item, k=0)
    elif sender == my_win.comboBox_kategor_sec:
        kat = my_win.comboBox_kategor_sec.currentText()
        item = my_win.comboBox_secretary.currentText()
        whrite_referee_to_db(kat, item, k=0)
    else:
        for k in range(2, count):
            item = my_win.tableWidget.cellWidget(k, 2).currentText()
            kat = my_win.tableWidget.cellWidget(k, 3).currentText()
            whrite_referee_to_db(kat, item, k)


def whrite_referee_to_db(kat, item, k=0):
    """запись рефери в db"""
    sender = my_win.sender()
    if kat != "":
        mark = item.find("/")
        family_referee = item[:mark]
        family_referee = family_referee.title()
        city_referee = item[mark + 2:]
        city_referee = city_referee.title()
        full_referee = f"{family_referee}/ {city_referee}"
        if sender == my_win.Button_made_page_pdf:
            f_referee = my_win.tableWidget.cellWidget(k, 2)
            f_referee.setCurrentText(full_referee)
        referees = Referee.select().where(Referee.family == family_referee)
        if len(referees) == 0:
            with db:
                ref = Referee(family=family_referee, city=city_referee, category=kat).save()


def view_all_page_pdf():    
    """просмотр все страниц соревнования pdf"""
    title = Title.get(Title.id == title_id())
    pdf_files_list = []
    rus_name_list = []
    stage_dict = {"table_group.pdf": "Предварительный",
                   "player_list.pdf": "Список участников",
                   "winners_list.pdf": "Список победителей и призеров",
                   "player_list_alf.pdf": "Список участников по алф",
                   "title.pdf": "Титульный лист",
                   "referee_list.pdf": "ГСК",
                   "regions_list.pdf": "Список субъктов РФ",
                   "1-semifinal.pdf": "1-й полуфинал",
                   "2-semifinal.pdf": "2-й полуфинал",
                   "1-final.pdf": "1-й финал",
                   "2-final.pdf": "2-й финал",
                   "3-final.pdf": "3-й финал",
                   "4-final.pdf": "4-й финал",
                   "5-final.pdf": "5-й финал",
                   "6-final.pdf": "6-й финал",
                   "7-final.pdf": "7-й финал",
                   "8-final.pdf": "8-й финал",
                   "one_table.pdf": "одна таблица"}
    short_name = title.short_name_comp
    count_mark = len(short_name)
    all_pdf_files_list = os.listdir("table_pdf")
    for name_files in all_pdf_files_list:
        text = name_files.find(short_name)
        text_stage = name_files[count_mark + 1:]
        if text == 0 and text_stage != "player_list_payment.pdf":
            pdf_files_list.append(name_files)
            for latin_name in stage_dict.keys():
                if text_stage == latin_name:
                    rus_name = stage_dict[text_stage]
                    rus_name_list.append(rus_name)
                    break

    row = len(pdf_files_list)
    my_win.tableWidget.setColumnCount(3) # устанавливает колво столбцов
    my_win.tableWidget.setRowCount(row)
    column_label = ["№", "Файлы", "Этапы"]
    for c in range(0, 3):
        my_win.tableWidget.showColumn(c)

    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
    my_win.tableWidget.setDragDropOverwriteMode(True)
    my_win.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
    my_win.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
    my_win.tableWidget.show()
    row_count = 0
    for item in pdf_files_list:
        item_name = rus_name_list[row_count]
        my_win.tableWidget.setItem(row_count, 0, (QTableWidgetItem(str(row_count + 1))))
        my_win.tableWidget.setItem(row_count, 1, (QTableWidgetItem(str(item))))
        my_win.tableWidget.setItem(row_count, 2, (QTableWidgetItem(str(item_name))))
        row_count += 1
    my_win.Button_made_one_file_pdf.setEnabled(True)


def made_list_regions():
    """создание списка регионов"""
    my_win.radioButton_regions.setChecked(True)
    my_win.Button_made_page_pdf.setEnabled(True)
    my_win.tableWidget.clear()
    region_list = []
    regions = Player.select().where(Player.title_id == title_id())

    for k in regions:
        reg = k.region
        if reg not in region_list:
            region_list.append(reg)
    count = len(region_list)
    region_list.sort()
    n = 0
    for l in region_list:
        my_win.tableWidget.setItem(n, 0, QTableWidgetItem(str(n + 1)))
        my_win.tableWidget.setItem(n, 1, QTableWidgetItem(str(l)))
        my_win.tableWidget.setColumnCount(2) # устанавливает колво столбцов
        my_win.tableWidget.setRowCount(count)
        column_label = ["№", "Субъекты РФ"]
        my_win.tableWidget.setColumnWidth(2, 10000)
        for i in range(0, 2):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
            my_win.tableWidget.showColumn(i)
            item = QtWidgets.QTableWidgetItem()
            brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
            brush.setStyle(QtCore.Qt.SolidPattern)
            item.setForeground(brush)
            my_win.tableWidget.setHorizontalHeaderItem(i, item)
        n += 1
    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget


def made_list_players_on_alf():
    """создание списка по алфавиту"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.player)
    short_name = tit.short_name_comp
    gamer = tit.gamer
    count = len(player_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in player_list:
        n += 1
        p = l.player
        b = l.bday
        r = l.rank
        c = l.city
        g = l.region
        z = l.razryad
        coach_id = l.coach_id
        t = coach_id.coach
        m = l.mesto
        data = [n, p, b, r, c, g, z, t, m]

        elements.append(data)
    elements.insert(0, ["№", "Фамилия, Имя", "Дата рожд.", "Рейтинг", "Город", "Регион", "Разряд", "Тренер(ы)",
                        "Место"])
    t = Table(elements,
              colWidths=(0.6 * cm, 3.9 * cm, 1.7 * cm, 1.2 * cm, 2.5 * cm, 3.1 * cm, 1.2 * cm, 4.8 * cm, 1.0 * cm),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список участников. {gamer}', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_player_list_alf.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def made_list_winners():
    """создание списка победителей и призеров"""
    my_win.radioButton_winner.setChecked(True)
    my_win.Button_made_page_pdf.setEnabled(True)
    my_win.tableWidget.clear()
    players = Player.select().where(Player.title_id == title_id())
    winners = players.select().where(Player.mesto < 4).order_by(Player.mesto)
    count = len(winners)
    
    n = 0
    for l in winners:
        my_win.tableWidget.setColumnCount(8) # устанавливает колво столбцов
        my_win.tableWidget.setRowCount(count)       
        column_label = ["Место", "Фамилия, Имя", "Дата рождения", "Рейтинг", "Город", "Регион", "Разряд", "Тренеры"]
        coachs = Coach.select().where(Coach.id == l.coach_id).get()
        family_coach = coachs.coach
        my_win.tableWidget.setItem(n, 0, QTableWidgetItem(str(f"{l.mesto} место")))
        my_win.tableWidget.setItem(n, 1, QTableWidgetItem(str(l.player)))
        my_win.tableWidget.setItem(n, 2, QTableWidgetItem(str(l.bday)))
        my_win.tableWidget.setItem(n, 3, QTableWidgetItem(str(l.rank)))
        my_win.tableWidget.setItem(n, 4, QTableWidgetItem(str(l.city)))
        my_win.tableWidget.setItem(n, 5, QTableWidgetItem(str(l.region)))
        my_win.tableWidget.setItem(n, 6, QTableWidgetItem(str(l.razryad)))
        my_win.tableWidget.setItem(n, 7, QTableWidgetItem(str(family_coach)))

        for i in range(0, 8):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
            my_win.tableWidget.showColumn(i)
            item = QtWidgets.QTableWidgetItem()
            brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
            brush.setStyle(QtCore.Qt.SolidPattern)
            item.setForeground(brush)
            my_win.tableWidget.setHorizontalHeaderItem(i, item)
        n += 1
    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget


def made_pdf_list():
    """создание страниц PDF соревнования"""
    if my_win.radioButton_GSK.isChecked():
        add_referee_to_db()
        list_referee_pdf()
    elif my_win.radioButton_regions.isChecked():
        list_regions_pdf()
    elif my_win.radioButton_winner.isChecked():
        list_winners_pdf()
    my_win.Button_made_page_pdf.setEnabled(False)


def check_pay():
    """список для отметки оплаты взноса"""
    from reportlab.platypus import Table
    from sys import platform
    elements = []
    story = []
    tit = Title.get(Title.id == title_id())
    view_sort = ["По алфавиту", "По региону"]
    view_sort, ok = QInputDialog.getItem(
                            my_win, "Сортировка", "Выберите вид сортировки,\n просмотра списка участников.", view_sort, 0, False)
    if view_sort == "По алфавиту": 
        player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.player)  # сортировка по алф
    elif view_sort == "По региону":
        player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.region)  # сортировка по региону

    short_name = tit.short_name_comp
    gamer = tit.gamer
    count = len(player_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in player_list:
        n += 1
        p = l.player
        c = l.city
        g = l.region
        coach_id = l.coach_id
        t = coach_id.coach

        data = [n, p, c, g, t]
        elements.append(data)

    elements.insert(0, ["№", "Фамилия, Имя", "Город", "Регион", "Тренер(ы)"])
    t = Table(elements, colWidths=(0.7 * cm, 5.0 * cm, 3.5 * cm, 4.5 * cm, 5.9 * cm), rowHeights=(0.6 * cm), repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            # ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 10),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))

    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список участников. {gamer}', h3))
    story.append(t)
    doc = SimpleDocTemplate(f"{short_name}_player_list_payment.pdf", pagesize=A4, 
                            rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm) # название, вид страницы, размер полей
    view_file = f"{short_name}_player_list_payment.pdf"
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok)
    if platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    change_dir(catalog)


def referee():
    """добавление судей в базу"""
    msgBox = QMessageBox()
    sender = my_win.sender()
    if sender == my_win.comboBox_referee: # комбобокс глав судьи
        text = my_win.comboBox_referee.currentText()
        index = my_win.comboBox_referee.findText(text)
        if index != -1:
            my_win.comboBox_referee.setCurrentIndex(index)
            my_win.comboBox_referee.lineEdit().setSelection(len(text), len(my_win.comboBox_referee.currentText()))
            category = find_referee_in_db(text)
            my_win.comboBox_kategor_ref.setCurrentText(category)
    elif sender == my_win.comboBox_secretary: # комбобокс глав секретарь:
        text = my_win.comboBox_secretary.currentText()
        index = my_win.comboBox_secretary.findText(text)
        if index != -1:
            my_win.comboBox_secretary.setCurrentIndex(index)
            my_win.comboBox_secretary.lineEdit().setSelection(len(text), len(my_win.comboBox_secretary.currentText()))
            category = find_referee_in_db(text)
            my_win.comboBox_kategor_sec.setCurrentText(category)
    elif sender == my_win.tableWidget.comboBox_family_city: # комбобокс выбора судей гск:
        text = my_win.tableWidget.comboBox_family_city.currentText()
        index = my_win.tableWidget.comboBox_family_city.findText(text)
        if index != -1:
            my_win.tableWidget.comboBox_family_city.setCurrentIndex(index)
            my_win.tableWidget.comboBox_family_city.lineEdit().setSelection(len(text), len(my_win.tableWidget.comboBox_family_city.currentText()))
            # category = find_referee_in_db(text)
            # my_win.comboBox_kategor_sec.setCurrentText(category)




def find_referee_in_db(text):
    """ищет фамилию судьи в базе данных и возвращает судейскую категорию"""
    mark = text.find("/")
    fio = text[:mark]
    referee = Referee.select().where(Referee.family == fio).get()
    category = referee.category
    return category


def open_close_file(view_file):
# Введите имя файла для проверки
# Импортировать модуль os для проверки существования файла
    # import os
    # Функция Drfine проверяет, закрыт ли файл или нет
    # if view_file.closed == False:
    # # Распечатать сообщение об успешном завершении
    #     print("Файл открыт для чтения.")
    # else:
    # # Распечатать сообщение об ошибке
    #     print(" Файл закрыт.")

    # Взять имя файла для проверки
    # Проверить, существует
    if os.path.exists(view_file):
    # Открыть файл для чтения
        view_file = open(view_file, "r")
    # Вызвать функцию
        open_close_file(view_file)
    else:
    # Вывести сообщение, если файл не существует
        print("Файл не существует.")


# def proba_pdf():
    # """проба пдф"""

# import itertools
# from random import randint
# from statistics import mean

# from reportlab.lib.pagesizes import A4
# from reportlab.pdfgen import canvas


# def grouper(iterable, n):
#     args = [iter(iterable)] * n
#     return itertools.zip_longest(*args)


# def export_to_pdf(data):
#     c = canvas.Canvas("grid-students.pdf", pagesize=A4)
#     w, h = landscape(A4)
#     max_rows_per_page = 45
#     # Margin.
#     x_offset = 50
#     y_offset = 50
#     # Space between rows.
#     padding = 15

#     xlist = [x + x_offset for x in [0, 200, 250, 300, 350, 400, 480]]
#     ylist = [h - y_offset - i*padding for i in range(max_rows_per_page + 1)]

    # for rows in grouper(data, max_rows_per_page):
    #     rows = tuple(filter(bool, rows))
    #     c.grid(xlist, ylist[:len(rows) + 1])
    #     for y, row in zip(ylist[:-1], rows):
    #         for x, cell in zip(xlist, row):
    #             c.drawString(x + 2, y - padding + 3, str(cell))
    #     c.showPage()

#     c.save()


# data = [("NAME", "GR. 1", "GR. 2", "GR. 3", "AVG", "STATUS")]

# for i in range(1, 101):
#     exams = [randint(0, 10) for _ in range(3)]
#     avg = round(mean(exams), 2)
#     state = "Approved" if avg >= 4 else "Disapproved"
#     data.append((f"Student {i}", *exams, avg, state))

# export_to_pdf(data)
# . =====

    # styles = getSampleStyleSheet()
    # styleN = styles['Normal']
    # styleH = styles['Heading1']
    # story = []
    #     #add some flowables
    # story.append(Paragraph("This is a Heading", styleH))
    # story.append(Paragraph("This is a paragraph in <i>Normal</i> style.", styleN))
    # c  = Canvas('mydoc.pdf', pagesize = landscape)
    # f = Frame(5* cm, 3 * cm, 6 * cm, 25 * cm, showBoundary=1) # высота прямоугольника  6 Х 25, showBoundary = 1, рамка 0- нет
    # f.addFromList(story, c)
    # c.save()
# =======        
# def proba():
# #     """добавление столбца в существующую таблицу, затем его добавить в -models- соответсвующую таблицу этот столбец"""
#     data = my_win.lineEd.text()
    # my_db = SqliteDatabase('comp_db.db')
    # migrator = SqliteMigrator(my_db)
    # multiregion = IntegerField(null=True)
#     # system_id = IntegerField(null=False)  # новый столбец, его поле и значение по умолчанию
#     # system_id = ForeignKeyField(System, field=System.id, null=True)

    # with db:
#         # migrate(migrator.drop_column('referees', 'signature')) # удаление столбца
# #         # migrate(migrator.alter_column_type('system', 'mesta_exit', IntegerField()))
#         migrate(migrator.rename_column('titles', 'kat_sek', 'kat_sec')) # Переименование столбца (таблица, старое название, новое название столбца)
        # migrate(migrator.add_column('titles', 'multiregion', multiregion)) # Добавление столбца (таблица, столбец, повтор название столбца)


# ===== переводит фокус на поле ввода счета в партии вкладки -группа-
my_win.lineEdit_pl1_s1_gr.returnPressed.connect(focus)
my_win.lineEdit_pl2_s1_gr.returnPressed.connect(focus)
my_win.lineEdit_pl1_s2_gr.returnPressed.connect(focus)
my_win.lineEdit_pl2_s2_gr.returnPressed.connect(focus)
my_win.lineEdit_pl1_s3_gr.returnPressed.connect(focus)
my_win.lineEdit_pl2_s3_gr.returnPressed.connect(focus)
my_win.lineEdit_pl1_s4_gr.returnPressed.connect(focus)
my_win.lineEdit_pl2_s4_gr.returnPressed.connect(focus)
my_win.lineEdit_pl1_s5_gr.returnPressed.connect(focus)
my_win.lineEdit_pl2_s5_gr.returnPressed.connect(focus)
# ===== проверка правильность ввода цифр
my_win.lineEdit_pl1_s1_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl2_s1_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl1_s2_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl2_s2_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl1_s3_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl2_s3_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl2_s4_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl1_s5_gr.textChanged.connect(control_mark_in_score)
my_win.lineEdit_pl2_s5_gr.textChanged.connect(control_mark_in_score)
# ===== переводит фокус на поле ввода счета в партии вкладки -полуфиналы-
my_win.lineEdit_pl1_s1_pf.returnPressed.connect(focus)
my_win.lineEdit_pl2_s1_pf.returnPressed.connect(focus)
my_win.lineEdit_pl1_s2_pf.returnPressed.connect(focus)
my_win.lineEdit_pl2_s2_pf.returnPressed.connect(focus)
my_win.lineEdit_pl1_s3_pf.returnPressed.connect(focus)
my_win.lineEdit_pl2_s3_pf.returnPressed.connect(focus)
my_win.lineEdit_pl1_s4_pf.returnPressed.connect(focus)
my_win.lineEdit_pl2_s4_pf.returnPressed.connect(focus)
my_win.lineEdit_pl1_s5_pf.returnPressed.connect(focus)
my_win.lineEdit_pl2_s5_pf.returnPressed.connect(focus)
# ===== переводит фокус на полее ввода счета в партии вкладки -финалы-
my_win.lineEdit_pl1_s1_fin.returnPressed.connect(focus)
my_win.lineEdit_pl2_s1_fin.returnPressed.connect(focus)
my_win.lineEdit_pl1_s2_fin.returnPressed.connect(focus)
my_win.lineEdit_pl2_s2_fin.returnPressed.connect(focus)
my_win.lineEdit_pl1_s3_fin.returnPressed.connect(focus)
my_win.lineEdit_pl2_s3_fin.returnPressed.connect(focus)
my_win.lineEdit_pl1_s4_fin.returnPressed.connect(focus)
my_win.lineEdit_pl2_s4_fin.returnPressed.connect(focus)
my_win.lineEdit_pl1_s5_fin.returnPressed.connect(focus)
my_win.lineEdit_pl2_s5_fin.returnPressed.connect(focus)

my_win.lineEdit_range_tours.returnPressed.connect(enter_print_begunki)
my_win.lineEdit_num_game_fin.returnPressed.connect(filter_fin)

my_win.lineEdit_pl1_score_total_gr.returnPressed.connect(enter_total_score)
my_win.lineEdit_pl2_score_total_gr.returnPressed.connect(enter_total_score)
my_win.lineEdit_pl1_score_total_pf.returnPressed.connect(enter_total_score)
my_win.lineEdit_pl2_score_total_pf.returnPressed.connect(enter_total_score)
my_win.lineEdit_pl1_score_total_fin.returnPressed.connect(enter_total_score)
my_win.lineEdit_pl2_score_total_fin.returnPressed.connect(enter_total_score)

my_win.lineEdit_Family_name.returnPressed.connect(input_player)
my_win.lineEdit_bday.returnPressed.connect(next_field)
my_win.lineEdit_city_list.returnPressed.connect(add_city)
# ====== отслеживание изменения текста в полях ============
my_win.lineEdit_find_player_in_system.textChanged.connect(find_player_on_tab_system)
my_win.lineEdit_Family_name.textChanged.connect(find_in_rlist)  # в поле поиска и вызов функции
my_win.lineEdit_find_player_in_R.textChanged.connect(find_in_player_rejting_list)
my_win.lineEdit_coach.textChanged.connect(find_coach)
my_win.lineEdit_city_list.textChanged.connect(find_city)

my_win.comboBox_region.currentTextChanged.connect(find_city)
# comboBox_family_city = QComboBox()
# comboBox_family_city.currentTextChanged.connect(referee)
# ============= двойной клик
# двойной клик по listWidget (рейтинг, тренеры)
my_win.listWidget.itemDoubleClicked.connect(dclick_in_listwidget)
# двойной клик по строке игроков в таблице -результаты-, -списки-
my_win.tableView.doubleClicked.connect(select_player_in_game)

my_win.tabWidget.currentChanged.connect(tab)
my_win.toolBox.currentChanged.connect(tool_page)
# ==================================
my_win.spinBox_kol_group.textChanged.connect(kol_player_in_group)
# ======== изменение индекса комбобоксов ===========

fir_window.comboBox.currentTextChanged.connect(change_sroki)

my_win.comboBox_table_1.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_2.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_3.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_4.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_5.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_6.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_7.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_8.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_9.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_10.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_11.currentTextChanged.connect(kol_player_in_final)

my_win.comboBox_etap.currentTextChanged.connect(made_system_load_combobox_etap)

my_win.comboBox_page_vid.currentTextChanged.connect(page_vid)
my_win.comboBox_filter_number_group_final.currentTextChanged.connect(filter_player_on_system)
my_win.comboBox_filter_choice_stage.currentTextChanged.connect(choice_filter_on_system)
my_win.comboBox_fltr_region.currentTextChanged.connect(change_city_from_region)
my_win.comboBox_select_stage_begunki.currentTextChanged.connect(select_stage_for_begunki)
my_win.comboBox_select_group_begunki.currentTextChanged.connect(select_tour_for_begunki)
my_win.comboBox_select_tours.currentTextChanged.connect(select_diapazon)
my_win.comboBox_edit_etap1.currentTextChanged.connect(select_stage_for_edit)
my_win.comboBox_edit_etap2.currentTextChanged.connect(select_stage_for_edit)
my_win.comboBox_first_group.currentTextChanged.connect(add_item_listwidget)
my_win.comboBox_second_group.currentTextChanged.connect(add_item_listwidget)

# my_win.comboBox_filter_group.currentTextChanged.connect(filter_gr)
# my_win.comboBox_filter_played.currentTextChanged.connect(filter_gr)
# my_win.comboBox_find_name.currentTextChanged.connect(filter_gr)
my_win.comboBox_filter_final.currentTextChanged.connect(filter_fin)
my_win.comboBox_choice_R.currentTextChanged.connect(r_list_load_tableView)
# my_win.comboBox_filter_region_in_R.currentTextChanged.connect(filter_rejting_list)
# my_win.comboBox_filter_city_in_R.currentTextChanged.connect(filter_rejting_list)
# my_win.comboBox_filter_date_in_R.currentTextChanged.connect(filter_rejting_list)

my_win.comboBox_referee.currentTextChanged.connect(referee)
my_win.comboBox_secretary.currentTextChanged.connect(referee)
my_win.comboBox_kategor_ref.currentTextChanged.connect(add_referee_to_db)
my_win.comboBox_kategor_sec.currentTextChanged.connect(add_referee_to_db)



# =======  отслеживание переключение чекбоксов =========
# my_win.radioButton_group.toggled.connect(load_combobox_filter_group)
# my_win.radioButton_PF.toggled.connect(load_combobox_filter_group_semifinal)
# my_win.radioButton_Fin.toggled.connect(load_combobox_filter_group)

my_win.radioButton_match_3.toggled.connect(change_status_visible_and_score_game)
my_win.radioButton_match_5.toggled.connect(change_status_visible_and_score_game)
my_win.radioButton_match_7.toggled.connect(change_status_visible_and_score_game)
my_win.radioButton_match_4.toggled.connect(change_status_visible_and_score_game)
my_win.radioButton_match_6.toggled.connect(change_status_visible_and_score_game)
my_win.radioButton_match_8.toggled.connect(change_status_visible_and_score_game)

my_win.radioButton_repeat_regions.toggled.connect(change_choice_group) 


# при изменении чекбокса активирует кнопку создать
my_win.checkBox.stateChanged.connect(button_title_made_enable)
# при изменении чекбокса активирует кнопку создать
my_win.checkBox_3.stateChanged.connect(button_system_made_enable)
# при изменении чекбокса показывает поля для ввода счета
my_win.checkBox_4.stateChanged.connect(change_status_visible_and_score_game)
# при изменении чекбокса показывает поля для ввода счета финала)
my_win.checkBox_5.stateChanged.connect(change_status_visible_and_score_game)
# при изменении чекбокса показывает список удаленных игроков
my_win.checkBox_6.stateChanged.connect(del_player_table)
my_win.checkBox_7.stateChanged.connect(no_play)  # поражение по неявке игрок 1 группа
my_win.checkBox_8.stateChanged.connect(no_play)  # поражение по неявке игрок 2 группа
my_win.checkBox_9.stateChanged.connect(no_play)  # поражение по неявке игрок 1 пф
my_win.checkBox_10.stateChanged.connect(no_play)  # поражение по неявке игрок 2 пф
my_win.checkBox_12.stateChanged.connect(no_play)  # поражение по неявке игрок 1 финал
my_win.checkBox_13.stateChanged.connect(no_play)  # поражение по неявке игрок 2 финал
my_win.checkBox_11.stateChanged.connect(debtor_R) # должники рейтинга оплаты
my_win.checkBox_15.stateChanged.connect(filter_player_list)
my_win.checkBox_find_player.stateChanged.connect(find_player)
my_win.checkBox_GSK.stateChanged.connect(made_list_GSK)
# my_win.checkBox_edit_etap.stateChanged.connect(change_player_in_etap )
# =======  нажатие кнопок =========


my_win.Button_Ok_gr.setAutoDefault(True)  # click on <Enter>
my_win.Button_Ok_pf.setAutoDefault(True)  # click on <Enter>
my_win.Button_Ok_fin.setAutoDefault(True)  # click on <Enter>
my_win.Button_pay_R.clicked.connect(save_in_db_pay_R)
my_win.Button_clear_del.clicked.connect(clear_del_player)
my_win.Button_reset_filter_gr.clicked.connect(reset_filter)
my_win.Button_reset_filter_fin.clicked.connect(reset_filter)
my_win.Button_reset_filter_sf.clicked.connect(reset_filter)
my_win.Button_filter_fin.clicked.connect(filter_fin)
my_win.Button_filter_sf.clicked.connect(filter_sf)
my_win.Button_filter_gr.clicked.connect(filter_gr)
my_win.Button_app.clicked.connect(check_real_player) # отмечает что игрок по заявке
# рисует таблицы группового этапа и заполняет game_list
my_win.Button_etap_made.clicked.connect(etap_made)
my_win.Button_add_edit_player.clicked.connect(add_player)  # добавляет игроков в список и базу
# записывает в базу или редактирует титул
my_win.Button_title_made.clicked.connect(title_made)
# записывает в базу счет в партии встречи
my_win.Button_Ok_gr.clicked.connect(enter_score)
my_win.Button_Ok_pf.clicked.connect(enter_score)
my_win.Button_Ok_fin.clicked.connect(enter_score)
my_win.Button_del_player.clicked.connect(delete_player)
my_win.Button_print_begunki.clicked.connect(begunki_made)

# my_win.Button_proba.clicked.connect(proba) # запуск пробной функции

my_win.Button_add_pl1.clicked.connect(list_player_in_group_after_draw)
my_win.Button_add_pl2.clicked.connect(list_player_in_group_after_draw)
my_win.Button_change_player.clicked.connect(change_player_between_group_after_draw)

my_win.Button_sort_mesto.clicked.connect(sort)
my_win.Button_sort_R.clicked.connect(sort)
my_win.Button_sort_Name.clicked.connect(sort)
my_win.Button_fltr_list.clicked.connect(filter_player_list)
my_win.Button_reset_fltr_list.clicked.connect(filter_player_list)
my_win.Button_reset_fltr_in_R.clicked.connect(clear_filter_rejting_list)
my_win.Button_sort_alf_R.clicked.connect(filter_rejting_list)
my_win.Button_sort_rejting_in_R.clicked.connect(filter_rejting_list)
my_win.Button_filter_R.clicked.connect(filter_rejting_list)
my_win.Button_made_R_file.clicked.connect(made_file_excel_for_rejting)
my_win.Button_made_one_file_pdf.clicked.connect(merdge_pdf_files)

my_win.Button_up.clicked.connect(move_row_in_tablewidget)
my_win.Button_down.clicked.connect(move_row_in_tablewidget)
# my_win.tableWidget.cellClicked.connect(button_move_enabled)
my_win.Button_list_referee.clicked.connect(made_list_referee)
my_win.Button_list_regions.clicked.connect(made_list_regions)
my_win.Button_list_winner.clicked.connect(made_list_winners)
my_win.Button_players_on_alf.clicked.connect(made_list_players_on_alf)
my_win.Button_made_page_pdf.clicked.connect(made_pdf_list)
my_win.Button_view_page_pdf.clicked.connect(view_all_page_pdf)

my_win.Button_pay.clicked.connect(check_pay)
sys.exit(app.exec())